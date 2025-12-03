# ==============================================================================
# Ø¨Ø®Ø´ Û±: ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø§ÙˆÙ„ÛŒÙ‡ Ù…Ø­ÛŒØ· (Ø¨Ø³ÛŒØ§Ø± Ù…Ù‡Ù…)
# ==============================================================================
import sys
import asyncio

# Ø§ÛŒÙ† Ø¨Ø®Ø´ Ø¨Ø±Ø§ÛŒ Ø³Ø§Ø²Ú¯Ø§Ø±ÛŒ Ø¨Ø§ ÙˆÛŒÙ†Ø¯ÙˆØ² Ø¶Ø±ÙˆØ±ÛŒ Ø§Ø³Øª Ùˆ Ø¨Ù‡ Ø¯Ø±Ø³ØªÛŒ Ù¾ÛŒØ§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø´Ø¯Ù‡.
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

# ==============================================================================
# Ø¨Ø®Ø´ Û²: ÙˆØ§Ø±Ø¯ Ú©Ø±Ø¯Ù† Ú©ØªØ§Ø¨Ø®Ø§Ù†Ù‡â€ŒÙ‡Ø§ÛŒ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯ Ùˆ Ø´Ø®Øµ Ø«Ø§Ù„Ø« (Standard & Third-party)
# ==============================================================================
import os
import re
import json
import base64
import traceback
import logging
from io import BytesIO
from collections import defaultdict
from datetime import datetime, date, time, timedelta
from enum import Enum as PyEnum
from typing import List, Optional, Any, Dict, Tuple
from contextlib import contextmanager
import numpy as np

# Ú©ØªØ§Ø¨Ø®Ø§Ù†Ù‡â€ŒÙ‡Ø§ÛŒ ÙˆØ¨ ÙØ±ÛŒÙ…ÙˆØ±Ú© (FastAPI)
from fastapi import (
    FastAPI, WebSocket, WebSocketDisconnect, Depends, HTTPException, 
    Response, status, File, UploadFile, Request, BackgroundTasks
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.httpsredirect import HTTPSRedirectMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.staticfiles import StaticFiles
# âœ… Ø§ØµÙ„Ø§Ø­: ÙˆØ§Ø±Ø¯ Ú©Ø±Ø¯Ù† JSONResponse Ø¨Ø±Ø§ÛŒ Ù…Ø¯ÛŒØ±ÛŒØª Ø¨Ù‡ØªØ± Ø®Ø·Ø§Ù‡Ø§ÛŒ HTTP
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

# Ú©ØªØ§Ø¨Ø®Ø§Ù†Ù‡â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒØªØ§Ø¨ÛŒØ³ (SQLAlchemy & Pydantic)
from sqlalchemy import (
    Column, Integer, String, DateTime, ForeignKey, Enum as SQLEnum, 
    create_engine, UniqueConstraint, event, Text, Date as SQLDateType, JSON,
    func, DDL, Index, text, case
)
from sqlalchemy.orm import sessionmaker, relationship, Session, declarative_base, joinedload
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from pydantic import BaseModel, Field, conint, field_validator, ValidationInfo, computed_field, ConfigDict

# Ú©ØªØ§Ø¨Ø®Ø§Ù†Ù‡â€ŒÙ‡Ø§ÛŒ Ú¯Ø²Ø§Ø±Ø´â€ŒÚ¯ÛŒØ±ÛŒ Ùˆ Ø§Ø¨Ø²Ø§Ø±Ù‡Ø§ÛŒ Ø¬Ø§Ù†Ø¨ÛŒ
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import qrcode
from jinja2 import Environment, FileSystemLoader, Template
#from xhtml2pdf import pisa
import jdatetime
from dotenv import load_dotenv

# Ú©ØªØ§Ø¨Ø®Ø§Ù†Ù‡â€ŒÙ‡Ø§ÛŒ Ø§Ù…Ù†ÛŒØªÛŒ Ùˆ Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª
from passlib.context import CryptContext
from jose import JWTError, jwt

# ==============================================================================
# Ø¨Ø®Ø´ Û³: ÙˆØ§Ø±Ø¯ Ú©Ø±Ø¯Ù† Ù…Ø§Ú˜ÙˆÙ„â€ŒÙ‡Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ù¾Ø±ÙˆÚ˜Ù‡
# ==============================================================================
# from naab_connector_final import NaabConnector # Ø§ÛŒÙ† Ø¨Ø®Ø´ Ø¨Ù‡ Ø¯Ø±Ø³ØªÛŒ Ú©Ø§Ù…Ù†Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.

# ==============================================================================
# Ø¨Ø®Ø´ Û´: ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ùˆ Ù¾ÛŒÚ©Ø±Ø¨Ù†Ø¯ÛŒ Ø§ÙˆÙ„ÛŒÙ‡ Ø¨Ø±Ù†Ø§Ù…Ù‡
# ==============================================================================
load_dotenv() # Ø®ÙˆØ§Ù†Ø¯Ù† Ù…ØªØºÛŒØ±Ù‡Ø§ Ø§Ø² ÙØ§ÛŒÙ„ .env

# âœ… Ø§ØµÙ„Ø§Ø­: ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù„Ø§Ú¯â€ŒÚ¯ÛŒØ±ÛŒ Ø¨Ù‡ Ø§Ø¨ØªØ¯Ø§ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡ Ù…Ù†ØªÙ‚Ù„ Ø´Ø¯ ØªØ§ Ù‚Ø¨Ù„ Ø§Ø² Ù‡Ø± Ø§Ø³ØªÙØ§Ø¯Ù‡â€ŒØ§ÛŒ Ù¾ÛŒÚ©Ø±Ø¨Ù†Ø¯ÛŒ Ø´Ø¯Ù‡ Ø¨Ø§Ø´Ø¯.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)

logger = logging.getLogger(__name__)

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø§Ù…Ù†ÛŒØªÛŒ - Ø§Ø¬Ø¨Ø§Ø±ÛŒ Ú©Ø±Ø¯Ù† Ù…ØªØºÛŒØ±Ù‡Ø§ÛŒ Ù…Ø­ÛŒØ·ÛŒ
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    # Ø­Ø§Ù„Ø§ Ø§ÛŒÙ† Ù„Ø§Ú¯ Ø¨Ù‡ Ø¯Ø±Ø³ØªÛŒ Ú©Ø§Ø± Ù…ÛŒâ€ŒÚ©Ù†Ø¯ Ú†ÙˆÙ† Ù„Ø§Ú¯Ø± Ù¾ÛŒÚ©Ø±Ø¨Ù†Ø¯ÛŒ Ø´Ø¯Ù‡ Ø§Ø³Øª.
    logger.error("CRITICAL: SECRET_KEY environment variable is not set. Application cannot start securely.")
    raise ValueError("SECRET_KEY must be set in environment variables for security.")

# ğŸ”¥ Ø§ØµÙ„Ø§Ø­ Ù‚Ø·Ø¹ÛŒ: Ø§Ù„Ú¯ÙˆØ±ÛŒØªÙ… Ù‡Ø´ÛŒÙ†Ú¯ JWT Ø¨Ù‡ ØµØ±Ø§Ø­Øª ØªØ¹Ø±ÛŒÙ Ø´Ø¯ ØªØ§ Ø§Ø² Ø¢Ø³ÛŒØ¨â€ŒÙ¾Ø°ÛŒØ±ÛŒ "alg:none" Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø´ÙˆØ¯.
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 Ø³Ø§Ø¹Øª

# Ú†Ø§Ù¾ ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø¨Ø±Ø§ÛŒ Ø¯ÛŒØ¨Ø§Ú¯ (ÙÙ‚Ø· Ø¯Ø± Ø­Ø§Ù„Øª ØªÙˆØ³Ø¹Ù‡)
if os.getenv("ENVIRONMENT") != "production":
    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ù†Ù…Ø§ÛŒØ´ Ø¨Ø®Ø´ÛŒ Ø§Ø² Ú©Ù„ÛŒØ¯ Ø¨Ø±Ø§ÛŒ ØªØ£ÛŒÛŒØ¯ ØµØ­Øª Ø¢Ù† Ø¨Ø¯ÙˆÙ† Ø§ÙØ´Ø§ÛŒ Ú©Ø§Ù…Ù„.
    print(f"ğŸ” SECRET_KEY: {'*' * (len(SECRET_KEY) - 4) + SECRET_KEY[-4:] if SECRET_KEY and len(SECRET_KEY) > 4 else 'SET (TOO SHORT)'}")
    print(f"ğŸ” ALGORITHM: {ALGORITHM}")
    print(f"ğŸ” ACCESS_TOKEN_EXPIRE_MINUTES: {ACCESS_TOKEN_EXPIRE_MINUTES}")

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª CORS
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(",")
# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÙØ²ÙˆØ¯Ù† '*' Ø¨Ù‡ Ù‡Ø§Ø³Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø¬Ø§Ø² Ø¨Ø±Ø§ÛŒ Ø³Ù‡ÙˆÙ„Øª Ø¯Ø± ØªÙˆØ³Ø¹Ù‡ (Ø¯Ø± Ù…Ø­ÛŒØ· ØªÙˆÙ„ÛŒØ¯ Ø¨Ø§ÛŒØ¯ Ù…Ø­Ø¯ÙˆØ¯ Ø´ÙˆØ¯).
ALLOWED_HOSTS = os.getenv("ALLOWED_HOSTS", "localhost,127.0.0.1,*").split(",")

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª NAAB (Ø¨Ø±Ø§ÛŒ Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ)
NAAB_USERNAME = os.getenv("NAAB_USERNAME")
NAAB_PASSWORD = os.getenv("NAAB_PASSWORD")

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ø«Ø§Ø¨Øªâ€ŒÙ‡Ø§ Ø¨Ø±Ø§ÛŒ Ù†Ø§Ù… Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ Ø¬Ù‡Øª Ø®ÙˆØ§Ù†Ø§ÛŒÛŒ Ùˆ Ù†Ú¯Ù‡Ø¯Ø§Ø±ÛŒ Ø¨Ù‡ØªØ±.
WEST_COMPANY = "Ø´Ø±Ú©Øª Ù¾Ø§ÛŒØ¯Ø§Ø± Ù†ÛŒØ±Ùˆ Ù†ÛŒÚ©Ø§"
EAST_COMPANY = "Ø´Ø±Ú©Øª Ú©Ø§Ù…ÛŒØ§Ø±Ø§Ù† Ø§Ø±Ù…"

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# ØªÙˆØ§Ø¨Ø¹ Ú©Ù…Ú©ÛŒ
def normalize_text(text: str) -> str:
    """Ù†Ø±Ù…Ø§Ù„â€ŒØ³Ø§Ø²ÛŒ Ùˆ Ù¾Ø§Ú©â€ŒØ³Ø§Ø²ÛŒ Ù…ØªÙ† ÙØ§Ø±Ø³ÛŒ Ø¨Ø±Ø§ÛŒ Ù…Ù‚Ø§ÛŒØ³Ù‡ Ø¨Ù‡ØªØ±."""
    if not isinstance(text, str): 
        return ""
    text = text.replace('ÙŠ', 'ÛŒ').replace('Ùƒ', 'Ú©')
    # Ø­Ø°Ù Ø¨Ø®Ø´â€ŒÙ‡Ø§ÛŒÛŒ Ù…Ø§Ù†Ù†Ø¯ "(Ø¨Ø¯ÙˆÙ†...)"
    text = text.split('(Ø¨Ø¯ÙˆÙ†')[0]
    # Ø­Ø°Ù ÙØ§ØµÙ„Ù‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¶Ø§ÙÛŒ
    text = " ".join(text.split())
    return text

def resource_path(relative_path: str) -> str:
    """
    Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…Ø³ÛŒØ± Ù…Ø·Ù„Ù‚ Ù…Ù†Ø§Ø¨Ø¹ Ø¨Ø±Ø§ÛŒ Ø³Ø§Ø²Ú¯Ø§Ø±ÛŒ Ø¨Ø§ Ù…Ø­ÛŒØ·â€ŒÙ‡Ø§ÛŒ Ù…Ø®ØªÙ„Ù (Ø¹Ø§Ø¯ÛŒ Ùˆ PyInstaller).
    """
    try: 
        # Ø­Ø§Ù„Øª Ø§Ø¬Ø±Ø§ Ø¯Ø± PyInstaller
        base_path = sys._MEIPASS
    except AttributeError: 
        # Ø­Ø§Ù„Øª Ø§Ø¬Ø±Ø§ÛŒ Ø¹Ø§Ø¯ÛŒ
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

# ØªØ¹Ø±ÛŒÙ Ù…Ø³ÛŒØ±Ù‡Ø§ Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² ØªØ§Ø¨Ø¹ Ú©Ù…Ú©ÛŒ
STATIC_DIR_NAME = "static"
STATIC_DIR = resource_path(STATIC_DIR_NAME)
DB_NAME = "projects.db"
DATABASE_URL = f"sqlite:///{resource_path(DB_NAME)}"

# Ø§ÛŒØ¬Ø§Ø¯ Ø¯Ø§ÛŒØ±Ú©ØªÙˆØ±ÛŒ Ø§Ø³ØªØ§ØªÛŒÚ© Ø§Ú¯Ø± ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯
if not os.path.exists(STATIC_DIR):
    try: 
        os.makedirs(STATIC_DIR, exist_ok=True)
        logger.info(f"Static directory created at: {STATIC_DIR}")
    except OSError as e: 
        logger.error(f"Could not create static directory {STATIC_DIR}: {e}")

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø¯ÛŒØªØ§Ø¨ÛŒØ³
engine = create_engine(
    DATABASE_URL, 
    # Ø§ÛŒÙ† Ø¢Ø±Ú¯ÙˆÙ…Ø§Ù† Ø¨Ø±Ø§ÛŒ SQLite Ø¯Ø± Ù…Ø­ÛŒØ· Ú†Ù†Ø¯Ù†Ø®ÛŒ Ø¶Ø±ÙˆØ±ÛŒ Ø§Ø³Øª.
    connect_args={"check_same_thread": False},
    # Ø¨Ø±Ø§ÛŒ Ø¨Ø±Ø±Ø³ÛŒ Ø³Ù„Ø§Ù…Øª Ú©Ø§Ù†Ú©Ø´Ù†â€ŒÙ‡Ø§ Ù‚Ø¨Ù„ Ø§Ø² Ø§Ø³ØªÙØ§Ø¯Ù‡ Ù…Ø¬Ø¯Ø¯
    pool_pre_ping=True,
    # ØºÛŒØ±ÙØ¹Ø§Ù„ Ú©Ø±Ø¯Ù† Ù„Ø§Ú¯ Ú©ÙˆØ¦Ø±ÛŒâ€ŒÙ‡Ø§ Ø¯Ø± Ù…Ø­ÛŒØ· ØªÙˆÙ„ÛŒØ¯ Ø¨Ø±Ø§ÛŒ Ø§ÙØ²Ø§ÛŒØ´ Ø³Ø±Ø¹Øª
    echo=(os.getenv("ENVIRONMENT") != "production")
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª Jinja2 Ø¨Ø±Ø§ÛŒ Ø±Ù†Ø¯Ø± Ù‚Ø§Ù„Ø¨â€ŒÙ‡Ø§ÛŒ HTML
jinja_env = Environment(loader=FileSystemLoader(STATIC_DIR))
templates = Jinja2Templates(directory=STATIC_DIR)

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Context manager Ø¨Ø±Ø§ÛŒ Ù…Ø¯ÛŒØ±ÛŒØª session Ø¯ÛŒØªØ§Ø¨ÛŒØ³ Ø¨Ù‡ÛŒÙ†Ù‡ Ø´Ø¯ ØªØ§ Ù„Ø§Ú¯ Ø¨Ù‡ØªØ±ÛŒ Ø¯Ø§Ø´ØªÙ‡ Ø¨Ø§Ø´Ø¯.
@contextmanager
def get_db_session():
    """Ù…Ø¯ÛŒØ±ÛŒØª session Ø¯ÛŒØªØ§Ø¨ÛŒØ³ Ø¨Ø§ context manager Ø¨Ø±Ø§ÛŒ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø¯Ø± ØªØ³Ú©â€ŒÙ‡Ø§ÛŒ Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡."""
    db = SessionLocal()
    try:
        yield db
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Database session error occurred: {e}", exc_info=True)
        raise
    finally:
        db.close()

def get_db():
    """Dependency Ø¨Ø±Ø§ÛŒ FastAPI Ø¬Ù‡Øª ØªØ²Ø±ÛŒÙ‚ session Ø¯ÛŒØªØ§Ø¨ÛŒØ³ Ø¨Ù‡ Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øªâ€ŒÙ‡Ø§."""
    db = SessionLocal()
    try: 
        yield db
    finally: 
        db.close()
# ==============================================================================
# Ø¨Ø®Ø´ Ûµ: EnumÙ‡Ø§ Ùˆ ØªÙ†Ø¸ÛŒÙ…Ø§Øª
# ==============================================================================

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² EnumÙ‡Ø§ Ø¨Ø±Ø§ÛŒ Ù…Ù‚Ø§Ø¯ÛŒØ± Ø«Ø§Ø¨ØªØŒ Ú©Ø¯ Ø±Ø§ Ø®ÙˆØ§Ù†Ø§ØªØ± Ùˆ Ø®Ø·Ø§Ù¾Ø°ÛŒØ±ØªØ± Ù…ÛŒâ€ŒÚ©Ù†Ø¯.
class StepNameKey(str, PyEnum):
    START_ASSEMBLY = "START_ASSEMBLY"
    END_ASSEMBLY = "END_ASSEMBLY"
    TEAM_LEAD_APPROVAL = "TEAM_LEAD_APPROVAL"
    TEST = "TEST"
    QUALITY_CONTROL = "QUALITY_CONTROL"
    SUPERVISOR_APPROVAL = "SUPERVISOR_APPROVAL"
    EXIT_PANEL = "EXIT_PANEL"
    
# ØªØ±ØªÛŒØ¨ Ù…Ø±Ø§Ø­Ù„ Ø¨Ø±Ø§ÛŒ Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ
ORDERED_MANUAL_STEP_KEYS = [
    StepNameKey.START_ASSEMBLY, StepNameKey.END_ASSEMBLY, StepNameKey.TEAM_LEAD_APPROVAL,
    StepNameKey.TEST, StepNameKey.QUALITY_CONTROL, StepNameKey.SUPERVISOR_APPROVAL
]

# Ù†Ú¯Ø§Ø´Øª Ú©Ù„ÛŒØ¯Ù‡Ø§ Ø¨Ù‡ Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ ÙØ§Ø±Ø³ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø± UI
STEP_KEY_TO_NAME_MAP = {
    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø³Ø§Ø®Øª Ø¯ÛŒÚ©Ø´Ù†Ø±ÛŒ Ø¨Ù‡ Ø±ÙˆØ´ÛŒ Ø®ÙˆØ§Ù†Ø§ØªØ±
    key.value: name for key, name in [
        (StepNameKey.START_ASSEMBLY, "Ø´Ø±ÙˆØ¹ Ù…ÙˆÙ†ØªØ§Ú˜"),
        (StepNameKey.END_ASSEMBLY, "Ù¾Ø§ÛŒØ§Ù† Ù…ÙˆÙ†ØªØ§Ú˜"),
        (StepNameKey.TEAM_LEAD_APPROVAL, "ØªØ§ÛŒÛŒØ¯ Ø³Ø±Ú¯Ø±ÙˆÙ‡"),
        (StepNameKey.TEST, "ØªØ³Øª Ø³Ù…Ø§Ú©"),
        (StepNameKey.QUALITY_CONTROL, "Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª"),
        (StepNameKey.SUPERVISOR_APPROVAL, "ØªØ£ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø±"),
        (StepNameKey.EXIT_PANEL, "Ø®Ø±ÙˆØ¬ ØªØ§Ø¨Ù„Ùˆ"),
    ]
}

class PanelTypeKey(str, PyEnum):
    FAHAM_WITH_FRAME = "FAHAM_WITH_FRAME"
    FAHAM_WITHOUT_FRAME = "FAHAM_WITHOUT_FRAME"
    ID2R = "ID2R"
    ID5R = "ID5R"
    ID116 = "ID116"
    ID6_1R = "ID6_1R"
    ID12_1R = "ID12_1R"
    ID18_1R = "ID18_1R"
    ID24_1R = "ID24_1R"
    ID101_1 = "ID101_1"
    ID101_3 = "ID101_3"
    ID102_1 = "ID102_1"
    ID102_3 = "ID102_3"
    ID104_1 = "ID104_1"
    ID104_3 = "ID104_3"
    ID105 = "ID105"
    ID107 = "ID107"
    ID115 = "ID115"
    ID108 = "ID108"
    ID109 = "ID109"
    ID110 = "ID110"
    ID111 = "ID111"
    ID112_STAR = "ID112_STAR"
    ID120 = "ID120"
    ID121 = "ID121"
    ID122 = "ID122"
    ID123 = "ID123"
    ID124_STAR = "ID124_STAR"
    ID211 = "ID211"
    ID212 = "ID212"
    ID213 = "ID213"
    ID214 = "ID214"
    ID215 = "ID215"
    ID216 = "ID216"
    ID218 = "ID218"

PANEL_TYPE_NAMES = {
    PanelTypeKey.FAHAM_WITH_FRAME: "ÙÙ‡Ø§Ù… Ø¨Ø§ Ù‚Ø§Ø¨",
    PanelTypeKey.FAHAM_WITHOUT_FRAME: "ÙÙ‡Ø§Ù… Ø¨Ø¯ÙˆÙ† Ù‚Ø§Ø¨",
    PanelTypeKey.ID2R: "ID2R - ØªØ§Ø¨Ù„Ùˆ Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ 1-2 Ú©Ù†ØªÙˆØ± ØªÚ©ÙØ§Ø² Ø±ÛŒÙ„ÛŒ",
    PanelTypeKey.ID5R: "ID5R - ØªØ§Ø¨Ù„Ùˆ Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ 3-5 Ú©Ù†ØªÙˆØ± ØªÚ©ÙØ§Ø² Ø±ÛŒÙ„ÛŒ",
    PanelTypeKey.ID116: "ID116 - ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ ØªÚ©ÙØ§Ø² Ø±ÛŒÙ„ÛŒ Ø±ÙˆÛŒ Ø¯ÛŒÙˆØ§Ø±",
    PanelTypeKey.ID6_1R: "ID6+1R - ØªØ§Ø¨Ù„Ùˆ 6 Ú©Ù†ØªÙˆØ± ÙÙ„Ø²ÛŒ Ø¯ÛŒÙˆØ§Ø±ÛŒ",
    PanelTypeKey.ID12_1R: "ID12+1R",
    PanelTypeKey.ID18_1R: "ID18+1R - ØªØ§Ø¨Ù„Ùˆ 18 Ú©Ù†ØªÙˆØ± ÙÙ„Ø²ÛŒ Ø¯ÛŒÙˆØ§Ø±ÛŒ",
    PanelTypeKey.ID24_1R: "ID24+1R - ØªØ§Ø¨Ù„Ùˆ 24 Ú©Ù†ØªÙˆØ± ÙÙ„Ø²ÛŒ Ø¯ÛŒÙˆØ§Ø±ÛŒ",
    PanelTypeKey.ID101_1: "ID101.1 - ØªÚ© Ú©Ù†ØªÙˆØ± Ù‡ÙˆØ§ÛŒÛŒ ØªÚ©ÙØ§Ø² (ÙÛŒÙˆØ² Ø¯Ø± Ù…Ø­Ù„)",
    PanelTypeKey.ID101_3: "ID101.3 - ØªÚ© Ú©Ù†ØªÙˆØ± Ù‡ÙˆØ§ÛŒÛŒ Ø³Ù‡ ÙØ§Ø² (ÙÛŒÙˆØ² Ø¯Ø± Ù…Ø­Ù„)",
    PanelTypeKey.ID102_1: "ID102.1 - ØªÚ© Ú©Ù†ØªÙˆØ± Ù‡ÙˆØ§ÛŒÛŒ ØªÚ©ÙØ§Ø² (ÙÛŒÙˆØ²Ø¯Ø§Ø±)",
    PanelTypeKey.ID102_3: "ID102.3 - ØªÚ© Ú©Ù†ØªÙˆØ± Ù‡ÙˆØ§ÛŒÛŒ Ø³Ù‡ ÙØ§Ø² (ÙÛŒÙˆØ²Ø¯Ø§Ø±)",
    PanelTypeKey.ID104_1: "ID104.1 - ØªÚ© Ú©Ù†ØªÙˆØ± Ø²Ù…ÛŒÙ†ÛŒ ØªÚ©ÙØ§Ø² (ÙÛŒÙˆØ² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡)",
    PanelTypeKey.ID104_3: "ID104.3 - ØªÚ© Ú©Ù†ØªÙˆØ± Ø²Ù…ÛŒÙ†ÛŒ Ø³Ù‡ ÙØ§Ø² (ÙÛŒÙˆØ² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡)",
    PanelTypeKey.ID105: "ID105 - ØªØ§Ø¨Ù„Ùˆ Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ Ø²Ù…ÛŒÙ†ÛŒ ØªÚ© Ú©Ù†ØªÙˆØ±Ù‡ ÛŒÚ©Ø·Ø±ÙÙ‡",
    PanelTypeKey.ID107: "ID107 - ØªØ§Ø¨Ù„Ùˆ Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ Ø¯ÛŒÙˆØ§Ø±ÛŒ ØªÚ© Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² Ø¨Ø§ ÙÛŒÙˆØ²",
    PanelTypeKey.ID115: "ID115 - ØªØ§Ø¨Ù„Ùˆ ØªÚ© Ú©Ù†ØªÙˆØ± Ø¯ÛŒÙˆØ§Ø±ÛŒ - ÙÛŒÙˆØ² Ø¯Ø± Ù…Ø­Ù„",
    PanelTypeKey.ID108: "ID108 - ØªØ§Ø¨Ù„Ùˆ Ø²Ù…ÛŒÙ†ÛŒ Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡ ÛŒÚ©Ø·Ø±ÙÙ‡",
    PanelTypeKey.ID109: "ID109 - ØªØ§Ø¨Ù„Ùˆ Ø²Ù…ÛŒÙ†ÛŒ Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡ Ø¯ÙˆØ·Ø±ÙÙ‡",
    PanelTypeKey.ID110: "ID110 - ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ ØªÚ©ÙØ§Ø² (Ù‡ÙˆØ§ÛŒÛŒ) - ÙÛŒÙˆØ² Ø¯Ø± Ù…Ø­Ù„",
    PanelTypeKey.ID111: "ID111 - ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ ØªÚ©ÙØ§Ø² (Ø²Ù…ÛŒÙ†ÛŒ) - ÙÛŒÙˆØ² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡",
    PanelTypeKey.ID112_STAR: "ID112* - ØªØ§Ø¨Ù„Ùˆ Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡ ØªÚ© ÙØ§Ø² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡ (Ú©Ù„ÛŒ)",
    PanelTypeKey.ID120: "ID120 - ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² (Ù‡ÙˆØ§ÛŒÛŒ) - ÙÛŒÙˆØ² Ø¯Ø± Ù…Ø­Ù„",
    PanelTypeKey.ID121: "ID121 - ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² (Ø²Ù…ÛŒÙ†ÛŒ) - ÙÛŒÙˆØ² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡",
    PanelTypeKey.ID122: "ID122 - 2x ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² - Ø¬Ø¹Ø¨Ù‡ 8 ÙÛŒÙˆØ²",
    PanelTypeKey.ID123: "ID123 - 2x ØªØ§Ø¨Ù„Ùˆ 2 Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² - Ø¬Ø¹Ø¨Ù‡ 16 ÙÛŒÙˆØ²",
    PanelTypeKey.ID124_STAR: "ID124* - ØªØ§Ø¨Ù„Ùˆ Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡ Ø³Ù‡ ÙØ§Ø² Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡ (Ú©Ù„ÛŒ)",
    PanelTypeKey.ID211: "ID211 - ØªØ§Ø¨Ù„Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ù‡ÙˆØ§ÛŒÛŒ 30-150 kW",
    PanelTypeKey.ID212: "ID212 - ØªØ§Ø¨Ù„Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ù‡ÙˆØ§ÛŒÛŒ 151-249 kW",
    PanelTypeKey.ID213: "ID213 - ØªØ§Ø¨Ù„Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ø²Ù…ÛŒÙ†ÛŒ ÛŒÚ©Ø·Ø±ÙÙ‡",
    PanelTypeKey.ID214: "ID214 - ØªØ§Ø¨Ù„Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ø²Ù…ÛŒÙ†ÛŒ Ø¯ÙˆØ·Ø±ÙÙ‡",
    PanelTypeKey.ID215: "ID215 - ØªØ§Ø¨Ù„Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ ÙÙ„Ø²ÛŒ Ø²Ù…ÛŒÙ†ÛŒ",
    PanelTypeKey.ID216: "ID216 - ØªØ§Ø¨Ù„Ùˆ Ø¯Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ù‡ÙˆØ§ÛŒÛŒ",
    PanelTypeKey.ID218: "ID218 - ØªØ§Ø¨Ù„Ùˆ Ú†Ù†Ø¯ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ Ø²Ù…ÛŒÙ†ÛŒ Ø¯ÙˆØ·Ø±ÙÙ‡",
}

class TransactionType(str, PyEnum):
    IN = "IN"    # ÙˆØ±ÙˆØ¯ Ø¨Ù‡ Ø§Ù†Ø¨Ø§Ø±
    OUT = "OUT"  # Ø®Ø±ÙˆØ¬ Ø§Ø² Ø§Ù†Ø¨Ø§Ø±

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÛŒØ¬Ø§Ø¯ Enum Ø¨Ø±Ø§ÛŒ Ù†Ù‚Ø´â€ŒÙ‡Ø§ÛŒ Ú©Ø§Ø±Ø¨Ø±Ø§Ù† Ø¬Ù‡Øª Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ø®Ø·Ø§ÛŒ ØªØ§ÛŒÙ¾ÛŒ Ùˆ Ø§ÙØ²Ø§ÛŒØ´ Ø®ÙˆØ§Ù†Ø§ÛŒÛŒ.
class PersonnelRole(str, PyEnum):
    EMPLOYEE = "employee"
    SUPERVISOR = "supervisor"
    
# ==============================================================================
# Ø¨Ø®Ø´ Û¶: Ù…Ø¯Ù„â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒØªØ§Ø¨ÛŒØ³
# ==============================================================================
# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÙØ²ÙˆØ¯Ù† Ú©Ø§Ù…Ù†Øªâ€ŒÙ‡Ø§ÛŒ ØªÙˆØ¶ÛŒØ­ÛŒ Ø¨Ù‡ Ù…Ø¯Ù„â€ŒÙ‡Ø§ Ø¨Ø±Ø§ÛŒ Ø¯Ø±Ú© Ø¨Ù‡ØªØ± Ø³Ø§Ø®ØªØ§Ø± Ø¯ÛŒØªØ§Ø¨ÛŒØ³.

class Comment(Base):
    """Ù…Ø¯Ù„ Ù†Ø¸Ø±Ø§Øª Ø«Ø¨Øªâ€ŒØ´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ù‡Ø± Ù¾Ø±ÙˆÚ˜Ù‡."""
    __tablename__ = "comments"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    text = Column(Text, nullable=False)
    author = Column(String, nullable=False, default="Ø§Ù¾Ø±Ø§ØªÙˆØ±")
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    
    project = relationship("Project", back_populates="comments")

class Project(Base):
    """Ù…Ø¯Ù„ Ø§ØµÙ„ÛŒ Ù¾Ø±ÙˆÚ˜Ù‡."""
    __tablename__ = "projects"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True, nullable=False)
    location = Column(String, nullable=True)
    request_id = Column(String, index=True, nullable=False, unique=True)
    customer_name = Column(String, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    barcode_payload = Column(String, nullable=True, unique=True, index=True)
    panel_type_key = Column(String, nullable=True)
    panel_code = Column(String, index=True, nullable=True)
    assembler_1 = Column(String, nullable=True)
    assembler_2 = Column(String, nullable=True)
    
    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² lazy="selectin" Ø¨Ø±Ø§ÛŒ Ù¾ÛŒØ´â€ŒØ¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø±ÙˆØ§Ø¨Ø· Ùˆ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² N+1 queries.
    steps = relationship("Step", back_populates="project", cascade="all, delete-orphan", lazy="selectin", order_by="Step.timestamp")
    equipment = relationship("EquipmentItem", back_populates="project", cascade="all, delete-orphan", lazy="selectin")
    comments = relationship("Comment", back_populates="project", cascade="all, delete-orphan", lazy="selectin", order_by="Comment.timestamp.desc()")
    
    __table_args__ = (
        UniqueConstraint('request_id', name='uq_project_request_id'),
        Index('idx_project_created_at', 'created_at'),
        Index('idx_project_panel_code', 'panel_code'),
    )

class Step(Base):
    """Ù…Ø¯Ù„ Ù…Ø±Ø§Ø­Ù„ Ø§Ù†Ø¬Ø§Ù…â€ŒØ´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ù‡Ø± Ù¾Ø±ÙˆÚ˜Ù‡."""
    __tablename__ = "steps"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    name_key = Column(SQLEnum(StepNameKey, name="step_name_key_enum"), nullable=False)
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    
    project = relationship("Project", back_populates="steps")
    
    __table_args__ = (
        UniqueConstraint('project_id', 'name_key', name='uq_project_step'),
        Index('idx_step_timestamp', 'timestamp'),
        Index('idx_step_name_key', 'name_key'),
    )

class EquipmentItem(Base):
    """Ù…Ø¯Ù„ ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù…ÙˆØ±Ø¯ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø¯Ø± Ù‡Ø± Ù¾Ø±ÙˆÚ˜Ù‡."""
    __tablename__ = "equipment_items"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    item_name = Column(Text, nullable=False)
    quantity = Column(Integer, nullable=False, default=1)
    
    project = relationship("Project", back_populates="equipment")
    
    __table_args__ = (
        Index('idx_equipment_project_id', 'project_id'),
    )

class Personnel(Base):
    """Ù…Ø¯Ù„ Ù¾Ø±Ø³Ù†Ù„ Ùˆ Ú©Ø§Ø±Ø¨Ø±Ø§Ù† Ø³ÛŒØ³ØªÙ…."""
    __tablename__ = "personnel"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False, unique=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    password_hash = Column(String, nullable=False)
    # âœ… Ø§ØµÙ„Ø§Ø­: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…Ù‚Ø¯Ø§Ø± Enum Ø¨Ø±Ø§ÛŒ Ø³ØªÙˆÙ† Ù†Ù‚Ø´
    role = Column(String, default=PersonnelRole.EMPLOYEE.value, nullable=False)
    is_active = Column(Integer, default=1, nullable=False)
    
    daily_reports = relationship("DailyWorkReport", back_populates="personnel")
    
    __table_args__ = (
        Index('idx_personnel_username', 'username'),
        Index('idx_personnel_role', 'role'),
    )

class DailyWorkReport(Base):
    """Ù…Ø¯Ù„ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ Ù¾Ø±Ø³Ù†Ù„."""
    __tablename__ = "daily_work_reports"
    id = Column(Integer, primary_key=True, index=True)
    personnel_id = Column(Integer, ForeignKey("personnel.id"), nullable=False)
    report_date = Column(SQLDateType, nullable=False)
    status = Column(String, default='submitted', nullable=False)  # 'submitted', 'approved', 'rejected'
    supervisor_notes = Column(Text, nullable=True)
    report_data = Column(JSON, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    personnel = relationship("Personnel", back_populates="daily_reports")
    
    __table_args__ = (
        UniqueConstraint('personnel_id', 'report_date', name='uq_personnel_date_report'),
        Index('idx_report_date', 'report_date'),
        Index('idx_report_status', 'status'),
    )

class Warehouse(Base):
    """Ù…Ø¯Ù„ Ø§Ù†Ø¨Ø§Ø±Ù‡Ø§."""
    __tablename__ = "warehouses"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, nullable=False, index=True)
    location = Column(String, nullable=True)
    description = Column(Text, nullable=True)

class WarehouseItem(Base):
    """Ù…Ø¯Ù„ Ú©Ø§Ù„Ø§Ù‡Ø§ÛŒ Ù‚Ø§Ø¨Ù„ Ø°Ø®ÛŒØ±Ù‡ Ø¯Ø± Ø§Ù†Ø¨Ø§Ø±."""
    __tablename__ = "warehouse_items"
    id = Column(Integer, primary_key=True, index=True)
    item_name = Column(String, unique=True, nullable=False, index=True)
    description = Column(Text, nullable=True)
    min_stock_level = Column(Integer, default=0, nullable=False)

class InventoryTransaction(Base):
    """Ù…Ø¯Ù„ ØªØ±Ø§Ú©Ù†Ø´â€ŒÙ‡Ø§ÛŒ Ø§Ù†Ø¨Ø§Ø± (ÙˆØ±ÙˆØ¯ Ùˆ Ø®Ø±ÙˆØ¬)."""
    __tablename__ = "inventory_transactions"
    id = Column(Integer, primary_key=True, index=True)
    warehouse_id = Column(Integer, ForeignKey("warehouses.id"), nullable=False)
    item_id = Column(Integer, ForeignKey("warehouse_items.id"), nullable=False)
    quantity = Column(Integer, nullable=False)
    transaction_type = Column(SQLEnum(TransactionType, name="transaction_type_enum"), nullable=False)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="SET NULL"), nullable=True)
    user_id = Column(Integer, ForeignKey("personnel.id"), nullable=False)
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    notes = Column(Text, nullable=True)

    warehouse = relationship("Warehouse")
    item = relationship("WarehouseItem")
    project = relationship("Project")
    user = relationship("Personnel")

class PanelCodeItems(Base):
    """Ù…Ø¯Ù„ ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ø¨Ø±Ø§ÛŒ Ù‡Ø± Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ (Bill of Materials - BOM)."""
    __tablename__ = "panel_code_items"
    id = Column(Integer, primary_key=True, index=True)
    panel_code = Column(String, nullable=False, index=True)
    item_name = Column(String, nullable=False)
    quantity_required = Column(Integer, nullable=False)
    
    __table_args__ = (
        UniqueConstraint('panel_code', 'item_name', name='uq_panel_item'),
    )

# Ø§ÛŒØ¬Ø§Ø¯ Ø¬Ø¯Ø§ÙˆÙ„ Ùˆ Ø§ÛŒÙ†Ø¯Ú©Ø³â€ŒÙ‡Ø§
def create_tables_and_indexes():
    """Ø§ÛŒØ¬Ø§Ø¯ ØªÙ…Ø§Ù… Ø¬Ø¯Ø§ÙˆÙ„ Ùˆ Ø§ÛŒÙ†Ø¯Ú©Ø³â€ŒÙ‡Ø§ÛŒ ØªØ¹Ø±ÛŒÙâ€ŒØ´Ø¯Ù‡ Ø¯Ø± Ø¯ÛŒØªØ§Ø¨ÛŒØ³."""
    try:
        Base.metadata.create_all(bind=engine)
        logger.info("Database tables and indexes created/verified successfully.")
    except Exception as e:
        logger.error(f"FATAL: Error creating database tables: {e}")
        raise

create_tables_and_indexes()

def create_default_supervisor():
    """Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ø±Ø¨Ø± supervisor Ù¾ÛŒØ´â€ŒÙØ±Ø¶ Ø¨Ø±Ø§ÛŒ Ø§ÙˆÙ„ÛŒÙ† Ø§Ø¬Ø±Ø§ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡"""
    DEFAULT_ADMIN_USER = os.getenv("DEFAULT_ADMIN_USER", "admin")
    DEFAULT_ADMIN_PASSWORD = os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123")
    
    try:
        with SessionLocal() as db:
            existing_supervisor = db.query(Personnel).filter(
                Personnel.role == PersonnelRole.SUPERVISOR.value,
                Personnel.is_active == 1
            ).first()
            
            if not existing_supervisor:
                supervisor = Personnel(
                    name="Ù…Ø¯ÛŒØ± Ø³ÛŒØ³ØªÙ…",
                    username=DEFAULT_ADMIN_USER,
                    password_hash=pwd_context.hash(DEFAULT_ADMIN_PASSWORD),
                    role=PersonnelRole.SUPERVISOR.value,
                    is_active=1
                )
                db.add(supervisor)
                db.commit()
                logger.info(f"Default supervisor user '{DEFAULT_ADMIN_USER}' created.")
                print("=" * 60)
                print("âœ… Ú©Ø§Ø±Ø¨Ø± supervisor Ù¾ÛŒØ´â€ŒÙØ±Ø¶ Ø§ÛŒØ¬Ø§Ø¯ Ø´Ø¯.")
                print(f"   Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ: {DEFAULT_ADMIN_USER}")
                print(f"   Ø±Ù…Ø² Ø¹Ø¨ÙˆØ±: {DEFAULT_ADMIN_PASSWORD}")
                print("=" * 60)
            else:
                logger.info(f"Default supervisor user already exists: {existing_supervisor.username}")
                
    except Exception as e:
        logger.error(f"Error creating default supervisor: {e}")
        print(f"âš ï¸ Ø®Ø·Ø§ Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ø±Ø¨Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶: {e}")

def create_default_warehouse():
    """Ø§ÛŒØ¬Ø§Ø¯ Ø§Ù†Ø¨Ø§Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶ Ø¨Ø±Ø§ÛŒ Ø§ÙˆÙ„ÛŒÙ† Ø§Ø¬Ø±Ø§ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡"""
    try:
        with SessionLocal() as db:
            existing_warehouse = db.query(Warehouse).first()
            
            if not existing_warehouse:
                warehouse = Warehouse(
                    name="Ø§Ù†Ø¨Ø§Ø± Ø§ØµÙ„ÛŒ",
                    location="Ø³Ø§Ø®ØªÙ…Ø§Ù† Ù…Ø±Ú©Ø²ÛŒ",
                    description="Ø§Ù†Ø¨Ø§Ø± Ø§ØµÙ„ÛŒ Ø´Ø±Ú©Øª"
                )
                db.add(warehouse)
                db.commit()
                logger.info("Default warehouse created.")
                print("=" * 50)
                print("âœ… Ø§Ù†Ø¨Ø§Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶ Ø§ÛŒØ¬Ø§Ø¯ Ø´Ø¯.")
                print("   Ù†Ø§Ù…: Ø§Ù†Ø¨Ø§Ø± Ø§ØµÙ„ÛŒ")
                print("   Ù…Ú©Ø§Ù†: Ø³Ø§Ø®ØªÙ…Ø§Ù† Ù…Ø±Ú©Ø²ÛŒ")
                print("=" * 50)
            else:
                logger.info(f"Default warehouse already exists: {existing_warehouse.name}")
                
    except Exception as e:
        logger.error(f"Error creating default warehouse: {e}")
        print(f"âš ï¸ Ø®Ø·Ø§ Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ø§Ù†Ø¨Ø§Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶: {e}")

# Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ø±Ø¨Ø± Ùˆ Ø§Ù†Ø¨Ø§Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶
create_default_supervisor()
create_default_warehouse()


# ==============================================================================
# Ø¨Ø®Ø´ Û·: Pydantic Schemas
# ==============================================================================
# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² ConfigDict Ø¨Ù‡ Ø¬Ø§ÛŒ Config class Ú©Ù‡ Ø¯Ø± Pydantic v2 Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯ Ø§Ø³Øª.
# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÙØ²ÙˆØ¯Ù† ØªÙˆØ¶ÛŒØ­Ø§Øª (description) Ø¨Ù‡ ÙÛŒÙ„Ø¯Ù‡Ø§ Ø¨Ø±Ø§ÛŒ Ù…Ø³ØªÙ†Ø¯Ø³Ø§Ø²ÛŒ Ø¨Ù‡ØªØ± API Ø¯Ø± Swagger/OpenAPI.

class ValidationResponse(BaseModel):
    has_discrepancy: bool
    message: str

class CommentCreate(BaseModel):
    text: str = Field(..., min_length=1, description="Ù…ØªÙ† Ù†Ø¸Ø±")

class CommentOut(BaseModel):
    id: int
    text: str
    author: str
    timestamp: datetime
    
    model_config = ConfigDict(from_attributes=True)

class StepOut(BaseModel):
    name_key: StepNameKey
    timestamp: datetime

    @computed_field
    @property
    def name(self) -> str:
        return STEP_KEY_TO_NAME_MAP.get(self.name_key.value, self.name_key.value)
    
    model_config = ConfigDict(from_attributes=True)

class EquipmentItemBase(BaseModel):
    item_name: str
    quantity: int

class EquipmentItemCreate(EquipmentItemBase): 
    pass

class EquipmentItemOut(EquipmentItemBase):
    id: int
    project_id: int
    
    model_config = ConfigDict(from_attributes=True)

class ProjectOut(BaseModel):
    id: int
    name: str
    location: Optional[str] = None
    request_id: str
    customer_name: str
    created_at: datetime
    barcode_payload: Optional[str] = None
    steps: List[StepOut] = []
    equipment: List[EquipmentItemOut] = []
    panel_type_key: Optional[str] = None
    panel_type_name: Optional[str] = None
    panel_code: Optional[str] = None
    assembler_1: Optional[str] = None
    assembler_2: Optional[str] = None
    comments: List[CommentOut] = []
    
    model_config = ConfigDict(from_attributes=True)
    
    @field_validator('panel_type_name', mode='before')
    @classmethod
    def set_panel_type_name(cls, v: Any, info: ValidationInfo) -> Optional[str]:
        key = info.data.get('panel_type_key')
        if key:
            try:
                # âœ… Ø§ØµÙ„Ø§Ø­: Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø¨Ø§ Enum Ø¨Ø±Ø§ÛŒ Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² ØµØ­Øª Ú©Ù„ÛŒØ¯
                enum_key = PanelTypeKey(key)
                return PANEL_TYPE_NAMES.get(enum_key, f"Ú©Ø¯ Ù†Ø§Ù…Ø¹ØªØ¨Ø±: {key}")
            except ValueError: 
                return f"Ú©Ø¯ Ù†Ø§Ù…Ø¹ØªØ¨Ø±: {key}"
        return None

class ProjectCreateFromExcelData(BaseModel):
    name: str
    request_id: str
    customer_name: str
    location: Optional[str] = None

class ProjectCreateManual(BaseModel):
    name: str
    location: Optional[str] = None
    customer_name: Optional[str] = Field(None)
    request_id: Optional[str] = Field(None)

class StepCreate(BaseModel):
    step: StepNameKey
    
    @field_validator('step', mode='before')
    @classmethod
    def validate_step_from_string(cls, v: Any) -> StepNameKey:
        if isinstance(v, StepNameKey): 
            return v
        if isinstance(v, str):
            try:
                # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…Ù‚Ø¯Ø§Ø± Enum Ø¨Ø±Ø§ÛŒ Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ
                return StepNameKey(v)
            except ValueError:
                raise ValueError(f"'{v}' is not a valid step name or key.")
        raise TypeError("Step must be a string or StepNameKey enum.")

class BarcodeExitPayload(BaseModel):
    barcode_data: str
class AssemblyDetailsUpdate(BaseModel):
    panel_type_key: str  # ÙÙ‚Ø· string Ù…Ø¹Ù…ÙˆÙ„ÛŒ
    assembler_1: str = Field(..., min_length=1)
    assembler_2: Optional[str] = None
    
    model_config = ConfigDict(from_attributes=True)
class AssemblerStatsOut(BaseModel):
    total_panels: int
    panels_by_type: Dict[str, int]

class PersonnelCreate(BaseModel):
    name: str
    username: str
    password: str
    # âœ… Ø§ØµÙ„Ø§Ø­: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Enum Ø¨Ø±Ø§ÛŒ Ù†Ù‚Ø´
    role: PersonnelRole = PersonnelRole.EMPLOYEE

class PersonnelUpdate(BaseModel):
    name: str
    username: str
    password: Optional[str] = None
    # âœ… Ø§ØµÙ„Ø§Ø­: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Enum Ø¨Ø±Ø§ÛŒ Ù†Ù‚Ø´
    role: PersonnelRole = PersonnelRole.EMPLOYEE

class PersonnelLogin(BaseModel):
    username: str
    password: str

class PersonnelOut(BaseModel):
    id: int
    name: str
    username: str
    role: str
    is_active: int

    model_config = ConfigDict(from_attributes=True)

class Token(BaseModel):
    access_token: str
    token_type: str
    user: PersonnelOut

class TokenData(BaseModel):
    username: Optional[str] = None
    role: Optional[str] = None

# Daily Work Report Schemas
class DailyWorkReportCreate(BaseModel):
    personnel_id: int
    report_date: date
    report_data: Dict[str, Any]

class DailyWorkReportOut(BaseModel):
    id: int
    personnel_id: int
    report_date: date
    status: str
    supervisor_notes: Optional[str] = None
    report_data: Dict[str, Any]
    created_at: datetime
    updated_at: datetime
    personnel: PersonnelOut

    model_config = ConfigDict(from_attributes=True)

class KpiSummary(BaseModel):
    total_projects: int
    completed_this_month: int
    avg_completion_time_days: Optional[float] = None
    bottleneck_step: Optional[str] = None
    step_durations: Dict[str, Optional[float]]

class WarehouseBase(BaseModel):
    name: str
    location: Optional[str] = None
    description: Optional[str] = None

class WarehouseCreate(WarehouseBase):
    pass

class WarehouseOut(WarehouseBase):
    id: int
    model_config = ConfigDict(from_attributes=True)

class WarehouseItemBase(BaseModel):
    item_name: str
    description: Optional[str] = None
    min_stock_level: int = Field(0, ge=0)

class WarehouseItemCreate(WarehouseItemBase):
    pass

class WarehouseItemOut(WarehouseItemBase):
    id: int
    model_config = ConfigDict(from_attributes=True)

class InventoryTransactionBase(BaseModel):
    warehouse_id: int
    item_name: str
    quantity: int = Field(..., gt=0)
    notes: Optional[str] = None

class InventoryTransactionIn(InventoryTransactionBase):
    pass

class InventoryTransactionOutManual(InventoryTransactionBase):
    project_id: Optional[int] = None

class InventoryTransactionOut(BaseModel):
    id: int
    warehouse: WarehouseOut
    item: WarehouseItemOut
    quantity: int
    transaction_type: TransactionType
    project_id: Optional[int] = None
    user: PersonnelOut
    timestamp: datetime
    notes: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)
    
class CurrentStockItem(BaseModel):
    item_id: int
    item_name: str
    warehouse_id: int
    warehouse_name: str
    current_stock: int
    min_stock_level: int

class PanelCodeItemBase(BaseModel):
    item_name: str
    quantity_required: int = Field(..., gt=0)

class PanelCodeItemCreate(PanelCodeItemBase):
    pass

class PanelCodeItemsDefinition(BaseModel):
    panel_code: str
    items: List[PanelCodeItemCreate]

class PanelCodeItemOut(PanelCodeItemBase):
    id: int
    panel_code: str
    model_config = ConfigDict(from_attributes=True)
    
# ==============================================================================
# Ø¨Ø®Ø´ Û¸: ØªÙˆØ§Ø¨Ø¹ Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª Ùˆ Ø§Ù…Ù†ÛŒØª (Ù†Ø³Ø®Ù‡ Ú©Ø§Ù…Ù„Ø§Ù‹ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡)
# ==============================================================================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """ØªØ£ÛŒÛŒØ¯ ØµØ­Øª Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø¨Ø§ Ù‡Ø´ Ø°Ø®ÛŒØ±Ù‡â€ŒØ´Ø¯Ù‡."""
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    """Ù‡Ø´ Ú©Ø±Ø¯Ù† Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø¨Ø±Ø§ÛŒ Ø°Ø®ÛŒØ±Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø§Ù…Ù†."""
    return pwd_context.hash(password)

def authenticate_user(db: Session, username: str, password: str) -> Optional[Personnel]:
    """Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª Ú©Ø§Ø±Ø¨Ø± Ø¨Ø± Ø§Ø³Ø§Ø³ Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ Ùˆ Ø±Ù…Ø² Ø¹Ø¨ÙˆØ±."""
    try:
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø¨Ø±Ø±Ø³ÛŒ ÙØ¹Ø§Ù„ Ø¨ÙˆØ¯Ù† Ú©Ø§Ø±Ø¨Ø± Ø¯Ø± Ù‡Ù…Ø§Ù† Ú©ÙˆØ¦Ø±ÛŒ
        user = db.query(Personnel).filter(Personnel.username == username, Personnel.is_active == 1).first()
        if not user or not verify_password(password, user.password_hash):
            logger.warning(f"Authentication failed for user: {username}")
            return None
        logger.info(f"Authentication successful for user: {username}")
        return user
    except Exception as e:
        logger.error(f"Error during authentication for user {username}: {e}", exc_info=True)
        return None

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    """Ø§ÛŒØ¬Ø§Ø¯ ØªÙˆÚ©Ù† Ø¯Ø³ØªØ±Ø³ÛŒ JWT."""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: ØªØ¹ÛŒÛŒÙ† Ø²Ù…Ø§Ù† Ø§Ù†Ù‚Ø¶Ø§ÛŒ Ù¾ÛŒØ´â€ŒÙØ±Ø¶
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    logger.debug(f"Access token created for: {data.get('sub')}")
    return encoded_jwt

# Ø§Ø³ØªØ«Ù†Ø§ÛŒ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯ Ø¨Ø±Ø§ÛŒ Ø®Ø·Ø§Ù‡Ø§ÛŒ Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª
credentials_exception = HTTPException(
    status_code=status.HTTP_401_UNAUTHORIZED,
    detail="Could not validate credentials",
    headers={"WWW-Authenticate": "Bearer"},
)

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: ÛŒÚ©Ù¾Ø§Ø±Ú†Ù‡â€ŒØ³Ø§Ø²ÛŒ ØªÙˆØ§Ø¨Ø¹ get_current_user. Ø§ÛŒÙ† ØªØ§Ø¨Ø¹ Ù¾Ø§ÛŒÙ‡ØŒ ØªÙˆÚ©Ù† Ø±Ø§ Ø±Ù…Ø²Ú¯Ø´Ø§ÛŒÛŒ Ú©Ø±Ø¯Ù‡ Ùˆ Ú©Ø§Ø±Ø¨Ø± Ø±Ø§ Ø¨Ø±Ù…ÛŒâ€ŒÚ¯Ø±Ø¯Ø§Ù†Ø¯.
async def get_current_active_user(
    credentials: HTTPAuthorizationCredentials = Depends(security), 
    db: Session = Depends(get_db)
) -> Personnel:
    """ÙˆØ§Ø¨Ø³ØªÚ¯ÛŒ (Dependency) Ø¨Ø±Ø§ÛŒ Ø¯Ø±ÛŒØ§ÙØª Ú©Ø§Ø±Ø¨Ø± ÙØ¹Ø§Ù„ ÙØ¹Ù„ÛŒ Ø§Ø² ØªÙˆÚ©Ù† JWT."""
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username, role=payload.get("role"))
    except JWTError as e:
        logger.warning(f"JWT decode error: {e}")
        raise credentials_exception
    
    user = db.query(Personnel).filter(Personnel.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    if user.is_active != 1:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Inactive user")
    return user

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: ÙˆØ§Ø¨Ø³ØªÚ¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø¬Ø¯ÛŒØ¯ Ø¨Ø±Ø§ÛŒ Ø¨Ø±Ø±Ø³ÛŒ Ù†Ù‚Ø´â€ŒÙ‡Ø§ Ú©Ù‡ Ø®ÙˆØ§Ù†Ø§ØªØ± Ùˆ Ù‚Ø§Ø¨Ù„ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ù…Ø¬Ø¯Ø¯ Ù‡Ø³ØªÙ†Ø¯.
async def get_current_user(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """ÙˆØ§Ø¨Ø³ØªÚ¯ÛŒ Ø¨Ø±Ø§ÛŒ Ø¯Ø±ÛŒØ§ÙØª Ú©Ø§Ø±Ø¨Ø± ÙØ¹Ù„ÛŒ (Ø¨Ø¯ÙˆÙ† Ø¨Ø±Ø±Ø³ÛŒ Ù†Ù‚Ø´). Ø¬Ø§ÛŒÚ¯Ø²ÛŒÙ† get_current_user Ù‚Ø¯ÛŒÙ…ÛŒ."""
    return current_user

async def get_current_employee(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """ÙˆØ§Ø¨Ø³ØªÚ¯ÛŒ Ø¨Ø±Ø§ÛŒ Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² Ø§ÛŒÙ†Ú©Ù‡ Ú©Ø§Ø±Ø¨Ø± ÙØ¹Ù„ÛŒ 'employee' ÛŒØ§ 'supervisor' Ø§Ø³Øª."""
    if current_user.role not in [PersonnelRole.EMPLOYEE.value, PersonnelRole.SUPERVISOR.value]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Operation not permitted for this role")
    return current_user

async def get_current_supervisor(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """ÙˆØ§Ø¨Ø³ØªÚ¯ÛŒ Ø¨Ø±Ø§ÛŒ Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² Ø§ÛŒÙ†Ú©Ù‡ Ú©Ø§Ø±Ø¨Ø± ÙØ¹Ù„ÛŒ 'supervisor' Ø§Ø³Øª."""
    if current_user.role != PersonnelRole.SUPERVISOR.value:
        logger.warning(f"Access denied for user '{current_user.username}'. Required role: 'supervisor', User role: '{current_user.role}'")
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: Supervisor role required.")
    return current_user

async def get_user_by_username(db: Session, username: str) -> Optional[Personnel]:
    """ÛŒØ§ÙØªÙ† Ú©Ø§Ø±Ø¨Ø± Ø¨Ø± Ø§Ø³Ø§Ø³ Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ."""
    return db.query(Personnel).filter(Personnel.username == username).first()

# ==============================================================================
# Ø¨Ø®Ø´ Û¹: WebSocket Connection Manager
# ==============================================================================

class ConnectionManager:
    def __init__(self): 
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø³Ø§Ø®ØªØ§Ø± Ø¯Ø§Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ø°Ø®ÛŒØ±Ù‡ WebSocket Ùˆ user_id Ù…Ø±ØªØ¨Ø· Ø¨Ø§ Ø¢Ù†.
        self.active_connections: Dict[WebSocket, int] = {} # WebSocket -> user_id (0 for anonymous)
    
    async def connect(self, websocket: WebSocket, user_id: int = 0):
        await websocket.accept()
        self.active_connections[websocket] = user_id
        logger.info(f"WebSocket connected. User ID: {user_id}. Total connections: {len(self.active_connections)}")
    
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            del self.active_connections[websocket]
            logger.info(f"WebSocket disconnected. Total connections: {len(self.active_connections)}")
    
    async def send_to_user(self, user_id: int, message: dict):
        # Ø§Ø±Ø³Ø§Ù„ Ù¾ÛŒØ§Ù… Ø¨Ù‡ ÛŒÚ© Ú©Ø§Ø±Ø¨Ø± Ø®Ø§Øµ
        for ws, uid in self.active_connections.items():
            if uid == user_id:
                try:
                    await ws.send_json(message)
                except Exception as e:
                    logger.error(f"Error sending message to user {user_id}: {e}")
                    self.disconnect(ws)

    async def broadcast_to_supervisors(self, message: dict):
        """Ø§Ø±Ø³Ø§Ù„ Ù¾ÛŒØ§Ù… Ø¨Ù‡ ØªÙ…Ø§Ù… Ú©Ø§Ø±Ø¨Ø±Ø§Ù† Ù…ØªØµÙ„. Ø¯Ø± Ø§ÛŒÙ† Ù¾ÛŒØ§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø³Ø§Ø¯Ù‡ØŒ ØªÙÚ©ÛŒÚ© Ù†Ù‚Ø´ Ø¯Ø± Ø³Ù…Øª Ú©Ù„Ø§ÛŒÙ†Øª Ø§Ù†Ø¬Ø§Ù… Ù…ÛŒâ€ŒØ´ÙˆØ¯."""
        await self.broadcast(message)

    async def broadcast(self, message: dict):
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² asyncio.gather Ø¨Ø±Ø§ÛŒ Ø§Ø±Ø³Ø§Ù„ Ù‡Ù…Ø²Ù…Ø§Ù† Ùˆ Ù…Ø¯ÛŒØ±ÛŒØª Ø¨Ù‡ØªØ± Ù‚Ø·Ø¹ Ø§ØªØµØ§Ù„.
        if not self.active_connections:
            return
            
        disconnected_sockets = []
        # Ø§ÛŒØ¬Ø§Ø¯ ÛŒÚ© Ú©Ù¾ÛŒ Ø¨Ø±Ø§ÛŒ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ø®Ø·Ø§ Ù‡Ù†Ú¯Ø§Ù… ØªØºÛŒÛŒØ± Ø¯ÛŒÚ©Ø´Ù†Ø±ÛŒ Ø¯Ø± Ø­ÛŒÙ† Ù¾ÛŒÙ…Ø§ÛŒØ´
        connections_to_send = list(self.active_connections.keys())

        # Ø§Ø±Ø³Ø§Ù„ Ù‡Ù…Ø²Ù…Ø§Ù† Ù¾ÛŒØ§Ù…â€ŒÙ‡Ø§
        results = await asyncio.gather(
            *[conn.send_json(message) for conn in connections_to_send], 
            return_exceptions=True
        )

        # Ø¨Ø±Ø±Ø³ÛŒ Ù†ØªØ§ÛŒØ¬ Ùˆ Ø­Ø°Ù Ø§ØªØµØ§Ù„Ø§Øª Ù‚Ø·Ø¹ Ø´Ø¯Ù‡
        for i, result in enumerate(results):
            if isinstance(result, Exception): 
                logger.error(f"WebSocket send error: {result}")
                disconnected_sockets.append(connections_to_send[i])
        
        for ws in disconnected_sockets:
            self.disconnect(ws)

manager = ConnectionManager()

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û°: FastAPI App Ùˆ Middleware
# ==============================================================================

app = FastAPI(
    title="Project Monitoring System",
    description="Ø³ÛŒØ³ØªÙ… Ù…Ø§Ù†ÛŒØªÙˆØ±ÛŒÙ†Ú¯ Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ÛŒ ØªÙˆÙ„ÛŒØ¯ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø¨Ø±Ù‚",
    version="2.1.0" # âœ… Ù†Ø³Ø®Ù‡ Ø¨Ù‡â€ŒØ±ÙˆØ² Ø´Ø¯.
)

# Ø§ÙØ²ÙˆØ¯Ù† middleware Ù‡Ø§ÛŒ Ø§Ù…Ù†ÛŒØªÛŒ
if os.getenv("ENVIRONMENT") == "production":
    app.add_middleware(HTTPSRedirectMiddleware)

app.add_middleware(
    TrustedHostMiddleware, 
    allowed_hosts=ALLOWED_HOSTS
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    max_age=3600,
)

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û±: ØªØ´Ø®ÛŒØµ Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ
# ==============================================================================
# ğŸ”¥ Ù†Ú©ØªÙ‡: Ø§ÛŒÙ† Ø¨Ø®Ø´ Ù…Ù†Ø·Ù‚ ØªØ¬Ø§Ø±ÛŒ Ø¨Ø³ÛŒØ§Ø± Ø´Ú©Ù†Ù†Ø¯Ù‡â€ŒØ§ÛŒ Ø¯Ø§Ø±Ø¯. Ø¨Ù‡ØªØ±ÛŒÙ† Ø±Ø§Ù‡Ú©Ø§Ø± Ø¯Ø± Ø¨Ù„Ù†Ø¯Ù…Ø¯ØªØŒ
# Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ú©Ø¯Ù‡Ø§ÛŒ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯ ØªØ¬Ù‡ÛŒØ²Ø§Øª Ø¨Ù‡ Ø¬Ø§ÛŒ ØªØ·Ø¨ÛŒÙ‚ Ø±Ø´ØªÙ‡ Ø§Ø³Øª.
# Ø§ÛŒÙ† Ø¨Ø®Ø´ Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ± Ø¨Ø§Ù‚ÛŒ Ù…ÛŒâ€ŒÙ…Ø§Ù†Ø¯ Ú†ÙˆÙ† Ù…Ù†Ø·Ù‚ ØªØ¬Ø§Ø±ÛŒ Ø®Ø§Øµ Ù¾Ø±ÙˆÚ˜Ù‡ Ø§Ø³Øª.

ITEM_ID_TO_KEYWORDS = {
    'ID2R': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ", "Ø¸Ø±ÙÛŒØª", "Ø¯Ùˆ Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²", "Ø±ÛŒÙ„ÛŒ"),
    '140': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ", "ØªÚ©ÙØ§Ø²"),
    '141': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ", "Ø³Ù‡ ÙØ§Ø²"),
    '146': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "ØªÚ©ÙØ§Ø²", "Ø¨Ø§ ÙÛŒÙˆØ²"),
    '147': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Ø³Ù‡ ÙØ§Ø²", "Ø¨Ø§ ÙÛŒÙˆØ²"),
    '120': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ©ÙØ§Ø²"),
    '121': ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø³Ù‡ ÙØ§Ø²"),
    'FAHAM_METER_ASSEMBLED': ("Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²", "Ù‡ÙˆØ´Ù…Ù†Ø¯", "Ù…ÙˆÙ†ØªØ§Ú˜", "Ø´Ø¯Ù‡"),
    'FAHAM_METER_GPRS': ("Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²", "Ù‡ÙˆØ´Ù…Ù†Ø¯", "Ù…Ø§Ú˜ÙˆÙ„", "GPRS"),
    'FUSE_SP_ANY': ("Ú©Ù„ÛŒØ¯", "ÙÛŒÙˆØ²", "Ù…ÛŒÙ†ÛŒØ§ØªÙˆØ±ÛŒ", "ØªÚ©ÙØ§Ø²"),
    '103': ("Ø¬Ø¹Ø¨Ù‡", "ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"),
    '104': ("Ø¬Ø¹Ø¨Ù‡", "ÙÛŒÙˆØ²", "Ø³Ù‡ ÙØ§Ø²", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"),
    '105': ("Ø³Ú©ÙˆÛŒ", "Ù†ØµØ¨"),
    '116': ("Ø¬Ø¹Ø¨Ù‡", "ÙÛŒÙˆØ²", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ©ÙØ§Ø²", "Ù†ØµØ¨ Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡"),
    '117': ("Ø¬Ø¹Ø¨Ù‡", "ÙÛŒÙˆØ²", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø³Ù‡ ÙØ§Ø²", "Ù†ØµØ¨ Ø±ÙˆÛŒ Ù¾Ø§ÛŒÙ‡"),
    '122': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ© Ú©Ù†ØªÙˆØ±Ù‡", "Ø³Ù‡ ÙØ§Ø²", "ØªÚ©ÙØ§Ø²", "Ø³Ø§ÛŒØ²", "Û³Û¶"),
    '130': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Ø¸Ø±ÙÛŒØª Û´ Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²", "Ø±ÛŒÙ„ÛŒ"),
    '134': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Û´ Ú©Ù†ØªÙˆØ±Ù‡", "Ø³Ù‡ ÙØ§Ø²", "Ú©Ù†ØªÙˆØ± Ù‡ÙˆØ´Ù…Ù†Ø¯"),
    '135': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "Ø¸Ø±ÙÛŒØª Û¸ ÙÛŒÙˆØ²"),
    '136': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "Ø¸Ø±ÙÛŒØª Û±Û¶ ÙÛŒÙˆØ²"),
    '137': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡", "ÛŒÚ©Ø·Ø±ÙÙ‡"),
    '138': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "Ú†Ù†Ø¯ Ú©Ù†ØªÙˆØ±Ù‡", "Ø¯ÙˆØ·Ø±ÙÙ‡"),
    '142': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ù†ØªÙˆØ±", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ", "Ø¸Ø±ÙÛŒØª Û´ Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²", "Ø±ÛŒÙ„ÛŒ"),
    '170': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ© Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Ù…ÙˆØªÙˆØ± Ø¯Ø§Ø±", "Û³Û° ØªØ§ Û±ÛµÛ°"),
    '171': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ© Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Ù…ÙˆØªÙˆØ± Ø¯Ø§Ø±", "Û±ÛµÛ± ØªØ§ Û²Û´Û¹"),
    '172': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ© Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "ÛŒÚ©Ø·Ø±ÙÙ‡"),
    '173': ("ØªØ§Ø¨Ù„Ùˆ", "ÙÙ„Ø²ÛŒ", "ØªÚ© Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "ÛŒÚ©Ø·Ø±ÙÙ‡"),
    '174': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "Ø¯Ùˆ Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ù‡ÙˆØ§ÛŒÛŒ", "Ù…ÙˆØªÙˆØ±Ø¯Ø§Ø±"),
    '176': ("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ø§Ù…Ù¾ÙˆØ²ÛŒØªÛŒ", "ØªÚ© Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "Ø²Ù…ÛŒÙ†ÛŒ", "Ø¯ÙˆØ·Ø±ÙÙ‡"),
}

PANEL_ID_RULES = {
    PanelTypeKey.FAHAM_WITH_FRAME: {"requires": {'FAHAM_METER_ASSEMBLED', 'FUSE_SP_ANY'}, "forbids": {'120', '121', '140', '141', '146', '147'}},
    PanelTypeKey.FAHAM_WITHOUT_FRAME: {"requires": {'FAHAM_METER_GPRS', 'FUSE_SP_ANY'}, "forbids": {'120', '121', '140', '141', '146', '147'}},
    PanelTypeKey.ID218: {"requires": {'176'}},
    PanelTypeKey.ID216: {"requires": {'174'}},
    PanelTypeKey.ID215: {"requires": {'173'}},
    PanelTypeKey.ID213: {"requires": {'172'}},
    PanelTypeKey.ID212: {"requires": {'171'}},
    PanelTypeKey.ID211: {"requires": {'170'}},
    PanelTypeKey.ID123: {"requires": {'134', '136'}},
    PanelTypeKey.ID122: {"requires": {'134', '135'}},
    PanelTypeKey.ID109: {"requires": {'138'}},
    PanelTypeKey.ID108: {"requires": {'137'}},
    PanelTypeKey.ID105: {"requires": {'122'}},
    PanelTypeKey.ID2R: {"requires": {'ID2R'}},
    PanelTypeKey.ID116: {"requires": {'142'}},
    PanelTypeKey.ID107: {"requires": {'141', '104'}}, 
    PanelTypeKey.ID115: {"requires": {'140', '103'}},
    PanelTypeKey.ID121: {"requires": {'134', '117'}},
    PanelTypeKey.ID120: {"requires": {'134', '104'}, "forbids": {'117'}},
    PanelTypeKey.ID111: {"requires": {'130', '116'}},
    PanelTypeKey.ID110: {"requires": {'130', '103'}, "forbids": {'116'}},
    PanelTypeKey.ID104_3: {"requires": {'121', '117'}, "count": {'105': 2}},
    PanelTypeKey.ID104_1: {"requires": {'120', '116'}, "count": {'105': 2}},
    PanelTypeKey.ID102_3: {"requires": {'147', '105'}},
    PanelTypeKey.ID102_1: {"requires": {'146', '105'}},
    PanelTypeKey.ID101_3: {"requires": {'121'}, "forbids": {'147'}, "count": {'105': 1}},
    PanelTypeKey.ID101_1: {"requires": {'120'}, "forbids": {'146'}, "count": {'105': 1}},
}

EQUIPMENT_KEYWORD_TO_PANEL_MAPPING = {
    ("6 Ú©Ù†ØªÙˆØ±", "ÙÙ„Ø²ÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID6_1R,
    ("12 Ú©Ù†ØªÙˆØ±", "ÙÙ„Ø²ÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID12_1R,
    ("18 Ú©Ù†ØªÙˆØ±", "ÙÙ„Ø²ÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID18_1R,
    ("24 Ú©Ù†ØªÙˆØ±", "ÙÙ„Ø²ÛŒ", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID24_1R,
    ("Ø¯Ùˆ Ú©Ù†ØªÙˆØ±", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID2R,
    ("5 Ú©Ù†ØªÙˆØ±", "Ø¯ÛŒÙˆØ§Ø±ÛŒ"): PanelTypeKey.ID5R,
}

def _find_panel_details_from_equipment(equipment_list: List[EquipmentItemBase]) -> Tuple[Optional[str], Optional[str]]:
    """
    ØªØ´Ø®ÛŒØµ Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù„ÛŒØ³Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù‚ÙˆØ§Ù†ÛŒÙ† Ù…Ø¨ØªÙ†ÛŒ Ø¨Ø± Ú©Ù„Ù…Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ.
    """
    if not equipment_list: 
        return None, None
    
    found_item_ids = set()
    item_counts = defaultdict(int)
    normalized_item_names = [normalize_text(eq.item_name) for eq in equipment_list]
    
    for eq in equipment_list:
        name = normalize_text(eq.item_name)
        for item_id, keywords in ITEM_ID_TO_KEYWORDS.items():
            if all(keyword in name for keyword in keywords):
                found_item_ids.add(item_id)
                item_counts[item_id] += eq.quantity
                break
                
    for panel_key, rule in PANEL_ID_RULES.items():
        required_items = rule.get("requires", set())
        forbidden_items = rule.get("forbids", set())
        count_rules = rule.get("count", {})
        
        if not required_items.issubset(found_item_ids) or not forbidden_items.isdisjoint(found_item_ids):
            continue

        counts_met = True
        for item_id, required_count in count_rules.items():
            if item_counts.get(item_id, 0) != required_count:
                counts_met = False
                break
        
        if counts_met:
            return panel_key.value, panel_key.value # âœ… Ø§ØµÙ„Ø§Ø­: Ø¨Ø§Ø²Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† Ù…Ù‚Ø¯Ø§Ø± enum
            
    for name in normalized_item_names:
        for keywords_tuple, panel_key in EQUIPMENT_KEYWORD_TO_PANEL_MAPPING.items():
            if all(keyword in name for keyword in keywords_tuple):
                return panel_key.value, panel_key.value # âœ… Ø§ØµÙ„Ø§Ø­: Ø¨Ø§Ø²Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† Ù…Ù‚Ø¯Ø§Ø± enum
                
    return None, None

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û²: ØªÙˆØ§Ø¨Ø¹ Ú©Ù…Ú©ÛŒ
# ==============================================================================
# Ø§ÛŒÙ† ØªÙˆØ§Ø¨Ø¹ Ù…Ù†Ø·Ù‚ ØªØ¬Ø§Ø±ÛŒ Ù‡Ø³ØªÙ†Ø¯ Ùˆ Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ± Ø¨Ø§Ù‚ÛŒ Ù…ÛŒâ€ŒÙ…Ø§Ù†Ù†Ø¯.

def normalize_panel_key(key: str) -> str:
    """Ù†Ø±Ù…Ø§Ù„â€ŒØ³Ø§Ø²ÛŒ Ú©Ù„ÛŒØ¯ Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ."""
    if not isinstance(key, str): 
        return ""
    return re.sub(r'\s+', '', key).upper()
    
def convert_project_orm_to_pydantic(p_orm: Project) -> ProjectOut:
    """ØªØ¨Ø¯ÛŒÙ„ Ù…Ø¯Ù„ ORM Ø¨Ù‡ Pydantic."""
    return ProjectOut.model_validate(p_orm)    

def calculate_reservation_details(project: Project) -> Dict[str, Optional[int]]:
    """Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø¬Ø²Ø¦ÛŒØ§Øª Ø±Ø²Ø±Ùˆ."""
    max_capacity = None
    purchased_single_phase = 0
    reserved_count = None
    ignore_keywords = ["ØªØ§Ø¨Ù„Ùˆ", "Ø¬Ø¹Ø¨Ù‡", "Ù‚Ø§Ø¨", "Ø¨Ø¯Ù†Ù‡"]
    
    for item in project.equipment:
        normalized_name = normalize_text(item.item_name)
        is_single_phase_meter = "Ú©Ù†ØªÙˆØ±" in normalized_name and "ØªÚ©ÙØ§Ø²" in normalized_name
        should_be_ignored = any(keyword in normalized_name for keyword in ignore_keywords)
        if is_single_phase_meter and not should_be_ignored:
            purchased_single_phase += item.quantity
            
    if project.panel_code and isinstance(project.panel_code, str):
        match = re.search(r'ID(\d+)\+', project.panel_code.strip().upper())
        if match:
            try:
                max_capacity = int(match.group(1))
                reserved_count = max(0, max_capacity - purchased_single_phase)
            except (ValueError, TypeError):
                max_capacity = None
                reserved_count = None
                
    return {
        "max_capacity": max_capacity, 
        "purchased_single_phase": purchased_single_phase, 
        "reserved_count": reserved_count
    }

def generate_project_summary_data(project: Project) -> Dict[str, Any]:
    """ØªÙˆÙ„ÛŒØ¯ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ Ø®Ù„Ø§ØµÙ‡ Ù¾Ø±ÙˆÚ˜Ù‡."""
    ALL_SUMMARY_KEYS = [
        "Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°", "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63", "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32", "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25",
        "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32", "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25", "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16", "Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²",
        "Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²", "Ø³Ú©Ùˆ", "ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„", "Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„", "Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ",
        "Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯" 
    ]
    summary = {key: 0 for key in ALL_SUMMARY_KEYS}
    PRIORITIZED_RULES = [
        ("IGNORED", [("ØªØ§Ø¨Ù„Ùˆ", "Ú©Ù†ØªÙˆØ±"), ("Ø¬Ø¹Ø¨Ù‡", "Ú©Ù†ØªÙˆØ±"), ("Ø¬Ø¹Ø¨Ù‡", "ÙÛŒÙˆØ²"), ("Ù‚Ø§Ø¨", "Ú©Ù†ØªÙˆØ±"), ("Ù…ÙˆØ¯Ù…",)]),
        ("Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°", [("Ù¾ÛŒÚ†", "Ù…Ù‡Ø±Ù‡", "16", "300"), ("Ù¾ÛŒÚ†", "16Ø¯Ø±300")]),
        ("ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63", [("ÙÛŒÙˆØ²", "Ø³Ù‡", "ÙØ§Ø²", "63"), ("Ú©Ù„ÛŒØ¯", "Ø³Ù‡", "ÙØ§Ø²", "63")]),
        ("ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32", [("ÙÛŒÙˆØ²", "Ø³Ù‡", "ÙØ§Ø²", "32"), ("Ú©Ù„ÛŒØ¯", "Ø³Ù‡", "ÙØ§Ø²", "32")]),
        ("ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25", [("ÙÛŒÙˆØ²", "Ø³Ù‡", "ÙØ§Ø²", "25"), ("Ú©Ù„ÛŒØ¯", "Ø³Ù‡", "ÙØ§Ø²", "25")]),
        ("ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32", [("ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "32"), ("Ú©Ù„ÛŒØ¯", "ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "32")]),
        ("ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25", [("ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "25"), ("Ú©Ù„ÛŒØ¯", "ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "25")]),
        ("ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16", [("ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "16"), ("Ú©Ù„ÛŒØ¯", "ÙÛŒÙˆØ²", "ØªÚ©ÙØ§Ø²", "16")]),
        ("Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²", [("Ú©Ù†ØªÙˆØ±", "ØªÚ©ÙØ§Ø²")]), 
        ("Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²", [("Ú©Ù†ØªÙˆØ±", "Ø³Ù‡", "ÙØ§Ø²")]),
        ("Ø³Ú©Ùˆ", [("Ø³Ú©ÙˆÛŒ", "Ù†ØµØ¨")]), 
        ("ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„", [("ØªØ³Ù…Ù‡", "Ø§Ø³ØªÛŒÙ„")]), 
        ("Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„", [("Ø¨Ø³Øª", "ØªØ³Ù…Ù‡")]),
        ("Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ", [("Ù„ÙˆÙ„Ù‡", "Ø®Ø±Ø·ÙˆÙ…ÛŒ")]), 
        ("Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯", [("Ù„ÙˆÙ„Ù‡", "Ù†ÛŒÙ… Ú¯Ø±Ø¯")]),
    ]
    
    for equipment_item in project.equipment:
        normalized_name = normalize_text(equipment_item.item_name)
        matched = False
        for category, keyword_sets in PRIORITIZED_RULES:
            if any(all(keyword in normalized_name for keyword in kw_set) for kw_set in keyword_sets):
                if category != "IGNORED":
                    summary[category] += equipment_item.quantity
                matched = True
                break
        if not matched:
            if "Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ" in normalized_name:
                summary["Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ"] += equipment_item.quantity
            elif "Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯" in normalized_name:
                summary["Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯"] += equipment_item.quantity

    reservation_details = calculate_reservation_details(project)
    summary["Ø¸Ø±ÙÛŒØª ØªØ§Ø¨Ù„Ùˆ (ØªÚ©ÙØ§Ø²)"] = reservation_details.get("max_capacity")
    summary["ØªØ¹Ø¯Ø§Ø¯ Ø®Ø±ÛŒØ¯Ø§Ø±ÛŒ Ø´Ø¯Ù‡ (ØªÚ©ÙØ§Ø²)"] = reservation_details.get("purchased_single_phase", 0)
    summary["ØªØ¹Ø¯Ø§Ø¯ Ø±Ø²Ø±Ùˆ"] = reservation_details.get("reserved_count")
    
    return summary

def get_direction(request_id: str) -> Optional[str]:
    """ØªØ¹ÛŒÛŒÙ† Ø¬Ù‡Øª (Ø´Ø±Ù‚/ØºØ±Ø¨) Ø¨Ø± Ø§Ø³Ø§Ø³ Ú©Ø¯ Ù…Ù†Ø·Ù‚Ù‡ Ø¯Ø± Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª."""
    if not isinstance(request_id, str) or len(request_id) < 5: 
        return None
    area_code = request_id.strip()[3:5]
    direction_map = {
        "west": {"03", "04", "05", "07", "09", "10", "11", "12", "15"}, 
        "east": {"01", "02", "06", "08", "16", "17", "18"}
    }
    if area_code in direction_map["west"]: 
        return "west"
    if area_code in direction_map["east"]: 
        return "east"
    return None

def get_projects_by_status(
    db: Session, 
    required_steps: set, 
    forbidden_steps: set = None
) -> List[Project]:
    """Ø¯Ø±ÛŒØ§ÙØª Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ Ø¨Ø± Ø§Ø³Ø§Ø³ ÙˆØ¶Ø¹ÛŒØª Ù…Ø±Ø§Ø­Ù„ Ø§Ù†Ø¬Ø§Ù…â€ŒØ´Ø¯Ù‡ Ùˆ Ø§Ù†Ø¬Ø§Ù…â€ŒÙ†Ø´Ø¯Ù‡."""
    query = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment)
    )
    
    for step_key in required_steps: 
        query = query.filter(Project.steps.any(Step.name_key == step_key))
        
    if forbidden_steps:
        for step_key in forbidden_steps: 
            query = query.filter(~Project.steps.any(Step.name_key == step_key))
            
    return query.order_by(Project.panel_code, Project.created_at).all()

def generate_detailed_report_data(projects: List[Project], direction: str) -> List[Dict[str, Any]]:
    """ØªÙˆÙ„ÛŒØ¯ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ø¨Ø±Ø§ÛŒ Ú¯Ø²Ø§Ø±Ø´ ØªÙØµÛŒÙ„ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„."""
    report_data = []
    filtered_projects = [p for p in projects if get_direction(p.request_id) == direction]
    
    for i, project in enumerate(filtered_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "Ù…ÙˆØ¯Ù…" in normalize_text(item.item_name)
        )
        
        row_data = {
            "project_id": project.id, 
            "Ø±Ø¯ÛŒÙ": i, 
            "Ù†Ø§Ù… Ù…Ø´ØªØ±Ú©": project.customer_name, 
            "Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§": project.request_id,
            "Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ": project.panel_code or project.panel_type_key or "-", 
            "Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²": summary.get('Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²', 0),
            "Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²": summary.get('Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²', 0), 
            "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16": summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16', 0),
            "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25": summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25', 0), 
            "ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32": summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32', 0),
            "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25": summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25', 0), 
            "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32": summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32', 0),
            "ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63": summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63', 0), 
            "Ù…ÙˆØ¯Ù…": modem_count, 
            "Ø¸Ø±ÙÛŒØª": reservation.get('max_capacity'),
            "Ø®Ø±ÛŒØ¯": reservation.get('purchased_single_phase'), 
            "Ø±Ø²Ø±Ùˆ": reservation.get('reserved_count'), 
            "Ø³Ú©Ùˆ": summary.get('Ø³Ú©Ùˆ', 0),
            "ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„": summary.get('ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0), 
            "Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„": summary.get('Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0),
            "Ù¾ÛŒÚ† ÛŒÚ©Ø³Ø±Ø±Ø²ÙˆÙ‡": 0, 
            "Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°": summary.get('Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°', 0), 
            "Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯": summary.get('Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯', 0),
            "Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ": summary.get('Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ', 0), 
            "ØªÙˆØ¶ÛŒØ­Ø§Øª": ""
        }
        report_data.append(row_data)
        
    return report_data

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û².Ûµ: ØªÙˆØ§Ø¨Ø¹ Ú©Ù…Ú©ÛŒ Ø§Ù†Ø¨Ø§Ø±Ø¯Ø§Ø±ÛŒ (Ø¨Ø®Ø´ Ø¬Ø¯ÛŒØ¯)
# ==============================================================================
async def check_stock_and_alert(db: Session, item: WarehouseItem, warehouse_id: int):
    """Ù…ÙˆØ¬ÙˆØ¯ÛŒ Ú©Ø§Ù„Ø§ Ø±Ø§ Ø¨Ø±Ø±Ø³ÛŒ Ú©Ø±Ø¯Ù‡ Ùˆ Ø¯Ø± ØµÙˆØ±Øª Ø±Ø³ÛŒØ¯Ù† Ø¨Ù‡ Ø­Ø¯Ø§Ù‚Ù„ØŒ Ù‡Ø´Ø¯Ø§Ø± WebSocket Ø§Ø±Ø³Ø§Ù„ Ù…ÛŒâ€ŒÚ©Ù†Ø¯."""
    stock_level_expr = func.sum(
        case(
            (InventoryTransaction.transaction_type == TransactionType.IN.value, InventoryTransaction.quantity),
            else_=-InventoryTransaction.quantity
        )
    ).label("current_stock")

    result = db.query(stock_level_expr).filter(
        InventoryTransaction.item_id == item.id,
        InventoryTransaction.warehouse_id == warehouse_id
    ).one_or_none()

    current_stock = result.current_stock if result and result.current_stock is not None else 0
    
    if current_stock <= item.min_stock_level:
        logger.warning(f"LOW STOCK ALERT: Item '{item.item_name}' (ID: {item.id}) is at {current_stock}, below/equal to threshold of {item.min_stock_level}")
        await manager.broadcast({
            "type": "low_stock_alert",
            "data": {
                "item_name": item.item_name,
                "current_stock": current_stock,
                "min_stock_level": item.min_stock_level
            }
        })

async def deduct_stock_for_project(db: Session, project: Project, user: Personnel):
    """Ù…ÙˆØ¬ÙˆØ¯ÛŒ Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ø¨Ø±Ø§ÛŒ ÛŒÚ© Ù¾Ø±ÙˆÚ˜Ù‡ Ø±Ø§ Ø¨Ù‡ ØµÙˆØ±Øª Ø®ÙˆØ¯Ú©Ø§Ø± Ø§Ø² Ø§Ù†Ø¨Ø§Ø± Ú©Ø³Ø± Ù…ÛŒâ€ŒÚ©Ù†Ø¯."""
    if not project.panel_code:
        logger.warning(f"Project {project.id} has no panel_code, skipping auto stock deduction.")
        return

    required_items = db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == project.panel_code).all()
    if not required_items:
        logger.info(f"No BOM defined for panel_code '{project.panel_code}'. Cannot auto-deduct stock.")
        return

    main_warehouse = db.query(Warehouse).order_by(Warehouse.id).first()
    if not main_warehouse:
        logger.error("Cannot deduct stock: No warehouse defined in the system.")
        raise HTTPException(status_code=500, detail="Ù‡ÛŒÚ† Ø§Ù†Ø¨Ø§Ø±ÛŒ Ø¯Ø± Ø³ÛŒØ³ØªÙ… ØªØ¹Ø±ÛŒÙ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")

    for required in required_items:
        item = db.query(WarehouseItem).filter(WarehouseItem.item_name == required.item_name).first()
        if not item:
            logger.error(f"Item '{required.item_name}' for panel '{project.panel_code}' not found in warehouse items.")
            continue

        transaction = InventoryTransaction(
            warehouse_id=main_warehouse.id,
            item_id=item.id,
            quantity=required.quantity_required,
            transaction_type=TransactionType.OUT,
            project_id=project.id,
            user_id=user.id,
            notes=f"Ø®Ø±ÙˆØ¬ Ø®ÙˆØ¯Ú©Ø§Ø± Ø¨Ø±Ø§ÛŒ Ù¾Ø±ÙˆÚ˜Ù‡ {project.request_id} Ø¨Ø§ Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ {project.panel_code}"
        )
        db.add(transaction)
    
    db.commit()
    logger.info(f"Auto-deducted stock for project {project.id} based on panel code '{project.panel_code}'.")

    # Ø¨Ø±Ø±Ø³ÛŒ Ù…ÙˆØ¬ÙˆØ¯ÛŒ Ù¾Ø³ Ø§Ø² Ø«Ø¨Øª Ù‡Ù…Ù‡ ØªØ±Ø§Ú©Ù†Ø´â€ŒÙ‡Ø§
    for required in required_items:
        item = db.query(WarehouseItem).filter(WarehouseItem.item_name == required.item_name).first()
        if item:
            await check_stock_and_alert(db, item, main_warehouse.id)
    
# ==============================================================================
# Ø¨Ø®Ø´ Û±Û³: Background Tasks
# ==============================================================================
# Ø§ÛŒÙ† Ø¨Ø®Ø´â€ŒÙ‡Ø§ Ø¨Ù‡ Ø®ÙˆØ¨ÛŒ Ù¾ÛŒØ§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø´Ø¯Ù‡â€ŒØ§Ù†Ø¯ Ùˆ ØªØºÛŒÛŒØ±Ø§Øª Ø¬Ø²Ø¦ÛŒ Ø¨Ø±Ø§ÛŒ Ø¨Ù‡Ø¨ÙˆØ¯ Ù„Ø§Ú¯ Ùˆ Ø®Ø·Ø§ Ø¯Ø§Ø±Ù†Ø¯.

def process_excel_in_background(file_contents: bytes, ws_manager: ConnectionManager):
    """Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡."""
    with get_db_session() as db:
        try:
            projects_payload_map: Dict[str, Dict[str, Any]] = defaultdict(
                lambda: {'base_info': None, 'equipment': []}
            )
            workbook = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)
            sheet = workbook.active
            header_row_values = [
                str(cell.value).strip().lower() if cell.value else "" 
                for cell in sheet[1]
            ]
            
            EXPECTED_COLS = { 
                "Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª ØªÙˆØ²ÛŒØ¹": "Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª ØªÙˆØ²ÛŒØ¹".lower(), 
                "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ Ø®Ø±ÛŒØ¯Ø§Ø±": "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ Ø®Ø±ÛŒØ¯Ø§Ø±".lower(), 
                "ØªØ¬Ù‡ÛŒØ²": "ØªØ¬Ù‡ÛŒØ²".lower(), 
                "ØªØ¹Ø¯Ø§Ø¯": "ØªØ¹Ø¯Ø§Ø¯".lower() 
            }
            
            if not all(col in header_row_values for col in EXPECTED_COLS.values()):
                missing = [k for k, v in EXPECTED_COLS.items() if v not in header_row_values]
                raise ValueError(f"Ø³ØªÙˆÙ†â€ŒÙ‡Ø§ÛŒ Ø¶Ø±ÙˆØ±ÛŒ ÛŒØ§ÙØª Ù†Ø´Ø¯: {', '.join(missing)}")
                
            RID_COL_IDX = header_row_values.index(EXPECTED_COLS["Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª ØªÙˆØ²ÛŒØ¹"])
            CUST_COL_IDX = header_row_values.index(EXPECTED_COLS["Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ Ø®Ø±ÛŒØ¯Ø§Ø±"])
            ITEM_NAME_COL_IDX = header_row_values.index(EXPECTED_COLS["ØªØ¬Ù‡ÛŒØ²"])
            ITEM_QTY_COL_IDX = header_row_values.index(EXPECTED_COLS["ØªØ¹Ø¯Ø§Ø¯"])
            
            last_valid_req_id = None
            for r_idx, row_tuple in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row = list(row_tuple)
                req_id = str(row[RID_COL_IDX]).strip() if RID_COL_IDX < len(row) and row[RID_COL_IDX] else last_valid_req_id
                if not req_id: 
                    continue
                last_valid_req_id = req_id
                
                if not projects_payload_map[req_id]['base_info']:
                    cust_name = str(row[CUST_COL_IDX]).strip() if CUST_COL_IDX < len(row) and row[CUST_COL_IDX] else "Ù†Ø§Ù…Ø´Ø®Øµ"
                    projects_payload_map[req_id]['base_info'] = ProjectCreateFromExcelData(
                        name=f"Ù¾Ø±ÙˆÚ˜Ù‡ {req_id}", 
                        request_id=req_id, 
                        customer_name=cust_name, 
                        location=None
                    )
                    
                item_name = str(row[ITEM_NAME_COL_IDX]).strip() if ITEM_NAME_COL_IDX < len(row) and row[ITEM_NAME_COL_IDX] else None
                if not item_name: 
                    continue
                    
                qty = int(float(str(row[ITEM_QTY_COL_IDX]).strip())) if ITEM_QTY_COL_IDX < len(row) and row[ITEM_QTY_COL_IDX] else 1
                projects_payload_map[req_id]['equipment'].append(
                    EquipmentItemCreate(item_name=item_name, quantity=qty)
                )
                
            processed_count = 0
            for req_id, data in projects_payload_map.items():
                if not data['base_info']: 
                    continue
                    
                db_proj = db.query(Project).filter(Project.request_id == req_id).first()
                if db_proj:
                    db.query(EquipmentItem).filter(EquipmentItem.project_id == db_proj.id).delete()
                    db_proj.name = data['base_info'].name
                    db_proj.customer_name = data['base_info'].customer_name
                else:
                    db_proj = Project(**data['base_info'].model_dump())
                    db.add(db_proj)
                    
                db.flush()
                
                for eq in data['equipment']: 
                    db.add(EquipmentItem(project_id=db_proj.id, **eq.model_dump()))
                    
                db.commit()
                processed_count += 1
                
            if processed_count > 0:
                asyncio.run(ws_manager.broadcast({
                    "type": "update", 
                    "source": "excel_upload", 
                    "count": processed_count
                }))
                logger.info(f"Excel processing completed: {processed_count} projects updated")
                
        except Exception as e:
            logger.error(f"Error processing Excel: {e}", exc_info=True)
            asyncio.run(ws_manager.broadcast({
                "type": "error", 
                "message": f"Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø§Ú©Ø³Ù„: {e}"
            }))

def process_detailed_excel_in_background(files_contents: List[Tuple[str, bytes]], ws_manager: ConnectionManager):
    """Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø§Ú©Ø³Ù„ ØªÙØµÛŒÙ„ÛŒ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡."""
    with get_db_session() as db:
        processed_count = 0
        errors = []
        
        def find_value_by_keyword(sheet, keyword: str):
            """Ù¾ÛŒØ¯Ø§ Ú©Ø±Ø¯Ù† Ù…Ù‚Ø¯Ø§Ø± Ø¨Ø± Ø§Ø³Ø§Ø³ Ú©Ù„ÛŒØ¯ÙˆØ§Ú˜Ù‡"""
            for row in sheet["A1:C10"]:
                cell = row[0]
                if cell.value and keyword in str(cell.value):
                    target_cell = sheet.cell(row=cell.row, column=3)
                    return str(target_cell.value).strip() if target_cell.value else None
            return None
            
        try:
            for filename, contents in files_contents:
                try:
                    workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
                    sheet = workbook.active
                    
                    request_id = find_value_by_keyword(sheet, "Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø´Ø±Ú©Øª ØªÙˆØ²ÛŒØ¹")
                    if not request_id:
                        errors.append(f"ÙØ§ÛŒÙ„ '{filename}': Ú©Ù„ÛŒØ¯ÙˆØ§Ú˜Ù‡ 'Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø´Ø±Ú©Øª ØªÙˆØ²ÛŒØ¹' ÛŒØ§ÙØª Ù†Ø´Ø¯.")
                        continue
                        
                    invoice_number = find_value_by_keyword(sheet, "Ø´Ù…Ø§Ø±Ù‡ Ø³ÙØ§Ø±Ø´")
                    customer_name = find_value_by_keyword(sheet, "Ù†Ø§Ù… Ù…Ø´ØªØ±ÛŒ") or "Ù†Ø§Ù…Ø´Ø®Øµ"
                    distribution_company = find_value_by_keyword(sheet, "Ø´Ø±Ú©Øª ØªÙˆØ²ÛŒØ¹") or ""
                    location = f"Ø´Ø±Ú©Øª: {distribution_company} - ÙØ§Ú©ØªÙˆØ±: {invoice_number}" if distribution_company or invoice_number else None
                    
                    equipment_start_row = -1
                    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, min_col=1, max_col=3, values_only=True), 1):
                        if row[0] == "Ø±Ø¯ÛŒÙ" and row[1] == "Ú©Ø¯ ØªØ¬Ù‡ÛŒØ²/Ø¹Ù…Ù„ÛŒØ§Øª" and row[2] == "ØªØ¬Ù‡ÛŒØ²/Ø¹Ù…Ù„ÛŒØ§Øª":
                            equipment_start_row = i + 1
                            break
                            
                    if equipment_start_row == -1:
                        errors.append(f"ÙØ§ÛŒÙ„ '{filename}': Ø¬Ø¯ÙˆÙ„ ØªØ¬Ù‡ÛŒØ²Ø§Øª (Ù‡Ø¯Ø± 'Ø±Ø¯ÛŒÙ', 'Ú©Ø¯ ØªØ¬Ù‡ÛŒØ²'...) ÛŒØ§ÙØª Ù†Ø´Ø¯.")
                        continue
                        
                    equip_list = []
                    for row in sheet.iter_rows(min_row=equipment_start_row, max_col=5, values_only=True):
                        item_name_raw, quantity_raw = row[2], row[4]
                        if not item_name_raw or str(item_name_raw).strip() == "": 
                            break
                            
                        full_item_name = str(item_name_raw).strip()
                        try: 
                            quantity = int(quantity_raw) if isinstance(quantity_raw, (int, float)) and quantity_raw > 0 else 1
                        except (ValueError, TypeError): 
                            quantity = 1
                            
                        equip_list.append(EquipmentItemCreate(item_name=full_item_name, quantity=quantity))
                        
                    if not equip_list:
                        errors.append(f"ÙØ§ÛŒÙ„ '{filename}': Ù‡ÛŒÚ† ØªØ¬Ù‡ÛŒØ²ÛŒ Ø¨Ø±Ø§ÛŒ Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª {request_id} ÛŒØ§ÙØª Ù†Ø´Ø¯.")
                        continue
                        
                    panel_code, panel_type_key = _find_panel_details_from_equipment(equip_list)
                    db_proj = db.query(Project).filter(Project.request_id == request_id).first()
                    project_name = f"Ù¾Ø±ÙˆÚ˜Ù‡ {panel_code} - {request_id}" if panel_code else f"Ù¾Ø±ÙˆÚ˜Ù‡ {request_id}"
                    
                    if db_proj:
                        db.query(EquipmentItem).filter(EquipmentItem.project_id == db_proj.id).delete()
                        db_proj.name = project_name
                        db_proj.customer_name = customer_name
                        db_proj.location = location
                        db_proj.panel_code = panel_code
                        db_proj.panel_type_key = panel_type_key
                    else:
                        db_proj = Project(
                            name=project_name, 
                            request_id=request_id, 
                            customer_name=customer_name, 
                            location=location, 
                            panel_code=panel_code, 
                            panel_type_key=panel_type_key, 
                            barcode_payload=request_id
                        )
                        db.add(db_proj)
                        
                    db.flush()
                    
                    for eq in equip_list: 
                        db.add(EquipmentItem(project_id=db_proj.id, **eq.model_dump()))
                        
                    db.commit()
                    processed_count += 1
                    
                except Exception as e:
                    db.rollback()
                    errors.append(f"ÙØ§ÛŒÙ„ '{filename}': Ø®Ø·Ø§ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ - {e}")
                    logger.error(f"Error processing file {filename}: {e}", exc_info=True)
                    
            if processed_count > 0:
                asyncio.run(ws_manager.broadcast({
                    "type": "update", 
                    "source": "detailed_excel_upload", 
                    "count": processed_count, 
                    "message": f"{processed_count} Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø´Ø¯."
                }))
                logger.info(f"Detailed Excel processing completed: {processed_count} projects")
                
            if errors:
                error_message = "Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø±Ø®ÛŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§:\n" + "\n".join(errors)
                logger.warning(f"Excel processing errors: {error_message}")
                asyncio.run(ws_manager.broadcast({
                    "type": "error", 
                    "message": error_message
                }))
                
        except Exception as e:
            logger.error(f"Error in detailed Excel processing: {e}", exc_info=True)
            asyncio.run(ws_manager.broadcast({
                "type": "error", 
                "message": f"Ø®Ø·Ø§ÛŒ Ú©Ù„ÛŒ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø§Ú©Ø³Ù„ (ØªÙØµÛŒÙ„ÛŒ): {e}"
            }))

# ØªÙˆØ§Ø¨Ø¹ Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ (Ø¨Ù‡ ØµÙˆØ±Øª Ù…ÙˆÙ‚Øª ØºÛŒØ±ÙØ¹Ø§Ù„)
async def run_branch_validation(project_id: int, request_id: str, purchase_list: list, ws_manager: ConnectionManager):
    """Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø§Ù†Ø´Ø¹Ø§Ø¨Ø§Øª (Ù…Ù†Ø·Ù‚ Ø®Ø§Ø±Ø¬ÛŒ)."""
    result = ValidationResponse(has_discrepancy=True, message="Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´.")
    
    try:
        # Ø¨Ù‡ ØµÙˆØ±Øª Ù…ÙˆÙ‚Øª ØºÛŒØ±ÙØ¹Ø§Ù„ Ø´Ø¯Ù‡
        # async with NaabConnector(NAAB_USERNAME, NAAB_PASSWORD) as connector:
        #     site_branch_count = await connector.get_site_branch_count(request_id)
        
        site_branch_count = 0  # Ù…Ù‚Ø¯Ø§Ø± Ù…ÙˆÙ‚Øª
        purchased_meter_count = sum(
            item['quantity'] for item in purchase_list 
            if "Ú©Ù†ØªÙˆØ±" in normalize_text(item['name'])
        )
        
        if site_branch_count == purchased_meter_count:
            result.has_discrepancy = False
            result.message = f"ØªØ·Ø§Ø¨Ù‚ Ù…ÙˆÙÙ‚: {site_branch_count} Ø§Ù†Ø´Ø¹Ø§Ø¨ Ø¯Ø± Ø³Ø§ÛŒØª Ùˆ {purchased_meter_count} Ú©Ù†ØªÙˆØ± Ø¯Ø± Ù¾Ø±ÙˆÚ˜Ù‡."
        else:
            result.has_discrepancy = True
            result.message = f"Ù…ØºØ§ÛŒØ±Øª Ø§Ù†Ø´Ø¹Ø§Ø¨: {site_branch_count} Ø§Ù†Ø´Ø¹Ø§Ø¨ Ø¯Ø± Ø³Ø§ÛŒØªØŒ Ø§Ù…Ø§ {purchased_meter_count} Ú©Ù†ØªÙˆØ± Ø¯Ø± Ù¾Ø±ÙˆÚ˜Ù‡ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª."
            
    except Exception as e:
        logger.error(f"Branch validation error for project {project_id}: {e}")
        result.message = f"Ø®Ø·Ø§ Ø¯Ø± Ø¨Ø±Ø±Ø³ÛŒ Ø§Ù†Ø´Ø¹Ø§Ø¨: {str(e)}"
    
    await ws_manager.broadcast({
        "type": "validation_result", 
        "project_id": project_id, 
        "result": result.model_dump()
    })

async def run_purchase_validation(project_id: int, request_id: str, purchase_list: list, ws_manager: ConnectionManager):
    """Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø®Ø±ÛŒØ¯Ù‡Ø§ (Ù…Ù†Ø·Ù‚ Ø®Ø§Ø±Ø¬ÛŒ)."""
    result = ValidationResponse(has_discrepancy=True, message="Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´.")

    try:
        # Ø¨Ù‡ ØµÙˆØ±Øª Ù…ÙˆÙ‚Øª ØºÛŒØ±ÙØ¹Ø§Ù„ Ø´Ø¯Ù‡
        # async with NaabConnector(NAAB_USERNAME, NAAB_PASSWORD) as connector:
        #     site_items = await connector.get_site_purchased_items(request_id)

        site_items = []  # Ù…Ù‚Ø¯Ø§Ø± Ù…ÙˆÙ‚Øª
        if not site_items:
            result.message = "Ù‡ÛŒÚ† ØªØ¬Ù‡ÛŒØ²ÛŒ Ø¯Ø± ØªØ¨ Ù…Ø§Ù„ÛŒ Ø³Ø§ÛŒØª Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† ØªÙ‚Ø§Ø¶Ø§ ÛŒØ§ÙØª Ù†Ø´Ø¯."
        else:
            site_dict = {normalize_text(item['name']): item['quantity'] for item in site_items}
            project_dict = {normalize_text(item['name']): item['quantity'] for item in purchase_list}
            discrepancies = []
            all_keys = set(site_dict.keys()) | set(project_dict.keys())
            
            for key in all_keys:
                if site_dict.get(key, 0) != project_dict.get(key, 0):
                    discrepancies.append(f"'{key}': Ø³Ø§ÛŒØª ({site_dict.get(key, 0)}) / Ù¾Ø±ÙˆÚ˜Ù‡ ({project_dict.get(key, 0)})")
            
            if not discrepancies:
                result.has_discrepancy = False
                result.message = "ØªØ·Ø§Ø¨Ù‚ Ù…ÙˆÙÙ‚: ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ ØªØ¨ Ù…Ø§Ù„ÛŒ Ø³Ø§ÛŒØª ÛŒÚ©Ø³Ø§Ù† Ø§Ø³Øª."
            else:
                result.has_discrepancy = True
                result.message = "Ù…ØºØ§ÛŒØ±Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª ÛŒØ§ÙØª Ø´Ø¯:\n" + "\n".join(discrepancies)

    except Exception as e:
        logger.error(f"Purchase validation error for project {project_id}: {e}")
        result.message = f"Ø®Ø·Ø§ Ø¯Ø± Ø¨Ø±Ø±Ø³ÛŒ ØªØ¬Ù‡ÛŒØ²Ø§Øª: {str(e)}"
    
    await ws_manager.broadcast({
        "type": "validation_result", 
        "project_id": project_id, 
        "result": result.model_dump()
    })

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û´: HTML Templates
# ==============================================================================
# Ø«Ø§Ø¨Øªâ€ŒÙ‡Ø§ÛŒ Ø±Ø´ØªÙ‡â€ŒØ§ÛŒ Ø¨Ø²Ø±Ú¯ Ø¨Ø±Ø§ÛŒ Ù‚Ø§Ù„Ø¨â€ŒÙ‡Ø§ÛŒ HTMLØŒ Ø¨Ù‡ØªØ± Ø§Ø³Øª Ø¯Ø± ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡ Ø¨Ø§Ø´Ù†Ø¯ØŒ
# Ø§Ù…Ø§ Ø·Ø¨Ù‚ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø¯Ø± Ú©Ø¯ Ø¨Ø§Ù‚ÛŒ Ù…ÛŒâ€ŒÙ…Ø§Ù†Ù†Ø¯. Ø§ÛŒÙ† Ø¨Ø®Ø´ Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ± Ø§Ø³Øª.

PROJECT_SLIP_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Ø¨Ø±Ú†Ø³Ø¨ Ù¾Ø±ÙˆÚ˜Ù‡ {{ project.name }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('{{ font_path }}'); }
        @page { size: 100mm 150mm; margin: 5mm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; width: 90mm; height: 140mm; margin: 0; padding: 0; display: flex; flex-direction: column; border: 1px solid #ccc; box-sizing: border-box; }
        header { text-align: center; padding: 4px 6px; border-bottom: 1px solid black; flex-shrink: 0; }
        header h1 { font-size: 12pt; margin: 0; }
        .details { padding: 4px 6px; line-height: 1.2; flex-shrink: 0; }
        .details p { margin: 2px 0; font-size: 11pt; }
        .request-info { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 4px; }
        .request-info div p { margin: 0; font-size: 10pt; line-height: 1.4; }
        .qr-code-inline { width: 22mm; height: 22mm; min-width: 22mm; min-height: 22mm; }
        .equipment-section { flex-grow: 1; padding: 0 6px; display: flex; flex-direction: column; overflow: hidden; }
        .equipment-header { display: flex; justify-content: space-between; align-items: center; padding: 2px 0; margin-bottom: 2px; }
        .direction-info { font-size: 10pt; font-weight: bold; border: 1px solid black; padding: 2px 6px; border-radius: 4px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #333; padding: 3px 4px; text-align: right; font-size: 8pt; word-break: break-word; }
        th { background-color: #f2f2f2; font-weight: bold; text-align: center; }
    </style>
</head>
<body>
    <header>
        <h1>Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†</h1>
        <p style="font-size: 8pt;">Ø¨Ø±Ú†Ø³Ø¨ Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ù¾Ø±ÙˆÚ˜Ù‡</p>
    </header>
    <section class="details">
        <div class="request-info">
            <div>
                <p><strong>Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª:</strong> {{ project.request_id }}</p>
                {% if project.panel_code %}<p><strong>Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ:</strong> {{ project.panel_code }}</p>{% endif %}
            </div>
            {% if qr_code_base64 %}
            <img class="qr-code-inline" src="data:image/png;base64,{{ qr_code_base64 }}" alt="QR Code">
            {% endif %}
        </div>
        <p><strong>Ù…Ø´ØªØ±ÛŒ:</strong> {{ project.customer_name }}</p>
        <p><strong>ØªØ§Ø±ÛŒØ® ØµØ¯ÙˆØ±:</strong> {{ report_date_jalali }}</p>
    </section>
    <section class="equipment-section">
        <div class="equipment-header">
            <span><strong>Ù„ÛŒØ³Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª:</strong></span>
            {% if direction %}<span class="direction-info">{{ direction }}</span>{% endif %}
        </div>
        <table>
            <thead><tr><th style="width: 10%;">Ø±Ø¯ÛŒÙ</th><th>Ù†Ø§Ù… ØªØ¬Ù‡ÛŒØ²</th><th style="width: 15%;">ØªØ¹Ø¯Ø§Ø¯</th></tr></thead>
            <tbody>
                {% for item in project.equipment[:10] %}
                <tr><td style="text-align: center;">{{ loop.index }}</td><td>{{ item.item_name }}</td><td style="text-align: center;">{{ item.quantity }}</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </section>
</body>
</html>
"""

DETAILED_EXIT_SLIP_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Ø¨Ø±Ú¯Ù‡ Ø®Ø±ÙˆØ¬ Ù¾Ø±ÙˆÚ˜Ù‡ {{ project.request_id }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('{{ font_path }}'); }
        @page { size: A5 portrait; margin: 0.7cm; }
        html, body { margin: 0; padding: 0; width: 100%; box-sizing: border-box; font-family: 'Vazirmatn', sans-serif; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { border: 1.5px solid black; padding: 0.5cm; display: flex; flex-direction: column; height: calc(210mm - 1.4cm); }
        header { flex-shrink: 0; text-align: center; margin-bottom: 5mm; border-bottom: 1px solid #666; padding-bottom: 2mm; }
        header h1 { font-size: 13pt; margin: 0; }
        header p { font-size: 8pt; margin: 1mm 0 0 0; }
        .slip-content { flex-grow: 1; overflow: hidden; font-size: 8pt; }
        .details-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0 10px; margin-bottom: 4mm; line-height: 1.3; }
        .details-grid p { margin: 1mm 0; }
        h2 { font-size: 10pt; border-bottom: 1px solid #ccc; padding-bottom: 1mm; margin-top: 0; margin-bottom: 2mm; text-align: center; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #333; padding: 1.2mm; text-align: right; vertical-align: middle; font-size: 7pt; }
        th { background-color: #f2f2f2; font-weight: bold; text-align: center; }
        .summary-table td:nth-child(2) { text-align: center; font-weight: bold; font-size: 8pt; }
        .signatures { flex-shrink: 0; margin-top: auto; padding-top: 4mm; border-top: 1px dashed #aaa; display: flex; justify-content: space-around; font-size: 8pt; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="slip-content">
        <header><h1>Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†</h1><p>Ø¨Ø±Ú¯Ù‡ Ø®Ø±ÙˆØ¬ Ùˆ ØªØ­ÙˆÛŒÙ„ ØªØ§Ø¨Ù„Ùˆ</p></header>
        <section class="details-grid">
            <p><strong>Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª:</strong> {{ project.request_id }}</p>
            <p><strong>Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ:</strong> {{ project.panel_code or '-' }}</p>
            <p><strong>Ù…Ø´ØªØ±ÛŒ:</strong> {{ project.customer_name }}</p>
            <p><strong>ØªØ§Ø±ÛŒØ® Ø®Ø±ÙˆØ¬:</strong> {{ report_date_jalali }}</p>
        </section>
        <h2>Ø¬Ø¯ÙˆÙ„ Ø®Ù„Ø§ØµÙ‡ ØªØ¬Ù‡ÛŒØ²Ø§Øª</h2>
        <table class="summary-table">
            <thead><tr><th style="width: 70%;">Ø¹Ù†ÙˆØ§Ù†</th><th style="width: 30%;">ØªØ¹Ø¯Ø§Ø¯</th></tr></thead>
            <tbody>
                {% if summary_data['Ø¸Ø±ÙÛŒØª ØªØ§Ø¨Ù„Ùˆ (ØªÚ©ÙØ§Ø²)'] is not none %}
                <tr><td style="background-color:#eef2ff;">Ø¸Ø±ÙÛŒØª ØªØ§Ø¨Ù„Ùˆ (ØªÚ©ÙØ§Ø²)</td><td style="background-color:#eef2ff;">{{ summary_data['Ø¸Ø±ÙÛŒØª ØªØ§Ø¨Ù„Ùˆ (ØªÚ©ÙØ§Ø²)'] | default(0) }}</td></tr>
                <tr><td style="background-color:#eef2ff;">Ú©Ù†ØªÙˆØ± ØªÚ©ÙØ§Ø² Ø®Ø±ÛŒØ¯Ø§Ø±ÛŒ Ø´Ø¯Ù‡</td><td style="background-color:#eef2ff;">{{ summary_data['ØªØ¹Ø¯Ø§Ø¯ Ø®Ø±ÛŒØ¯Ø§Ø±ÛŒ Ø´Ø¯Ù‡ (ØªÚ©ÙØ§Ø²)'] | default(0) }}</td></tr>
                <tr><td style="background-color:#dbeafe;color:#1e40af;">ØªØ¹Ø¯Ø§Ø¯ Ø±Ø²Ø±Ùˆ</td><td style="background-color:#dbeafe;color:#1e40af;">{{ summary_data['ØªØ¹Ø¯Ø§Ø¯ Ø±Ø²Ø±Ùˆ'] | default(0) }}</td></tr>
                {% endif %}
                <tr><td>Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²</td><td>{{ summary_data.get('Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²', 0) }}</td></tr>
                <tr><td>Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²</td><td>{{ summary_data.get('Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">ÙÛŒÙˆØ²Ù‡Ø§ÛŒ ØªÚ© ÙØ§Ø²</td></tr>
                <tr><td>ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16</td><td>{{ summary_data.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16', 0) }}</td></tr>
                <tr><td>ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25</td><td>{{ summary_data.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25', 0) }}</td></tr>
                <tr><td>ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32</td><td>{{ summary_data.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">ÙÛŒÙˆØ²Ù‡Ø§ÛŒ Ø³Ù‡ ÙØ§Ø²</td></tr>
                <tr><td>ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25</td><td>{{ summary_data.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25', 0) }}</td></tr>
                <tr><td>ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32</td><td>{{ summary_data.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32', 0) }}</td></tr>
                <tr><td>ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63</td><td>{{ summary_data.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">Ù…Ù„Ø²ÙˆÙ…Ø§Øª</td></tr>
                <tr><td>Ø³Ú©Ùˆ</td><td>{{ summary_data.get('Ø³Ú©Ùˆ', 0) }}</td></tr>
                <tr><td>ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„</td><td>{{ summary_data.get('ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0) }}</td></tr>
                <tr><td>Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„</td><td>{{ summary_data.get('Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0) }}</td></tr>
                <tr><td>Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ</td><td>{{ summary_data.get('Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ', 0) }}</td></tr>
                <tr><td>Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯</td><td>{{ summary_data.get('Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯', 0) }}</td></tr>
                <tr><td>Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°</td><td>{{ summary_data.get('Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°', 0) }}</td></tr>
            </tbody>
        </table>
    </div>
    <div class="signatures">
        <div class="signature-box">Ø§Ù…Ø¶Ø§Ø¡ ØªØ­ÙˆÛŒÙ„ Ø¯Ù‡Ù†Ø¯Ù‡<br><br>...........................</div>
        <div class="signature-box">Ø§Ù…Ø¶Ø§Ø¡ ØªØ­ÙˆÛŒÙ„ Ú¯ÛŒØ±Ù†Ø¯Ù‡<br><br>...........................</div>
    </div>
    {% if auto_print %}<script>window.onload = function() { setTimeout(function() { window.print(); }, 500); };</script>{% endif %}
</body>
</html>
"""

SUPERVISOR_APPROVAL_SIMPLE_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Ú¯Ø²Ø§Ø±Ø´ ØªØ§ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø±</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4; margin: 1.5cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 11pt; }
        .header { text-align: center; margin-bottom: 20px; }
        .header h1 { font-size: 16pt; margin: 0; }
        .header h2 { font-size: 14pt; margin: 5px 0; }
        .report-date { text-align: left; font-size: 10pt; margin-bottom: 15px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 8px; text-align: center; vertical-align: middle; }
        th { background-color: #f2f2f2; font-weight: bold; }
        .signatures { margin-top: 80px; display: flex; justify-content: space-around; font-size: 12pt; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header"><h1>Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†</h1><h2>Ù„ÛŒØ³Øª ØªØ§Ø¨Ù„Ùˆ Ù‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„</h2></div>
    <div class="report-date">ØªØ§Ø±ÛŒØ®: {{ jalali_date }}</div>
    <table>
        <thead>
            <tr>
                <th style="width: 5%;">Ø±Ø¯ÛŒÙ</th>
                <th style="width: 35%;">Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ</th>
                <th style="width: 20%;">Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§</th>
                <th style="width: 15%;">Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ</th>
                <th style="width: 15%;">ØªØ§ÛŒÛŒØ¯</th>
                <th style="width: 10%;">ØªÙˆØ¶ÛŒØ­Ø§Øª</th>
            </tr>
        </thead>
        <tbody>
            {% for p in projects %}
            <tr>
                <td>{{ loop.index }}</td>
                <td style="text-align: right;">{{ p.customer_name }}</td>
                <td>{{ p.request_id }}</td>
                <td>{{ p.panel_code or '-' }}</td>
                <td>âˆš</td>
                <td></td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ú©Ø§Ø±Ú¯Ø§Ù‡<br><br>...........................</div>
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ø¯ÙØªØ± Ù†Ø¸Ø§Ø±Øª<br><br>...........................</div>
    </div>
</body>
</html>
"""

SUPERVISOR_CHECKLIST_ITEMS = [
    "Ù…ØºØ§ÛŒØ±Øª Ø§Ø¨Ø¹Ø§Ø¯ÛŒ Ùˆ Ú©ÛŒÙÛŒØª Ø³Ø§Ø®Øª ØªØ§Ø¨Ù„Ùˆ Ø¨Ø§ Ù…Ø´Ø®ØµØ§Øª ÙÙ†ÛŒ Ùˆ Ø³ÙØ§Ø±Ø´ Ø³Ø§Ø®Øª", "Ú©ÛŒÙÛŒØª Ø±Ù†Ú¯", "Ø¢Ø³ÛŒØ¨ Ø¯ÛŒØ¯Ú¯ÛŒ Ø¸Ø§Ù‡Ø±ÛŒ Ø¨Ø¯Ù†Ù‡ Ùˆ Ø±Ù†Ú¯",
    "Ù…ØºØ§ÛŒØ±Øª Ø±Ù†Ø¬ Ú©Ù„ÛŒØ¯ Ø§ØªÙˆÙ…Ø§Øª Ø¨Ø§ Ù…Ø´Ø®ØµØ§Øª ÙÙ†ÛŒ", "Ù…ØºØ§ÛŒØ±Øª Ø±Ù†Ø¬ Ú©Ù„ÛŒØ¯ ÙÛŒÙˆØ²Ù‡Ø§ÛŒ Ù…ÛŒÙ†ÛŒØ§ØªÙˆØ±ÛŒ Ø¨Ø§ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ù…Ø´ØªØ±Ú©", "Ù…ØºØ§ÛŒØ±Øª Ø¨Ø±Ù†Ø¯ ØªØ¬Ù‡ÛŒØ²Ø§Øª Ø¨Ø§ Ùˆ Ù†Ø¯ÙˆØ± Ù„ÛŒØ³Øª",
    "Ù…ØºØ§ÛŒØ±Øª ØªØ¹Ø¯Ø§Ø¯ Ú©Ù†ØªÙˆØ±Ù‡Ø§ Ùˆ ÙÛŒÙˆØ²Ù‡Ø§ Ø¨Ø§ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ù…Ø´ØªØ±Ú©", "Ù…ØºØ§ÛŒØ±Øª Ø±Ø¯ÛŒÙ Ú©Ù†ØªÙˆØ±Ù‡Ø§ Ø¨Ø§ ØªØ³Øª Ø³Ù…Ø§Ú©", "Ø±Ø¹Ø§ÙŠØª Ø³ÙŠÙ… Ø¨Ù†Ø¯ÙŠ Ù…Ø®Ø§Ø¨Ø±Ø§ØªÙŠ ØµØ­ÙŠØ­ Ø¬Ù‡Øª ØªØ³Øª Ø³Ù…Ø§Ùƒ",
    "Ø±Ø¹Ø§ÙŠØª Ø§Ø¨Ø¹Ø§Ø¯ÙŠ Ø´ÙŠÙ†Ù‡ Ù‡Ø§ Ùˆ Ø³ÙŠÙ… Ù‡Ø§", "Ø§Ø±Ø³Ø§Ù„ ØªØ§Ø¨Ù„Ùˆ Ø¨Ø¯ÙˆÙ† Ø¨Ø±Ú†Ø³Ø¨ Ù†Ø§Ø¸Ø± Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ø¯ÙØªØ± Ù†Ø¸Ø§Ø±Øª Ø¨Ø± Ø³Ø§Ø²Ù†Ø¯Ú¯Ø§Ù†",
    "ØªØ§ÙŠÙŠØ¯ ØªØ³ØªÙ‡Ø§ÙŠ Ø§Ù„ÙƒØªØ±ÙŠÙƒØ§Ù„ ØªÙˆØ³Ø· Ù†Ø§Ø¸Ø± ÙƒÙ†ØªØ±Ù„ ÙƒÙŠÙÙŠØª Ø¯Ø± Ø²Ù…Ø§Ù† ØªØ§ÙŠÙŠØ¯ ØªØ§Ø¨Ù„Ùˆ", "Ø¨Ø±Ú¯Ø´Øª ØªØ§Ø¨Ù„Ùˆ Ø§Ø² Ù…Ø±Ø§Ú©Ø² ØªØ¬Ù…ÛŒØ¹ Ø¨Ù‡ Ø¯Ù„Ø§ÛŒÙ„ ÙÙ†ÛŒ",
    "ØªØ·Ø§Ø¨Ù‚ Ùˆ ØªÙ†Ø§Ø¸Ø± ØµÙØ­Ù‡ Ú¯Ù„Ù†Ø¯ Ø¨Ø§ Ù…Ø´Ø®ØµØ§Øª ÙÙ†ÙŠ Ùˆ ØªØ±Ù…ÙŠÙ†Ø§Ù„ Ø®Ø±ÙˆØ¬ÙŠ Ù…Ø´ØªØ±Ùƒ", "Ø±Ø¹Ø§ÙŠØª ÙØ§ØµÙ„Ù‡ ØªØ¬Ù‡ÙŠØ²Ø§Øª Ø¨Ø§ Ø¨Ø¯Ù†Ù‡ Ù…Ø·Ø§Ø¨Ù‚ Ø¨Ø§ Ù…Ø´Ø®ØµØ§Øª ÙÙ†ÙŠ",
    "Ù†ØµØ¨ Ù¾Ù„Ø§Ùƒ Ù…Ø´Ø®ØµØ§Øª Ø¯Ø± Ø¯Ø§Ø®Ù„ Ùˆ Ø®Ø§Ø±Ø¬ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§", "Ø±Ø¹Ø§ÙŠØª Ø´Ù…Ø§Ø±Ù‡ Ú¯Ø°Ø§Ø±ÙŠ Ø³Ø± Ø³ÙŠÙ… Ù†ØµØ¨ Ø¯Ø³ØªÙˆØ± Ø§Ù„Ø¹Ù…Ù„ Ùˆ Ù†Ù‚Ø´Ù‡ Ùˆ ... Ø¯Ø± ØªØ§Ø¨Ù„ÙˆÙ‡Ø§",
    "ØªÙƒÙ…ÙŠÙ„ ØªØ§Ø¨Ù„Ùˆ Ø¯Ø± Ø²Ù…Ø§Ù† Ø¨Ø§Ø²Ø¯ÙŠØ¯ Ùˆ Ù†ØµØ¨ Ø¨Ø±Ú†Ø³Ø¨ ÙƒÙ†ØªØ±Ù„ ÙƒÙŠÙÙŠØª ØªØ§Ø¨Ù„ÙˆØ³Ø§Ø²", "Ø§Ø±Ø§Ø¦Ù‡ ØªØ³Øª Ù‡Ø§ÙŠ Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§",
    "Ø§Ø±Ø³Ø§Ù„ ÙØ±Ù…Ù‡Ø§ÙŠ Ø¢Ù…Ø§Ø¯Ù‡ Ø¨Ø§Ø²Ø¯ÙŠØ¯ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ Ù…Ø·Ø§Ø¨Ù‚ Ø¨Ø§ Ø²Ù…Ø§Ù†Ø¨Ù†Ø¯ÙŠ Ù…Ø´Ø®Øµ Ø´Ø¯Ù‡", "Ù†ØµØ¨ Ù…Ù†Ø§Ø³Ø¨ Ø¢Ù†ØªÙ† ÙƒÙ†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø² Ù‡ÙˆØ´Ù…Ù†Ø¯ ÙŠØ§ Ù…ÙˆØ¯Ù…",
    "Ø±Ø¹Ø§ÙŠØª Ø¢Ø±Ø§ÙŠØ´ Ù…Ù†Ø§Ø³Ø¨ Ø³ÙŠÙ… Ø¨Ù†Ø¯ÙŠ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ Ùˆ Ø¬Ø§Ù†Ù…Ø§ÙŠÙŠ ØªØ¬Ù‡ÙŠØ²Ø§Øª", "Ø±Ø¹Ø§ÙŠØª Ù‡Ù…Ø¨Ù†Ø¯ÙŠ Ù…Ù†Ø§Ø³Ø¨", "Ø±Ú¯Ù„Ø§Ú˜ Ø¯Ø±Ø¨ Ù‡Ø§",
    "Ø§Ø±Ø³Ø§Ù„ Ù†Ø§Ù‚Øµ ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ù†ØµØ¨ ØªØ§Ø¨Ù„Ùˆ", "Ø±Ø¹Ø§ÙŠØª Ù†Ø¸Ø§ÙØª Ø¯Ø§Ø®Ù„ÙŠ Ùˆ Ø®Ø§Ø±Ø¬ÙŠ ØªØ§Ø¨Ù„Ùˆ", "Ø¨Ø³ØªÙ‡ Ø¨Ù†Ø¯ÙŠ Ù…Ù†Ø§Ø³Ø¨",
]

INDIVIDUAL_QC_CHECKLIST_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Ú†Ú© Ù„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ù¾Ø±ÙˆÚ˜Ù‡ {{ project.request_id }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4; margin: 0.5cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; }
        .header { text-align: center; margin-bottom: 15px; }
        .header h1 { font-size: 15pt; margin: 0; }
        .header h2 { font-size: 13pt; margin: 4px 0; }
        .project-info { border: 1px solid #ccc; padding: 8px; margin-bottom: 15px; border-radius: 8px; display: grid; grid-template-columns: 1fr 1fr; gap: 8px; font-size: 9pt; }
        .project-info p { margin: 0; }
        .report-date { text-align: left; font-size: 9pt; margin-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 4px 6px; text-align: center; vertical-align: middle; line-height: 1.4; }
        th { background-color: #f2f2f2; font-weight: bold; font-size: 9pt;}
        .description-col { text-align: right; width: 70%; }
        .result-col { width: 15%; }
        .signatures { margin-top: 30px; display: flex; justify-content: space-around; font-size: 11pt; page-break-inside: avoid; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header">
        <h1>Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†</h1>
        <h2>ÙØ±Ù… Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ù†Ù‡Ø§ÛŒÛŒ ØªØ§Ø¨Ù„Ùˆ</h2>
    </div>
    <div class="report-date">ØªØ§Ø±ÛŒØ®: {{ jalali_date }}</div>
    <div class="project-info">
        <p><strong>Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§:</strong> {{ project.request_id }}</p>
        <p><strong>Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ:</strong> {{ project.panel_code or '-' }}</p>
        <p><strong>Ù†Ø§Ù… Ù…Ø´ØªØ±Ú©:</strong> {{ project.customer_name }}</p>
        <p><strong>Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±Ø§Ù†:</strong> {{ project.assembler_1 }}{% if project.assembler_2 %}ØŒ {{ project.assembler_2 }}{% endif %}</p>
    </div>
    <table>
        <thead>
            <tr>
                <th style="width: 5%;">Ø±Ø¯ÛŒÙ</th>
                <th class="description-col">Ø´Ø±Ø­ Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª</th>
                <th class="result-col">Ù†ØªÛŒØ¬Ù‡</th>
                <th class="result-col">ØªÙˆØ¶ÛŒØ­Ø§Øª</th>
            </tr>
        </thead>
        <tbody>
            {% for item in checklist_items %}
            <tr>
                <td>{{ loop.index }}</td>
                <td class="description-col">{{ item }}</td>
                <td></td>
                <td></td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ù…ÙˆÙ†ØªØ§Ú˜<br><br>...........................</div>
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª<br><br>...........................</div>
    </div>
    <script>
        window.onload = function() {
            setTimeout(function() { window.print(); }, 500);
        };
    </script>
</body>
</html>
"""

SUPERVISOR_CHECKLIST_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Ú†Ú© Ù„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ø¬Ù‡Øª ØªØ§ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø±</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4 landscape; margin: 1cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; }
        .header { text-align: center; margin-bottom: 15px; }
        .header h1 { font-size: 15pt; margin: 0; }
        .header h2 { font-size: 13pt; margin: 5px 0; }
        .report-date { text-align: left; font-size: 10pt; margin-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 5px; text-align: center; vertical-align: middle; }
        th { background-color: #f2f2f2; font-weight: bold; font-size: 9pt; }
        .description-col { text-align: right; width: 35%; font-size: 9pt;}
        .project-col { min-width: 80px; font-size: 8pt; word-break: break-all; }
        tbody td:first-child { width: 3%; }
        .signatures { margin-top: 50px; display: flex; justify-content: space-around; font-size: 11pt; page-break-inside: avoid; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header"><h1>Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†</h1><h2>Ú†Ú© Ù„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„</h2></div>
    <div class="report-date">ØªØ§Ø±ÛŒØ®: {{ jalali_date }}</div>
    <table>
        <thead>
            <tr>
                <th rowspan="3" style="vertical-align: middle;">Ø±Ø¯ÛŒÙ</th>
                <th rowspan="3" style="vertical-align: middle;">Ø´Ø±Ø­ Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª</th>
                {% for p in projects %}<th class="project-col">Ø´. ØªÙ‚Ø§Ø¶Ø§</th>{% endfor %}
            </tr>
            <tr>{% for p in projects %}<th class="project-col">{{ p.request_id }}</th>{% endfor %}</tr>
            <tr>{% for p in projects %}<th class="project-col">Ú©Ø¯: {{ p.panel_code or '-' }}</th>{% endfor %}</tr>
        </thead>
        <tbody>
            {% for item in checklist_items %}
            <tr>
                <td>{{ loop.index }}</td>
                <td class="description-col">{{ item }}</td>
                {% for p in projects %}<td></td>{% endfor %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ú©Ø§Ø±Ú¯Ø§Ù‡<br><br>...........................</div>
        <div class="signature-box">Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ø¯ÙØªØ± Ù†Ø¸Ø§Ø±Øª<br><br>...........................</div>
    </div>
</body>
</html>
"""

# ==============================================================================
# Ø¨Ø®Ø´ Û±Ûµ: Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ - ØªÙ†Ø¸ÛŒÙ…Ø§Øª
# ==============================================================================
# Ø§ÛŒÙ† Ø¨Ø®Ø´ ÛŒÚ© Ú©Ø§Ù†ÙÛŒÚ¯ Ø§Ø³ØªØ§ØªÛŒÚ© Ø§Ø³Øª Ùˆ Ø¨Ù‡ Ø¯Ø±Ø³ØªÛŒ Ù¾ÛŒØ§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø´Ø¯Ù‡.

REPORT_CONFIG = {
    "headers": ["Ø¢ÛŒØªÙ… Ù‡Ø§ÛŒ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ Ù¾Ø±Ø³Ù†Ù„", "ID 2R", "ID 5R", "ID 6+1R", "ID 12+1R", "ID 18+1R"],
    "rows": [
        {"title": "Ø¢Ù…Ø§Ø¯Ù‡ Ø³Ø§Ø²ÛŒ Ø¯Ø³ØªÙ‡ Ø³ÛŒÙ…", "items": [{"label": "ID2R", "type": "number"}, {"label": "ID5R", "type": "number"}, {"label": "ID6+1R", "type": "number"}, {"label": "ID12+1R", "type": "number"}, {"label": "ID18+1R", "type": "number"}]},
        {"title": "Ø¢Ù…Ø§Ø¯Ù‡ Ø³Ø§Ø²ÛŒ Ø´Ù…Ø´", "items": [{"label": "Ø¯ÛŒÙ…Ø§Ù†Ø¯ÛŒ", "type": "number"}, {"label": "ID5R", "type": "number"}, {"label": "ID6+1R", "type": "number"}, {"label": "ID12+1R", "type": "number"}, {"label": "ID18+1R", "type": "number"}]},
        {"title": "Ø¨Ø±Ø´ Ú©Ø§Ø±ÛŒ Ùˆ Ù„Ø¨Ù‡ Ú¯ÛŒØ±", "items": [{"label": "Ø±ÛŒÙ„ ID2R", "type": "number"}, {"label": "Ø³Ú©Ùˆ ID2R", "type": "number"}, {"label": "Ø¯Ø§Ú©Øª", "type": "number"}, {"label": "Ø±ÛŒÙ„ ID5R", "type": "number"}, {"label": "Ø³Ú©Ùˆ ID5R", "type": "number"}]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Û±)", "items": [{"label": "ID101.1", "type": "number"}, {"label": "ID101.3", "type": "number"}, {"label": "ID102.1", "type": "number"}, {"label": "ID102.3", "type": "number"}, None]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Û²)", "items": [{"label": "ID104.1", "type": "number"}, {"label": "ID104.3", "type": "number"}, {"label": "ID105.1", "type": "number"}, {"label": "ID105.3", "type": "number"}, None]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Û³)", "items": [{"label": "ID109", "type": "number"}, {"label": "ID108", "type": "number"}, {"label": "ID111", "type": "number"}, {"label": "ID112", "type": "number"}, None]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Û´)", "items": [{"label": "ID115.1", "type": "number"}, {"label": "ID115.3", "type": "number"}, {"label": "ID116", "type": "number"}, {"label": "ID120", "type": "number"}, {"label": "ID218", "type": "number"}]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Ûµ)", "items": [{"label": "ID212", "type": "number"}, {"label": "ID213", "type": "number"}, {"label": "ID214", "type": "number"}, {"label": "ID215", "type": "number"}, {"label": "ID216", "type": "number"}]},
        {"title": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ø§Ù…Ù„ ØªØ§Ø¨Ù„Ùˆ (Û¶)", "items": [{"label": "ID121", "type": "number"}, {"label": "ID122", "type": "number"}, {"label": "ID123", "type": "number"}, {"label": "ID124", "type": "number"}, {"label": "ID211", "type": "number"}]},
        {
            "title": "Ù¾ÛŒØ´Ø±ÙØª Ù…ÙˆÙ†ØªØ§Ú˜ ÛŒÚ© ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø¯Ø±ØµØ¯ ID 2R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "Ø³ÛŒÙ… Ø¨Ù†Ø¯ÛŒ", "value": 50},
                {"label": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ù†ØªÙˆØ± Ùˆ ÙÛŒÙˆØ²", "value": 30},
                {"label": "Ø¢Ù…Ø§Ø¯Ù‡ Ø³Ø§Ø²ÛŒ ØµÙØ­Ù‡ Ø²ÛŒØ±", "value": 20}
            ]
        },
        {
            "title": "Ù¾ÛŒØ´Ø±ÙØª Ù…ÙˆÙ†ØªØ§Ú˜ ÛŒÚ© ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø¯Ø±ØµØ¯ ID 5R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "Ø³ÛŒÙ… Ø¨Ù†Ø¯ÛŒ", "value": 30},
                {"label": "Ù…ÙˆÙ†ØªØ§Ú˜ Ú©Ù†ØªÙˆØ± Ùˆ ÙÛŒÙˆØ²", "value": 20},
                {"label": "Ø¢Ù…Ø§Ø¯Ù‡ Ø³Ø§Ø²ÛŒ ØµÙØ­Ù‡ Ø²ÛŒØ±", "value": 50}
            ]
        },
        {
            "title": "Ù¾ÛŒØ´Ø±ÙØª Ù…ÙˆÙ†ØªØ§Ú˜ ÛŒÚ© ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø¯Ø±ØµØ¯ ID 6+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "ØµÙØ­Ù‡ Ú¯Ù„Ù†Ø¯", "value": 5},
                {"label": "ØµÙØ­Ù‡ Ú©Ù†ØªÙˆØ±", "value": 20},
                {"label": "Ø³ÛŒÙ… Ø¨Ù†Ø¯ÛŒ", "value": 55},
                {"label": "ØµÙØ­Ù‡ ÙÛŒÙˆØ²", "value": 15},
                {"label": "Ø¯Ø±Ø¨ Ù‡Ø§", "value": 5}
            ]
        },
        {
            "title": "Ù¾ÛŒØ´Ø±ÙØª Ù…ÙˆÙ†ØªØ§Ú˜ ÛŒÚ© ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø¯Ø±ØµØ¯ ID 12+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "ØµÙØ­Ù‡ Ú¯Ù„Ù†Ø¯", "value": 5},
                {"label": "ØµÙØ­Ù‡ Ú©Ù†ØªÙˆØ±", "value": 20},
                {"label": "Ø³ÛŒÙ… Ø¨Ù†Ø¯ÛŒ", "value": 55},
                {"label": "ØµÙØ­Ù‡ ÙÛŒÙˆØ²", "value": 15},
                {"label": "Ø¯Ø±Ø¨ Ù‡Ø§", "value": 5}
            ]
        },
        {
            "title": "Ù¾ÛŒØ´Ø±ÙØª Ù…ÙˆÙ†ØªØ§Ú˜ ÛŒÚ© ØªØ§Ø¨Ù„Ùˆ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø¯Ø±ØµØ¯ ID 18+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "ØµÙØ­Ù‡ Ú¯Ù„Ù†Ø¯", "value": 5},
                {"label": "ØµÙØ­Ù‡ Ú©Ù†ØªÙˆØ±", "value": 20},
                {"label": "Ø³ÛŒÙ… Ø¨Ù†Ø¯ÛŒ", "value": 55},
                {"label": "ØµÙØ­Ù‡ ÙÛŒÙˆØ²", "value": 15},
                {"label": "Ø¯Ø±Ø¨ Ù‡Ø§", "value": 5}
            ]
        },
        {"title": "Ù†Ø¸Ø§ÙØª Ùˆ Ø¬Ù…Ø¹ Ø¢ÙˆØ±ÛŒ Ø¶Ø§ÛŒØ¹Ø§Øª Ù…Ø­ÛŒØ· Ú©Ø§Ø±Ú¯Ø§Ù‡", "type": "checkbox"},
        {"title": "Ø§Ù…ÙˆØ± Ø®Ø¯Ù…Ø§ØªÛŒ", "type": "checkbox"},
        {"title": "Ø¨Ø§Ø±Ú¯ÛŒØ±ÛŒ Ùˆ ØªØ®Ù„ÛŒÙ‡ Ù„ÙˆØ§Ø²Ù…", "type": "checkbox"},
        {"title": "Ù…Ø§Ù…ÙˆØ±ÛŒØª Ø®Ø§Ø±Ø¬ Ø§Ø² Ú©Ø§Ø±Ú¯Ø§Ù‡", "type": "checkbox"},
        {"title": "Ø§Ù…ÙˆØ± Ù…ØªÙØ±Ù‚Ù‡ Ù…Ø±ØªØ¨Ø· Ø¨Ø§ Ú©Ø§Ø±Ú¯Ø§Ù‡", "type": "checkbox"},
        {"title": "Ù†Ø¸Ø§ÙØª", "type": "checkbox"},
    ]
}

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û¶: Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øªâ€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ
# ==============================================================================

# Health Check
@app.get("/health", summary="Health Check Endpoint")
def health_check():
    """Ø¨Ø±Ø±Ø³ÛŒ Ø³Ù„Ø§Ù…Øª Ø³ÛŒØ³ØªÙ…ØŒ Ø´Ø§Ù…Ù„ Ø§ØªØµØ§Ù„ Ø¨Ù‡ Ø¯ÛŒØªØ§Ø¨ÛŒØ³."""
    try:
        with SessionLocal() as db:
            db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected", "version": "2.1.0"}
    except Exception as e:
        logger.critical(f"Health check failed: Database connection error: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Service unavailable: Database connection failed."
        )

# Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª
@app.post("/auth/login", response_model=Token, summary="User Login")
def login_for_access_token(login_data: PersonnelLogin, db: Session = Depends(get_db)):
    """ÙˆØ±ÙˆØ¯ Ø¨Ù‡ Ø³ÛŒØ³ØªÙ… Ùˆ Ø¯Ø±ÛŒØ§ÙØª ØªÙˆÚ©Ù† JWT."""
    user = authenticate_user(db, login_data.username, login_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ ÛŒØ§ Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø§Ø´ØªØ¨Ø§Ù‡ Ø§Ø³Øª",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role},
        expires_delta=access_token_expires
    )
    
    logger.info(f"User '{user.username}' logged in successfully.")
    return {"access_token": access_token, "token_type": "bearer", "user": user}

@app.get("/auth/me", response_model=PersonnelOut, summary="Get Current User")
def read_users_me(current_user: Personnel = Depends(get_current_active_user)):
    """Ø¯Ø±ÛŒØ§ÙØª Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ú©Ø§Ø±Ø¨Ø± Ù„Ø§Ú¯ÛŒÙ† Ú©Ø±Ø¯Ù‡."""
    return current_user

# Ù…Ø¯ÛŒØ±ÛŒØª Ù¾Ø±Ø³Ù†Ù„
@app.post("/personnel/", response_model=PersonnelOut, status_code=status.HTTP_201_CREATED)
def create_personnel(
    personnel: PersonnelCreate, 
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±Ø³Ù†Ù„ Ø¬Ø¯ÛŒØ¯ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    if db.query(Personnel).filter(Personnel.username == personnel.username).first():
        logger.warning(f"Attempt to create personnel with existing username: {personnel.username}")
        raise HTTPException(status_code=409, detail="Ø§ÛŒÙ† Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ Ù‚Ø¨Ù„Ø§Ù‹ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    
    db_personnel = Personnel(
        name=personnel.name,
        username=personnel.username,
        password_hash=get_password_hash(personnel.password),
        role=personnel.role.value
    )
    
    try:
        db.add(db_personnel)
        db.commit()
        db.refresh(db_personnel)
        logger.info(f"Personnel '{personnel.username}' created by '{current_user.username}'.")
        return db_personnel
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Ø§ÛŒÙ† Ù†Ø§Ù… ÛŒØ§ Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ Ù‚Ø¨Ù„Ø§Ù‹ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating personnel {personnel.username}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ø³Ø±ÙˆØ± Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±Ø³Ù†Ù„.")

@app.get("/personnel/", response_model=List[PersonnelOut])
def list_personnel(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ù„ÛŒØ³Øª ØªÙ…Ø§Ù… Ù¾Ø±Ø³Ù†Ù„ ÙØ¹Ø§Ù„ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    return db.query(Personnel).filter(Personnel.is_active == 1).order_by(Personnel.name).all()

@app.put("/personnel/{personnel_id}", response_model=PersonnelOut)
def update_personnel(
    personnel_id: int, 
    personnel_data: PersonnelUpdate,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø¨Ø±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ù¾Ø±Ø³Ù†Ù„ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    db_personnel = db.query(Personnel).filter(Personnel.id == personnel_id).first()
    if not db_personnel:
        raise HTTPException(status_code=404, detail="Ù¾Ø±Ø³Ù†Ù„ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    db_personnel.name = personnel_data.name
    db_personnel.username = personnel_data.username
    db_personnel.role = personnel_data.role.value
    
    # ÙÙ‚Ø· Ø¯Ø± ØµÙˆØ±ØªÛŒ Ú©Ù‡ Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø¬Ø¯ÛŒØ¯ÛŒ Ø§Ø±Ø³Ø§Ù„ Ø´Ø¯Ù‡ Ø¨Ø§Ø´Ø¯ØŒ Ø¢Ù† Ø±Ø§ Ù‡Ø´ Ùˆ Ø°Ø®ÛŒØ±Ù‡ Ú©Ù†
    if personnel_data.password:
        db_personnel.password_hash = get_password_hash(personnel_data.password)
        
    try:
        db.commit()
        db.refresh(db_personnel)
        logger.info(f"Personnel ID {personnel_id} updated by '{current_user.username}'.")
        return db_personnel
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Ù¾Ø±Ø³Ù†Ù„ Ø¯ÛŒÚ¯Ø±ÛŒ Ø¨Ø§ Ø§ÛŒÙ† Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ ÙˆØ¬ÙˆØ¯ Ø¯Ø§Ø±Ø¯.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating personnel {personnel_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø¨Ø±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ù¾Ø±Ø³Ù†Ù„.")

@app.delete("/personnel/{personnel_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_personnel(
    personnel_id: int,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø­Ø°Ù (ØºÛŒØ±ÙØ¹Ø§Ù„ Ú©Ø±Ø¯Ù†) Ù¾Ø±Ø³Ù†Ù„ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    db_personnel = db.query(Personnel).filter(Personnel.id == personnel_id).first()
    if not db_personnel:
        raise HTTPException(status_code=404, detail="Ù¾Ø±Ø³Ù†Ù„ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    if db_personnel.id == current_user.id:
        raise HTTPException(status_code=400, detail="Ø´Ù…Ø§ Ù†Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø­Ø³Ø§Ø¨ Ú©Ø§Ø±Ø¨Ø±ÛŒ Ø®ÙˆØ¯ Ø±Ø§ Ø­Ø°Ù Ú©Ù†ÛŒØ¯.")
    
    db_personnel.is_active = 0
    try:
        db.commit()
        logger.info(f"Personnel '{db_personnel.username}' deactivated by '{current_user.username}'.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error deactivating personnel {personnel_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø­Ø°Ù Ù¾Ø±Ø³Ù†Ù„.")
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡
@app.get("/daily-reports/config")
def get_report_config():
    """Ø¯Ø±ÛŒØ§ÙØª ØªÙ†Ø¸ÛŒÙ…Ø§Øª ÙØ±Ù… Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡."""
    return REPORT_CONFIG

@app.post("/daily-reports/", response_model=DailyWorkReportOut, status_code=status.HTTP_201_CREATED)
async def create_daily_work_report(
    report: DailyWorkReportCreate, 
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø§ÛŒØ¬Ø§Ø¯ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ Ø¬Ø¯ÛŒØ¯."""
    if report.personnel_id != current_user.id:
        logger.warning(f"User '{current_user.username}' attempted to create report for another user (ID: {report.personnel_id})")
        raise HTTPException(status_code=403, detail="Ø´Ù…Ø§ ÙÙ‚Ø· Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø¨Ø±Ø§ÛŒ Ø®ÙˆØ¯ØªØ§Ù† Ú¯Ø²Ø§Ø±Ø´ Ø§ÛŒØ¬Ø§Ø¯ Ú©Ù†ÛŒØ¯.")
    
    existing_report = db.query(DailyWorkReport).filter(
        DailyWorkReport.personnel_id == current_user.id,
        DailyWorkReport.report_date == report.report_date
    ).first()
    
    if existing_report:
        raise HTTPException(status_code=409, detail="Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† ØªØ§Ø±ÛŒØ® Ù‚Ø¨Ù„Ø§Ù‹ Ú¯Ø²Ø§Ø±Ø´ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    
    db_report = DailyWorkReport(**report.model_dump())
    try:
        db.add(db_report)
        db.commit()
        db.refresh(db_report)
        
        await manager.broadcast_to_supervisors({
            "type": "new_daily_report",
            "data": {
                "report_id": db_report.id,
                "personnel_name": current_user.name,
                "report_date": report.report_date.isoformat()
            }
        })
        
        logger.info(f"Daily work report created for user '{current_user.username}' on {report.report_date}")
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† Ù…Ø¯Ù„ Ú©Ø§Ù…Ù„ Ø¨Ø§ relation Ù„ÙˆØ¯ Ø´Ø¯Ù‡
        loaded_report = db.query(DailyWorkReport).options(joinedload(DailyWorkReport.personnel)).filter(DailyWorkReport.id == db_report.id).first()
        return loaded_report
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating daily report for user '{current_user.username}': {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ú¯Ø²Ø§Ø±Ø´.")

@app.get("/daily-reports/my-reports", response_model=List[DailyWorkReportOut])
def get_my_daily_reports(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ÛŒ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ Ú©Ø§Ø±Ø¨Ø± Ø¬Ø§Ø±ÛŒ."""
    return db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.personnel_id == current_user.id)\
        .order_by(DailyWorkReport.report_date.desc(), DailyWorkReport.created_at.desc())\
        .all()

@app.get("/daily-reports/all", response_model=List[DailyWorkReportOut])
def get_all_daily_reports(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø¯Ø±ÛŒØ§ÙØª ØªÙ…Ø§Ù… Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ÛŒ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    return db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .order_by(DailyWorkReport.report_date.desc(), DailyWorkReport.created_at.desc())\
        .all()

@app.put("/daily-reports/{report_id}/approve", response_model=DailyWorkReportOut)
async def approve_daily_work_report(
    report_id: int, 
    supervisor_notes: Optional[str] = None,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """ØªØ§ÛŒÛŒØ¯ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    report = db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.id == report_id).first()
        
    if not report:
        raise HTTPException(status_code=404, detail="Ú¯Ø²Ø§Ø±Ø´ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    report.status = 'approved'
    report.supervisor_notes = supervisor_notes
    report.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(report)
        
        await manager.send_to_user(report.personnel_id, {
            "type": "report_approved",
            "data": {
                "report_id": report.id,
                "supervisor_name": current_user.name,
                "notes": supervisor_notes
            }
        })
        
        logger.info(f"Daily work report {report_id} approved by '{current_user.username}'.")
        return report
    except Exception as e:
        db.rollback()
        logger.error(f"Error approving report {report_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± ØªØ§ÛŒÛŒØ¯ Ú¯Ø²Ø§Ø±Ø´.")

@app.put("/daily-reports/{report_id}/reject", response_model=DailyWorkReportOut)
async def reject_daily_work_report(
    report_id: int, 
    supervisor_notes: str,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø±Ø¯ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    report = db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.id == report_id).first()
        
    if not report:
        raise HTTPException(status_code=404, detail="Ú¯Ø²Ø§Ø±Ø´ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    if not supervisor_notes:
        raise HTTPException(status_code=400, detail="Ù„Ø·ÙØ§Ù‹ Ø¯Ù„ÛŒÙ„ Ø±Ø¯ Ú¯Ø²Ø§Ø±Ø´ Ø±Ø§ ÙˆØ§Ø±Ø¯ Ú©Ù†ÛŒØ¯.")
    
    report.status = 'rejected'
    report.supervisor_notes = supervisor_notes
    report.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(report)
        
        await manager.send_to_user(report.personnel_id, {
            "type": "report_rejected",
            "data": {
                "report_id": report.id,
                "supervisor_name": current_user.name,
                "notes": supervisor_notes
            }
        })
        
        logger.info(f"Daily work report {report_id} rejected by '{current_user.username}'.")
        return report
    except Exception as e:
        db.rollback()
        logger.error(f"Error rejecting report {report_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø±Ø¯ Ú¯Ø²Ø§Ø±Ø´.")

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û·: Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øªâ€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ Ù¾Ø±ÙˆÚ˜Ù‡
# ==============================================================================

# Ø¢Ù¾Ù„ÙˆØ¯ Ø§Ú©Ø³Ù„
@app.post("/projects/upload-excel/", status_code=status.HTTP_202_ACCEPTED)
async def upload_projects_from_excel(
    background_tasks: BackgroundTasks, 
    file: UploadFile = File(...)
):
    """Ø¢Ù¾Ù„ÙˆØ¯ ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„ Ùˆ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡."""
    if not file.filename.endswith(('.xlsx', '.xls')): 
        raise HTTPException(status_code=400, detail="ÙØ±Ù…Øª ÙØ§ÛŒÙ„ Ù†Ø§Ù…Ø¹ØªØ¨Ø± Ø§Ø³Øª. ÙÙ‚Ø· ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ .xlsx Ùˆ .xls Ù…Ø¬Ø§Ø² Ù‡Ø³ØªÙ†Ø¯.")
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"Ø­Ø¬Ù… ÙØ§ÛŒÙ„ Ø¨ÛŒØ´ Ø§Ø² Ø­Ø¯ Ù…Ø¬Ø§Ø² ({MAX_FILE_SIZE / 1024 / 1024}MB) Ø§Ø³Øª.")
    
    await file.close()
    background_tasks.add_task(process_excel_in_background, contents, manager)
    logger.info(f"Excel file '{file.filename}' uploaded for background processing.")
    return {"message": "ÙØ§ÛŒÙ„ Ø¨Ø±Ø§ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡ Ø§Ø±Ø³Ø§Ù„ Ø´Ø¯. Ù†ØªÛŒØ¬Ù‡ Ø§Ø² Ø·Ø±ÛŒÙ‚ WebSocket Ø§Ø·Ù„Ø§Ø¹â€ŒØ±Ø³Ø§Ù†ÛŒ Ù…ÛŒâ€ŒØ´ÙˆØ¯."}

@app.post("/projects/upload-detailed-excel/", status_code=status.HTTP_202_ACCEPTED)
async def upload_detailed_excel(
    background_tasks: BackgroundTasks, 
    files: List[UploadFile] = File(...)
):
    """Ø¢Ù¾Ù„ÙˆØ¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø§Ú©Ø³Ù„ ØªÙØµÛŒÙ„ÛŒ Ùˆ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡."""
    files_contents = []
    MAX_FILE_SIZE = 10 * 1024 * 1024 # 10MB per file
    
    for file in files:
        if not file.filename.endswith(('.xlsx', '.xls')): 
            raise HTTPException(status_code=400, detail=f"ÙØ±Ù…Øª ÙØ§ÛŒÙ„ '{file.filename}' Ù†Ø§Ù…Ø¹ØªØ¨Ø± Ø§Ø³Øª.")
        
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail=f"Ø­Ø¬Ù… ÙØ§ÛŒÙ„ '{file.filename}' Ø¨ÛŒØ´ Ø§Ø² Ø­Ø¯ Ù…Ø¬Ø§Ø² Ø§Ø³Øª.")
        
        await file.close()
        files_contents.append((file.filename, contents))
    
    background_tasks.add_task(process_detailed_excel_in_background, files_contents, manager)
    logger.info(f"{len(files_contents)} detailed Excel files uploaded for background processing.")
    return {"message": f"{len(files_contents)} ÙØ§ÛŒÙ„ Ø¨Ø±Ø§ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡ Ø§Ø±Ø³Ø§Ù„ Ø´Ø¯."}

# Ù…Ø¯ÛŒØ±ÛŒØª Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§
@app.post("/projects/", response_model=ProjectOut, status_code=status.HTTP_201_CREATED)
async def create_project_manual(
    project_in: ProjectCreateManual, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor) 
):
    """Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ù‡ ØµÙˆØ±Øª Ø¯Ø³ØªÛŒ ØªÙˆØ³Ø· Ù†Ø§Ø¸Ø±ØŒ Ø¨Ø§ Ú©Ø³Ø± Ø®ÙˆØ¯Ú©Ø§Ø± Ø§Ø² Ø§Ù†Ø¨Ø§Ø±."""
    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ request_id Ø§Ù„Ø²Ø§Ù…ÛŒ Ø´Ø¯Ù‡ Ø§Ø³Øª
    if not project_in.request_id:
        raise HTTPException(status_code=422, detail="Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø§Ù„Ø²Ø§Ù…ÛŒ Ø§Ø³Øª.")
        
    if db.query(Project.id).filter(Project.request_id == project_in.request_id).first():
        raise HTTPException(status_code=409, detail=f"Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª '{project_in.request_id}' Ù…ÙˆØ¬ÙˆØ¯ Ø§Ø³Øª.")
    
    db_proj = Project(
        name=project_in.name,
        location=project_in.location,
        customer_name=project_in.customer_name or "Ù†Ø§Ù…Ø´Ø®Øµ",
        request_id=project_in.request_id,
        barcode_payload=project_in.request_id
    )
    
    try:
        db.add(db_proj)
        db.commit()
        db.refresh(db_proj)
        
        # Ú©Ø³Ø± Ø®ÙˆØ¯Ú©Ø§Ø± Ø§Ø² Ø§Ù†Ø¨Ø§Ø± Ù¾Ø³ Ø§Ø² Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±ÙˆÚ˜Ù‡ (Ø§Ú¯Ø± BOM ØªØ¹Ø±ÛŒÙ Ø´Ø¯Ù‡ Ø¨Ø§Ø´Ø¯)
        await deduct_stock_for_project(db, db_proj, current_user)
        
        await manager.broadcast({"type": "update", "source": "new_manual_project", "project_id": db_proj.id})
        logger.info(f"Manual project '{project_in.request_id}' created by '{current_user.username}'.")
        return convert_project_orm_to_pydantic(db_proj)
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating manual project: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ø³Ø±ÙˆØ± Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±ÙˆÚ˜Ù‡.")

@app.get("/projects/", response_model=List[ProjectOut])
def list_projects(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """Ù„ÛŒØ³Øª Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª ÙÛŒÙ„ØªØ± Ø¨Ø± Ø§Ø³Ø§Ø³ ØªØ§Ø±ÛŒØ® Ø§ÛŒØ¬Ø§Ø¯."""
    query = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment), 
        joinedload(Project.comments)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects_orm = query.order_by(Project.created_at.desc()).all()
    return [convert_project_orm_to_pydantic(p) for p in projects_orm]

@app.get("/projects/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, db: Session = Depends(get_db)):
    """Ø¯Ø±ÛŒØ§ÙØª Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ú©Ø§Ù…Ù„ ÛŒÚ© Ù¾Ø±ÙˆÚ˜Ù‡."""
    p = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment), 
        joinedload(Project.comments)
    ).filter(Project.id == project_id).first()
    
    if not p: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯")
    return convert_project_orm_to_pydantic(p)

@app.delete("/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_project(
    project_id: int, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø­Ø°Ù Ù¾Ø±ÙˆÚ˜Ù‡ (ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ù†Ø§Ø¸Ø±Ø§Ù†)."""
    
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯")
    
    try:
        db.delete(p)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø­Ø°Ù Ù¾Ø±ÙˆÚ˜Ù‡ Ø§Ø² Ø¯ÛŒØªØ§Ø¨ÛŒØ³.")
    
    # Ø§Ø±Ø³Ø§Ù„ Ù¾ÛŒØ§Ù… ÙˆØ¨â€ŒØ³ÙˆÚ©ØªØŒ Ø­ØªÛŒ Ø§Ú¯Ø± Ø®Ø·Ø§ Ø¯Ù‡Ø¯ Ø­Ø°Ù Ø§Ù†Ø¬Ø§Ù… Ù…ÛŒâ€ŒØ´ÙˆØ¯
    try:
        await manager.broadcast({
            "type": "delete_project",
            "data": {"project_id": project_id}
        })
    except Exception as ws_err:
        logger.error(f"WebSocket broadcast failed: {ws_err}")

    logger.info(f"Project ID {project_id} deleted by '{current_user.username}'.")
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# Ú©Ø§Ù…Ù†Øªâ€ŒÙ‡Ø§
@app.post("/projects/{project_id}/comments", response_model=CommentOut, status_code=status.HTTP_201_CREATED)
async def create_comment_for_project(
    project_id: int, 
    comment_in: CommentCreate, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee) # âœ… Ù†ÛŒØ§Ø² Ø¨Ù‡ Ù„Ø§Ú¯ÛŒÙ†
):
    """Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ù…Ù†Øª Ø¨Ø±Ø§ÛŒ ÛŒÚ© Ù¾Ø±ÙˆÚ˜Ù‡."""
    project = db.query(Project).get(project_id)
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯")
    
    new_comment = Comment(
        project_id=project_id, 
        text=comment_in.text, 
        author=current_user.name # âœ… Ù†ÙˆÛŒØ³Ù†Ø¯Ù‡ Ú©Ø§Ù…Ù†ØªØŒ Ú©Ø§Ø±Ø¨Ø± Ù„Ø§Ú¯ÛŒÙ† Ú©Ø±Ø¯Ù‡ Ø§Ø³Øª
    )
    
    try:
        db.add(new_comment)
        db.commit()
        db.refresh(new_comment)
        
        await manager.broadcast({
            "type": "update", 
            "source": "new_comment",
            "project_id": project_id
        })
        logger.info(f"Comment added to project {project_id} by '{current_user.username}'.")
        return new_comment
    except Exception as e:
        db.rollback()
        logger.error(f"Error adding comment to project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Ú©Ø§Ù…Ù†Øª.")

# Ù…Ø±Ø§Ø­Ù„ Ù¾Ø±ÙˆÚ˜Ù‡
@app.post("/projects/{project_id}/steps", response_model=StepOut, status_code=status.HTTP_201_CREATED)
async def add_step_to_project(
    project_id: int, 
    step_in: StepCreate, 
    db: Session = Depends(get_db)
):
    """Ø§ÙØ²ÙˆØ¯Ù† ÛŒÚ© Ù…Ø±Ø­Ù„Ù‡ Ø¬Ø¯ÛŒØ¯ Ø¨Ù‡ Ù¾Ø±ÙˆÚ˜Ù‡."""
    project_orm = db.query(Project).options(joinedload(Project.steps)).filter(Project.id == project_id).first()
    if not project_orm: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯")
    
    step_to_add = step_in.step
    if step_to_add == StepNameKey.EXIT_PANEL: 
        raise HTTPException(status_code=403, detail="Ø®Ø±ÙˆØ¬ ØªØ§Ø¨Ù„Ùˆ ÙÙ‚Ø· Ø¨Ø§ Ø§Ø³Ú©Ù† Ø¨Ø§Ø±Ú©Ø¯ Ø«Ø¨Øª Ù…ÛŒâ€ŒØ´ÙˆØ¯.")
    
    if db.query(Step.id).filter(Step.project_id == project_id, Step.name_key == step_to_add).first(): 
        raise HTTPException(status_code=409, detail="Ø§ÛŒÙ† Ù…Ø±Ø­Ù„Ù‡ Ù‚Ø¨Ù„Ø§Ù‹ Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† Ù¾Ø±ÙˆÚ˜Ù‡ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    
    completed_keys = {s.name_key for s in project_orm.steps}
    try:
        current_idx = ORDERED_MANUAL_STEP_KEYS.index(step_to_add)
        if current_idx > 0:
            previous_step_key = ORDERED_MANUAL_STEP_KEYS[current_idx - 1]
            if previous_step_key not in completed_keys: 
                raise HTTPException(
                    status_code=412, 
                    detail=f"Ù…Ø±Ø­Ù„Ù‡ Ù¾ÛŒØ´â€ŒÙ†ÛŒØ§Ø² '{STEP_KEY_TO_NAME_MAP.get(previous_step_key.value)}' Ø§Ù†Ø¬Ø§Ù… Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª."
                )
    except ValueError: 
        raise HTTPException(status_code=400, detail=f"Ù…Ø±Ø­Ù„Ù‡ '{step_to_add.value}' ÛŒÚ© Ù…Ø±Ø­Ù„Ù‡ Ø¯Ø³ØªÛŒ Ù…Ø¬Ø§Ø² Ù†ÛŒØ³Øª.")
    
    new_s = Step(project_id=project_id, name_key=step_to_add)
    
    try:
        db.add(new_s)
        db.commit()
        db.refresh(new_s)
        
        await manager.broadcast({"type": "update", "project_id": project_id})
        logger.info(f"Step '{step_to_add.value}' added to project {project_id}.")
        return StepOut.model_validate(new_s)
    except Exception as e:
        db.rollback()
        logger.error(f"Error adding step to project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø§ÙØ²ÙˆØ¯Ù† Ù…Ø±Ø­Ù„Ù‡.")

@app.delete("/projects/{project_id}/steps/{step_name}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_step_from_project(
    project_id: int, 
    step_name: str,  # ØªØºÛŒÛŒØ± Ø§Ø² StepNameKey Ø¨Ù‡ string
    db: Session = Depends(get_db)
):
    """Ø­Ø°Ù ÛŒÚ© Ù…Ø±Ø­Ù„Ù‡ Ø§Ø² Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ Ù¾Ø´ØªÛŒØ¨Ø§Ù†ÛŒ Ø§Ø² Ù†Ø§Ù… ÙØ§Ø±Ø³ÛŒ"""
    
    # Ù†Ú¯Ø§Ø´Øª Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ ÙØ§Ø±Ø³ÛŒ Ø¨Ù‡ Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ enum
    persian_to_enum_map = {
        "Ø´Ø±ÙˆØ¹ Ù…ÙˆÙ†ØªØ§Ú˜": StepNameKey.START_ASSEMBLY,
        "Ù¾Ø§ÛŒØ§Ù† Ù…ÙˆÙ†ØªØ§Ú˜": StepNameKey.END_ASSEMBLY,
        "ØªØ§ÛŒÛŒØ¯ Ø³Ø±Ú¯Ø±ÙˆÙ‡": StepNameKey.TEAM_LEAD_APPROVAL,
        "ØªØ³Øª Ø³Ù…Ø§Ú©": StepNameKey.TEST,
        "Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª": StepNameKey.QUALITY_CONTROL,
        "ØªØ£ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø±": StepNameKey.SUPERVISOR_APPROVAL,
        "Ø®Ø±ÙˆØ¬ ØªØ§Ø¨Ù„Ùˆ": StepNameKey.EXIT_PANEL,
    }
    
    # Ù‡Ù…Ú†Ù†ÛŒÙ† Ù¾Ø´ØªÛŒØ¨Ø§Ù†ÛŒ Ø§Ø² Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ enum
    enum_to_enum_map = {
        "START_ASSEMBLY": StepNameKey.START_ASSEMBLY,
        "END_ASSEMBLY": StepNameKey.END_ASSEMBLY,
        "TEAM_LEAD_APPROVAL": StepNameKey.TEAM_LEAD_APPROVAL,
        "TEST": StepNameKey.TEST,
        "QUALITY_CONTROL": StepNameKey.QUALITY_CONTROL,
        "SUPERVISOR_APPROVAL": StepNameKey.SUPERVISOR_APPROVAL,
        "EXIT_PANEL": StepNameKey.EXIT_PANEL,
    }
    
    # Ù¾ÛŒØ¯Ø§ Ú©Ø±Ø¯Ù† Ú©Ù„ÛŒØ¯ enum
    step_enum = None
    if step_name in persian_to_enum_map:
        step_enum = persian_to_enum_map[step_name]
    elif step_name.upper() in enum_to_enum_map:
        step_enum = enum_to_enum_map[step_name.upper()]
    else:
        # ØªÙ„Ø§Ø´ Ø¨Ø±Ø§ÛŒ ØªØ¨Ø¯ÛŒÙ„ Ù…Ø³ØªÙ‚ÛŒÙ…
        try:
            step_enum = StepNameKey(step_name.upper())
        except ValueError:
            raise HTTPException(
                status_code=422, 
                detail=f"Ù†Ø§Ù… Ù…Ø±Ø­Ù„Ù‡ '{step_name}' Ù†Ø§Ù…Ø¹ØªØ¨Ø± Ø§Ø³Øª. Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ Ù…Ø¬Ø§Ø²: {', '.join(persian_to_enum_map.keys())}"
            )
    
    # Ø¨Ù‚ÛŒÙ‡ Ú©Ø¯ Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ±...
    project_orm = db.query(Project).options(joinedload(Project.steps)).filter(Project.id == project_id).first()
    if not project_orm: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
        
    step_to_delete = db.query(Step).filter(
        Step.project_id == project_id, 
        Step.name_key == step_enum  # Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² enum
    ).first()
    
    if not step_to_delete: 
        raise HTTPException(
            status_code=404, 
            detail=f"Ù…Ø±Ø­Ù„Ù‡ '{STEP_KEY_TO_NAME_MAP.get(step_enum.value)}' Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯."
        )
        
    # Ø§Ø¯Ø§Ù…Ù‡ Ù…Ù†Ø·Ù‚ Ø¨Ø±Ø±Ø³ÛŒ Ù¾ÛŒØ´â€ŒÙ†ÛŒØ§Ø²Ù‡Ø§...
    completed_keys = {s.name_key for s in project_orm.steps}
    try:
        all_ordered_keys = ORDERED_MANUAL_STEP_KEYS + [StepNameKey.EXIT_PANEL]
        current_idx = all_ordered_keys.index(step_enum)
        if current_idx < len(all_ordered_keys) - 1:
            next_step_key = all_ordered_keys[current_idx + 1]
            if next_step_key in completed_keys: 
                raise HTTPException(
                    status_code=409, 
                    detail=f"Ø§Ø¨ØªØ¯Ø§ Ø¨Ø§ÛŒØ¯ Ù…Ø±Ø­Ù„Ù‡ Ø¨Ø¹Ø¯ÛŒ ('{STEP_KEY_TO_NAME_MAP.get(next_step_key.value)}') Ø±Ø§ Ù„ØºÙˆ Ú©Ù†ÛŒØ¯."
                )
    except ValueError: 
        pass 
        
    try:
        db.delete(step_to_delete)
        db.commit()
        
        await manager.broadcast({"type": "update", "project_id": project_id})
        logger.info(f"Step '{step_enum.value}' deleted from project {project_id}.")
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting step from project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø­Ø°Ù Ù…Ø±Ø­Ù„Ù‡.")
# Ø®Ø±ÙˆØ¬ Ø¨Ø§ Ø¨Ø§Ø±Ú©Ø¯
@app.post("/projects/exit-by-barcode/", response_model=StepOut, status_code=status.HTTP_201_CREATED)
async def exit_project_by_barcode(payload: BarcodeExitPayload, db: Session = Depends(get_db)):
    """Ø«Ø¨Øª Ø®Ø±ÙˆØ¬ ØªØ§Ø¨Ù„Ùˆ Ø¨Ø§ Ø§Ø³Ú©Ù† Ø¨Ø§Ø±Ú©Ø¯ (Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª)."""
    barcode_data = payload.barcode_data
    proj = db.query(Project).options(joinedload(Project.steps)).filter(Project.request_id == barcode_data).first()
    if not proj: 
        raise HTTPException(status_code=404, detail=f"Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª '{barcode_data}' ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    completed_step_keys = {s.name_key for s in proj.steps}
    required_step_keys = set(ORDERED_MANUAL_STEP_KEYS)
    missing_prerequisites = required_step_keys - completed_step_keys
    
    if missing_prerequisites:
        missing_names = sorted(
            [STEP_KEY_TO_NAME_MAP.get(s.value) for s in missing_prerequisites], 
            key=lambda x: [STEP_KEY_TO_NAME_MAP.get(step.value) for step in ORDERED_MANUAL_STEP_KEYS].index(x)
        )
        raise HTTPException(
            status_code=400, 
            detail=f"ØªÙ…Ø§Ù… Ù¾ÛŒØ´â€ŒÙ†ÛŒØ§Ø²Ù‡Ø§ Ø§Ù†Ø¬Ø§Ù… Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª. Ù…Ø±Ø§Ø­Ù„ Ø¨Ø§Ù‚ÛŒâ€ŒÙ…Ø§Ù†Ø¯Ù‡: {', '.join(missing_names)}"
        )
    
    if StepNameKey.EXIT_PANEL in completed_step_keys: 
        raise HTTPException(status_code=409, detail="Ø®Ø±ÙˆØ¬ Ù‚Ø¨Ù„Ø§Ù‹ Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† Ù¾Ø±ÙˆÚ˜Ù‡ Ø«Ø¨Øª Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    
    exit_s = Step(project_id=proj.id, name_key=StepNameKey.EXIT_PANEL)
    
    try:
        db.add(exit_s)
        db.commit()
        db.refresh(exit_s)
        
        await manager.broadcast({"type": "update", "project_id": proj.id})
        logger.info(f"Project {proj.id} exited by barcode: {barcode_data}")
        return StepOut.model_validate(exit_s)
    except Exception as e:
        db.rollback()
        logger.error(f"Error exiting project by barcode {barcode_data}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø«Ø¨Øª Ø®Ø±ÙˆØ¬.")
# Ø¬Ø²Ø¦ÛŒØ§Øª Ù…ÙˆÙ†ØªØ§Ú˜
@app.put("/projects/{project_id}/assembly-details/", response_model=ProjectOut)
async def update_assembly_details(
    project_id: int, 
    details: AssemblyDetailsUpdate, 
    db: Session = Depends(get_db)
):
    """Ø¨Ø±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ø¬Ø²Ø¦ÛŒØ§Øª Ù…ÙˆÙ†ØªØ§Ú˜ (Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ Ùˆ Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±Ø§Ù†)."""
    logger.info(f"Updating assembly details for project {project_id}")
    logger.info(f"Received data: {details.model_dump()}")
    
    # Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø¯Ø³ØªÛŒ panel_type_key
    panel_key = details.panel_type_key.strip()
    
    # âœ… Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡: Ù†Ú¯Ø§Ø´Øª ÙˆØ±ÙˆØ¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø§Ø­ØªÙ…Ø§Ù„ÛŒ (Ù…Ø«Ù„Ø§Ù‹ Ø¨Ø§ +) Ø¨Ù‡ Ù…Ù‚Ø§Ø¯ÛŒØ± ØµØ­ÛŒØ­ Enum (Ø¨Ø§ _)
    key_mapping = {
        # ÙˆØ±ÙˆØ¯ÛŒ Ø§Ø´ØªØ¨Ø§Ù‡ -> Ù…Ù‚Ø¯Ø§Ø± ØµØ­ÛŒØ­ Ø¯Ø± Enum
        "ID12+1R": "ID12_1R",
        "ID18+1R": "ID18_1R", 
        "ID24+1R": "ID24_1R",
        "ID6+1R":  "ID6_1R",
        
        # ÙˆØ±ÙˆØ¯ÛŒâ€ŒÙ‡Ø§ÛŒ ÙØ´Ø±Ø¯Ù‡ -> Ù…Ù‚Ø¯Ø§Ø± ØµØ­ÛŒØ­ Ø¯Ø± Enum
        "ID121R": "ID12_1R",
        "ID181R": "ID18_1R", 
        "ID241R": "ID24_1R",
        "ID61R":  "ID6_1R",

        # Ù…Ù‚Ø§Ø¯ÛŒØ±ÛŒ Ú©Ù‡ Ø¯Ø± Ú©Ø¯ Ø´Ù…Ø§ Ø¨Ø±Ø¹Ú©Ø³ Ø¨ÙˆØ¯Ù†Ø¯ Ùˆ Ø¨Ø§Ø¹Ø« Ø®Ø·Ø§ Ù…ÛŒâ€ŒØ´Ø¯Ù†Ø¯ Ø±Ø§ Ø­Ø°Ù ÛŒØ§ Ù…Ø¹Ú©ÙˆØ³ Ú©Ø±Ø¯ÛŒÙ…
        # Ø§Ú¯Ø± ÙˆØ±ÙˆØ¯ÛŒ Ø®ÙˆØ¯Ø´ ID6_1R Ø¨Ø§Ø´Ø¯ØŒ Ú†ÙˆÙ† Ø¯Ø± Ø§ÛŒÙ† Ø¯ÛŒÚ©Ø´Ù†Ø±ÛŒ Ù†ÛŒØ³ØªØŒ Ø®ÙˆØ¯Ø´ Ø¨Ø§Ù‚ÛŒ Ù…ÛŒâ€ŒÙ…Ø§Ù†Ø¯ Ú©Ù‡ ØµØ­ÛŒØ­ Ø§Ø³Øª.
    }
    
    # ØªØ¨Ø¯ÛŒÙ„ Ø¨Ù‡ Ù…Ù‚Ø¯Ø§Ø± ØµØ­ÛŒØ­
    final_panel_key = key_mapping.get(panel_key, panel_key)
    
    # Ø¨Ø±Ø±Ø³ÛŒ ÙˆØ¬ÙˆØ¯ Ø¯Ø± enum
    valid_keys = [e.value for e in PanelTypeKey]
    if final_panel_key not in valid_keys:
        # Ø¨Ø±Ø§ÛŒ Ù„Ø§Ú¯ Ø¨Ù‡ØªØ±ØŒ Ù…Ù‚Ø¯Ø§Ø± ØªØ¨Ø¯ÛŒÙ„ Ø´Ø¯Ù‡ Ø±Ø§ Ù‡Ù… Ú†Ø§Ù¾ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
        logger.error(f"Invalid panel key provided: '{panel_key}' -> Converted to: '{final_panel_key}'")
        raise HTTPException(
            status_code=422,
            detail=f"'{panel_key}' ÛŒÚ© Ú©Ù„ÛŒØ¯ Ù…Ø¹ØªØ¨Ø± Ø¨Ø±Ø§ÛŒ Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ Ù†ÛŒØ³Øª. Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ù…Ø¬Ø§Ø²: {', '.join(valid_keys)}"
        )
    
    p_orm = db.query(Project).get(project_id)
    if not p_orm: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    if p_orm.steps: 
        raise HTTPException(status_code=403, detail="Ø§Ù…Ú©Ø§Ù† ØªØºÛŒÛŒØ± Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ù…ÙˆÙ†ØªØ§Ú˜ Ù¾Ø³ Ø§Ø² Ø´Ø±ÙˆØ¹ Ù…Ø±Ø§Ø­Ù„ ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯.")
    
    p_orm.panel_type_key = final_panel_key
    p_orm.panel_code = final_panel_key
    p_orm.assembler_1 = details.assembler_1.strip()
    p_orm.assembler_2 = details.assembler_2.strip() if details.assembler_2 else None
    
    try:
        db.commit()
        db.refresh(p_orm)
        
        await manager.broadcast({"type": "update", "project_id": p_orm.id})
        logger.info(f"Assembly details updated successfully for project {project_id}")
        return convert_project_orm_to_pydantic(p_orm)
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating assembly details for project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ø¨Ø±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ø¬Ø²Ø¦ÛŒØ§Øª Ù…ÙˆÙ†ØªØ§Ú˜.")

# Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ
@app.post("/projects/{project_id}/validate-branches", status_code=status.HTTP_202_ACCEPTED)
async def validate_branches_endpoint(
    project_id: int, 
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db)
):
    """Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø§Ù†Ø´Ø¹Ø§Ø¨Ø§Øª Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø§ Ø³Ø±ÙˆÛŒØ³ Ø®Ø§Ø±Ø¬ÛŒ (Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡)."""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    purchase_list = [{"name": eq.item_name, "quantity": eq.quantity} for eq in project.equipment]
    background_tasks.add_task(run_branch_validation, project.id, project.request_id, purchase_list, manager)
    logger.info(f"Branch validation requested for project {project_id}")
    return {"message": "Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø¨Ø±Ø±Ø³ÛŒ Ù…ØºØ§ÛŒØ±Øª Ø§Ù†Ø´Ø¹Ø§Ø¨ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡ Ø«Ø¨Øª Ø´Ø¯."}

@app.post("/projects/{project_id}/validate-purchases", status_code=status.HTTP_202_ACCEPTED)
async def validate_purchases_endpoint(
    project_id: int, 
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db)
):
    """Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ø®Ø±ÛŒØ¯Ù‡Ø§ Ø¨Ø§ Ø³Ø±ÙˆÛŒØ³ Ø®Ø§Ø±Ø¬ÛŒ (Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡)."""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")

    purchase_list = [{"name": eq.item_name, "quantity": eq.quantity} for eq in project.equipment]
    background_tasks.add_task(run_purchase_validation, project.id, project.request_id, purchase_list, manager)
    logger.info(f"Purchase validation requested for project {project_id}")
    return {"message": "Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø¨Ø±Ø±Ø³ÛŒ Ù…ØºØ§ÛŒØ±Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡ Ø«Ø¨Øª Ø´Ø¯."}

# Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…Ø¬Ø¯Ø¯ Ú©Ø¯Ù‡Ø§ÛŒ ØªØ§Ø¨Ù„Ùˆ
@app.post("/projects/recalculate-all-panel-codes/", status_code=200)
async def recalculate_all_panel_codes(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…Ø¬Ø¯Ø¯ Ú©Ø¯Ù‡Ø§ÛŒ ØªØ§Ø¨Ù„Ùˆ Ø¨Ø±Ø§ÛŒ ØªÙ…Ø§Ù… Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ Ø¨Ø± Ø§Ø³Ø§Ø³ ØªØ¬Ù‡ÛŒØ²Ø§Øª ÙØ¹Ù„ÛŒ."""
    all_projects = db.query(Project).options(joinedload(Project.equipment)).all()
    updated_count = 0
    
    for project in all_projects:
        equipment_pydantic = [
            EquipmentItemBase(item_name=eq.item_name, quantity=eq.quantity) 
            for eq in project.equipment
        ]
        panel_code, panel_type_key = _find_panel_details_from_equipment(equipment_pydantic)
        
        if project.panel_code != panel_code or project.panel_type_key != panel_type_key:
            project.panel_code = panel_code
            project.panel_type_key = panel_type_key
            updated_count += 1
    
    if updated_count > 0:
        try:
            db.commit()
            await manager.broadcast({"type": "update", "source": "recalculation"})
            logger.info(f"Panel codes recalculated by '{current_user.username}': {updated_count} projects updated.")
        except Exception as e:
            db.rollback()
            logger.error(f"Error recalculating panel codes: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ø®Ø·Ø§ Ø¯Ø± Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…Ø¬Ø¯Ø¯ Ú©Ø¯Ù‡Ø§ÛŒ ØªØ§Ø¨Ù„Ùˆ.")
    
    return {"message": f"Ø¹Ù…Ù„ÛŒØ§Øª Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯. {updated_count} Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø±ÙˆØ²Ø±Ø³Ø§Ù†ÛŒ Ø´Ø¯."}

# Ø¢Ù…Ø§Ø± Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±Ø§Ù†
@app.get("/reports/assembler-stats/", response_model=Dict[str, AssemblerStatsOut])
def get_assembler_stats(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """Ø¯Ø±ÛŒØ§ÙØª Ø¢Ù…Ø§Ø± Ø¹Ù…Ù„Ú©Ø±Ø¯ Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±Ø§Ù†."""
    query = db.query(Project).filter(
        Project.assembler_1.isnot(None), 
        Project.panel_type_key.isnot(None)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects = query.all()
    stats = defaultdict(lambda: {"total_panels": 0, "panels_by_type": defaultdict(int)})
    
    for p in projects:
        if not p.panel_type_key or not isinstance(p.panel_type_key, str): 
            continue
        
        panel_key = re.sub(r'\s+', '', p.panel_type_key).strip().upper()
        if not panel_key: 
            continue
        
        for assembler_name in [p.assembler_1, p.assembler_2]:
            if assembler_name and assembler_name.strip():
                clean_name = assembler_name.strip()
                stats[clean_name]["panels_by_type"][panel_key] += 1
    
    final_stats = {}
    for assembler, data in stats.items():
        total = sum(data["panels_by_type"].values())
        final_stats[assembler] = {
            "total_panels": total, 
            "panels_by_type": dict(data["panels_by_type"])
        }
    
    return final_stats
    
# ==============================================================================
# Ø¨Ø®Ø´ Û±Û·.Ûµ: Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øªâ€ŒÙ‡Ø§ÛŒ Ø§Ù†Ø¨Ø§Ø±Ø¯Ø§Ø±ÛŒ
# ==============================================================================

# --- Ù…Ø¯ÛŒØ±ÛŒØª Ø§Ù†Ø¨Ø§Ø±Ù‡Ø§ ---
@app.post("/warehouses/", response_model=WarehouseOut, status_code=status.HTTP_201_CREATED)
def create_warehouse(
    warehouse: WarehouseCreate,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """Ø§ÛŒØ¬Ø§Ø¯ ÛŒÚ© Ø§Ù†Ø¨Ø§Ø± Ø¬Ø¯ÛŒØ¯."""
    db_warehouse = Warehouse(**warehouse.model_dump())
    db.add(db_warehouse)
    db.commit()
    db.refresh(db_warehouse)
    logger.info(f"Warehouse '{warehouse.name}' created by '{current_user.username}'.")
    return db_warehouse

@app.get("/warehouses/", response_model=List[WarehouseOut])
def list_warehouses(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª Ù„ÛŒØ³Øª ØªÙ…Ø§Ù… Ø§Ù†Ø¨Ø§Ø±Ù‡Ø§."""
    return db.query(Warehouse).all()

# --- Ù…Ø¯ÛŒØ±ÛŒØª Ú©Ø§Ù„Ø§Ù‡Ø§ Ø¯Ø± Ø§Ù†Ø¨Ø§Ø± ---
@app.post("/warehouse-items/", response_model=WarehouseItemOut, status_code=status.HTTP_201_CREATED)
def create_warehouse_item(
    item: WarehouseItemCreate,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """ØªØ¹Ø±ÛŒÙ ÛŒÚ© Ú©Ø§Ù„Ø§ÛŒ Ø¬Ø¯ÛŒØ¯ Ø¯Ø± Ø³ÛŒØ³ØªÙ… Ø§Ù†Ø¨Ø§Ø±."""
    if db.query(WarehouseItem).filter(WarehouseItem.item_name == item.item_name).first():
        raise HTTPException(status_code=409, detail="Ø§ÛŒÙ† Ú©Ø§Ù„Ø§ Ù‚Ø¨Ù„Ø§Ù‹ ØªØ¹Ø±ÛŒÙ Ø´Ø¯Ù‡ Ø§Ø³Øª.")
    db_item = WarehouseItem(**item.model_dump())
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    logger.info(f"Warehouse item '{item.item_name}' created by '{current_user.username}'.")
    return db_item

@app.get("/warehouse-items/", response_model=List[WarehouseItemOut])
def list_warehouse_items(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª Ù„ÛŒØ³Øª ØªÙ…Ø§Ù… Ú©Ø§Ù„Ø§Ù‡Ø§ÛŒ ØªØ¹Ø±ÛŒÙâ€ŒØ´Ø¯Ù‡."""
    return db.query(WarehouseItem).order_by(WarehouseItem.item_name).all()
    
# --- Ø«Ø¨Øª ØªØ±Ø§Ú©Ù†Ø´â€ŒÙ‡Ø§ ---
@app.post("/inventory/in/", response_model=InventoryTransactionOut)
async def log_inventory_in(
    transaction_in: InventoryTransactionIn,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø«Ø¨Øª ÙˆØ±ÙˆØ¯ Ú©Ø§Ù„Ø§ Ø¨Ù‡ Ø§Ù†Ø¨Ø§Ø±."""
    item = db.query(WarehouseItem).filter(WarehouseItem.item_name == transaction_in.item_name).first()
    if not item:
        raise HTTPException(status_code=404, detail=f"Ú©Ø§Ù„Ø§ÛŒÛŒ Ø¨Ø§ Ù†Ø§Ù… '{transaction_in.item_name}' ÛŒØ§ÙØª Ù†Ø´Ø¯. Ø§Ø¨ØªØ¯Ø§ Ø¢Ù† Ø±Ø§ ØªØ¹Ø±ÛŒÙ Ú©Ù†ÛŒØ¯.")

    transaction = InventoryTransaction(
        warehouse_id=transaction_in.warehouse_id,
        item_id=item.id,
        quantity=transaction_in.quantity,
        transaction_type=TransactionType.IN,
        user_id=current_user.id,
        notes=transaction_in.notes
    )
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    logger.info(f"{transaction_in.quantity} of '{item.item_name}' logged IN by '{current_user.username}'.")
    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† Ù…Ø¯Ù„ Ú©Ø§Ù…Ù„ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø± UI
    return db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    ).filter(InventoryTransaction.id == transaction.id).first()


@app.post("/inventory/out/", response_model=InventoryTransactionOut)
async def log_inventory_out(
    transaction_out: InventoryTransactionOutManual,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø«Ø¨Øª Ø®Ø±ÙˆØ¬ Ø¯Ø³ØªÛŒ Ú©Ø§Ù„Ø§ Ø§Ø² Ø§Ù†Ø¨Ø§Ø±."""
    item = db.query(WarehouseItem).filter(WarehouseItem.item_name == transaction_out.item_name).first()
    if not item:
        raise HTTPException(status_code=404, detail=f"Ú©Ø§Ù„Ø§ÛŒÛŒ Ø¨Ø§ Ù†Ø§Ù… '{transaction_out.item_name}' ÛŒØ§ÙØª Ù†Ø´Ø¯.")

    transaction = InventoryTransaction(
        warehouse_id=transaction_out.warehouse_id,
        item_id=item.id,
        quantity=transaction_out.quantity,
        transaction_type=TransactionType.OUT,
        user_id=current_user.id,
        project_id=transaction_out.project_id,
        notes=transaction_out.notes
    )
    db.add(transaction)
    db.commit()
    await check_stock_and_alert(db, item, transaction_out.warehouse_id)
    db.refresh(transaction)
    logger.info(f"{transaction_out.quantity} of '{item.item_name}' logged OUT by '{current_user.username}'.")
    return db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    ).filter(InventoryTransaction.id == transaction.id).first()

# --- Ù…Ø´Ø§Ù‡Ø¯Ù‡ Ùˆ Ø¬Ø³ØªØ¬ÙˆÛŒ Ù…ÙˆØ¬ÙˆØ¯ÛŒ ---
@app.get("/inventory/items/", response_model=List[CurrentStockItem])
def get_current_inventory(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª Ù…ÙˆØ¬ÙˆØ¯ÛŒ ÙØ¹Ù„ÛŒ ØªÙ…Ø§Ù… Ú©Ø§Ù„Ø§Ù‡Ø§ Ø¯Ø± ØªÙ…Ø§Ù… Ø§Ù†Ø¨Ø§Ø±Ù‡Ø§."""
    stock_level_expr = func.sum(
        case(
            (InventoryTransaction.transaction_type == TransactionType.IN.value, InventoryTransaction.quantity),
            else_=-InventoryTransaction.quantity
        )
    ).label("current_stock")

    # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÛŒÙ† Ú©ÙˆØ¦Ø±ÛŒ ØªÙ…Ø§Ù… Ú©Ø§Ù„Ø§Ù‡Ø§ Ø±Ø§ Ø¯Ø± ØªÙ…Ø§Ù… Ø§Ù†Ø¨Ø§Ø±Ù‡Ø§ Ø¨Ø§ Ù…ÙˆØ¬ÙˆØ¯ÛŒâ€ŒØ´Ø§Ù† Ø¨Ø±Ù…ÛŒâ€ŒÚ¯Ø±Ø¯Ø§Ù†Ø¯.
    results = db.query(
        WarehouseItem.id.label("item_id"),
        WarehouseItem.item_name,
        Warehouse.id.label("warehouse_id"),
        Warehouse.name.label("warehouse_name"),
        WarehouseItem.min_stock_level,
        stock_level_expr
    ).join(InventoryTransaction, InventoryTransaction.item_id == WarehouseItem.id)\
     .join(Warehouse, Warehouse.id == InventoryTransaction.warehouse_id)\
     .group_by(WarehouseItem.id, Warehouse.id)\
     .all()

    return [CurrentStockItem(**row._asdict()) for row in results]
    
@app.get("/inventory/history/", response_model=List[InventoryTransactionOut])
def get_inventory_history(
    db: Session = Depends(get_db),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    transaction_type: Optional[TransactionType] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None,
    item_id: Optional[int] = None,
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª ØªØ§Ø±ÛŒØ®Ú†Ù‡ ØªØ±Ø§Ú©Ù†Ø´â€ŒÙ‡Ø§ÛŒ Ø§Ù†Ø¨Ø§Ø± Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª ÙÛŒÙ„ØªØ±."""
    query = db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    )
    if start_date:
        query = query.filter(InventoryTransaction.timestamp >= datetime.combine(start_date, time.min))
    if end_date:
        query = query.filter(InventoryTransaction.timestamp <= datetime.combine(end_date, time.max))
    if transaction_type:
        query = query.filter(InventoryTransaction.transaction_type == transaction_type)
    if project_id:
        query = query.filter(InventoryTransaction.project_id == project_id)
    if user_id:
        query = query.filter(InventoryTransaction.user_id == user_id)
    if item_id:
        query = query.filter(InventoryTransaction.item_id == item_id)
    
    return query.order_by(InventoryTransaction.timestamp.desc()).all()


# --- Ù…Ø¯ÛŒØ±ÛŒØª ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ù‡Ø± ØªØ§Ø¨Ù„Ùˆ ---
@app.post("/inventory/panel-items/", status_code=status.HTTP_201_CREATED)
def define_panel_code_items(
    definition: PanelCodeItemsDefinition,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """ØªØ¹Ø±ÛŒÙ ÛŒØ§ Ø¨Ø§Ø²Ù†ÙˆÛŒØ³ÛŒ Ù„ÛŒØ³Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª (BOM) Ø¨Ø±Ø§ÛŒ ÛŒÚ© Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ."""
    # Ø­Ø°Ù ØªØ¹Ø§Ø±ÛŒÙ Ù‚Ø¨Ù„ÛŒ Ø¨Ø±Ø§ÛŒ Ø§ÛŒÙ† Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ
    db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == definition.panel_code).delete()
    
    for item in definition.items:
        # Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² ÙˆØ¬ÙˆØ¯ Ú©Ø§Ù„Ø§ Ø¯Ø± Ø§Ù†Ø¨Ø§Ø±
        if not db.query(WarehouseItem).filter(WarehouseItem.item_name == item.item_name).first():
            raise HTTPException(status_code=400, detail=f"Ú©Ø§Ù„Ø§ÛŒ '{item.item_name}' Ø¯Ø± Ø§Ù†Ø¨Ø§Ø± ØªØ¹Ø±ÛŒÙ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")
        
        db_item = PanelCodeItems(
            panel_code=definition.panel_code,
            item_name=item.item_name,
            quantity_required=item.quantity_required
        )
        db.add(db_item)
    
    db.commit()
    logger.info(f"BOM for panel code '{definition.panel_code}' defined/updated by '{current_user.username}'.")
    return {"message": f"ØªØ¬Ù‡ÛŒØ²Ø§Øª Ø¨Ø±Ø§ÛŒ Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ '{definition.panel_code}' Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ØªØ¹Ø±ÛŒÙ Ø´Ø¯."}

@app.get("/inventory/panel-items/{panel_code}", response_model=List[PanelCodeItemOut])
def get_panel_code_items(
    panel_code: str,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """Ø¯Ø±ÛŒØ§ÙØª Ù„ÛŒØ³Øª ØªØ¬Ù‡ÛŒØ²Ø§Øª (BOM) Ø¨Ø±Ø§ÛŒ ÛŒÚ© Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ."""
    return db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == panel_code).all()    

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û¸: Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§
# ==============================================================================

# Ú¯Ø²Ø§Ø±Ø´ Ø®Ù„Ø§ØµÙ‡ Ø®Ø±ÛŒØ¯
@app.get("/reports/procurement-summary/excel", response_class=StreamingResponse)
def get_procurement_summary_report(
    start_date: date, 
    end_date: date, 
    db: Session = Depends(get_db)
):
    """Ú¯Ø²Ø§Ø±Ø´ Ø®Ù„Ø§ØµÙ‡ Ø®Ø±ÛŒØ¯ Ø¨Ù‡ ØµÙˆØ±Øª Ø§Ú©Ø³Ù„"""
    query = db.query(Project).options(joinedload(Project.equipment)).filter(
        Project.created_at >= datetime.combine(start_date, time.min),
        Project.created_at <= datetime.combine(end_date, time.max)
    )
    projects = query.all()

    if not projects:
        raise HTTPException(status_code=404, detail="Ù‡ÛŒÚ† Ù¾Ø±ÙˆÚ˜Ù‡â€ŒØ§ÛŒ Ø¯Ø± Ø§ÛŒÙ† Ø¨Ø§Ø²Ù‡ Ø²Ù…Ø§Ù†ÛŒ ÛŒØ§ÙØª Ù†Ø´Ø¯.")

    summary_by_panel = defaultdict(lambda: defaultdict(int))
    overall_summary = defaultdict(int)

    for project in projects:
        panel_key = project.panel_code or project.panel_type_key or "Ù†Ø§Ù…Ø´Ø®Øµ"
        
        for item in project.equipment:
            normalized_name = normalize_text(item.item_name)
            summary_by_panel[panel_key][normalized_name] += item.quantity
            overall_summary[normalized_name] += item.quantity

    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    summary_sheet = workbook.create_sheet(title="Ø®Ù„Ø§ØµÙ‡ Ú©Ù„ÛŒ ØªØ¬Ù‡ÛŒØ²Ø§Øª")
    summary_sheet.sheet_view.rightToLeft = True
    summary_sheet.append(["Ù†Ø§Ù… ØªØ¬Ù‡ÛŒØ²", "ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø²"])
    
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for item_name, total_quantity in sorted(overall_summary.items()):
        summary_sheet.append([item_name, total_quantity])
    
    summary_sheet.column_dimensions['A'].width = 60
    summary_sheet.column_dimensions['B'].width = 20

    for panel_key, items in sorted(summary_by_panel.items()):
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', panel_key)[:30]
        sheet = workbook.create_sheet(title=safe_sheet_name)
        sheet.sheet_view.rightToLeft = True
        
        sheet.append([f"ØªØ¬Ù‡ÛŒØ²Ø§Øª Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø² Ø¨Ø±Ø§ÛŒ ØªØ§Ø¨Ù„Ùˆ: {panel_key}"])
        sheet.merge_cells('A1:B1')
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = center_align
        
        sheet.append(["Ù†Ø§Ù… ØªØ¬Ù‡ÛŒØ²", "ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ±Ø¯ Ù†ÛŒØ§Ø²"])
        for cell in sheet[2]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        for item_name, quantity in sorted(items.items()):
            row = [item_name, quantity]
            sheet.append(row)
            for cell in sheet[sheet.max_row]:
                cell.border = thin_border
        
        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['B'].width = 20

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Procurement_Summary_{start_date}_to_{end_date}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    
    logger.info(f"Procurement summary report generated: {filename}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

# Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ÛŒ Ø®Ø±ÙˆØ¬
@app.get("/reports/exited-panels/simple-excel", response_class=StreamingResponse)
def get_exited_panels_simple_report(report_date: date, db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ Ø³Ø§Ø¯Ù‡ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø®Ø±ÙˆØ¬ÛŒ Ø¯Ø± ÛŒÚ© ØªØ§Ø±ÛŒØ® Ù…Ø´Ø®Øµ"""
    start_of_day, end_of_day = datetime.combine(report_date, time.min), datetime.combine(report_date, time.max)
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    headers = ["Ø±Ø¯ÛŒÙ", "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ", "Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§", "Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ", "ØªØ§Ø±ÛŒØ® Ùˆ Ø³Ø§Ø¹Øª Ø®Ø±ÙˆØ¬"]
    sheet.append(headers)
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for idx, project in enumerate(exited_projects, 1):
        exit_step = next((step for step in project.steps if step.name_key == StepNameKey.EXIT_PANEL), None)
        exit_time_str = jdatetime.datetime.fromgregorian(datetime=exit_step.timestamp).strftime('%Y/%m/%d - %H:%M:%S') if exit_step else "Ù†Ø§Ù…Ø´Ø®Øµ"
        row_data = [
            idx, 
            project.customer_name, 
            project.request_id, 
            project.panel_code or project.panel_type_key or "-", 
            exit_time_str
        ]
        sheet.append(row_data)
        
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 25
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Simple_Exited_Report_{report_date.strftime('%Y-%m-%d')}.xlsx"
    logger.info(f"Simple exit report generated for date: {report_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/exited-panels/simple-excel-range", response_class=StreamingResponse)
def get_exited_panels_simple_report_range(start_date: date, end_date: date, db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ Ø³Ø§Ø¯Ù‡ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø®Ø±ÙˆØ¬ÛŒ Ø¯Ø± Ø¨Ø§Ø²Ù‡ Ø²Ù…Ø§Ù†ÛŒ"""
    start_of_day = datetime.combine(start_date, time.min)
    end_of_day = datetime.combine(end_date, time.max)
    
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    headers = ["Ø±Ø¯ÛŒÙ", "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ", "Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§", "Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ", "ØªØ§Ø±ÛŒØ® Ùˆ Ø³Ø§Ø¹Øª Ø®Ø±ÙˆØ¬"]
    sheet.append(headers)
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for idx, project in enumerate(exited_projects, 1):
        exit_step = next((step for step in project.steps if step.name_key == StepNameKey.EXIT_PANEL), None)
        exit_time_str = jdatetime.datetime.fromgregorian(datetime=exit_step.timestamp).strftime('%Y/%m/%d - %H:%M:%S') if exit_step else "Ù†Ø§Ù…Ø´Ø®Øµ"
        row_data = [
            idx, 
            project.customer_name, 
            project.request_id, 
            project.panel_code or project.panel_type_key or "-", 
            exit_time_str
        ]
        sheet.append(row_data)
        
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 25
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Exited_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    
    logger.info(f"Exit report generated for range: {start_date} to {end_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

# Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„
@app.get("/reports/ready-for-delivery/detailed-excel", response_class=StreamingResponse)
def get_detailed_delivery_report_excel(direction: str, db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ ØªÙØµÛŒÙ„ÛŒ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„"""
    if direction not in ["west", "east"]: 
        raise HTTPException(status_code=400, detail="Ø¬Ù‡Øª Ø¨Ø§ÛŒØ¯ 'west' ÛŒØ§ 'east' Ø¨Ø§Ø´Ø¯.")
    
    company_name = WEST_COMPANY if direction == "west" else EAST_COMPANY
    jalali_date_str = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    ready_projects = get_projects_by_status(
        db, 
        required_steps=set(ORDERED_MANUAL_STEP_KEYS), 
        forbidden_steps={StepNameKey.EXIT_PANEL}
    )
    filtered_projects = [p for p in ready_projects if get_direction(p.request_id) == direction]
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    header_font_large = Font(bold=True, name='B Titr', size=14)
    header_font_small = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.merge_cells('A1:U1')
    sheet['A1'].value = "Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†"
    sheet['A1'].font = header_font_large
    sheet['A1'].alignment = center_align

    sheet.merge_cells('A2:U2')
    sheet['A2'].value = f"Ø´Ø±Ú©Øª ØªØ­ÙˆÛŒÙ„ Ú¯ÛŒØ±Ù†Ø¯Ù‡: {company_name}"
    sheet['A2'].font = header_font_small
    sheet['A2'].alignment = center_align

    sheet.merge_cells('S3:U3')
    sheet['S3'].value = f"ØªØ§Ø±ÛŒØ® Ú¯Ø²Ø§Ø±Ø´: {jalali_date_str}"
    sheet['S3'].font = Font(name='B Nazanin', size=11)
    sheet['S3'].alignment = center_align

    table_start_row = 5
    
    header_font = Font(bold=True, name='B Nazanin', size=11)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    headers_l1 = {
        'A': 'Ø±Ø¯ÛŒÙ', 
        'B': 'Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ Ù…ØªÙ‚Ø§Ø¶ÛŒ', 
        'C': 'Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§', 
        'D': 'Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ', 
        'E': 'ØªØ¹Ø¯Ø§Ø¯ Ú©Ù†ØªÙˆØ±', 
        'G': 'ØªØ¹Ø¯Ø§Ø¯ ÙÛŒÙˆØ²', 
        'J': 'ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ¯Ù…', 
        'K': 'Ú©Ù†ØªÙˆØ± Ù¾ÛŒØ´ Ø¨ÛŒÙ†ÛŒ', 
        'M': 'ÙÛŒÙˆØ² Ù¾ÛŒØ´ Ø¨ÛŒÙ†ÛŒ', 
        'O': 'Ø³Ú©Ùˆ', 
        'P': 'ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 
        'Q': 'Ø¨Ø³Øª ØªØ³Ù…Ù‡', 
        'R': 'Ù¾ÛŒÚ†', 
        'T': 'Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯', 
        'U': 'Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ'
    }
    
    for col, value in headers_l1.items(): 
        sheet[f'{col}{table_start_row}'] = value
    
    sheet.merge_cells(f'E{table_start_row}:F{table_start_row}')
    sheet.merge_cells(f'G{table_start_row}:I{table_start_row}')
    sheet.merge_cells(f'K{table_start_row}:L{table_start_row}')
    sheet.merge_cells(f'M{table_start_row}:N{table_start_row}')
    sheet.merge_cells(f'R{table_start_row}:S{table_start_row}')
    
    table_header_row_2 = table_start_row + 1
    headers_l2 = {
        'E': 'ØªÚ©ÙØ§Ø²', 
        'F': 'Ø³Ù‡ ÙØ§Ø²', 
        'G': 'ØªÚ©ÙØ§Ø²', 
        'H': 'Ø³Ù‡ ÙØ§Ø²', 
        'I': None, 
        'K': 'ØªÚ©ÙØ§Ø²', 
        'L': 'Ø³Ù‡ ÙØ§Ø²', 
        'M': 'ØªÚ©ÙØ§Ø²', 
        'N': 'Ø³Ù‡ ÙØ§Ø²', 
        'R': 'ÛŒÚ©Ø³Ø±Ø±Ø²ÙˆÙ‡', 
        'S': 'Û±Û¶*Û³Û°Û°', 
        'T': None, 
        'U': None
    }
    
    for col, value in headers_l2.items(): 
        if value:
            sheet[f'{col}{table_header_row_2}'] = value
    
    for col in ['A', 'B', 'C', 'D', 'J', 'O', 'P', 'Q', 'T', 'U']: 
        sheet.merge_cells(f'{col}{table_start_row}:{col}{table_header_row_2}')
    
    for row in sheet.iter_rows(min_row=table_start_row, max_row=table_header_row_2):
        for cell in row: 
            if cell.value:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
    
    row_num = table_header_row_2 + 1
    for idx, project in enumerate(filtered_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "Ù…ÙˆØ¯Ù…" in normalize_text(item.item_name)
        )
        
        total_single_phase_fuses = sum([
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16', 0), 
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25', 0), 
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32', 0)
        ])
        total_three_phase_fuses = sum([
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25', 0), 
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32', 0), 
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63', 0)
        ])
        
        row_data = {
            'A': idx, 
            'B': project.customer_name, 
            'C': project.request_id, 
            'D': project.panel_code or '-',
            'E': summary.get('Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²', 0), 
            'F': summary.get('Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²', 0), 
            'G': total_single_phase_fuses, 
            'H': total_three_phase_fuses, 
            'I': None, 
            'J': modem_count,
            'K': reservation.get('purchased_single_phase', 0), 
            'L': 0, 
            'M': 0, 
            'N': 0,
            'O': summary.get('Ø³Ú©Ùˆ', 0), 
            'P': summary.get('ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0), 
            'Q': summary.get('Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0),
            'R': 0, 
            'S': summary.get('Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°', 0), 
            'T': summary.get('Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯', 0), 
            'U': summary.get('Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ', 0)
        }
        
        for col, value in row_data.items():
            cell = sheet[f'{col}{row_num}']
            cell.value = value
            cell.alignment = center_align
            cell.border = thin_border
        
        row_num += 1
    
    footer_row = sheet.max_row + 3
    sheet.cell(row=footer_row, column=3, value="ØªØ­ÙˆÛŒÙ„ Ø¯Ù‡Ù†Ø¯Ù‡:").font = header_font_small
    sheet.cell(row=footer_row, column=10, value="ØªØ­ÙˆÛŒÙ„ Ú¯ÛŒØ±Ù†Ø¯Ù‡:").font = header_font_small
    sheet.cell(row=footer_row, column=18, value=f"ØªØ§Ø±ÛŒØ®: {jalali_date_str}").font = header_font_small

    for col, width in {'B': 35, 'C': 18, 'D': 18}.items(): 
        sheet.column_dimensions[col].width = width
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Ready_For_Delivery_{direction}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info(f"Detailed delivery report generated for direction: {direction}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

# Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ÛŒ ØªØ§ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø±
@app.get("/reports/supervisor-checklist-items", response_model=List[str])
def get_supervisor_checklist_items():
    """Ø¯Ø±ÛŒØ§ÙØª Ø¢ÛŒØªÙ…â€ŒÙ‡Ø§ÛŒ Ú†Ú©â€ŒÙ„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ù†Ø§Ø¸Ø±"""
    return SUPERVISOR_CHECKLIST_ITEMS

@app.get("/reports/supervisor-approval/simple/excel", response_class=StreamingResponse)
def get_supervisor_approval_simple_report_excel(db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ Ø³Ø§Ø¯Ù‡ ØªØ§ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø± Ø¨Ù‡ ØµÙˆØ±Øª Ø§Ú©Ø³Ù„"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    sheet.merge_cells('A1:F1')
    cell_a1 = sheet['A1']
    cell_a1.value = "Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†"
    cell_a1.font = Font(bold=True, size=14)
    cell_a1.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('A2:F2')
    cell_a2 = sheet['A2']
    cell_a2.value = "Ù„ÛŒØ³Øª ØªØ§Ø¨Ù„Ùˆ Ù‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„"
    cell_a2.font = Font(bold=True, size=12)
    cell_a2.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('E3:F3')
    sheet['E3'].value = f"ØªØ§Ø±ÛŒØ®: {jalali_date}"
    sheet['E3'].alignment = Alignment(horizontal='center')
    
    headers = ["Ø±Ø¯ÛŒÙ", "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ", "Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§", "Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ", "ØªØ§ÛŒÛŒØ¯", "ØªÙˆØ¶ÛŒØ­Ø§Øª"]
    sheet.append(headers)
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    header_row_num = sheet.max_row
    for cell in sheet[header_row_num]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    
    for idx, p in enumerate(projects, 1):
        sheet.append([idx, p.customer_name, p.request_id, p.panel_code or '-', 'âˆš', ''])
    
    for row in sheet.iter_rows(min_row=header_row_num + 1, max_row=sheet.max_row, min_col=1, max_col=6):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            cell.alignment = right_align if col_idx == 2 else center_align
    
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    
    last_row = sheet.max_row + 3
    sheet.cell(row=last_row, column=2, value="Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ú©Ø§Ø±Ú¯Ø§Ù‡:").font = Font(bold=True)
    sheet.cell(row=last_row, column=4, value="Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ø¯ÙØªØ± Ù†Ø¸Ø§Ø±Øª:").font = Font(bold=True)
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"SupervisorApprovalReport_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info("Supervisor approval report generated")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/supervisor-approval/simple/html", response_class=HTMLResponse)
def get_supervisor_approval_simple_report_html(db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ Ø³Ø§Ø¯Ù‡ ØªØ§ÛŒÛŒØ¯ Ù†Ø§Ø¸Ø± Ø¨Ù‡ ØµÙˆØ±Øª HTML"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    template = Template(SUPERVISOR_APPROVAL_SIMPLE_TEMPLATE_STR)
    html_content = template.render(projects=projects, jalali_date=jalali_date)
    return HTMLResponse(content=html_content)

@app.get("/reports/supervisor-approval/checklist/excel", response_class=StreamingResponse)
def get_supervisor_qc_checklist_excel(db: Session = Depends(get_db)):
    """Ú†Ú©â€ŒÙ„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ù†Ø§Ø¸Ø± Ø¨Ù‡ ØµÙˆØ±Øª Ø§Ú©Ø³Ù„"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    sheet.merge_cells('A1:E1')
    sheet['A1'].value = "Ø´Ø±Ú©Øª ÙØ±Ø¯Ø§Ø¯ Ø³Ø§Ø²Ù‡ Ú¯Ù„Ø´Ù†"
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = center_align
    
    sheet.merge_cells('A2:E2')
    sheet['A2'].value = "Ú†Ú© Ù„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„"
    sheet['A2'].font = Font(bold=True, size=12)
    sheet['A2'].alignment = center_align
    
    sheet.cell(row=2, column=len(projects) + 2).value = f"ØªØ§Ø±ÛŒØ®: {jalali_date}"
    
    header_request_id = ["Ø±Ø¯ÛŒÙ", "Ø´Ø±Ø­ Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª"] + [p.request_id for p in projects]
    header_panel_code = ["", ""] + [p.panel_code or '-' for p in projects]
    
    sheet.append(header_request_id)
    sheet.append(header_panel_code)
    
    for row_idx in range(4, 6):
        for col_idx in range(1, len(header_request_id) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
    
    for idx, item in enumerate(SUPERVISOR_CHECKLIST_ITEMS, 1):
        sheet.append([idx, item] + ([''] * len(projects)))
    
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row):
        for cell in row: 
            cell.border = thin_border
        row[1].alignment = right_align
    
    sheet.column_dimensions['B'].width = 50
    for i in range(len(projects)): 
        sheet.column_dimensions[get_column_letter(i + 3)].width = 15
    
    last_row = sheet.max_row + 3
    sheet.cell(row=last_row, column=2, value="Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ú©Ø§Ø±Ú¯Ø§Ù‡:").font = bold_font
    sheet.cell(row=last_row, column=5, value="Ù†Ø§Ù… Ùˆ Ø§Ù…Ø¶Ø§ Ù…Ø³Ø¦ÙˆÙ„ Ø¯ÙØªØ± Ù†Ø¸Ø§Ø±Øª:").font = bold_font
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"SupervisorChecklist_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info("Supervisor checklist report generated")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/supervisor-approval/checklist/html", response_class=HTMLResponse)
def get_supervisor_qc_checklist_html(db: Session = Depends(get_db)):
    """Ú†Ú©â€ŒÙ„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª Ù†Ø§Ø¸Ø± Ø¨Ù‡ ØµÙˆØ±Øª HTML"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    template = Template(SUPERVISOR_CHECKLIST_TEMPLATE_STR)
    html_content = template.render(
        projects=projects, 
        checklist_items=SUPERVISOR_CHECKLIST_ITEMS, 
        jalali_date=jalali_date
    )
    return HTMLResponse(content=html_content)

# Ú¯Ø²Ø§Ø±Ø´ KPI
@app.get("/reports/kpi-summary", response_model=KpiSummary)
def get_kpi_summary(db: Session = Depends(get_db)):
    """Ø¯Ø±ÛŒØ§ÙØª Ø®Ù„Ø§ØµÙ‡ Ø´Ø§Ø®Øµâ€ŒÙ‡Ø§ÛŒ Ø¹Ù…Ù„Ú©Ø±Ø¯ (Ù†Ø³Ø®Ù‡ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡)"""
    total_projects = db.query(func.count(Project.id)).scalar() or 0

    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    completed_this_month = db.query(func.count(Project.id)).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL,
        Step.timestamp >= start_of_month
    ).scalar() or 0

    all_projects_with_steps = db.query(Project).options(joinedload(Project.steps)).all()
    
    durations = defaultdict(list)
    completion_times = []

    for p in all_projects_with_steps:
        # Ù…Ø±Ø§Ø­Ù„ Ø±Ø§ Ø¨Ø± Ø§Ø³Ø§Ø³ Ø²Ù…Ø§Ù† Ù…Ø±ØªØ¨ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
        sorted_steps = sorted(p.steps, key=lambda s: s.timestamp)
        if not sorted_steps:
            continue
            
        # ÛŒÚ© ØªØ§ÛŒÙ…â€ŒÙ„Ø§ÛŒÙ† Ø§Ø² Ø²Ù…Ø§Ù† Ø§ÛŒØ¬Ø§Ø¯ Ù¾Ø±ÙˆÚ˜Ù‡ ØªØ§ Ù¾Ø§ÛŒØ§Ù† Ù‡Ø± Ù…Ø±Ø­Ù„Ù‡ Ù…ÛŒâ€ŒØ³Ø§Ø²ÛŒÙ…
        # âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø³Ø§Ø®ØªØ§Ø± ØªØ§ÛŒÙ…â€ŒÙ„Ø§ÛŒÙ† Ø¨Ø±Ø§ÛŒ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø¢Ø³Ø§Ù†â€ŒØªØ±
        timeline = [(p.created_at, "PROJECT_CREATION")] + [(s.timestamp, s.name_key) for s in sorted_steps]

        # Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…Ø¯Øª Ø²Ù…Ø§Ù† Ø¨ÛŒÙ† Ù‡Ø± Ø¯Ùˆ Ù†Ù‚Ø·Ù‡ Ù…ØªÙˆØ§Ù„ÛŒ Ø¯Ø± ØªØ§ÛŒÙ…â€ŒÙ„Ø§ÛŒÙ†
        for i in range(len(timeline) - 1):
            start_time, start_key_or_event = timeline[i]
            end_time, end_key = timeline[i+1]
            
            # Ù…Ø¯Øª Ø²Ù…Ø§Ù† Ø¨Ù‡ Ù…Ø±Ø­Ù„Ù‡ Ù‚Ø¨Ù„ÛŒ ØªØ¹Ù„Ù‚ Ø¯Ø§Ø±Ø¯
            # Ù…Ø«Ø§Ù„: Ø²Ù…Ø§Ù† Ø¨ÛŒÙ† START_ASSEMBLY Ùˆ END_ASSEMBLY Ù…ØªØ¹Ù„Ù‚ Ø¨Ù‡ Ù…Ø±Ø­Ù„Ù‡ START_ASSEMBLY Ø§Ø³Øª
            # âœ… Ø§ØµÙ„Ø§Ø­ Ù…Ù†Ø·Ù‚ Ø§ØµÙ„ÛŒ: Ù…Ø¯Øª Ø²Ù…Ø§Ù† Ø¨Ù‡ Ù…Ø±Ø­Ù„Ù‡ Ø´Ø±ÙˆØ¹ ØªØ¹Ù„Ù‚ Ù…ÛŒâ€ŒÚ¯ÛŒØ±Ø¯
            if start_key_or_event != "PROJECT_CREATION":
                duration_hours = (end_time - start_time).total_seconds() / 3600
                if duration_hours >= 0:
                     durations[start_key_or_event].append(duration_hours)

        # Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø²Ù…Ø§Ù† ØªÚ©Ù…ÛŒÙ„ Ú©Ù„ Ù¾Ø±ÙˆÚ˜Ù‡
        exit_step = next((s for s in sorted_steps if s.name_key == StepNameKey.EXIT_PANEL), None)
        if exit_step:
            total_duration_days = (exit_step.timestamp - p.created_at).total_seconds() / (3600 * 24)
            if total_duration_days > 0:
                completion_times.append(total_duration_days)

    # Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…ÛŒØ§Ù†Ú¯ÛŒÙ†â€ŒÙ‡Ø§
    avg_step_durations = {}
    for key in ORDERED_MANUAL_STEP_KEYS:
        if durations.get(key):
            # Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…ÛŒØ§Ù†Ù‡ (median) Ø¨Ù‡ Ø¬Ø§ÛŒ Ù…ÛŒØ§Ù†Ú¯ÛŒÙ† (mean) Ø¨Ø±Ø§ÛŒ Ú©Ø§Ù‡Ø´ ØªØ§Ø«ÛŒØ± Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ Ù¾Ø±Øª
            avg_hours = round(float(np.median(durations[key])), 2)
            avg_step_durations[key.value] = avg_hours
        else:
            avg_step_durations[key.value] = None

    # Ù¾ÛŒØ¯Ø§ Ú©Ø±Ø¯Ù† Ú¯Ù„ÙˆÚ¯Ø§Ù‡
    bottleneck_step_key = None
    if any(v is not None for v in avg_step_durations.values()):
        bottleneck_step_key = max(
            avg_step_durations, 
            key=lambda k: avg_step_durations.get(k) or -1
        )
    bottleneck_step_name = STEP_KEY_TO_NAME_MAP.get(bottleneck_step_key, None) if bottleneck_step_key else None

    avg_completion_time = round(float(np.mean(completion_times)), 2) if completion_times else None

    logger.info("KPI summary generated successfully.")
    return KpiSummary(
        total_projects=total_projects,
        completed_this_month=completed_this_month,
        avg_completion_time_days=avg_completion_time,
        bottleneck_step=bottleneck_step_name,
        step_durations=avg_step_durations
    )    

# ==============================================================================
# Ø¨Ø®Ø´ Û±Û¹: Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øªâ€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ú†Ø³Ø¨ Ùˆ Ú†Ø§Ù¾
# ==============================================================================

async def get_project_label_data(project_id: int, db: Session) -> Dict[str, Any]:
    """Ø¯Ø±ÛŒØ§ÙØª Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ú†Ø³Ø¨ Ù¾Ø±ÙˆÚ˜Ù‡"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    direction = get_direction(project.request_id)
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=2)
    qr.add_data(project.request_id)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    qr_img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
    
    jalali_date_str = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return {
        "project": project, 
        "qr_code_base64": qr_img_str, 
        "report_date_jalali": jalali_date_str, 
        "direction": "Ø´Ø±Ù‚" if direction == "east" else "ØºØ±Ø¨" if direction == "west" else None
    }

@app.get("/projects/{project_id}/qc-label", response_class=HTMLResponse, include_in_schema=False)
async def get_qc_label(request: Request, project_id: int, db: Session = Depends(get_db)):
    """Ø¨Ø±Ú†Ø³Ø¨ Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª"""
    project = db.query(Project).get(project_id)
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    panel_id = (project.panel_code or "NO-ID").replace("/", "-")
    try: 
        jalali_dt = jdatetime.datetime.fromgregorian(datetime=project.created_at)
        creation_month = f"{jalali_dt.month:02d}"
    except Exception: 
        creation_month = "00"
    
    serial_number = f"09-{panel_id}-{creation_month}-{project.request_id.strip()[-10:]}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=1)
    qr.add_data(serial_number)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    qr_code_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
    
    context = {
        "request": request, 
        "serial_number": serial_number, 
        "qr_code_base64": qr_code_base64
    }
    
    return templates.TemplateResponse("qc_label_template.html", context)

@app.get("/projects/{project_id}/label", response_class=HTMLResponse, include_in_schema=False)
async def get_project_label_for_printing(project_id: int, db: Session = Depends(get_db)):
    """Ø¨Ø±Ú†Ø³Ø¨ Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ø±Ø§ÛŒ Ú†Ø§Ù¾"""
    data = await get_project_label_data(project_id, db)
    data["font_path"] = "/static/Vazirmatn-regular.ttf"
    template_str_with_print = PROJECT_SLIP_TEMPLATE_STR.replace(
        "</body>", 
        "<script>window.onload = function() { window.print(); };</script></body>"
    )
    template = Template(template_str_with_print)
    return HTMLResponse(content=template.render(data))

@app.get("/projects/{project_id}/qc-checklist", response_class=HTMLResponse)
def get_individual_qc_checklist(project_id: int, db: Session = Depends(get_db)):
    """Ú†Ú©â€ŒÙ„ÛŒØ³Øª Ú©Ù†ØªØ±Ù„ Ú©ÛŒÙÛŒØª ÙØ±Ø¯ÛŒ"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    template = Template(INDIVIDUAL_QC_CHECKLIST_TEMPLATE_STR)
    html_content = template.render(
        project=project, 
        checklist_items=SUPERVISOR_CHECKLIST_ITEMS, 
        jalali_date=jalali_date
    )
    return HTMLResponse(content=html_content)

@app.get("/projects/{project_id}/download-pdf", response_class=StreamingResponse)
async def download_project_label_pdf(project_id: int, db: Session = Depends(get_db)):
    """Ø¯Ø§Ù†Ù„ÙˆØ¯ Ø¨Ø±Ú†Ø³Ø¨ Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ù‡ ØµÙˆØ±Øª PDF"""
    data = await get_project_label_data(project_id, db)
    data["font_path"] = "Vazirmatn-regular.ttf"
    template = Template(PROJECT_SLIP_TEMPLATE_STR)
    html_content = template.render(data)
    
    pdf_stream = BytesIO()
    
    def fetch_resources(uri, rel):
        static_path = resource_path(os.path.join(STATIC_DIR_NAME, uri))
        if os.path.exists(static_path): 
            return static_path
        root_path = resource_path(uri)
        if os.path.exists(root_path): 
            return root_path
        return None
    
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_stream, link_callback=fetch_resources, encoding='utf-8')
    if pisa_status.err: 
        logger.error(f"PDF generation error: {pisa_status.err}")
        raise HTTPException(status_code=500, detail=f"Ø®Ø·Ø§ Ø¯Ø± ØªÙˆÙ„ÛŒØ¯ ÙØ§ÛŒÙ„ PDF: {pisa_status.err}")
    
    pdf_stream.seek(0)
    filename = f"ProjectLabel_{data['project'].request_id}.pdf"
    
    logger.info(f"PDF label generated for project {project_id}")
    return StreamingResponse(
        pdf_stream, 
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/projects/{project_id}/download-excel", response_class=StreamingResponse)
async def download_project_label_excel(project_id: int, db: Session = Depends(get_db)):
    """Ø¯Ø§Ù†Ù„ÙˆØ¯ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ù¾Ø±ÙˆÚ˜Ù‡ Ø¨Ù‡ ØµÙˆØ±Øª Ø§Ú©Ø³Ù„"""
    project_data = await get_project_label_data(project_id, db)
    project = project_data['project']
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    sheet.append(["Ù†Ø§Ù… Ù¾Ø±ÙˆÚ˜Ù‡", project.name])
    sheet.append(["Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª", project.request_id])
    sheet.append(["Ù…Ø´ØªØ±ÛŒ", project.customer_name])
    sheet.append(["ØªØ§Ø±ÛŒØ® ØµØ¯ÙˆØ±", project_data['report_date_jalali']])
    
    if project.panel_code: 
        sheet.append(["Ú©Ø¯ ØªØ§Ø¨Ù„Ùˆ", project.panel_code])
    if project_data['direction']: 
        sheet.append(["Ø¬Ù‡Øª Ø§Ø±Ø³Ø§Ù„", project_data['direction']])
    
    sheet.append([])
    sheet.append(["Ø±Ø¯ÛŒÙ", "Ù†Ø§Ù… ØªØ¬Ù‡ÛŒØ²", "ØªØ¹Ø¯Ø§Ø¯"])
    
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for i, item in enumerate(project.equipment, start=1):
        sheet.append([i, item.item_name, item.quantity])
    
    for col in ['A', 'B', 'C']:
        sheet.column_dimensions[col].width = 20
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"ProjectEquipment_{project.request_id}.xlsx"
    logger.info(f"Excel equipment list generated for project {project_id}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/projects/{project_id}/exit-slip", response_class=HTMLResponse)
async def get_project_exit_slip(project_id: int, db: Session = Depends(get_db)):
    """Ø¨Ø±Ú¯Ù‡ Ø®Ø±ÙˆØ¬ Ù¾Ø±ÙˆÚ˜Ù‡"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    try:
        summary_data = generate_project_summary_data(project)
        data = {
            "project": project, 
            "summary_data": summary_data, 
            "report_date_jalali": jdatetime.datetime.now().strftime('%Y/%m/%d'), 
            "font_path": "/static/Vazirmatn-regular.ttf"
        }
        
        template_str_with_print = DETAILED_EXIT_SLIP_TEMPLATE_STR.replace(
            "</body>", 
            "<script>window.onload = function() { setTimeout(function(){ window.print(); }, 500); };</script></body>"
        )
        template = Template(template_str_with_print)
        return HTMLResponse(content=template.render(data))
    except Exception as e:
        logger.error(f"ERROR rendering exit slip for project {project_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ø®Ø·Ø§ Ø¯Ø± ØªÙˆÙ„ÛŒØ¯ Ø¨Ø±Ú¯Ù‡ Ø®Ø±ÙˆØ¬. Ø¬Ø²Ø¦ÛŒØ§Øª Ø®Ø·Ø§: {e}")

@app.get("/projects/{project_id}/exit-slip-raw", response_class=HTMLResponse)
async def get_project_exit_slip_raw_html(project_id: int, request: Request, db: Session = Depends(get_db)):
    """Ø¨Ø±Ú¯Ù‡ Ø®Ø±ÙˆØ¬ Ù¾Ø±ÙˆÚ˜Ù‡ (Ù‚Ø§Ù„Ø¨ Ø®Ø§Ù…)"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="Ù¾Ø±ÙˆÚ˜Ù‡ ÛŒØ§ÙØª Ù†Ø´Ø¯.")
    
    summary_data = generate_project_summary_data(project)
    context = {
        "request": request, 
        "project": project, 
        "summary_data": summary_data, 
        "report_date_jalali": jdatetime.datetime.now().strftime('%Y/%m/%d')
    }
    
    return templates.TemplateResponse("slip_instance_template.html", context)

@app.get("/reports/exited-projects/excel", response_class=StreamingResponse)
def get_exited_projects_report(report_date: date, db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ ØªØ§Ø¨Ù„ÙˆÙ‡Ø§ÛŒ Ø®Ø±ÙˆØ¬ÛŒ"""
    start_of_day, end_of_day = datetime.combine(report_date, time.min), datetime.combine(report_date, time.max)
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.equipment), joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    sheet.merge_cells('A1:V1')
    cell_a1 = sheet['A1']
    cell_a1.value = "ØµÙˆØ±ØªØ¬Ù„Ø³Ù‡ ØªØ­ÙˆÛŒÙ„ ØªØ§Ø¨Ù„Ùˆ"
    cell_a1.font = Font(bold=True, name='B Nazanin', size=16)
    cell_a1.alignment = center_align
    
    sheet['B2'] = "Ø´Ø±Ú©Øª :"
    sheet['B2'].font = header_font
    sheet['B2'].alignment = right_align
    
    sheet['O2'] = f"ØªØ§Ø±ÛŒØ® ØªØ­ÙˆÛŒÙ„ : {jdatetime.date.fromgregorian(date=report_date).strftime('%Y/%m/%d')}"
    sheet['O2'].font = header_font
    sheet['O2'].alignment = right_align
    
    headers_l1 = [
        "Ø±Ø¯ÛŒÙ", "Ù†Ø§Ù… Ùˆ Ù†Ø§Ù… Ø®Ø§Ù†ÙˆØ§Ø¯Ú¯ÛŒ Ù…ØªÙ‚Ø§Ø¶ÛŒ", "Ø´Ù…Ø§Ø±Ù‡ ØªÙ‚Ø§Ø¶Ø§", "Ù†ÙˆØ¹ ØªØ§Ø¨Ù„Ùˆ", "ØªØ¹Ø¯Ø§Ø¯ Ú©Ù†ØªÙˆØ±", None, 
        "ØªØ¹Ø¯Ø§Ø¯ ÙÛŒÙˆØ²", None, "ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ¯Ù…", "Ú©Ù†ØªÙˆØ± Ù¾ÛŒØ´ Ø¨ÛŒÙ†ÛŒ", None, "ÙÛŒÙˆØ² Ù¾ÛŒØ´ Ø¨ÛŒÙ†ÛŒ", None, 
        "Ø³Ú©Ùˆ", "ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„", "Ø¨Ø³Øª ØªØ³Ù…Ù‡", "Ù¾ÛŒÚ† ÛŒÚ©Ø³Ø±Ø±Ø²ÙˆÙ‡", "Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯", "Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ", "Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°"
    ]
    headers_l2 = [
        None, None, None, None, "ØªÚ©ÙØ§Ø²", "Ø³Ù‡ ÙØ§Ø²", "ØªÚ©ÙØ§Ø²", "Ø³Ù‡ ÙØ§Ø²", None, 
        "ØªÚ©ÙØ§Ø²", "Ø³Ù‡ ÙØ§Ø²", "ØªÚ©ÙØ§Ø²", "Ø³Ù‡ ÙØ§Ø²", None, None, None, None, None, None, None
    ]
    
    sheet.append(headers_l1)
    sheet.append(headers_l2)
    
    sheet.merge_cells('A3:A4')
    sheet.merge_cells('B3:B4')
    sheet.merge_cells('C3:C4')
    sheet.merge_cells('D3:D4')
    sheet.merge_cells('E3:F3')
    sheet.merge_cells('G3:H3')
    sheet.merge_cells('I3:I4')
    sheet.merge_cells('J3:K3')
    sheet.merge_cells('L3:M3')
    sheet.merge_cells('N3:N4')
    sheet.merge_cells('O3:O4')
    sheet.merge_cells('P3:P4')
    sheet.merge_cells('Q3:Q4')
    sheet.merge_cells('R3:R4')
    sheet.merge_cells('S3:S4')
    sheet.merge_cells('T3:T4')
    
    for row in sheet.iter_rows(min_row=3, max_row=4):
        for cell in row: 
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
    
    start_row = 5
    for idx, project in enumerate(exited_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "Ù…ÙˆØ¯Ù…" in normalize_text(item.item_name)
        )
        
        total_single_phase_fuses = sum([
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 16', 0), 
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 25', 0), 
            summary.get('ÙÛŒÙˆØ² ØªÚ© ÙØ§Ø² 32', 0)
        ])
        total_three_phase_fuses = sum([
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 25', 0), 
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 32', 0), 
            summary.get('ÙÛŒÙˆØ² Ø³Ù‡ ÙØ§Ø² 63', 0)
        ])
        
        row_data = [
            idx, project.customer_name, project.request_id, project.panel_code or '-',
            summary.get('Ú©Ù†ØªÙˆØ± ØªÚ© ÙØ§Ø²', 0), summary.get('Ú©Ù†ØªÙˆØ± Ø³Ù‡ ÙØ§Ø²', 0),
            total_single_phase_fuses, total_three_phase_fuses, modem_count,
            reservation.get('purchased_single_phase', 0), 0, 0, 0,
            summary.get('Ø³Ú©Ùˆ', 0), summary.get('ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0), summary.get('Ø¨Ø³Øª ØªØ³Ù…Ù‡ Ø§Ø³ØªÛŒÙ„', 0),
            0, summary.get('Ù„ÙˆÙ„Ù‡ Ù†ÛŒÙ… Ú¯Ø±Ø¯', 0), summary.get('Ù„ÙˆÙ„Ù‡ Ø®Ø±Ø·ÙˆÙ…ÛŒ', 0), summary.get('Ù¾ÛŒÚ† Û±Û¶*Û³Û°Û°', 0)
        ]
        sheet.append(row_data)
    
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
        for cell in row: 
            cell.alignment = center_align
            cell.border = thin_border
    
    footer_row = sheet.max_row + 2
    sheet.cell(row=footer_row, column=2, value="ØªØ­ÙˆÛŒÙ„ Ø¯Ù‡Ù†Ø¯Ù‡ :").font = header_font
    sheet.cell(row=footer_row, column=14, value="ØªØ­ÙˆÛŒÙ„ Ú¯ÛŒØ±Ù†Ø¯Ù‡ :").font = header_font
    
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 15
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Exited_Projects_Summary_{report_date.strftime('%Y-%m-%d')}.xlsx"
    logger.info(f"Exited projects report generated for date: {report_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/ready-for-delivery/detailed-html", response_class=HTMLResponse)
def get_detailed_delivery_report_html(direction: str, db: Session = Depends(get_db)):
    """Ú¯Ø²Ø§Ø±Ø´ ØªÙØµÛŒÙ„ÛŒ Ø¢Ù…Ø§Ø¯Ù‡ ØªØ­ÙˆÛŒÙ„ Ø¨Ù‡ ØµÙˆØ±Øª HTML"""
    if direction not in ["west", "east"]: 
        raise HTTPException(status_code=400, detail="Ø¬Ù‡Øª Ø¨Ø§ÛŒØ¯ 'west' ÛŒØ§ 'east' Ø¨Ø§Ø´Ø¯.")
    
    company_name = WEST_COMPANY if direction == "west" else EAST_COMPANY
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    ready_projects = get_projects_by_status(
        db, 
        required_steps=set(ORDERED_MANUAL_STEP_KEYS), 
        forbidden_steps={StepNameKey.EXIT_PANEL}
    )
    report_data = generate_detailed_report_data(ready_projects, direction)
    
    try:
        html_template = jinja_env.get_template("detailed_report_template.html")
        return HTMLResponse(content=html_template.render(
            company_name=company_name, 
            jalali_date=jalali_date, 
            report_data=report_data
        ))
    except Exception as e:
        logger.error(f"Error loading detailed report template: {e}")
        return HTMLResponse(
            content=f"<h1>Ø®Ø·Ø§ Ø¯Ø± Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ù‚Ø§Ù„Ø¨ Ú¯Ø²Ø§Ø±Ø´</h1><p>ÙØ§ÛŒÙ„ 'detailed_report_template.html' ÛŒØ§ÙØª Ù†Ø´Ø¯ ÛŒØ§ Ø®Ø·Ø§ÛŒÛŒ Ø¯Ø± Ø¢Ù† ÙˆØ¬ÙˆØ¯ Ø¯Ø§Ø±Ø¯.</p><p>Ø¬Ø²Ø¦ÛŒØ§Øª Ø®Ø·Ø§: {e}</p>", 
            status_code=500
        )

# Ø¯ÛŒØ¨Ø§Ú¯
@app.get("/debug-assembler-query", response_model=List[Dict[str, Any]])
def debug_assembler_query(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øª Ø¯ÛŒØ¨Ø§Ú¯ Ø¨Ø±Ø§ÛŒ Ú©ÙˆØ¦Ø±ÛŒ Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±Ø§Ù†"""
    query = db.query(Project).filter(
        Project.assembler_1.isnot(None), 
        Project.panel_type_key.isnot(None)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects = query.order_by(Project.created_at.desc()).all()
    result = []
    
    for p in projects:
        result.append({
            "project_id": p.id, 
            "request_id": p.request_id, 
            "panel_type_key": p.panel_type_key,
            "assembler_1": p.assembler_1, 
            "assembler_2": p.assembler_2, 
            "created_at": p.created_at.isoformat()
        })
    
    return result



# ==============================================================================
# Ø¨Ø®Ø´ Û²Û°: WebSocket Endpoint
# ==============================================================================

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, token: Optional[str] = None):
    """Ø§ØªØµØ§Ù„ WebSocket Ø¨Ø±Ø§ÛŒ Ø§Ø±ØªØ¨Ø§Ø· Ø¨Ù„Ø§Ø¯Ø±Ù†Ú¯ Ø¨Ø§ Ø³Ø±ÙˆØ±."""
    user_id = 0
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            username: str = payload.get("sub")
            if username:
                with get_db_session() as db:
                    user = await get_user_by_username(db, username)
                    if user and user.is_active:
                        user_id = user.id
        except JWTError as e:
            logger.warning(f"WebSocket connection attempted with invalid token: {e}")
    
    await manager.connect(websocket, user_id)
    
    try:
        while True: 
            # Ù†Ú¯Ù‡ Ø¯Ø§Ø´ØªÙ† Ø§Ø±ØªØ¨Ø§Ø· Ø¨Ø§Ø². Ù…ÛŒâ€ŒØªÙˆØ§Ù† Ù…Ù†Ø·Ù‚ Ø¯Ø±ÛŒØ§ÙØª Ù¾ÛŒØ§Ù… Ø±Ø§ Ø§ÛŒÙ†Ø¬Ø§ Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯.
            await websocket.receive_text()
    except WebSocketDisconnect: 
        manager.disconnect(websocket)
        logger.info(f"WebSocket with user_id {user_id} disconnected.")
    except Exception as e: 
        logger.error(f"WebSocket error for user_id {user_id}: {e}", exc_info=True)
        manager.disconnect(websocket)

# ==============================================================================
# Ø¨Ø®Ø´ Û²Û±: Static Files Ùˆ RouteÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ
# ==============================================================================

# âœ… Ø¨Ù‡Ø¨ÙˆØ¯: Ø§ÙØ²ÙˆØ¯Ù† Ù‡Ø¯Ø±Ù‡Ø§ÛŒ Ø§Ù…Ù†ÛŒØªÛŒ Ø¨Ù‡ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø§Ø³ØªØ§ØªÛŒÚ©
class SecureStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        # Ú©Ø´ Ú©Ø±Ø¯Ù† Ø¨Ø±Ø§ÛŒ ÛŒÚ© Ù‡ÙØªÙ‡ Ø¬Ù‡Øª Ø§ÙØ²Ø§ÛŒØ´ Ø³Ø±Ø¹Øª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø¨Ø±Ø§ÛŒ Ú©Ø§Ø±Ø¨Ø±Ø§Ù† ØªÚ©Ø±Ø§Ø±ÛŒ
        response.headers["Cache-Control"] = "public, max-age=604800, immutable" 
        return response

app.mount("/static", SecureStaticFiles(directory=STATIC_DIR, html=True), name="static")

# Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† headers Ø§Ù…Ù†ÛŒØªÛŒ Ø¨Ø±Ø§ÛŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø§Ø³ØªØ§ØªÛŒÚ©
class SecureStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        # Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† headers Ø§Ù…Ù†ÛŒØªÛŒ
        response.headers["Cache-Control"] = "public, max-age=3600"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        return response

app.mount("/static", SecureStaticFiles(directory=STATIC_DIR, html=True), name="static")

# --- ØµÙØ­Ù‡ Ù„Ø§Ú¯ÛŒÙ† (Ø¯Ø³ØªØ±Ø³ÛŒ Ø¹Ù…ÙˆÙ…ÛŒ) ---
@app.get("/login", response_class=FileResponse, include_in_schema=False)
async def read_login():
    """ØµÙØ­Ù‡ ÙˆØ±ÙˆØ¯ Ø¨Ù‡ Ø³ÛŒØ³ØªÙ…"""
    return FileResponse(os.path.join(STATIC_DIR, "login.html"))

# --- ØµÙØ­Ø§Øª Ø¯Ø§Ø®Ù„ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡ ---
@app.get("/dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_dashboard_page(): 
    """Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ø§ØµÙ„ÛŒ"""
    return FileResponse(os.path.join(STATIC_DIR, "dashboard.html"))

@app.get("/manager", response_class=HTMLResponse, include_in_schema=False)
async def read_manager_ui_page(): 
    """ØµÙØ­Ù‡ Ù…Ø¯ÛŒØ±ÛŒØª"""
    return FileResponse(os.path.join(STATIC_DIR, "manager.html"))

@app.get("/workshop", response_class=HTMLResponse, include_in_schema=False)
async def read_workshop_ui_page(): 
    """ØµÙØ­Ù‡ Ú©Ø§Ø±Ú¯Ø§Ù‡"""
    return FileResponse(os.path.join(STATIC_DIR, "workshop.html"))

@app.get("/personnel-management", response_class=HTMLResponse, include_in_schema=False)
async def read_personnel_management_page():
    """ØµÙØ­Ù‡ Ù…Ø¯ÛŒØ±ÛŒØª Ù¾Ø±Ø³Ù†Ù„"""
    return FileResponse(os.path.join(STATIC_DIR, "personnel_management.html"))

@app.get("/supervisor-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_supervisor_dashboard_page():
    """Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ù†Ø§Ø¸Ø±"""
    return FileResponse(os.path.join(STATIC_DIR, "supervisor_dashboard.html"))

# ==========================================================
# âœ…âœ…âœ… Ø§Ù†Ø¯Ù¾ÙˆÛŒÙ†Øª Ø¬Ø¯ÛŒØ¯ Ø¨Ø±Ø§ÛŒ Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ø§Ù†Ø¨Ø§Ø± âœ…âœ…âœ…
# ==========================================================
@app.get("/warehouse-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_warehouse_dashboard_page():
    """ØµÙØ­Ù‡ Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ø§Ù†Ø¨Ø§Ø±"""
    return FileResponse(os.path.join(STATIC_DIR, "warehouse_dashboard.html"))
# ==========================================================

@app.get("/reports", response_class=HTMLResponse, include_in_schema=False)
async def read_reports_page_html():
    """ØµÙØ­Ù‡ Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§"""
    fp = os.path.join(STATIC_DIR, "reports.html")
    if os.path.exists(fp):
        return FileResponse(fp)
    raise HTTPException(status_code=404, detail="reports.html not found")

@app.get("/assembler-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_assembler_dashboard_page():
    """Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ù…ÙˆÙ†ØªØ§Ú˜Ú©Ø§Ø±"""
    fp = os.path.join(STATIC_DIR, "assembler_dashboard.html")
    if os.path.exists(fp): 
        return FileResponse(fp)
    raise HTTPException(status_code=404, detail="assembler_dashboard.html not found")

@app.get("/daily-work-report", response_class=HTMLResponse, include_in_schema=False)
async def read_daily_work_report_page():
    """ØµÙØ­Ù‡ Ú¯Ø²Ø§Ø±Ø´ Ú©Ø§Ø± Ø±ÙˆØ²Ø§Ù†Ù‡"""
    return FileResponse(os.path.join(STATIC_DIR, "daily_work_report.html"))

@app.get("/employee-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_employee_dashboard_page():
    """Ø¯Ø§Ø´Ø¨ÙˆØ±Ø¯ Ú©Ø§Ø±Ù…Ù†Ø¯"""
    return FileResponse(os.path.join(STATIC_DIR, "employee_dashboard.html"))

# --- Ø±ÙˆØª Ø§ØµÙ„ÛŒ Ùˆ catch-all Ø¨Ø±Ø§ÛŒ Ø§Ù¾Ù„ÛŒÚ©ÛŒØ´Ù†â€ŒÙ‡Ø§ÛŒ ØªÚ©â€ŒØµÙØ­Ù‡â€ŒØ§ÛŒ (SPA) ---
@app.get("/", response_class=FileResponse, include_in_schema=False)
async def read_index():
    """ØµÙØ­Ù‡ Ø§ØµÙ„ÛŒ"""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="index.html not found.")
    return FileResponse(index_path)

@app.get("/{full_path:path}", response_class=FileResponse, include_in_schema=False)
async def serve_frontend_catch_all(request: Request, full_path: str):
    """Ù…Ø³ÛŒØ±Ù‡Ø§ÛŒ catch-all Ø¨Ø±Ø§ÛŒ SPA"""
    # Ø§ÛŒÙ† Ø±ÙˆØª ÙØ§ÛŒÙ„ index.html Ø±Ø§ Ø¨Ø±Ø§ÛŒ Ù…Ø³ÛŒØ±Ù‡Ø§ÛŒ Ù†Ø§Ø´Ù†Ø§Ø®ØªÙ‡ Ø¨Ø±Ù…ÛŒâ€ŒÚ¯Ø±Ø¯Ø§Ù†Ø¯
    # ØªØ§ Ù…Ù†Ø·Ù‚ Ø±ÙˆØªÛŒÙ†Ú¯ Ø¯Ø± ÙØ±Ø§Ù†Øªâ€ŒØ§Ù†Ø¯ Ø§Ù†Ø¬Ø§Ù… Ø´ÙˆØ¯.
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="index.html not found.")
    return FileResponse(index_path)

# ==============================================================================
# Ø¨Ø®Ø´ Û²Û²: Ø±Ø§Ù‡â€ŒØ§Ù†Ø¯Ø§Ø²ÛŒ Ø³Ø±ÙˆØ± Ùˆ Ù‡Ù†Ø¯Ù„Ø±Ù‡Ø§ÛŒ Ø®Ø·Ø§
# ==============================================================================

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Ù…Ø¯ÛŒØ±ÛŒØª ÛŒÚ©Ù¾Ø§Ø±Ú†Ù‡ Ø®Ø·Ø§Ù‡Ø§ÛŒ HTTP Ø¨Ø§ Ù„Ø§Ú¯â€ŒÚ¯ÛŒØ±ÛŒ Ø¯Ù‚ÛŒÙ‚."""
    logger.warning(f"HTTP Exception: {exc.status_code} {exc.detail} for URL: {request.url}")
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    """Ù…Ø¯ÛŒØ±ÛŒØª Ø®Ø·Ø§Ù‡Ø§ÛŒ Ù¾ÛŒØ´â€ŒØ¨ÛŒÙ†ÛŒ Ù†Ø´Ø¯Ù‡ Ø³Ø±ÙˆØ± (500)."""
    logger.error(f"Unhandled Exception for URL: {request.url}", exc_info=True)
    return JSONResponse(status_code=500, content={"detail": "An internal server error occurred."})

@app.on_event("startup")
async def startup_event():
    """Ø±ÙˆÛŒØ¯Ø§Ø¯Ù‡Ø§ÛŒ Ø²Ù…Ø§Ù† Ø±Ø§Ù‡â€ŒØ§Ù†Ø¯Ø§Ø²ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡."""
    logger.info("Application starting up...")
    try:
        with SessionLocal() as db:
            # âœ… Ø§ØµÙ„Ø§Ø­: Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² text() Ø¨Ø±Ø§ÛŒ Ø§Ø¬Ø±Ø§ÛŒ Ú©ÙˆØ¦Ø±ÛŒ Ø®Ø§Ù… Ø¬Ù‡Øª Ø§Ù…Ù†ÛŒØª Ùˆ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯ Ø¨ÙˆØ¯Ù†.
            db.execute(text("SELECT 1"))
        logger.info("Database connection verified successfully.")
    except Exception as e:
        logger.critical(f"FATAL: Database connection failed on startup: {e}")
        # Ø¯Ø± Ù…Ø­ÛŒØ· ÙˆØ§Ù‚Ø¹ÛŒØŒ Ù…Ù…Ú©Ù† Ø§Ø³Øª Ø¨Ø®ÙˆØ§Ù‡ÛŒØ¯ Ø¨Ø±Ù†Ø§Ù…Ù‡ Ø¯Ø± Ø§ÛŒÙ† Ø­Ø§Ù„Øª Ø®Ø§Ø±Ø¬ Ø´ÙˆØ¯.
        # raise SystemExit("Could not connect to the database.")

@app.on_event("shutdown")
async def shutdown_event():
    """Ø±ÙˆÛŒØ¯Ø§Ø¯Ù‡Ø§ÛŒ Ø²Ù…Ø§Ù† Ø®Ø§Ù…ÙˆØ´ Ø´Ø¯Ù† Ø¨Ø±Ù†Ø§Ù…Ù‡."""
    logger.info("Application shutting down...")

# ==============================================================================
# ØªØºÛŒÛŒØ±Ø§Øª Ø¨Ø±Ø§ÛŒ Liara/Production
# ==============================================================================

# ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù…Ø®ØµÙˆØµ production
if os.getenv("ENVIRONMENT") == "production":
    # ØºÛŒØ±ÙØ¹Ø§Ù„ Ú©Ø±Ø¯Ù† debug
    app.debug = False
    
    # ØªÙ†Ø¸ÛŒÙ… trusted hosts
    app.add_middleware(
        TrustedHostMiddleware, 
        allowed_hosts=os.getenv("ALLOWED_HOSTS", "localhost,127.0.0.1").split(",")
    )
    
    # CORS Ù…Ø­Ø¯ÙˆØ¯
    app.add_middleware(
        CORSMiddleware,
        allow_origins=os.getenv("ALLOWED_ORIGINS", "").split(","),
        allow_credentials=True,
        allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        allow_headers=["*"],
        max_age=3600,
    )
    
    # Ø§ÙØ²Ø§ÛŒØ´ timeoutÙ‡Ø§
    import uvicorn.config
    uvicorn.config.LOGGING_CONFIG["formatters"]["default"]["fmt"] = "%(asctime)s - %(name)s - %(levelname)s - %(message)s"