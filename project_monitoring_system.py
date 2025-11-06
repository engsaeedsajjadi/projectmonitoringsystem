# ==============================================================================
# بخش ۱: تنظیمات اولیه محیط (بسیار مهم)
# ==============================================================================
import sys
import asyncio

# این بخش برای سازگاری با ویندوز ضروری است و به درستی پیاده‌سازی شده.
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

# ==============================================================================
# بخش ۲: وارد کردن کتابخانه‌های استاندارد و شخص ثالث (Standard & Third-party)
# ==============================================================================
import os
import re
import json
import base64
import traceback
import logging
from io import BytesIO
from collections import defaultdict
from datetime import datetime, date, time, timedelta
from enum import Enum as PyEnum
from typing import List, Optional, Any, Dict, Tuple
from contextlib import contextmanager
import numpy as np

# کتابخانه‌های وب فریمورک (FastAPI)
from fastapi import (
    FastAPI, WebSocket, WebSocketDisconnect, Depends, HTTPException, 
    Response, status, File, UploadFile, Request, BackgroundTasks
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.httpsredirect import HTTPSRedirectMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.staticfiles import StaticFiles
# ✅ اصلاح: وارد کردن JSONResponse برای مدیریت بهتر خطاهای HTTP
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

# کتابخانه‌های دیتابیس (SQLAlchemy & Pydantic)
from sqlalchemy import (
    Column, Integer, String, DateTime, ForeignKey, Enum as SQLEnum, 
    create_engine, UniqueConstraint, event, Text, Date as SQLDateType, JSON,
    func, DDL, Index, text, case
)
from sqlalchemy.orm import sessionmaker, relationship, Session, declarative_base, joinedload
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from pydantic import BaseModel, Field, conint, field_validator, ValidationInfo, computed_field, ConfigDict

# کتابخانه‌های گزارش‌گیری و ابزارهای جانبی
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import qrcode
from jinja2 import Environment, FileSystemLoader, Template
from xhtml2pdf import pisa
import jdatetime
from dotenv import load_dotenv

# کتابخانه‌های امنیتی و احراز هویت
from passlib.context import CryptContext
from jose import JWTError, jwt

# ==============================================================================
# بخش ۳: وارد کردن ماژول‌های داخلی پروژه
# ==============================================================================
# from naab_connector_final import NaabConnector # این بخش به درستی کامنت شده است.

# ==============================================================================
# بخش ۴: تنظیمات و پیکربندی اولیه برنامه
# ==============================================================================
load_dotenv() # خواندن متغیرها از فایل .env

# ✅ اصلاح: تنظیمات لاگ‌گیری به ابتدای برنامه منتقل شد تا قبل از هر استفاده‌ای پیکربندی شده باشد.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# تنظیمات امنیتی - اجباری کردن متغیرهای محیطی
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    # حالا این لاگ به درستی کار می‌کند چون لاگر پیکربندی شده است.
    logger.error("CRITICAL: SECRET_KEY environment variable is not set. Application cannot start securely.")
    raise ValueError("SECRET_KEY must be set in environment variables for security.")

# 🔥 اصلاح قطعی: الگوریتم هشینگ JWT به صراحت تعریف شد تا از آسیب‌پذیری "alg:none" جلوگیری شود.
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 ساعت

# چاپ تنظیمات برای دیباگ (فقط در حالت توسعه)
if os.getenv("ENVIRONMENT") != "production":
    # ✅ بهبود: نمایش بخشی از کلید برای تأیید صحت آن بدون افشای کامل.
    print(f"🔐 SECRET_KEY: {'*' * (len(SECRET_KEY) - 4) + SECRET_KEY[-4:] if SECRET_KEY and len(SECRET_KEY) > 4 else 'SET (TOO SHORT)'}")
    print(f"🔐 ALGORITHM: {ALGORITHM}")
    print(f"🔐 ACCESS_TOKEN_EXPIRE_MINUTES: {ACCESS_TOKEN_EXPIRE_MINUTES}")

# تنظیمات CORS
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(",")
# ✅ بهبود: افزودن '*' به هاست‌های مجاز برای سهولت در توسعه (در محیط تولید باید محدود شود).
ALLOWED_HOSTS = os.getenv("ALLOWED_HOSTS", "localhost,127.0.0.1,*").split(",")

# تنظیمات NAAB (برای اعتبارسنجی)
NAAB_USERNAME = os.getenv("NAAB_USERNAME")
NAAB_PASSWORD = os.getenv("NAAB_PASSWORD")

# ✅ بهبود: استفاده از ثابت‌ها برای نام شرکت‌ها جهت خوانایی و نگهداری بهتر.
WEST_COMPANY = "شرکت پایدار نیرو نیکا"
EAST_COMPANY = "شرکت کامیاران ارم"

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# توابع کمکی
def normalize_text(text: str) -> str:
    """نرمال‌سازی و پاک‌سازی متن فارسی برای مقایسه بهتر."""
    if not isinstance(text, str): 
        return ""
    text = text.replace('ي', 'ی').replace('ك', 'ک')
    # حذف بخش‌هایی مانند "(بدون...)"
    text = text.split('(بدون')[0]
    # حذف فاصله‌های اضافی
    text = " ".join(text.split())
    return text

def resource_path(relative_path: str) -> str:
    """
    محاسبه مسیر مطلق منابع برای سازگاری با محیط‌های مختلف (عادی و PyInstaller).
    """
    try: 
        # حالت اجرا در PyInstaller
        base_path = sys._MEIPASS
    except AttributeError: 
        # حالت اجرای عادی
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

# تعریف مسیرها با استفاده از تابع کمکی
STATIC_DIR_NAME = "static"
STATIC_DIR = resource_path(STATIC_DIR_NAME)
DB_NAME = "projects.db"
DATABASE_URL = f"sqlite:///{resource_path(DB_NAME)}"

# ایجاد دایرکتوری استاتیک اگر وجود ندارد
if not os.path.exists(STATIC_DIR):
    try: 
        os.makedirs(STATIC_DIR, exist_ok=True)
        logger.info(f"Static directory created at: {STATIC_DIR}")
    except OSError as e: 
        logger.error(f"Could not create static directory {STATIC_DIR}: {e}")

# تنظیمات دیتابیس
engine = create_engine(
    DATABASE_URL, 
    # این آرگومان برای SQLite در محیط چندنخی ضروری است.
    connect_args={"check_same_thread": False},
    # برای بررسی سلامت کانکشن‌ها قبل از استفاده مجدد
    pool_pre_ping=True,
    # غیرفعال کردن لاگ کوئری‌ها در محیط تولید برای افزایش سرعت
    echo=(os.getenv("ENVIRONMENT") != "production")
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# تنظیمات Jinja2 برای رندر قالب‌های HTML
jinja_env = Environment(loader=FileSystemLoader(STATIC_DIR))
templates = Jinja2Templates(directory=STATIC_DIR)

# ✅ بهبود: Context manager برای مدیریت session دیتابیس بهینه شد تا لاگ بهتری داشته باشد.
@contextmanager
def get_db_session():
    """مدیریت session دیتابیس با context manager برای استفاده در تسک‌های پس‌زمینه."""
    db = SessionLocal()
    try:
        yield db
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Database session error occurred: {e}", exc_info=True)
        raise
    finally:
        db.close()

def get_db():
    """Dependency برای FastAPI جهت تزریق session دیتابیس به اندپوینت‌ها."""
    db = SessionLocal()
    try: 
        yield db
    finally: 
        db.close()
# ==============================================================================
# بخش ۵: Enumها و تنظیمات
# ==============================================================================

# ✅ بهبود: استفاده از Enumها برای مقادیر ثابت، کد را خواناتر و خطاپذیرتر می‌کند.
class StepNameKey(str, PyEnum):
    START_ASSEMBLY = "START_ASSEMBLY"
    END_ASSEMBLY = "END_ASSEMBLY"
    TEAM_LEAD_APPROVAL = "TEAM_LEAD_APPROVAL"
    TEST = "TEST"
    QUALITY_CONTROL = "QUALITY_CONTROL"
    SUPERVISOR_APPROVAL = "SUPERVISOR_APPROVAL"
    EXIT_PANEL = "EXIT_PANEL"
    
# ترتیب مراحل برای اعتبارسنجی
ORDERED_MANUAL_STEP_KEYS = [
    StepNameKey.START_ASSEMBLY, StepNameKey.END_ASSEMBLY, StepNameKey.TEAM_LEAD_APPROVAL,
    StepNameKey.TEST, StepNameKey.QUALITY_CONTROL, StepNameKey.SUPERVISOR_APPROVAL
]

# نگاشت کلیدها به نام‌های فارسی برای نمایش در UI
STEP_KEY_TO_NAME_MAP = {
    # ✅ بهبود: ساخت دیکشنری به روشی خواناتر
    key.value: name for key, name in [
        (StepNameKey.START_ASSEMBLY, "شروع مونتاژ"),
        (StepNameKey.END_ASSEMBLY, "پایان مونتاژ"),
        (StepNameKey.TEAM_LEAD_APPROVAL, "تایید سرگروه"),
        (StepNameKey.TEST, "تست سماک"),
        (StepNameKey.QUALITY_CONTROL, "کنترل کیفیت"),
        (StepNameKey.SUPERVISOR_APPROVAL, "تأیید ناظر"),
        (StepNameKey.EXIT_PANEL, "خروج تابلو"),
    ]
}

class PanelTypeKey(str, PyEnum):
    FAHAM_WITH_FRAME = "FAHAM_WITH_FRAME"
    FAHAM_WITHOUT_FRAME = "FAHAM_WITHOUT_FRAME"
    ID2R = "ID2R"
    ID5R = "ID5R"
    ID116 = "ID116"
    ID6_1R = "ID6+1R"
    ID12_1R = "ID12+1R"
    ID18_1R = "ID18+1R"
    ID24_1R = "ID24+1R"
    ID101_1 = "ID101.1"
    ID101_3 = "ID101.3"
    ID102_1 = "ID102.1"
    ID102_3 = "ID102.3"
    ID104_1 = "ID104.1"
    ID104_3 = "ID104.3"
    ID105 = "ID105"
    ID107 = "ID107"
    ID115 = "ID115"
    ID108 = "ID108"
    ID109 = "ID109"
    ID110 = "ID110"
    ID111 = "ID111"
    ID112_STAR = "ID112_STAR"
    ID120 = "ID120"
    ID121 = "ID121"
    ID122 = "ID122"
    ID123 = "ID123"
    ID124_STAR = "ID124_STAR"
    ID211 = "ID211"
    ID212 = "ID212"
    ID213 = "ID213"
    ID214 = "ID214"
    ID215 = "ID215"
    ID216 = "ID216"
    ID218 = "ID218"

PANEL_TYPE_NAMES = {
    PanelTypeKey.FAHAM_WITH_FRAME: "فهام با قاب",
    PanelTypeKey.FAHAM_WITHOUT_FRAME: "فهام بدون قاب",
    PanelTypeKey.ID2R: "ID2R - تابلو کامپوزیتی 1-2 کنتور تکفاز ریلی",
    PanelTypeKey.ID5R: "ID5R - تابلو کامپوزیتی 3-5 کنتور تکفاز ریلی",
    PanelTypeKey.ID116: "ID116 - تابلو 2 کنتوره تکفاز ریلی روی دیوار",
    PanelTypeKey.ID6_1R: "ID6+1R - تابلو 6 کنتور فلزی دیواری",
    PanelTypeKey.ID12_1R: "ID12+1R - تابلو 12 کنتور فلزی دیواری",
    PanelTypeKey.ID18_1R: "ID18+1R - تابلو 18 کنتور فلزی دیواری",
    PanelTypeKey.ID24_1R: "ID24+1R - تابلو 24 کنتور فلزی دیواری",
    PanelTypeKey.ID101_1: "ID101.1 - تک کنتور هوایی تکفاز (فیوز در محل)",
    PanelTypeKey.ID101_3: "ID101.3 - تک کنتور هوایی سه فاز (فیوز در محل)",
    PanelTypeKey.ID102_1: "ID102.1 - تک کنتور هوایی تکفاز (فیوزدار)",
    PanelTypeKey.ID102_3: "ID102.3 - تک کنتور هوایی سه فاز (فیوزدار)",
    PanelTypeKey.ID104_1: "ID104.1 - تک کنتور زمینی تکفاز (فیوز روی پایه)",
    PanelTypeKey.ID104_3: "ID104.3 - تک کنتور زمینی سه فاز (فیوز روی پایه)",
    PanelTypeKey.ID105: "ID105 - تابلو کامپوزیتی زمینی تک کنتوره یکطرفه",
    PanelTypeKey.ID107: "ID107 - تابلو کامپوزیتی دیواری تک کنتوره سه فاز با فیوز",
    PanelTypeKey.ID115: "ID115 - تابلو تک کنتور دیواری - فیوز در محل",
    PanelTypeKey.ID108: "ID108 - تابلو زمینی چند کنتوره یکطرفه",
    PanelTypeKey.ID109: "ID109 - تابلو زمینی چند کنتوره دوطرفه",
    PanelTypeKey.ID110: "ID110 - تابلو 2 کنتوره تکفاز (هوایی) - فیوز در محل",
    PanelTypeKey.ID111: "ID111 - تابلو 2 کنتوره تکفاز (زمینی) - فیوز روی پایه",
    PanelTypeKey.ID112_STAR: "ID112* - تابلو چند کنتوره تک فاز روی پایه (کلی)",
    PanelTypeKey.ID120: "ID120 - تابلو 2 کنتوره سه فاز (هوایی) - فیوز در محل",
    PanelTypeKey.ID121: "ID121 - تابلو 2 کنتوره سه فاز (زمینی) - فیوز روی پایه",
    PanelTypeKey.ID122: "ID122 - 2x تابلو 2 کنتوره سه فاز - جعبه 8 فیوز",
    PanelTypeKey.ID123: "ID123 - 2x تابلو 2 کنتوره سه فاز - جعبه 16 فیوز",
    PanelTypeKey.ID124_STAR: "ID124* - تابلو چند کنتوره سه فاز روی پایه (کلی)",
    PanelTypeKey.ID211: "ID211 - تابلو دیماندی هوایی 30-150 kW",
    PanelTypeKey.ID212: "ID212 - تابلو دیماندی هوایی 151-249 kW",
    PanelTypeKey.ID213: "ID213 - تابلو دیماندی زمینی یکطرفه",
    PanelTypeKey.ID214: "ID214 - تابلو دیماندی زمینی دوطرفه",
    PanelTypeKey.ID215: "ID215 - تابلو دیماندی فلزی زمینی",
    PanelTypeKey.ID216: "ID216 - تابلو دو دیماندی هوایی",
    PanelTypeKey.ID218: "ID218 - تابلو چند دیماندی زمینی دوطرفه",
}

class TransactionType(str, PyEnum):
    IN = "IN"    # ورود به انبار
    OUT = "OUT"  # خروج از انبار

# ✅ بهبود: ایجاد Enum برای نقش‌های کاربران جهت جلوگیری از خطای تایپی و افزایش خوانایی.
class PersonnelRole(str, PyEnum):
    EMPLOYEE = "employee"
    SUPERVISOR = "supervisor"
    
# ==============================================================================
# بخش ۶: مدل‌های دیتابیس
# ==============================================================================
# ✅ بهبود: افزودن کامنت‌های توضیحی به مدل‌ها برای درک بهتر ساختار دیتابیس.

class Comment(Base):
    """مدل نظرات ثبت‌شده برای هر پروژه."""
    __tablename__ = "comments"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    text = Column(Text, nullable=False)
    author = Column(String, nullable=False, default="اپراتور")
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    
    project = relationship("Project", back_populates="comments")

class Project(Base):
    """مدل اصلی پروژه."""
    __tablename__ = "projects"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True, nullable=False)
    location = Column(String, nullable=True)
    request_id = Column(String, index=True, nullable=False, unique=True)
    customer_name = Column(String, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    barcode_payload = Column(String, nullable=True, unique=True, index=True)
    panel_type_key = Column(String, nullable=True)
    panel_code = Column(String, index=True, nullable=True)
    assembler_1 = Column(String, nullable=True)
    assembler_2 = Column(String, nullable=True)
    
    # ✅ بهبود: استفاده از lazy="selectin" برای پیش‌بارگذاری روابط و جلوگیری از N+1 queries.
    steps = relationship("Step", back_populates="project", cascade="all, delete-orphan", lazy="selectin", order_by="Step.timestamp")
    equipment = relationship("EquipmentItem", back_populates="project", cascade="all, delete-orphan", lazy="selectin")
    comments = relationship("Comment", back_populates="project", cascade="all, delete-orphan", lazy="selectin", order_by="Comment.timestamp.desc()")
    
    __table_args__ = (
        UniqueConstraint('request_id', name='uq_project_request_id'),
        Index('idx_project_created_at', 'created_at'),
        Index('idx_project_panel_code', 'panel_code'),
    )

class Step(Base):
    """مدل مراحل انجام‌شده برای هر پروژه."""
    __tablename__ = "steps"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    name_key = Column(SQLEnum(StepNameKey, name="step_name_key_enum"), nullable=False)
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    
    project = relationship("Project", back_populates="steps")
    
    __table_args__ = (
        UniqueConstraint('project_id', 'name_key', name='uq_project_step'),
        Index('idx_step_timestamp', 'timestamp'),
        Index('idx_step_name_key', 'name_key'),
    )

class EquipmentItem(Base):
    """مدل تجهیزات مورد استفاده در هر پروژه."""
    __tablename__ = "equipment_items"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    item_name = Column(Text, nullable=False)
    quantity = Column(Integer, nullable=False, default=1)
    
    project = relationship("Project", back_populates="equipment")
    
    __table_args__ = (
        Index('idx_equipment_project_id', 'project_id'),
    )

class Personnel(Base):
    """مدل پرسنل و کاربران سیستم."""
    __tablename__ = "personnel"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False, unique=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    password_hash = Column(String, nullable=False)
    # ✅ اصلاح: استفاده از مقدار Enum برای ستون نقش
    role = Column(String, default=PersonnelRole.EMPLOYEE.value, nullable=False)
    is_active = Column(Integer, default=1, nullable=False)
    
    daily_reports = relationship("DailyWorkReport", back_populates="personnel")
    
    __table_args__ = (
        Index('idx_personnel_username', 'username'),
        Index('idx_personnel_role', 'role'),
    )

class DailyWorkReport(Base):
    """مدل گزارش کار روزانه پرسنل."""
    __tablename__ = "daily_work_reports"
    id = Column(Integer, primary_key=True, index=True)
    personnel_id = Column(Integer, ForeignKey("personnel.id"), nullable=False)
    report_date = Column(SQLDateType, nullable=False)
    status = Column(String, default='submitted', nullable=False)  # 'submitted', 'approved', 'rejected'
    supervisor_notes = Column(Text, nullable=True)
    report_data = Column(JSON, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    personnel = relationship("Personnel", back_populates="daily_reports")
    
    __table_args__ = (
        UniqueConstraint('personnel_id', 'report_date', name='uq_personnel_date_report'),
        Index('idx_report_date', 'report_date'),
        Index('idx_report_status', 'status'),
    )

class Warehouse(Base):
    """مدل انبارها."""
    __tablename__ = "warehouses"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, nullable=False, index=True)
    location = Column(String, nullable=True)
    description = Column(Text, nullable=True)

class WarehouseItem(Base):
    """مدل کالاهای قابل ذخیره در انبار."""
    __tablename__ = "warehouse_items"
    id = Column(Integer, primary_key=True, index=True)
    item_name = Column(String, unique=True, nullable=False, index=True)
    description = Column(Text, nullable=True)
    min_stock_level = Column(Integer, default=0, nullable=False)

class InventoryTransaction(Base):
    """مدل تراکنش‌های انبار (ورود و خروج)."""
    __tablename__ = "inventory_transactions"
    id = Column(Integer, primary_key=True, index=True)
    warehouse_id = Column(Integer, ForeignKey("warehouses.id"), nullable=False)
    item_id = Column(Integer, ForeignKey("warehouse_items.id"), nullable=False)
    quantity = Column(Integer, nullable=False)
    transaction_type = Column(SQLEnum(TransactionType, name="transaction_type_enum"), nullable=False)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="SET NULL"), nullable=True)
    user_id = Column(Integer, ForeignKey("personnel.id"), nullable=False)
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    notes = Column(Text, nullable=True)

    warehouse = relationship("Warehouse")
    item = relationship("WarehouseItem")
    project = relationship("Project")
    user = relationship("Personnel")

class PanelCodeItems(Base):
    """مدل تجهیزات مورد نیاز برای هر کد تابلو (Bill of Materials - BOM)."""
    __tablename__ = "panel_code_items"
    id = Column(Integer, primary_key=True, index=True)
    panel_code = Column(String, nullable=False, index=True)
    item_name = Column(String, nullable=False)
    quantity_required = Column(Integer, nullable=False)
    
    __table_args__ = (
        UniqueConstraint('panel_code', 'item_name', name='uq_panel_item'),
    )

# ایجاد جداول و ایندکس‌ها
def create_tables_and_indexes():
    """ایجاد تمام جداول و ایندکس‌های تعریف‌شده در دیتابیس."""
    try:
        Base.metadata.create_all(bind=engine)
        logger.info("Database tables and indexes created/verified successfully.")
    except Exception as e:
        logger.error(f"FATAL: Error creating database tables: {e}")
        raise

create_tables_and_indexes()

def create_default_supervisor():
    """ایجاد کاربر supervisor پیش‌فرض برای اولین اجرای برنامه"""
    DEFAULT_ADMIN_USER = os.getenv("DEFAULT_ADMIN_USER", "admin")
    DEFAULT_ADMIN_PASSWORD = os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123")
    
    try:
        with SessionLocal() as db:
            existing_supervisor = db.query(Personnel).filter(
                Personnel.role == PersonnelRole.SUPERVISOR.value,
                Personnel.is_active == 1
            ).first()
            
            if not existing_supervisor:
                supervisor = Personnel(
                    name="مدیر سیستم",
                    username=DEFAULT_ADMIN_USER,
                    password_hash=pwd_context.hash(DEFAULT_ADMIN_PASSWORD),
                    role=PersonnelRole.SUPERVISOR.value,
                    is_active=1
                )
                db.add(supervisor)
                db.commit()
                logger.info(f"Default supervisor user '{DEFAULT_ADMIN_USER}' created.")
                print("=" * 60)
                print("✅ کاربر supervisor پیش‌فرض ایجاد شد.")
                print(f"   نام کاربری: {DEFAULT_ADMIN_USER}")
                print(f"   رمز عبور: {DEFAULT_ADMIN_PASSWORD}")
                print("=" * 60)
            else:
                logger.info(f"Default supervisor user already exists: {existing_supervisor.username}")
                
    except Exception as e:
        logger.error(f"Error creating default supervisor: {e}")
        print(f"⚠️ خطا در ایجاد کاربر پیش‌فرض: {e}")

def create_default_warehouse():
    """ایجاد انبار پیش‌فرض برای اولین اجرای برنامه"""
    try:
        with SessionLocal() as db:
            existing_warehouse = db.query(Warehouse).first()
            
            if not existing_warehouse:
                warehouse = Warehouse(
                    name="انبار اصلی",
                    location="ساختمان مرکزی",
                    description="انبار اصلی شرکت"
                )
                db.add(warehouse)
                db.commit()
                logger.info("Default warehouse created.")
                print("=" * 50)
                print("✅ انبار پیش‌فرض ایجاد شد.")
                print("   نام: انبار اصلی")
                print("   مکان: ساختمان مرکزی")
                print("=" * 50)
            else:
                logger.info(f"Default warehouse already exists: {existing_warehouse.name}")
                
    except Exception as e:
        logger.error(f"Error creating default warehouse: {e}")
        print(f"⚠️ خطا در ایجاد انبار پیش‌فرض: {e}")

# ایجاد کاربر و انبار پیش‌فرض
create_default_supervisor()
create_default_warehouse()


# ==============================================================================
# بخش ۷: Pydantic Schemas
# ==============================================================================
# ✅ بهبود: استفاده از ConfigDict به جای Config class که در Pydantic v2 استاندارد است.
# ✅ بهبود: افزودن توضیحات (description) به فیلدها برای مستندسازی بهتر API در Swagger/OpenAPI.

class ValidationResponse(BaseModel):
    has_discrepancy: bool
    message: str

class CommentCreate(BaseModel):
    text: str = Field(..., min_length=1, description="متن نظر")

class CommentOut(BaseModel):
    id: int
    text: str
    author: str
    timestamp: datetime
    
    model_config = ConfigDict(from_attributes=True)

class StepOut(BaseModel):
    name_key: StepNameKey
    timestamp: datetime

    @computed_field
    @property
    def name(self) -> str:
        return STEP_KEY_TO_NAME_MAP.get(self.name_key.value, self.name_key.value)
    
    model_config = ConfigDict(from_attributes=True)

class EquipmentItemBase(BaseModel):
    item_name: str
    quantity: int

class EquipmentItemCreate(EquipmentItemBase): 
    pass

class EquipmentItemOut(EquipmentItemBase):
    id: int
    project_id: int
    
    model_config = ConfigDict(from_attributes=True)

class ProjectOut(BaseModel):
    id: int
    name: str
    location: Optional[str] = None
    request_id: str
    customer_name: str
    created_at: datetime
    barcode_payload: Optional[str] = None
    steps: List[StepOut] = []
    equipment: List[EquipmentItemOut] = []
    panel_type_key: Optional[str] = None
    panel_type_name: Optional[str] = None
    panel_code: Optional[str] = None
    assembler_1: Optional[str] = None
    assembler_2: Optional[str] = None
    comments: List[CommentOut] = []
    
    model_config = ConfigDict(from_attributes=True)
    
    @field_validator('panel_type_name', mode='before')
    @classmethod
    def set_panel_type_name(cls, v: Any, info: ValidationInfo) -> Optional[str]:
        key = info.data.get('panel_type_key')
        if key:
            try:
                # ✅ اصلاح: اعتبارسنجی با Enum برای اطمینان از صحت کلید
                enum_key = PanelTypeKey(key)
                return PANEL_TYPE_NAMES.get(enum_key, f"کد نامعتبر: {key}")
            except ValueError: 
                return f"کد نامعتبر: {key}"
        return None

class ProjectCreateFromExcelData(BaseModel):
    name: str
    request_id: str
    customer_name: str
    location: Optional[str] = None

class ProjectCreateManual(BaseModel):
    name: str
    location: Optional[str] = None
    customer_name: Optional[str] = Field(None)
    request_id: Optional[str] = Field(None)

class StepCreate(BaseModel):
    step: StepNameKey
    
    @field_validator('step', mode='before')
    @classmethod
    def validate_step_from_string(cls, v: Any) -> StepNameKey:
        if isinstance(v, StepNameKey): 
            return v
        if isinstance(v, str):
            try:
                # ✅ بهبود: استفاده از مقدار Enum برای اعتبارسنجی
                return StepNameKey(v)
            except ValueError:
                raise ValueError(f"'{v}' is not a valid step name or key.")
        raise TypeError("Step must be a string or StepNameKey enum.")

class BarcodeExitPayload(BaseModel):
    barcode_data: str

class AssemblyDetailsUpdate(BaseModel):
    panel_type_key: PanelTypeKey
    assembler_1: str = Field(..., min_length=1)
    assembler_2: Optional[str] = None
    
    @field_validator('panel_type_key', mode='before')
    @classmethod
    def validate_panel_type_from_key(cls, v: str):
        try: 
            return PanelTypeKey(v)
        except ValueError: 
            raise ValueError(f"'{v}' یک کلید معتبر برای نوع تابلو نیست.")

class AssemblerStatsOut(BaseModel):
    total_panels: int
    panels_by_type: Dict[str, int]

class PersonnelCreate(BaseModel):
    name: str
    username: str
    password: str
    # ✅ اصلاح: استفاده از Enum برای نقش
    role: PersonnelRole = PersonnelRole.EMPLOYEE

class PersonnelUpdate(BaseModel):
    name: str
    username: str
    password: Optional[str] = None
    # ✅ اصلاح: استفاده از Enum برای نقش
    role: PersonnelRole = PersonnelRole.EMPLOYEE

class PersonnelLogin(BaseModel):
    username: str
    password: str

class PersonnelOut(BaseModel):
    id: int
    name: str
    username: str
    role: str
    is_active: int

    model_config = ConfigDict(from_attributes=True)

class Token(BaseModel):
    access_token: str
    token_type: str
    user: PersonnelOut

class TokenData(BaseModel):
    username: Optional[str] = None
    role: Optional[str] = None

# Daily Work Report Schemas
class DailyWorkReportCreate(BaseModel):
    personnel_id: int
    report_date: date
    report_data: Dict[str, Any]

class DailyWorkReportOut(BaseModel):
    id: int
    personnel_id: int
    report_date: date
    status: str
    supervisor_notes: Optional[str] = None
    report_data: Dict[str, Any]
    created_at: datetime
    updated_at: datetime
    personnel: PersonnelOut

    model_config = ConfigDict(from_attributes=True)

class KpiSummary(BaseModel):
    total_projects: int
    completed_this_month: int
    avg_completion_time_days: Optional[float] = None
    bottleneck_step: Optional[str] = None
    step_durations: Dict[str, Optional[float]]

class WarehouseBase(BaseModel):
    name: str
    location: Optional[str] = None
    description: Optional[str] = None

class WarehouseCreate(WarehouseBase):
    pass

class WarehouseOut(WarehouseBase):
    id: int
    model_config = ConfigDict(from_attributes=True)

class WarehouseItemBase(BaseModel):
    item_name: str
    description: Optional[str] = None
    min_stock_level: int = Field(0, ge=0)

class WarehouseItemCreate(WarehouseItemBase):
    pass

class WarehouseItemOut(WarehouseItemBase):
    id: int
    model_config = ConfigDict(from_attributes=True)

class InventoryTransactionBase(BaseModel):
    warehouse_id: int
    item_name: str
    quantity: int = Field(..., gt=0)
    notes: Optional[str] = None

class InventoryTransactionIn(InventoryTransactionBase):
    pass

class InventoryTransactionOutManual(InventoryTransactionBase):
    project_id: Optional[int] = None

class InventoryTransactionOut(BaseModel):
    id: int
    warehouse: WarehouseOut
    item: WarehouseItemOut
    quantity: int
    transaction_type: TransactionType
    project_id: Optional[int] = None
    user: PersonnelOut
    timestamp: datetime
    notes: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)
    
class CurrentStockItem(BaseModel):
    item_id: int
    item_name: str
    warehouse_id: int
    warehouse_name: str
    current_stock: int
    min_stock_level: int

class PanelCodeItemBase(BaseModel):
    item_name: str
    quantity_required: int = Field(..., gt=0)

class PanelCodeItemCreate(PanelCodeItemBase):
    pass

class PanelCodeItemsDefinition(BaseModel):
    panel_code: str
    items: List[PanelCodeItemCreate]

class PanelCodeItemOut(PanelCodeItemBase):
    id: int
    panel_code: str
    model_config = ConfigDict(from_attributes=True)
    
# ==============================================================================
# بخش ۸: توابع احراز هویت و امنیت (نسخه کاملاً اصلاح شده)
# ==============================================================================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """تأیید صحت رمز عبور با هش ذخیره‌شده."""
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    """هش کردن رمز عبور برای ذخیره‌سازی امن."""
    return pwd_context.hash(password)

def authenticate_user(db: Session, username: str, password: str) -> Optional[Personnel]:
    """احراز هویت کاربر بر اساس نام کاربری و رمز عبور."""
    try:
        # ✅ بهبود: بررسی فعال بودن کاربر در همان کوئری
        user = db.query(Personnel).filter(Personnel.username == username, Personnel.is_active == 1).first()
        if not user or not verify_password(password, user.password_hash):
            logger.warning(f"Authentication failed for user: {username}")
            return None
        logger.info(f"Authentication successful for user: {username}")
        return user
    except Exception as e:
        logger.error(f"Error during authentication for user {username}: {e}", exc_info=True)
        return None

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    """ایجاد توکن دسترسی JWT."""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        # ✅ بهبود: تعیین زمان انقضای پیش‌فرض
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    logger.debug(f"Access token created for: {data.get('sub')}")
    return encoded_jwt

# استثنای استاندارد برای خطاهای احراز هویت
credentials_exception = HTTPException(
    status_code=status.HTTP_401_UNAUTHORIZED,
    detail="Could not validate credentials",
    headers={"WWW-Authenticate": "Bearer"},
)

# ✅ بهبود: یکپارچه‌سازی توابع get_current_user. این تابع پایه، توکن را رمزگشایی کرده و کاربر را برمی‌گرداند.
async def get_current_active_user(
    credentials: HTTPAuthorizationCredentials = Depends(security), 
    db: Session = Depends(get_db)
) -> Personnel:
    """وابستگی (Dependency) برای دریافت کاربر فعال فعلی از توکن JWT."""
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username, role=payload.get("role"))
    except JWTError as e:
        logger.warning(f"JWT decode error: {e}")
        raise credentials_exception
    
    user = db.query(Personnel).filter(Personnel.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    if user.is_active != 1:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Inactive user")
    return user

# ✅ بهبود: وابستگی‌های جدید برای بررسی نقش‌ها که خواناتر و قابل استفاده مجدد هستند.
async def get_current_user(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """وابستگی برای دریافت کاربر فعلی (بدون بررسی نقش). جایگزین get_current_user قدیمی."""
    return current_user

async def get_current_employee(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """وابستگی برای اطمینان از اینکه کاربر فعلی 'employee' یا 'supervisor' است."""
    if current_user.role not in [PersonnelRole.EMPLOYEE.value, PersonnelRole.SUPERVISOR.value]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Operation not permitted for this role")
    return current_user

async def get_current_supervisor(current_user: Personnel = Depends(get_current_active_user)) -> Personnel:
    """وابستگی برای اطمینان از اینکه کاربر فعلی 'supervisor' است."""
    if current_user.role != PersonnelRole.SUPERVISOR.value:
        logger.warning(f"Access denied for user '{current_user.username}'. Required role: 'supervisor', User role: '{current_user.role}'")
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: Supervisor role required.")
    return current_user

async def get_user_by_username(db: Session, username: str) -> Optional[Personnel]:
    """یافتن کاربر بر اساس نام کاربری."""
    return db.query(Personnel).filter(Personnel.username == username).first()

# ==============================================================================
# بخش ۹: WebSocket Connection Manager
# ==============================================================================

class ConnectionManager:
    def __init__(self): 
        # ✅ بهبود: ساختار داده برای ذخیره WebSocket و user_id مرتبط با آن.
        self.active_connections: Dict[WebSocket, int] = {} # WebSocket -> user_id (0 for anonymous)
    
    async def connect(self, websocket: WebSocket, user_id: int = 0):
        await websocket.accept()
        self.active_connections[websocket] = user_id
        logger.info(f"WebSocket connected. User ID: {user_id}. Total connections: {len(self.active_connections)}")
    
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            del self.active_connections[websocket]
            logger.info(f"WebSocket disconnected. Total connections: {len(self.active_connections)}")
    
    async def send_to_user(self, user_id: int, message: dict):
        # ارسال پیام به یک کاربر خاص
        for ws, uid in self.active_connections.items():
            if uid == user_id:
                try:
                    await ws.send_json(message)
                except Exception as e:
                    logger.error(f"Error sending message to user {user_id}: {e}")
                    self.disconnect(ws)

    async def broadcast_to_supervisors(self, message: dict):
        """ارسال پیام به تمام کاربران متصل. در این پیاده‌سازی ساده، تفکیک نقش در سمت کلاینت انجام می‌شود."""
        await self.broadcast(message)

    async def broadcast(self, message: dict):
        # ✅ بهبود: استفاده از asyncio.gather برای ارسال همزمان و مدیریت بهتر قطع اتصال.
        if not self.active_connections:
            return
            
        disconnected_sockets = []
        # ایجاد یک کپی برای جلوگیری از خطا هنگام تغییر دیکشنری در حین پیمایش
        connections_to_send = list(self.active_connections.keys())

        # ارسال همزمان پیام‌ها
        results = await asyncio.gather(
            *[conn.send_json(message) for conn in connections_to_send], 
            return_exceptions=True
        )

        # بررسی نتایج و حذف اتصالات قطع شده
        for i, result in enumerate(results):
            if isinstance(result, Exception): 
                logger.error(f"WebSocket send error: {result}")
                disconnected_sockets.append(connections_to_send[i])
        
        for ws in disconnected_sockets:
            self.disconnect(ws)

manager = ConnectionManager()

# ==============================================================================
# بخش ۱۰: FastAPI App و Middleware
# ==============================================================================

app = FastAPI(
    title="Project Monitoring System",
    description="سیستم مانیتورینگ پروژه‌های تولید تابلوهای برق",
    version="2.1.0" # ✅ نسخه به‌روز شد.
)

# افزودن middleware های امنیتی
if os.getenv("ENVIRONMENT") == "production":
    app.add_middleware(HTTPSRedirectMiddleware)

app.add_middleware(
    TrustedHostMiddleware, 
    allowed_hosts=ALLOWED_HOSTS
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    max_age=3600,
)

# ==============================================================================
# بخش ۱۱: تشخیص نوع تابلو
# ==============================================================================
# 🔥 نکته: این بخش منطق تجاری بسیار شکننده‌ای دارد. بهترین راهکار در بلندمدت،
# استفاده از کدهای استاندارد تجهیزات به جای تطبیق رشته است.
# این بخش بدون تغییر باقی می‌ماند چون منطق تجاری خاص پروژه است.

ITEM_ID_TO_KEYWORDS = {
    'ID2R': ("جعبه", "کامپوزیتی", "دیواری", "ظرفیت", "دو کنتور", "تکفاز", "ریلی"),
    '140': ("جعبه", "کنتور", "کامپوزیتی", "دیواری", "تکفاز"),
    '141': ("جعبه", "کنتور", "کامپوزیتی", "دیواری", "سه فاز"),
    '146': ("جعبه", "کنتور", "کامپوزیتی", "هوایی", "تکفاز", "با فیوز"),
    '147': ("جعبه", "کنتور", "کامپوزیتی", "هوایی", "سه فاز", "با فیوز"),
    '120': ("جعبه", "کنتور", "کامپوزیتی", "تکفاز"),
    '121': ("جعبه", "کنتور", "کامپوزیتی", "سه فاز"),
    'FAHAM_METER_ASSEMBLED': ("کنتور", "تکفاز", "هوشمند", "مونتاژ", "شده"),
    'FAHAM_METER_GPRS': ("کنتور", "تکفاز", "هوشمند", "ماژول", "GPRS"),
    'FUSE_SP_ANY': ("کلید", "فیوز", "مینیاتوری", "تکفاز"),
    '103': ("جعبه", "فیوز", "تکفاز", "دیواری"),
    '104': ("جعبه", "فیوز", "سه فاز", "دیواری"),
    '105': ("سکوی", "نصب"),
    '116': ("جعبه", "فیوز", "کامپوزیتی", "تکفاز", "نصب روی پایه"),
    '117': ("جعبه", "فیوز", "کامپوزیتی", "سه فاز", "نصب روی پایه"),
    '122': ("تابلو", "کامپوزیتی", "تک کنتوره", "سه فاز", "تکفاز", "سایز", "۳۶"),
    '130': ("تابلو", "کنتور", "کامپوزیتی", "هوایی", "ظرفیت ۴ کنتور", "تکفاز", "ریلی"),
    '134': ("تابلو", "کامپوزیتی", "هوایی", "۴ کنتوره", "سه فاز", "کنتور هوشمند"),
    '135': ("تابلو", "کامپوزیتی", "زمینی", "ظرفیت ۸ فیوز"),
    '136': ("تابلو", "کامپوزیتی", "زمینی", "ظرفیت ۱۶ فیوز"),
    '137': ("تابلو", "کامپوزیتی", "زمینی", "چند کنتوره", "یکطرفه"),
    '138': ("تابلو", "کامپوزیتی", "زمینی", "چند کنتوره", "دوطرفه"),
    '142': ("تابلو", "کنتور", "کامپوزیتی", "دیواری", "ظرفیت ۴ کنتور", "تکفاز", "ریلی"),
    '170': ("تابلو", "کامپوزیتی", "تک دیماندی", "هوایی", "موتور دار", "۳۰ تا ۱۵۰"),
    '171': ("تابلو", "کامپوزیتی", "تک دیماندی", "هوایی", "موتور دار", "۱۵۱ تا ۲۴۹"),
    '172': ("تابلو", "کامپوزیتی", "تک دیماندی", "زمینی", "یکطرفه"),
    '173': ("تابلو", "فلزی", "تک دیماندی", "زمینی", "یکطرفه"),
    '174': ("تابلو", "کامپوزیتی", "دو دیماندی", "هوایی", "موتوردار"),
    '176': ("تابلو", "کامپوزیتی", "تک دیماندی", "زمینی", "دوطرفه"),
}

PANEL_ID_RULES = {
    PanelTypeKey.FAHAM_WITH_FRAME: {"requires": {'FAHAM_METER_ASSEMBLED', 'FUSE_SP_ANY'}, "forbids": {'120', '121', '140', '141', '146', '147'}},
    PanelTypeKey.FAHAM_WITHOUT_FRAME: {"requires": {'FAHAM_METER_GPRS', 'FUSE_SP_ANY'}, "forbids": {'120', '121', '140', '141', '146', '147'}},
    PanelTypeKey.ID218: {"requires": {'176'}},
    PanelTypeKey.ID216: {"requires": {'174'}},
    PanelTypeKey.ID215: {"requires": {'173'}},
    PanelTypeKey.ID213: {"requires": {'172'}},
    PanelTypeKey.ID212: {"requires": {'171'}},
    PanelTypeKey.ID211: {"requires": {'170'}},
    PanelTypeKey.ID123: {"requires": {'134', '136'}},
    PanelTypeKey.ID122: {"requires": {'134', '135'}},
    PanelTypeKey.ID109: {"requires": {'138'}},
    PanelTypeKey.ID108: {"requires": {'137'}},
    PanelTypeKey.ID105: {"requires": {'122'}},
    PanelTypeKey.ID2R: {"requires": {'ID2R'}},
    PanelTypeKey.ID116: {"requires": {'142'}},
    PanelTypeKey.ID107: {"requires": {'141', '104'}}, 
    PanelTypeKey.ID115: {"requires": {'140', '103'}},
    PanelTypeKey.ID121: {"requires": {'134', '117'}},
    PanelTypeKey.ID120: {"requires": {'134', '104'}, "forbids": {'117'}},
    PanelTypeKey.ID111: {"requires": {'130', '116'}},
    PanelTypeKey.ID110: {"requires": {'130', '103'}, "forbids": {'116'}},
    PanelTypeKey.ID104_3: {"requires": {'121', '117'}, "count": {'105': 2}},
    PanelTypeKey.ID104_1: {"requires": {'120', '116'}, "count": {'105': 2}},
    PanelTypeKey.ID102_3: {"requires": {'147', '105'}},
    PanelTypeKey.ID102_1: {"requires": {'146', '105'}},
    PanelTypeKey.ID101_3: {"requires": {'121'}, "forbids": {'147'}, "count": {'105': 1}},
    PanelTypeKey.ID101_1: {"requires": {'120'}, "forbids": {'146'}, "count": {'105': 1}},
}

EQUIPMENT_KEYWORD_TO_PANEL_MAPPING = {
    ("6 کنتور", "فلزی", "دیواری"): PanelTypeKey.ID6_1R,
    ("12 کنتور", "فلزی", "دیواری"): PanelTypeKey.ID12_1R,
    ("18 کنتور", "فلزی", "دیواری"): PanelTypeKey.ID18_1R,
    ("24 کنتور", "فلزی", "دیواری"): PanelTypeKey.ID24_1R,
    ("دو کنتور", "دیواری"): PanelTypeKey.ID2R,
    ("5 کنتور", "دیواری"): PanelTypeKey.ID5R,
}

def _find_panel_details_from_equipment(equipment_list: List[EquipmentItemBase]) -> Tuple[Optional[str], Optional[str]]:
    """
    تشخیص نوع تابلو بر اساس لیست تجهیزات با استفاده از قوانین مبتنی بر کلمات کلیدی.
    """
    if not equipment_list: 
        return None, None
    
    found_item_ids = set()
    item_counts = defaultdict(int)
    normalized_item_names = [normalize_text(eq.item_name) for eq in equipment_list]
    
    for eq in equipment_list:
        name = normalize_text(eq.item_name)
        for item_id, keywords in ITEM_ID_TO_KEYWORDS.items():
            if all(keyword in name for keyword in keywords):
                found_item_ids.add(item_id)
                item_counts[item_id] += eq.quantity
                break
                
    for panel_key, rule in PANEL_ID_RULES.items():
        required_items = rule.get("requires", set())
        forbidden_items = rule.get("forbids", set())
        count_rules = rule.get("count", {})
        
        if not required_items.issubset(found_item_ids) or not forbidden_items.isdisjoint(found_item_ids):
            continue

        counts_met = True
        for item_id, required_count in count_rules.items():
            if item_counts.get(item_id, 0) != required_count:
                counts_met = False
                break
        
        if counts_met:
            return panel_key.value, panel_key.value # ✅ اصلاح: بازگرداندن مقدار enum
            
    for name in normalized_item_names:
        for keywords_tuple, panel_key in EQUIPMENT_KEYWORD_TO_PANEL_MAPPING.items():
            if all(keyword in name for keyword in keywords_tuple):
                return panel_key.value, panel_key.value # ✅ اصلاح: بازگرداندن مقدار enum
                
    return None, None

# ==============================================================================
# بخش ۱۲: توابع کمکی
# ==============================================================================
# این توابع منطق تجاری هستند و بدون تغییر باقی می‌مانند.

def normalize_panel_key(key: str) -> str:
    """نرمال‌سازی کلید نوع تابلو."""
    if not isinstance(key, str): 
        return ""
    return re.sub(r'\s+', '', key).upper()
    
def convert_project_orm_to_pydantic(p_orm: Project) -> ProjectOut:
    """تبدیل مدل ORM به Pydantic."""
    return ProjectOut.model_validate(p_orm)    

def calculate_reservation_details(project: Project) -> Dict[str, Optional[int]]:
    """محاسبه جزئیات رزرو."""
    max_capacity = None
    purchased_single_phase = 0
    reserved_count = None
    ignore_keywords = ["تابلو", "جعبه", "قاب", "بدنه"]
    
    for item in project.equipment:
        normalized_name = normalize_text(item.item_name)
        is_single_phase_meter = "کنتور" in normalized_name and "تکفاز" in normalized_name
        should_be_ignored = any(keyword in normalized_name for keyword in ignore_keywords)
        if is_single_phase_meter and not should_be_ignored:
            purchased_single_phase += item.quantity
            
    if project.panel_code and isinstance(project.panel_code, str):
        match = re.search(r'ID(\d+)\+', project.panel_code.strip().upper())
        if match:
            try:
                max_capacity = int(match.group(1))
                reserved_count = max(0, max_capacity - purchased_single_phase)
            except (ValueError, TypeError):
                max_capacity = None
                reserved_count = None
                
    return {
        "max_capacity": max_capacity, 
        "purchased_single_phase": purchased_single_phase, 
        "reserved_count": reserved_count
    }

def generate_project_summary_data(project: Project) -> Dict[str, Any]:
    """تولید داده‌های خلاصه پروژه."""
    ALL_SUMMARY_KEYS = [
        "پیچ ۱۶*۳۰۰", "فیوز سه فاز 63", "فیوز سه فاز 32", "فیوز سه فاز 25",
        "فیوز تک فاز 32", "فیوز تک فاز 25", "فیوز تک فاز 16", "کنتور تک فاز",
        "کنتور سه فاز", "سکو", "تسمه استیل", "بست تسمه استیل", "لوله خرطومی",
        "لوله نیم گرد" 
    ]
    summary = {key: 0 for key in ALL_SUMMARY_KEYS}
    PRIORITIZED_RULES = [
        ("IGNORED", [("تابلو", "کنتور"), ("جعبه", "کنتور"), ("جعبه", "فیوز"), ("قاب", "کنتور"), ("مودم",)]),
        ("پیچ ۱۶*۳۰۰", [("پیچ", "مهره", "16", "300"), ("پیچ", "16در300")]),
        ("فیوز سه فاز 63", [("فیوز", "سه", "فاز", "63"), ("کلید", "سه", "فاز", "63")]),
        ("فیوز سه فاز 32", [("فیوز", "سه", "فاز", "32"), ("کلید", "سه", "فاز", "32")]),
        ("فیوز سه فاز 25", [("فیوز", "سه", "فاز", "25"), ("کلید", "سه", "فاز", "25")]),
        ("فیوز تک فاز 32", [("فیوز", "تکفاز", "32"), ("کلید", "فیوز", "تکفاز", "32")]),
        ("فیوز تک فاز 25", [("فیوز", "تکفاز", "25"), ("کلید", "فیوز", "تکفاز", "25")]),
        ("فیوز تک فاز 16", [("فیوز", "تکفاز", "16"), ("کلید", "فیوز", "تکفاز", "16")]),
        ("کنتور تک فاز", [("کنتور", "تکفاز")]), 
        ("کنتور سه فاز", [("کنتور", "سه", "فاز")]),
        ("سکو", [("سکوی", "نصب")]), 
        ("تسمه استیل", [("تسمه", "استیل")]), 
        ("بست تسمه استیل", [("بست", "تسمه")]),
        ("لوله خرطومی", [("لوله", "خرطومی")]), 
        ("لوله نیم گرد", [("لوله", "نیم گرد")]),
    ]
    
    for equipment_item in project.equipment:
        normalized_name = normalize_text(equipment_item.item_name)
        matched = False
        for category, keyword_sets in PRIORITIZED_RULES:
            if any(all(keyword in normalized_name for keyword in kw_set) for kw_set in keyword_sets):
                if category != "IGNORED":
                    summary[category] += equipment_item.quantity
                matched = True
                break
        if not matched:
            if "لوله خرطومی" in normalized_name:
                summary["لوله خرطومی"] += equipment_item.quantity
            elif "لوله نیم گرد" in normalized_name:
                summary["لوله نیم گرد"] += equipment_item.quantity

    reservation_details = calculate_reservation_details(project)
    summary["ظرفیت تابلو (تکفاز)"] = reservation_details.get("max_capacity")
    summary["تعداد خریداری شده (تکفاز)"] = reservation_details.get("purchased_single_phase", 0)
    summary["تعداد رزرو"] = reservation_details.get("reserved_count")
    
    return summary

def get_direction(request_id: str) -> Optional[str]:
    """تعیین جهت (شرق/غرب) بر اساس کد منطقه در شماره درخواست."""
    if not isinstance(request_id, str) or len(request_id) < 5: 
        return None
    area_code = request_id.strip()[3:5]
    direction_map = {
        "west": {"03", "04", "05", "07", "09", "10", "11", "12", "15"}, 
        "east": {"01", "02", "06", "08", "16", "17", "18"}
    }
    if area_code in direction_map["west"]: 
        return "west"
    if area_code in direction_map["east"]: 
        return "east"
    return None

def get_projects_by_status(
    db: Session, 
    required_steps: set, 
    forbidden_steps: set = None
) -> List[Project]:
    """دریافت پروژه‌ها بر اساس وضعیت مراحل انجام‌شده و انجام‌نشده."""
    query = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment)
    )
    
    for step_key in required_steps: 
        query = query.filter(Project.steps.any(Step.name_key == step_key))
        
    if forbidden_steps:
        for step_key in forbidden_steps: 
            query = query.filter(~Project.steps.any(Step.name_key == step_key))
            
    return query.order_by(Project.panel_code, Project.created_at).all()

def generate_detailed_report_data(projects: List[Project], direction: str) -> List[Dict[str, Any]]:
    """تولید داده‌های مورد نیاز برای گزارش تفصیلی آماده تحویل."""
    report_data = []
    filtered_projects = [p for p in projects if get_direction(p.request_id) == direction]
    
    for i, project in enumerate(filtered_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "مودم" in normalize_text(item.item_name)
        )
        
        row_data = {
            "project_id": project.id, 
            "ردیف": i, 
            "نام مشترک": project.customer_name, 
            "شماره تقاضا": project.request_id,
            "کد تابلو": project.panel_code or project.panel_type_key or "-", 
            "کنتور تک فاز": summary.get('کنتور تک فاز', 0),
            "کنتور سه فاز": summary.get('کنتور سه فاز', 0), 
            "فیوز تک فاز 16": summary.get('فیوز تک فاز 16', 0),
            "فیوز تک فاز 25": summary.get('فیوز تک فاز 25', 0), 
            "فیوز تک فاز 32": summary.get('فیوز تک فاز 32', 0),
            "فیوز سه فاز 25": summary.get('فیوز سه فاز 25', 0), 
            "فیوز سه فاز 32": summary.get('فیوز سه فاز 32', 0),
            "فیوز سه فاز 63": summary.get('فیوز سه فاز 63', 0), 
            "مودم": modem_count, 
            "ظرفیت": reservation.get('max_capacity'),
            "خرید": reservation.get('purchased_single_phase'), 
            "رزرو": reservation.get('reserved_count'), 
            "سکو": summary.get('سکو', 0),
            "تسمه استیل": summary.get('تسمه استیل', 0), 
            "بست تسمه استیل": summary.get('بست تسمه استیل', 0),
            "پیچ یکسررزوه": 0, 
            "پیچ ۱۶*۳۰۰": summary.get('پیچ ۱۶*۳۰۰', 0), 
            "لوله نیم گرد": summary.get('لوله نیم گرد', 0),
            "لوله خرطومی": summary.get('لوله خرطومی', 0), 
            "توضیحات": ""
        }
        report_data.append(row_data)
        
    return report_data

# ==============================================================================
# بخش ۱۲.۵: توابع کمکی انبارداری (بخش جدید)
# ==============================================================================
async def check_stock_and_alert(db: Session, item: WarehouseItem, warehouse_id: int):
    """موجودی کالا را بررسی کرده و در صورت رسیدن به حداقل، هشدار WebSocket ارسال می‌کند."""
    stock_level_expr = func.sum(
        case(
            (InventoryTransaction.transaction_type == TransactionType.IN.value, InventoryTransaction.quantity),
            else_=-InventoryTransaction.quantity
        )
    ).label("current_stock")

    result = db.query(stock_level_expr).filter(
        InventoryTransaction.item_id == item.id,
        InventoryTransaction.warehouse_id == warehouse_id
    ).one_or_none()

    current_stock = result.current_stock if result and result.current_stock is not None else 0
    
    if current_stock <= item.min_stock_level:
        logger.warning(f"LOW STOCK ALERT: Item '{item.item_name}' (ID: {item.id}) is at {current_stock}, below/equal to threshold of {item.min_stock_level}")
        await manager.broadcast({
            "type": "low_stock_alert",
            "data": {
                "item_name": item.item_name,
                "current_stock": current_stock,
                "min_stock_level": item.min_stock_level
            }
        })

async def deduct_stock_for_project(db: Session, project: Project, user: Personnel):
    """موجودی مورد نیاز برای یک پروژه را به صورت خودکار از انبار کسر می‌کند."""
    if not project.panel_code:
        logger.warning(f"Project {project.id} has no panel_code, skipping auto stock deduction.")
        return

    required_items = db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == project.panel_code).all()
    if not required_items:
        logger.info(f"No BOM defined for panel_code '{project.panel_code}'. Cannot auto-deduct stock.")
        return

    main_warehouse = db.query(Warehouse).order_by(Warehouse.id).first()
    if not main_warehouse:
        logger.error("Cannot deduct stock: No warehouse defined in the system.")
        raise HTTPException(status_code=500, detail="هیچ انباری در سیستم تعریف نشده است.")

    for required in required_items:
        item = db.query(WarehouseItem).filter(WarehouseItem.item_name == required.item_name).first()
        if not item:
            logger.error(f"Item '{required.item_name}' for panel '{project.panel_code}' not found in warehouse items.")
            continue

        transaction = InventoryTransaction(
            warehouse_id=main_warehouse.id,
            item_id=item.id,
            quantity=required.quantity_required,
            transaction_type=TransactionType.OUT,
            project_id=project.id,
            user_id=user.id,
            notes=f"خروج خودکار برای پروژه {project.request_id} با کد تابلو {project.panel_code}"
        )
        db.add(transaction)
    
    db.commit()
    logger.info(f"Auto-deducted stock for project {project.id} based on panel code '{project.panel_code}'.")

    # بررسی موجودی پس از ثبت همه تراکنش‌ها
    for required in required_items:
        item = db.query(WarehouseItem).filter(WarehouseItem.item_name == required.item_name).first()
        if item:
            await check_stock_and_alert(db, item, main_warehouse.id)
    
# ==============================================================================
# بخش ۱۳: Background Tasks
# ==============================================================================
# این بخش‌ها به خوبی پیاده‌سازی شده‌اند و تغییرات جزئی برای بهبود لاگ و خطا دارند.

def process_excel_in_background(file_contents: bytes, ws_manager: ConnectionManager):
    """پردازش فایل اکسل در پس‌زمینه."""
    with get_db_session() as db:
        try:
            projects_payload_map: Dict[str, Dict[str, Any]] = defaultdict(
                lambda: {'base_info': None, 'equipment': []}
            )
            workbook = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)
            sheet = workbook.active
            header_row_values = [
                str(cell.value).strip().lower() if cell.value else "" 
                for cell in sheet[1]
            ]
            
            EXPECTED_COLS = { 
                "شماره درخواست توزیع": "شماره درخواست توزیع".lower(), 
                "نام و نام خانوادگی خریدار": "نام و نام خانوادگی خریدار".lower(), 
                "تجهیز": "تجهیز".lower(), 
                "تعداد": "تعداد".lower() 
            }
            
            if not all(col in header_row_values for col in EXPECTED_COLS.values()):
                missing = [k for k, v in EXPECTED_COLS.items() if v not in header_row_values]
                raise ValueError(f"ستون‌های ضروری یافت نشد: {', '.join(missing)}")
                
            RID_COL_IDX = header_row_values.index(EXPECTED_COLS["شماره درخواست توزیع"])
            CUST_COL_IDX = header_row_values.index(EXPECTED_COLS["نام و نام خانوادگی خریدار"])
            ITEM_NAME_COL_IDX = header_row_values.index(EXPECTED_COLS["تجهیز"])
            ITEM_QTY_COL_IDX = header_row_values.index(EXPECTED_COLS["تعداد"])
            
            last_valid_req_id = None
            for r_idx, row_tuple in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row = list(row_tuple)
                req_id = str(row[RID_COL_IDX]).strip() if RID_COL_IDX < len(row) and row[RID_COL_IDX] else last_valid_req_id
                if not req_id: 
                    continue
                last_valid_req_id = req_id
                
                if not projects_payload_map[req_id]['base_info']:
                    cust_name = str(row[CUST_COL_IDX]).strip() if CUST_COL_IDX < len(row) and row[CUST_COL_IDX] else "نامشخص"
                    projects_payload_map[req_id]['base_info'] = ProjectCreateFromExcelData(
                        name=f"پروژه {req_id}", 
                        request_id=req_id, 
                        customer_name=cust_name, 
                        location=None
                    )
                    
                item_name = str(row[ITEM_NAME_COL_IDX]).strip() if ITEM_NAME_COL_IDX < len(row) and row[ITEM_NAME_COL_IDX] else None
                if not item_name: 
                    continue
                    
                qty = int(float(str(row[ITEM_QTY_COL_IDX]).strip())) if ITEM_QTY_COL_IDX < len(row) and row[ITEM_QTY_COL_IDX] else 1
                projects_payload_map[req_id]['equipment'].append(
                    EquipmentItemCreate(item_name=item_name, quantity=qty)
                )
                
            processed_count = 0
            for req_id, data in projects_payload_map.items():
                if not data['base_info']: 
                    continue
                    
                db_proj = db.query(Project).filter(Project.request_id == req_id).first()
                if db_proj:
                    db.query(EquipmentItem).filter(EquipmentItem.project_id == db_proj.id).delete()
                    db_proj.name = data['base_info'].name
                    db_proj.customer_name = data['base_info'].customer_name
                else:
                    db_proj = Project(**data['base_info'].model_dump())
                    db.add(db_proj)
                    
                db.flush()
                
                for eq in data['equipment']: 
                    db.add(EquipmentItem(project_id=db_proj.id, **eq.model_dump()))
                    
                db.commit()
                processed_count += 1
                
            if processed_count > 0:
                asyncio.run(ws_manager.broadcast({
                    "type": "update", 
                    "source": "excel_upload", 
                    "count": processed_count
                }))
                logger.info(f"Excel processing completed: {processed_count} projects updated")
                
        except Exception as e:
            logger.error(f"Error processing Excel: {e}", exc_info=True)
            asyncio.run(ws_manager.broadcast({
                "type": "error", 
                "message": f"خطا در پردازش اکسل: {e}"
            }))

def process_detailed_excel_in_background(files_contents: List[Tuple[str, bytes]], ws_manager: ConnectionManager):
    """پردازش فایل‌های اکسل تفصیلی در پس‌زمینه."""
    with get_db_session() as db:
        processed_count = 0
        errors = []
        
        def find_value_by_keyword(sheet, keyword: str):
            """پیدا کردن مقدار بر اساس کلیدواژه"""
            for row in sheet["A1:C10"]:
                cell = row[0]
                if cell.value and keyword in str(cell.value):
                    target_cell = sheet.cell(row=cell.row, column=3)
                    return str(target_cell.value).strip() if target_cell.value else None
            return None
            
        try:
            for filename, contents in files_contents:
                try:
                    workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
                    sheet = workbook.active
                    
                    request_id = find_value_by_keyword(sheet, "شماره درخواست شرکت توزیع")
                    if not request_id:
                        errors.append(f"فایل '{filename}': کلیدواژه 'شماره درخواست شرکت توزیع' یافت نشد.")
                        continue
                        
                    invoice_number = find_value_by_keyword(sheet, "شماره سفارش")
                    customer_name = find_value_by_keyword(sheet, "نام مشتری") or "نامشخص"
                    distribution_company = find_value_by_keyword(sheet, "شرکت توزیع") or ""
                    location = f"شرکت: {distribution_company} - فاکتور: {invoice_number}" if distribution_company or invoice_number else None
                    
                    equipment_start_row = -1
                    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, min_col=1, max_col=3, values_only=True), 1):
                        if row[0] == "ردیف" and row[1] == "کد تجهیز/عملیات" and row[2] == "تجهیز/عملیات":
                            equipment_start_row = i + 1
                            break
                            
                    if equipment_start_row == -1:
                        errors.append(f"فایل '{filename}': جدول تجهیزات (هدر 'ردیف', 'کد تجهیز'...) یافت نشد.")
                        continue
                        
                    equip_list = []
                    for row in sheet.iter_rows(min_row=equipment_start_row, max_col=5, values_only=True):
                        item_name_raw, quantity_raw = row[2], row[4]
                        if not item_name_raw or str(item_name_raw).strip() == "": 
                            break
                            
                        full_item_name = str(item_name_raw).strip()
                        try: 
                            quantity = int(quantity_raw) if isinstance(quantity_raw, (int, float)) and quantity_raw > 0 else 1
                        except (ValueError, TypeError): 
                            quantity = 1
                            
                        equip_list.append(EquipmentItemCreate(item_name=full_item_name, quantity=quantity))
                        
                    if not equip_list:
                        errors.append(f"فایل '{filename}': هیچ تجهیزی برای شماره درخواست {request_id} یافت نشد.")
                        continue
                        
                    panel_code, panel_type_key = _find_panel_details_from_equipment(equip_list)
                    db_proj = db.query(Project).filter(Project.request_id == request_id).first()
                    project_name = f"پروژه {panel_code} - {request_id}" if panel_code else f"پروژه {request_id}"
                    
                    if db_proj:
                        db.query(EquipmentItem).filter(EquipmentItem.project_id == db_proj.id).delete()
                        db_proj.name = project_name
                        db_proj.customer_name = customer_name
                        db_proj.location = location
                        db_proj.panel_code = panel_code
                        db_proj.panel_type_key = panel_type_key
                    else:
                        db_proj = Project(
                            name=project_name, 
                            request_id=request_id, 
                            customer_name=customer_name, 
                            location=location, 
                            panel_code=panel_code, 
                            panel_type_key=panel_type_key, 
                            barcode_payload=request_id
                        )
                        db.add(db_proj)
                        
                    db.flush()
                    
                    for eq in equip_list: 
                        db.add(EquipmentItem(project_id=db_proj.id, **eq.model_dump()))
                        
                    db.commit()
                    processed_count += 1
                    
                except Exception as e:
                    db.rollback()
                    errors.append(f"فایل '{filename}': خطای پردازش - {e}")
                    logger.error(f"Error processing file {filename}: {e}", exc_info=True)
                    
            if processed_count > 0:
                asyncio.run(ws_manager.broadcast({
                    "type": "update", 
                    "source": "detailed_excel_upload", 
                    "count": processed_count, 
                    "message": f"{processed_count} پروژه با موفقیت پردازش شد."
                }))
                logger.info(f"Detailed Excel processing completed: {processed_count} projects")
                
            if errors:
                error_message = "خطا در پردازش برخی فایل‌ها:\n" + "\n".join(errors)
                logger.warning(f"Excel processing errors: {error_message}")
                asyncio.run(ws_manager.broadcast({
                    "type": "error", 
                    "message": error_message
                }))
                
        except Exception as e:
            logger.error(f"Error in detailed Excel processing: {e}", exc_info=True)
            asyncio.run(ws_manager.broadcast({
                "type": "error", 
                "message": f"خطای کلی در پردازش اکسل (تفصیلی): {e}"
            }))

# توابع اعتبارسنجی (به صورت موقت غیرفعال)
async def run_branch_validation(project_id: int, request_id: str, purchase_list: list, ws_manager: ConnectionManager):
    """اعتبارسنجی انشعابات (منطق خارجی)."""
    result = ValidationResponse(has_discrepancy=True, message="خطا در پردازش.")
    
    try:
        # به صورت موقت غیرفعال شده
        # async with NaabConnector(NAAB_USERNAME, NAAB_PASSWORD) as connector:
        #     site_branch_count = await connector.get_site_branch_count(request_id)
        
        site_branch_count = 0  # مقدار موقت
        purchased_meter_count = sum(
            item['quantity'] for item in purchase_list 
            if "کنتور" in normalize_text(item['name'])
        )
        
        if site_branch_count == purchased_meter_count:
            result.has_discrepancy = False
            result.message = f"تطابق موفق: {site_branch_count} انشعاب در سایت و {purchased_meter_count} کنتور در پروژه."
        else:
            result.has_discrepancy = True
            result.message = f"مغایرت انشعاب: {site_branch_count} انشعاب در سایت، اما {purchased_meter_count} کنتور در پروژه ثبت شده است."
            
    except Exception as e:
        logger.error(f"Branch validation error for project {project_id}: {e}")
        result.message = f"خطا در بررسی انشعاب: {str(e)}"
    
    await ws_manager.broadcast({
        "type": "validation_result", 
        "project_id": project_id, 
        "result": result.model_dump()
    })

async def run_purchase_validation(project_id: int, request_id: str, purchase_list: list, ws_manager: ConnectionManager):
    """اعتبارسنجی خریدها (منطق خارجی)."""
    result = ValidationResponse(has_discrepancy=True, message="خطا در پردازش.")

    try:
        # به صورت موقت غیرفعال شده
        # async with NaabConnector(NAAB_USERNAME, NAAB_PASSWORD) as connector:
        #     site_items = await connector.get_site_purchased_items(request_id)

        site_items = []  # مقدار موقت
        if not site_items:
            result.message = "هیچ تجهیزی در تب مالی سایت برای این تقاضا یافت نشد."
        else:
            site_dict = {normalize_text(item['name']): item['quantity'] for item in site_items}
            project_dict = {normalize_text(item['name']): item['quantity'] for item in purchase_list}
            discrepancies = []
            all_keys = set(site_dict.keys()) | set(project_dict.keys())
            
            for key in all_keys:
                if site_dict.get(key, 0) != project_dict.get(key, 0):
                    discrepancies.append(f"'{key}': سایت ({site_dict.get(key, 0)}) / پروژه ({project_dict.get(key, 0)})")
            
            if not discrepancies:
                result.has_discrepancy = False
                result.message = "تطابق موفق: تجهیزات پروژه با تب مالی سایت یکسان است."
            else:
                result.has_discrepancy = True
                result.message = "مغایرت تجهیزات یافت شد:\n" + "\n".join(discrepancies)

    except Exception as e:
        logger.error(f"Purchase validation error for project {project_id}: {e}")
        result.message = f"خطا در بررسی تجهیزات: {str(e)}"
    
    await ws_manager.broadcast({
        "type": "validation_result", 
        "project_id": project_id, 
        "result": result.model_dump()
    })

# ==============================================================================
# بخش ۱۴: HTML Templates
# ==============================================================================
# ثابت‌های رشته‌ای بزرگ برای قالب‌های HTML، بهتر است در فایل‌های جداگانه باشند،
# اما طبق درخواست در کد باقی می‌مانند. این بخش بدون تغییر است.

PROJECT_SLIP_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>برچسب پروژه {{ project.name }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('{{ font_path }}'); }
        @page { size: 100mm 150mm; margin: 5mm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; width: 90mm; height: 140mm; margin: 0; padding: 0; display: flex; flex-direction: column; border: 1px solid #ccc; box-sizing: border-box; }
        header { text-align: center; padding: 4px 6px; border-bottom: 1px solid black; flex-shrink: 0; }
        header h1 { font-size: 12pt; margin: 0; }
        .details { padding: 4px 6px; line-height: 1.2; flex-shrink: 0; }
        .details p { margin: 2px 0; font-size: 11pt; }
        .request-info { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 4px; }
        .request-info div p { margin: 0; font-size: 10pt; line-height: 1.4; }
        .qr-code-inline { width: 22mm; height: 22mm; min-width: 22mm; min-height: 22mm; }
        .equipment-section { flex-grow: 1; padding: 0 6px; display: flex; flex-direction: column; overflow: hidden; }
        .equipment-header { display: flex; justify-content: space-between; align-items: center; padding: 2px 0; margin-bottom: 2px; }
        .direction-info { font-size: 10pt; font-weight: bold; border: 1px solid black; padding: 2px 6px; border-radius: 4px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #333; padding: 3px 4px; text-align: right; font-size: 8pt; word-break: break-word; }
        th { background-color: #f2f2f2; font-weight: bold; text-align: center; }
    </style>
</head>
<body>
    <header>
        <h1>شرکت فرداد سازه گلشن</h1>
        <p style="font-size: 8pt;">برچسب شناسایی پروژه</p>
    </header>
    <section class="details">
        <div class="request-info">
            <div>
                <p><strong>شماره درخواست:</strong> {{ project.request_id }}</p>
                {% if project.panel_code %}<p><strong>کد تابلو:</strong> {{ project.panel_code }}</p>{% endif %}
            </div>
            {% if qr_code_base64 %}
            <img class="qr-code-inline" src="data:image/png;base64,{{ qr_code_base64 }}" alt="QR Code">
            {% endif %}
        </div>
        <p><strong>مشتری:</strong> {{ project.customer_name }}</p>
        <p><strong>تاریخ صدور:</strong> {{ report_date_jalali }}</p>
    </section>
    <section class="equipment-section">
        <div class="equipment-header">
            <span><strong>لیست تجهیزات:</strong></span>
            {% if direction %}<span class="direction-info">{{ direction }}</span>{% endif %}
        </div>
        <table>
            <thead><tr><th style="width: 10%;">ردیف</th><th>نام تجهیز</th><th style="width: 15%;">تعداد</th></tr></thead>
            <tbody>
                {% for item in project.equipment[:10] %}
                <tr><td style="text-align: center;">{{ loop.index }}</td><td>{{ item.item_name }}</td><td style="text-align: center;">{{ item.quantity }}</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </section>
</body>
</html>
"""

DETAILED_EXIT_SLIP_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>برگه خروج پروژه {{ project.request_id }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('{{ font_path }}'); }
        @page { size: A5 portrait; margin: 0.7cm; }
        html, body { margin: 0; padding: 0; width: 100%; box-sizing: border-box; font-family: 'Vazirmatn', sans-serif; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { border: 1.5px solid black; padding: 0.5cm; display: flex; flex-direction: column; height: calc(210mm - 1.4cm); }
        header { flex-shrink: 0; text-align: center; margin-bottom: 5mm; border-bottom: 1px solid #666; padding-bottom: 2mm; }
        header h1 { font-size: 13pt; margin: 0; }
        header p { font-size: 8pt; margin: 1mm 0 0 0; }
        .slip-content { flex-grow: 1; overflow: hidden; font-size: 8pt; }
        .details-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0 10px; margin-bottom: 4mm; line-height: 1.3; }
        .details-grid p { margin: 1mm 0; }
        h2 { font-size: 10pt; border-bottom: 1px solid #ccc; padding-bottom: 1mm; margin-top: 0; margin-bottom: 2mm; text-align: center; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #333; padding: 1.2mm; text-align: right; vertical-align: middle; font-size: 7pt; }
        th { background-color: #f2f2f2; font-weight: bold; text-align: center; }
        .summary-table td:nth-child(2) { text-align: center; font-weight: bold; font-size: 8pt; }
        .signatures { flex-shrink: 0; margin-top: auto; padding-top: 4mm; border-top: 1px dashed #aaa; display: flex; justify-content: space-around; font-size: 8pt; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="slip-content">
        <header><h1>شرکت فرداد سازه گلشن</h1><p>برگه خروج و تحویل تابلو</p></header>
        <section class="details-grid">
            <p><strong>شماره درخواست:</strong> {{ project.request_id }}</p>
            <p><strong>کد تابلو:</strong> {{ project.panel_code or '-' }}</p>
            <p><strong>مشتری:</strong> {{ project.customer_name }}</p>
            <p><strong>تاریخ خروج:</strong> {{ report_date_jalali }}</p>
        </section>
        <h2>جدول خلاصه تجهیزات</h2>
        <table class="summary-table">
            <thead><tr><th style="width: 70%;">عنوان</th><th style="width: 30%;">تعداد</th></tr></thead>
            <tbody>
                {% if summary_data['ظرفیت تابلو (تکفاز)'] is not none %}
                <tr><td style="background-color:#eef2ff;">ظرفیت تابلو (تکفاز)</td><td style="background-color:#eef2ff;">{{ summary_data['ظرفیت تابلو (تکفاز)'] | default(0) }}</td></tr>
                <tr><td style="background-color:#eef2ff;">کنتور تکفاز خریداری شده</td><td style="background-color:#eef2ff;">{{ summary_data['تعداد خریداری شده (تکفاز)'] | default(0) }}</td></tr>
                <tr><td style="background-color:#dbeafe;color:#1e40af;">تعداد رزرو</td><td style="background-color:#dbeafe;color:#1e40af;">{{ summary_data['تعداد رزرو'] | default(0) }}</td></tr>
                {% endif %}
                <tr><td>کنتور تک فاز</td><td>{{ summary_data.get('کنتور تک فاز', 0) }}</td></tr>
                <tr><td>کنتور سه فاز</td><td>{{ summary_data.get('کنتور سه فاز', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">فیوزهای تک فاز</td></tr>
                <tr><td>فیوز تک فاز 16</td><td>{{ summary_data.get('فیوز تک فاز 16', 0) }}</td></tr>
                <tr><td>فیوز تک فاز 25</td><td>{{ summary_data.get('فیوز تک فاز 25', 0) }}</td></tr>
                <tr><td>فیوز تک فاز 32</td><td>{{ summary_data.get('فیوز تک فاز 32', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">فیوزهای سه فاز</td></tr>
                <tr><td>فیوز سه فاز 25</td><td>{{ summary_data.get('فیوز سه فاز 25', 0) }}</td></tr>
                <tr><td>فیوز سه فاز 32</td><td>{{ summary_data.get('فیوز سه فاز 32', 0) }}</td></tr>
                <tr><td>فیوز سه فاز 63</td><td>{{ summary_data.get('فیوز سه فاز 63', 0) }}</td></tr>
                <tr><td colspan="2" style="background-color:#f9fafb;text-align:center;font-weight:bold;">ملزومات</td></tr>
                <tr><td>سکو</td><td>{{ summary_data.get('سکو', 0) }}</td></tr>
                <tr><td>تسمه استیل</td><td>{{ summary_data.get('تسمه استیل', 0) }}</td></tr>
                <tr><td>بست تسمه استیل</td><td>{{ summary_data.get('بست تسمه استیل', 0) }}</td></tr>
                <tr><td>لوله خرطومی</td><td>{{ summary_data.get('لوله خرطومی', 0) }}</td></tr>
                <tr><td>لوله نیم گرد</td><td>{{ summary_data.get('لوله نیم گرد', 0) }}</td></tr>
                <tr><td>پیچ ۱۶*۳۰۰</td><td>{{ summary_data.get('پیچ ۱۶*۳۰۰', 0) }}</td></tr>
            </tbody>
        </table>
    </div>
    <div class="signatures">
        <div class="signature-box">امضاء تحویل دهنده<br><br>...........................</div>
        <div class="signature-box">امضاء تحویل گیرنده<br><br>...........................</div>
    </div>
    {% if auto_print %}<script>window.onload = function() { setTimeout(function() { window.print(); }, 500); };</script>{% endif %}
</body>
</html>
"""

SUPERVISOR_APPROVAL_SIMPLE_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>گزارش تایید ناظر</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4; margin: 1.5cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 11pt; }
        .header { text-align: center; margin-bottom: 20px; }
        .header h1 { font-size: 16pt; margin: 0; }
        .header h2 { font-size: 14pt; margin: 5px 0; }
        .report-date { text-align: left; font-size: 10pt; margin-bottom: 15px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 8px; text-align: center; vertical-align: middle; }
        th { background-color: #f2f2f2; font-weight: bold; }
        .signatures { margin-top: 80px; display: flex; justify-content: space-around; font-size: 12pt; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header"><h1>شرکت فرداد سازه گلشن</h1><h2>لیست تابلو های آماده تحویل</h2></div>
    <div class="report-date">تاریخ: {{ jalali_date }}</div>
    <table>
        <thead>
            <tr>
                <th style="width: 5%;">ردیف</th>
                <th style="width: 35%;">نام و نام خانوادگی</th>
                <th style="width: 20%;">شماره تقاضا</th>
                <th style="width: 15%;">کد تابلو</th>
                <th style="width: 15%;">تایید</th>
                <th style="width: 10%;">توضیحات</th>
            </tr>
        </thead>
        <tbody>
            {% for p in projects %}
            <tr>
                <td>{{ loop.index }}</td>
                <td style="text-align: right;">{{ p.customer_name }}</td>
                <td>{{ p.request_id }}</td>
                <td>{{ p.panel_code or '-' }}</td>
                <td>√</td>
                <td></td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">نام و امضا مسئول کارگاه<br><br>...........................</div>
        <div class="signature-box">نام و امضا مسئول دفتر نظارت<br><br>...........................</div>
    </div>
</body>
</html>
"""

SUPERVISOR_CHECKLIST_ITEMS = [
    "مغایرت ابعادی و کیفیت ساخت تابلو با مشخصات فنی و سفارش ساخت", "کیفیت رنگ", "آسیب دیدگی ظاهری بدنه و رنگ",
    "مغایرت رنج کلید اتومات با مشخصات فنی", "مغایرت رنج کلید فیوزهای مینیاتوری با درخواست مشترک", "مغایرت برند تجهیزات با و ندور لیست",
    "مغایرت تعداد کنتورها و فیوزها با درخواست مشترک", "مغایرت ردیف کنتورها با تست سماک", "رعايت سيم بندي مخابراتي صحيح جهت تست سماك",
    "رعايت ابعادي شينه ها و سيم ها", "ارسال تابلو بدون برچسب ناظر کنترل کیفیت دفتر نظارت بر سازندگان",
    "تاييد تستهاي الكتريكال توسط ناظر كنترل كيفيت در زمان تاييد تابلو", "برگشت تابلو از مراکز تجمیع به دلایل فنی",
    "تطابق و تناظر صفحه گلند با مشخصات فني و ترمينال خروجي مشترك", "رعايت فاصله تجهيزات با بدنه مطابق با مشخصات فني",
    "نصب پلاك مشخصات در داخل و خارج تابلوها", "رعايت شماره گذاري سر سيم نصب دستور العمل و نقشه و ... در تابلوها",
    "تكميل تابلو در زمان بازديد و نصب برچسب كنترل كيفيت تابلوساز", "ارائه تست هاي مربوط به تابلوها",
    "ارسال فرمهاي آماده بازديد تابلوها مطابق با زمانبندي مشخص شده", "نصب مناسب آنتن كنتور سه فاز هوشمند يا مودم",
    "رعايت آرايش مناسب سيم بندي تابلوها و جانمايي تجهيزات", "رعايت همبندي مناسب", "رگلاژ درب ها",
    "ارسال ناقص تجهیزات مورد نیاز نصب تابلو", "رعايت نظافت داخلي و خارجي تابلو", "بسته بندي مناسب",
]

INDIVIDUAL_QC_CHECKLIST_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>چک لیست کنترل کیفیت پروژه {{ project.request_id }}</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4; margin: 0.5cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; }
        .header { text-align: center; margin-bottom: 15px; }
        .header h1 { font-size: 15pt; margin: 0; }
        .header h2 { font-size: 13pt; margin: 4px 0; }
        .project-info { border: 1px solid #ccc; padding: 8px; margin-bottom: 15px; border-radius: 8px; display: grid; grid-template-columns: 1fr 1fr; gap: 8px; font-size: 9pt; }
        .project-info p { margin: 0; }
        .report-date { text-align: left; font-size: 9pt; margin-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 4px 6px; text-align: center; vertical-align: middle; line-height: 1.4; }
        th { background-color: #f2f2f2; font-weight: bold; font-size: 9pt;}
        .description-col { text-align: right; width: 70%; }
        .result-col { width: 15%; }
        .signatures { margin-top: 30px; display: flex; justify-content: space-around; font-size: 11pt; page-break-inside: avoid; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header">
        <h1>شرکت فرداد سازه گلشن</h1>
        <h2>فرم کنترل کیفیت نهایی تابلو</h2>
    </div>
    <div class="report-date">تاریخ: {{ jalali_date }}</div>
    <div class="project-info">
        <p><strong>شماره تقاضا:</strong> {{ project.request_id }}</p>
        <p><strong>کد تابلو:</strong> {{ project.panel_code or '-' }}</p>
        <p><strong>نام مشترک:</strong> {{ project.customer_name }}</p>
        <p><strong>مونتاژکاران:</strong> {{ project.assembler_1 }}{% if project.assembler_2 %}، {{ project.assembler_2 }}{% endif %}</p>
    </div>
    <table>
        <thead>
            <tr>
                <th style="width: 5%;">ردیف</th>
                <th class="description-col">شرح کنترل کیفیت</th>
                <th class="result-col">نتیجه</th>
                <th class="result-col">توضیحات</th>
            </tr>
        </thead>
        <tbody>
            {% for item in checklist_items %}
            <tr>
                <td>{{ loop.index }}</td>
                <td class="description-col">{{ item }}</td>
                <td></td>
                <td></td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">نام و امضا مسئول مونتاژ<br><br>...........................</div>
        <div class="signature-box">نام و امضا مسئول کنترل کیفیت<br><br>...........................</div>
    </div>
    <script>
        window.onload = function() {
            setTimeout(function() { window.print(); }, 500);
        };
    </script>
</body>
</html>
"""

SUPERVISOR_CHECKLIST_TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>چک لیست کنترل کیفیت جهت تایید ناظر</title>
    <style>
        @font-face { font-family: 'Vazirmatn'; src: url('/static/Vazirmatn-regular.ttf'); }
        @page { size: A4 landscape; margin: 1cm; }
        @media print { body { -webkit-print-color-adjust: exact; color-adjust: exact; } }
        body { font-family: 'Vazirmatn', sans-serif; font-size: 10pt; }
        .header { text-align: center; margin-bottom: 15px; }
        .header h1 { font-size: 15pt; margin: 0; }
        .header h2 { font-size: 13pt; margin: 5px 0; }
        .report-date { text-align: left; font-size: 10pt; margin-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid black; padding: 5px; text-align: center; vertical-align: middle; }
        th { background-color: #f2f2f2; font-weight: bold; font-size: 9pt; }
        .description-col { text-align: right; width: 35%; font-size: 9pt;}
        .project-col { min-width: 80px; font-size: 8pt; word-break: break-all; }
        tbody td:first-child { width: 3%; }
        .signatures { margin-top: 50px; display: flex; justify-content: space-around; font-size: 11pt; page-break-inside: avoid; }
        .signature-box { text-align: center; }
    </style>
</head>
<body>
    <div class="header"><h1>شرکت فرداد سازه گلشن</h1><h2>چک لیست کنترل کیفیت تابلوهای آماده تحویل</h2></div>
    <div class="report-date">تاریخ: {{ jalali_date }}</div>
    <table>
        <thead>
            <tr>
                <th rowspan="3" style="vertical-align: middle;">ردیف</th>
                <th rowspan="3" style="vertical-align: middle;">شرح کنترل کیفیت</th>
                {% for p in projects %}<th class="project-col">ش. تقاضا</th>{% endfor %}
            </tr>
            <tr>{% for p in projects %}<th class="project-col">{{ p.request_id }}</th>{% endfor %}</tr>
            <tr>{% for p in projects %}<th class="project-col">کد: {{ p.panel_code or '-' }}</th>{% endfor %}</tr>
        </thead>
        <tbody>
            {% for item in checklist_items %}
            <tr>
                <td>{{ loop.index }}</td>
                <td class="description-col">{{ item }}</td>
                {% for p in projects %}<td></td>{% endfor %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
    <div class="signatures">
        <div class="signature-box">نام و امضا مسئول کارگاه<br><br>...........................</div>
        <div class="signature-box">نام و امضا مسئول دفتر نظارت<br><br>...........................</div>
    </div>
</body>
</html>
"""

# ==============================================================================
# بخش ۱۵: گزارش کار روزانه - تنظیمات
# ==============================================================================
# این بخش یک کانفیگ استاتیک است و به درستی پیاده‌سازی شده.

REPORT_CONFIG = {
    "headers": ["آیتم های گزارش کار روزانه پرسنل", "ID 2R", "ID 5R", "ID 6+1R", "ID 12+1R", "ID 18+1R"],
    "rows": [
        {"title": "آماده سازی دسته سیم", "items": [{"label": "ID2R", "type": "number"}, {"label": "ID5R", "type": "number"}, {"label": "ID6+1R", "type": "number"}, {"label": "ID12+1R", "type": "number"}, {"label": "ID18+1R", "type": "number"}]},
        {"title": "آماده سازی شمش", "items": [{"label": "دیماندی", "type": "number"}, {"label": "ID5R", "type": "number"}, {"label": "ID6+1R", "type": "number"}, {"label": "ID12+1R", "type": "number"}, {"label": "ID18+1R", "type": "number"}]},
        {"title": "برش کاری و لبه گیر", "items": [{"label": "ریل ID2R", "type": "number"}, {"label": "سکو ID2R", "type": "number"}, {"label": "داکت", "type": "number"}, {"label": "ریل ID5R", "type": "number"}, {"label": "سکو ID5R", "type": "number"}]},
        {"title": "مونتاژ کامل تابلو (۱)", "items": [{"label": "ID101.1", "type": "number"}, {"label": "ID101.3", "type": "number"}, {"label": "ID102.1", "type": "number"}, {"label": "ID102.3", "type": "number"}, None]},
        {"title": "مونتاژ کامل تابلو (۲)", "items": [{"label": "ID104.1", "type": "number"}, {"label": "ID104.3", "type": "number"}, {"label": "ID105.1", "type": "number"}, {"label": "ID105.3", "type": "number"}, None]},
        {"title": "مونتاژ کامل تابلو (۳)", "items": [{"label": "ID109", "type": "number"}, {"label": "ID108", "type": "number"}, {"label": "ID111", "type": "number"}, {"label": "ID112", "type": "number"}, None]},
        {"title": "مونتاژ کامل تابلو (۴)", "items": [{"label": "ID115.1", "type": "number"}, {"label": "ID115.3", "type": "number"}, {"label": "ID116", "type": "number"}, {"label": "ID120", "type": "number"}, {"label": "ID218", "type": "number"}]},
        {"title": "مونتاژ کامل تابلو (۵)", "items": [{"label": "ID212", "type": "number"}, {"label": "ID213", "type": "number"}, {"label": "ID214", "type": "number"}, {"label": "ID215", "type": "number"}, {"label": "ID216", "type": "number"}]},
        {"title": "مونتاژ کامل تابلو (۶)", "items": [{"label": "ID121", "type": "number"}, {"label": "ID122", "type": "number"}, {"label": "ID123", "type": "number"}, {"label": "ID124", "type": "number"}, {"label": "ID211", "type": "number"}]},
        {
            "title": "پیشرفت مونتاژ یک تابلو بر اساس درصد ID 2R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "سیم بندی", "value": 50},
                {"label": "مونتاژ کنتور و فیوز", "value": 30},
                {"label": "آماده سازی صفحه زیر", "value": 20}
            ]
        },
        {
            "title": "پیشرفت مونتاژ یک تابلو بر اساس درصد ID 5R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "سیم بندی", "value": 30},
                {"label": "مونتاژ کنتور و فیوز", "value": 20},
                {"label": "آماده سازی صفحه زیر", "value": 50}
            ]
        },
        {
            "title": "پیشرفت مونتاژ یک تابلو بر اساس درصد ID 6+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "صفحه گلند", "value": 5},
                {"label": "صفحه کنتور", "value": 20},
                {"label": "سیم بندی", "value": 55},
                {"label": "صفحه فیوز", "value": 15},
                {"label": "درب ها", "value": 5}
            ]
        },
        {
            "title": "پیشرفت مونتاژ یک تابلو بر اساس درصد ID 12+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "صفحه گلند", "value": 5},
                {"label": "صفحه کنتور", "value": 20},
                {"label": "سیم بندی", "value": 55},
                {"label": "صفحه فیوز", "value": 15},
                {"label": "درب ها", "value": 5}
            ]
        },
        {
            "title": "پیشرفت مونتاژ یک تابلو بر اساس درصد ID 18+1R",
            "type": "checkbox_group_with_values",
            "items": [
                {"label": "صفحه گلند", "value": 5},
                {"label": "صفحه کنتور", "value": 20},
                {"label": "سیم بندی", "value": 55},
                {"label": "صفحه فیوز", "value": 15},
                {"label": "درب ها", "value": 5}
            ]
        },
        {"title": "نظافت و جمع آوری ضایعات محیط کارگاه", "type": "checkbox"},
        {"title": "امور خدماتی", "type": "checkbox"},
        {"title": "بارگیری و تخلیه لوازم", "type": "checkbox"},
        {"title": "ماموریت خارج از کارگاه", "type": "checkbox"},
        {"title": "امور متفرقه مرتبط با کارگاه", "type": "checkbox"},
        {"title": "نظافت", "type": "checkbox"},
    ]
}

# ==============================================================================
# بخش ۱۶: اندپوینت‌های اصلی
# ==============================================================================

# Health Check
@app.get("/health", summary="Health Check Endpoint")
def health_check():
    """بررسی سلامت سیستم، شامل اتصال به دیتابیس."""
    try:
        with SessionLocal() as db:
            db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected", "version": "2.1.0"}
    except Exception as e:
        logger.critical(f"Health check failed: Database connection error: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Service unavailable: Database connection failed."
        )

# احراز هویت
@app.post("/auth/login", response_model=Token, summary="User Login")
def login_for_access_token(login_data: PersonnelLogin, db: Session = Depends(get_db)):
    """ورود به سیستم و دریافت توکن JWT."""
    user = authenticate_user(db, login_data.username, login_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="نام کاربری یا رمز عبور اشتباه است",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role},
        expires_delta=access_token_expires
    )
    
    logger.info(f"User '{user.username}' logged in successfully.")
    return {"access_token": access_token, "token_type": "bearer", "user": user}

@app.get("/auth/me", response_model=PersonnelOut, summary="Get Current User")
def read_users_me(current_user: Personnel = Depends(get_current_active_user)):
    """دریافت اطلاعات کاربر لاگین کرده."""
    return current_user

# مدیریت پرسنل
@app.post("/personnel/", response_model=PersonnelOut, status_code=status.HTTP_201_CREATED)
def create_personnel(
    personnel: PersonnelCreate, 
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """ایجاد پرسنل جدید (فقط برای ناظران)."""
    if db.query(Personnel).filter(Personnel.username == personnel.username).first():
        logger.warning(f"Attempt to create personnel with existing username: {personnel.username}")
        raise HTTPException(status_code=409, detail="این نام کاربری قبلاً ثبت شده است.")
    
    db_personnel = Personnel(
        name=personnel.name,
        username=personnel.username,
        password_hash=get_password_hash(personnel.password),
        role=personnel.role.value
    )
    
    try:
        db.add(db_personnel)
        db.commit()
        db.refresh(db_personnel)
        logger.info(f"Personnel '{personnel.username}' created by '{current_user.username}'.")
        return db_personnel
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="این نام یا نام کاربری قبلاً ثبت شده است.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating personnel {personnel.username}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطای داخلی سرور در ایجاد پرسنل.")

@app.get("/personnel/", response_model=List[PersonnelOut])
def list_personnel(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """لیست تمام پرسنل فعال (فقط برای ناظران)."""
    return db.query(Personnel).filter(Personnel.is_active == 1).order_by(Personnel.name).all()

@app.put("/personnel/{personnel_id}", response_model=PersonnelOut)
def update_personnel(
    personnel_id: int, 
    personnel_data: PersonnelUpdate,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """بروزرسانی اطلاعات پرسنل (فقط برای ناظران)."""
    db_personnel = db.query(Personnel).filter(Personnel.id == personnel_id).first()
    if not db_personnel:
        raise HTTPException(status_code=404, detail="پرسنل یافت نشد.")
    
    db_personnel.name = personnel_data.name
    db_personnel.username = personnel_data.username
    db_personnel.role = personnel_data.role.value
    
    # فقط در صورتی که رمز عبور جدیدی ارسال شده باشد، آن را هش و ذخیره کن
    if personnel_data.password:
        db_personnel.password_hash = get_password_hash(personnel_data.password)
        
    try:
        db.commit()
        db.refresh(db_personnel)
        logger.info(f"Personnel ID {personnel_id} updated by '{current_user.username}'.")
        return db_personnel
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="پرسنل دیگری با این نام کاربری وجود دارد.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating personnel {personnel_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در بروزرسانی پرسنل.")

@app.delete("/personnel/{personnel_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_personnel(
    personnel_id: int,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """حذف (غیرفعال کردن) پرسنل (فقط برای ناظران)."""
    db_personnel = db.query(Personnel).filter(Personnel.id == personnel_id).first()
    if not db_personnel:
        raise HTTPException(status_code=404, detail="پرسنل یافت نشد.")
    
    if db_personnel.id == current_user.id:
        raise HTTPException(status_code=400, detail="شما نمی‌توانید حساب کاربری خود را حذف کنید.")
    
    db_personnel.is_active = 0
    try:
        db.commit()
        logger.info(f"Personnel '{db_personnel.username}' deactivated by '{current_user.username}'.")
    except Exception as e:
        db.rollback()
        logger.error(f"Error deactivating personnel {personnel_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در حذف پرسنل.")
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# گزارش کار روزانه
@app.get("/daily-reports/config")
def get_report_config():
    """دریافت تنظیمات فرم گزارش کار روزانه."""
    return REPORT_CONFIG

@app.post("/daily-reports/", response_model=DailyWorkReportOut, status_code=status.HTTP_201_CREATED)
async def create_daily_work_report(
    report: DailyWorkReportCreate, 
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_employee)
):
    """ایجاد گزارش کار روزانه جدید."""
    if report.personnel_id != current_user.id:
        logger.warning(f"User '{current_user.username}' attempted to create report for another user (ID: {report.personnel_id})")
        raise HTTPException(status_code=403, detail="شما فقط می‌توانید برای خودتان گزارش ایجاد کنید.")
    
    existing_report = db.query(DailyWorkReport).filter(
        DailyWorkReport.personnel_id == current_user.id,
        DailyWorkReport.report_date == report.report_date
    ).first()
    
    if existing_report:
        raise HTTPException(status_code=409, detail="برای این تاریخ قبلاً گزارش ثبت شده است.")
    
    db_report = DailyWorkReport(**report.model_dump())
    try:
        db.add(db_report)
        db.commit()
        db.refresh(db_report)
        
        await manager.broadcast_to_supervisors({
            "type": "new_daily_report",
            "data": {
                "report_id": db_report.id,
                "personnel_name": current_user.name,
                "report_date": report.report_date.isoformat()
            }
        })
        
        logger.info(f"Daily work report created for user '{current_user.username}' on {report.report_date}")
        # ✅ بهبود: برگرداندن مدل کامل با relation لود شده
        loaded_report = db.query(DailyWorkReport).options(joinedload(DailyWorkReport.personnel)).filter(DailyWorkReport.id == db_report.id).first()
        return loaded_report
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating daily report for user '{current_user.username}': {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در ایجاد گزارش.")

@app.get("/daily-reports/my-reports", response_model=List[DailyWorkReportOut])
def get_my_daily_reports(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت گزارش‌های کار روزانه کاربر جاری."""
    return db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.personnel_id == current_user.id)\
        .order_by(DailyWorkReport.report_date.desc(), DailyWorkReport.created_at.desc())\
        .all()

@app.get("/daily-reports/all", response_model=List[DailyWorkReportOut])
def get_all_daily_reports(
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """دریافت تمام گزارش‌های کار روزانه (فقط برای ناظران)."""
    return db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .order_by(DailyWorkReport.report_date.desc(), DailyWorkReport.created_at.desc())\
        .all()

@app.put("/daily-reports/{report_id}/approve", response_model=DailyWorkReportOut)
async def approve_daily_work_report(
    report_id: int, 
    supervisor_notes: Optional[str] = None,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """تایید گزارش کار روزانه (فقط برای ناظران)."""
    report = db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.id == report_id).first()
        
    if not report:
        raise HTTPException(status_code=404, detail="گزارش یافت نشد.")
    
    report.status = 'approved'
    report.supervisor_notes = supervisor_notes
    report.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(report)
        
        await manager.send_to_user(report.personnel_id, {
            "type": "report_approved",
            "data": {
                "report_id": report.id,
                "supervisor_name": current_user.name,
                "notes": supervisor_notes
            }
        })
        
        logger.info(f"Daily work report {report_id} approved by '{current_user.username}'.")
        return report
    except Exception as e:
        db.rollback()
        logger.error(f"Error approving report {report_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در تایید گزارش.")

@app.put("/daily-reports/{report_id}/reject", response_model=DailyWorkReportOut)
async def reject_daily_work_report(
    report_id: int, 
    supervisor_notes: str,
    db: Session = Depends(get_db), 
    current_user: Personnel = Depends(get_current_supervisor)
):
    """رد گزارش کار روزانه (فقط برای ناظران)."""
    report = db.query(DailyWorkReport)\
        .options(joinedload(DailyWorkReport.personnel))\
        .filter(DailyWorkReport.id == report_id).first()
        
    if not report:
        raise HTTPException(status_code=404, detail="گزارش یافت نشد.")
    
    if not supervisor_notes:
        raise HTTPException(status_code=400, detail="لطفاً دلیل رد گزارش را وارد کنید.")
    
    report.status = 'rejected'
    report.supervisor_notes = supervisor_notes
    report.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(report)
        
        await manager.send_to_user(report.personnel_id, {
            "type": "report_rejected",
            "data": {
                "report_id": report.id,
                "supervisor_name": current_user.name,
                "notes": supervisor_notes
            }
        })
        
        logger.info(f"Daily work report {report_id} rejected by '{current_user.username}'.")
        return report
    except Exception as e:
        db.rollback()
        logger.error(f"Error rejecting report {report_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در رد گزارش.")

# ==============================================================================
# بخش ۱۷: اندپوینت‌های اصلی پروژه
# ==============================================================================

# آپلود اکسل
@app.post("/projects/upload-excel/", status_code=status.HTTP_202_ACCEPTED)
async def upload_projects_from_excel(
    background_tasks: BackgroundTasks, 
    file: UploadFile = File(...)
):
    """آپلود فایل اکسل و پردازش پروژه‌ها در پس‌زمینه."""
    if not file.filename.endswith(('.xlsx', '.xls')): 
        raise HTTPException(status_code=400, detail="فرمت فایل نامعتبر است. فقط فایل‌های .xlsx و .xls مجاز هستند.")
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"حجم فایل بیش از حد مجاز ({MAX_FILE_SIZE / 1024 / 1024}MB) است.")
    
    await file.close()
    background_tasks.add_task(process_excel_in_background, contents, manager)
    logger.info(f"Excel file '{file.filename}' uploaded for background processing.")
    return {"message": "فایل برای پردازش در پس‌زمینه ارسال شد. نتیجه از طریق WebSocket اطلاع‌رسانی می‌شود."}

@app.post("/projects/upload-detailed-excel/", status_code=status.HTTP_202_ACCEPTED)
async def upload_detailed_excel(
    background_tasks: BackgroundTasks, 
    files: List[UploadFile] = File(...)
):
    """آپلود فایل‌های اکسل تفصیلی و پردازش در پس‌زمینه."""
    files_contents = []
    MAX_FILE_SIZE = 10 * 1024 * 1024 # 10MB per file
    
    for file in files:
        if not file.filename.endswith(('.xlsx', '.xls')): 
            raise HTTPException(status_code=400, detail=f"فرمت فایل '{file.filename}' نامعتبر است.")
        
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail=f"حجم فایل '{file.filename}' بیش از حد مجاز است.")
        
        await file.close()
        files_contents.append((file.filename, contents))
    
    background_tasks.add_task(process_detailed_excel_in_background, files_contents, manager)
    logger.info(f"{len(files_contents)} detailed Excel files uploaded for background processing.")
    return {"message": f"{len(files_contents)} فایل برای پردازش در پس‌زمینه ارسال شد."}

# مدیریت پروژه‌ها
@app.post("/projects/", response_model=ProjectOut, status_code=status.HTTP_201_CREATED)
async def create_project_manual(
    project_in: ProjectCreateManual, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor) 
):
    """ایجاد پروژه به صورت دستی توسط ناظر، با کسر خودکار از انبار."""
    # ✅ بهبود: اعتبارسنجی request_id الزامی شده است
    if not project_in.request_id:
        raise HTTPException(status_code=422, detail="شماره درخواست الزامی است.")
        
    if db.query(Project.id).filter(Project.request_id == project_in.request_id).first():
        raise HTTPException(status_code=409, detail=f"پروژه با شماره درخواست '{project_in.request_id}' موجود است.")
    
    db_proj = Project(
        name=project_in.name,
        location=project_in.location,
        customer_name=project_in.customer_name or "نامشخص",
        request_id=project_in.request_id,
        barcode_payload=project_in.request_id
    )
    
    try:
        db.add(db_proj)
        db.commit()
        db.refresh(db_proj)
        
        # کسر خودکار از انبار پس از ایجاد پروژه (اگر BOM تعریف شده باشد)
        await deduct_stock_for_project(db, db_proj, current_user)
        
        await manager.broadcast({"type": "update", "source": "new_manual_project", "project_id": db_proj.id})
        logger.info(f"Manual project '{project_in.request_id}' created by '{current_user.username}'.")
        return convert_project_orm_to_pydantic(db_proj)
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating manual project: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطای داخلی سرور در ایجاد پروژه.")

@app.get("/projects/", response_model=List[ProjectOut])
def list_projects(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """لیست پروژه‌ها با قابلیت فیلتر بر اساس تاریخ ایجاد."""
    query = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment), 
        joinedload(Project.comments)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects_orm = query.order_by(Project.created_at.desc()).all()
    return [convert_project_orm_to_pydantic(p) for p in projects_orm]

@app.get("/projects/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, db: Session = Depends(get_db)):
    """دریافت اطلاعات کامل یک پروژه."""
    p = db.query(Project).options(
        joinedload(Project.steps), 
        joinedload(Project.equipment), 
        joinedload(Project.comments)
    ).filter(Project.id == project_id).first()
    
    if not p: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")
    return convert_project_orm_to_pydantic(p)

@app.delete("/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_project(
    project_id: int, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """حذف پروژه (فقط برای ناظران)."""
    p = db.query(Project).get(project_id)
    if not p: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")
    
    try:
        db.delete(p)
        db.commit()
        
        await manager.broadcast({
            "type": "delete_project", 
            "data": {"project_id": project_id}
        })
        logger.info(f"Project ID {project_id} deleted by '{current_user.username}'.")
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در حذف پروژه.")

# کامنت‌ها
@app.post("/projects/{project_id}/comments", response_model=CommentOut, status_code=status.HTTP_201_CREATED)
async def create_comment_for_project(
    project_id: int, 
    comment_in: CommentCreate, 
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee) # ✅ نیاز به لاگین
):
    """ایجاد کامنت برای یک پروژه."""
    project = db.query(Project).get(project_id)
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")
    
    new_comment = Comment(
        project_id=project_id, 
        text=comment_in.text, 
        author=current_user.name # ✅ نویسنده کامنت، کاربر لاگین کرده است
    )
    
    try:
        db.add(new_comment)
        db.commit()
        db.refresh(new_comment)
        
        await manager.broadcast({
            "type": "update", 
            "source": "new_comment",
            "project_id": project_id
        })
        logger.info(f"Comment added to project {project_id} by '{current_user.username}'.")
        return new_comment
    except Exception as e:
        db.rollback()
        logger.error(f"Error adding comment to project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در ایجاد کامنت.")

# مراحل پروژه
@app.post("/projects/{project_id}/steps", response_model=StepOut, status_code=status.HTTP_201_CREATED)
async def add_step_to_project(
    project_id: int, 
    step_in: StepCreate, 
    db: Session = Depends(get_db)
):
    """افزودن یک مرحله جدید به پروژه."""
    project_orm = db.query(Project).options(joinedload(Project.steps)).filter(Project.id == project_id).first()
    if not project_orm: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")
    
    step_to_add = step_in.step
    if step_to_add == StepNameKey.EXIT_PANEL: 
        raise HTTPException(status_code=403, detail="خروج تابلو فقط با اسکن بارکد ثبت می‌شود.")
    
    if db.query(Step.id).filter(Step.project_id == project_id, Step.name_key == step_to_add).first(): 
        raise HTTPException(status_code=409, detail="این مرحله قبلاً برای این پروژه ثبت شده است.")
    
    completed_keys = {s.name_key for s in project_orm.steps}
    try:
        current_idx = ORDERED_MANUAL_STEP_KEYS.index(step_to_add)
        if current_idx > 0:
            previous_step_key = ORDERED_MANUAL_STEP_KEYS[current_idx - 1]
            if previous_step_key not in completed_keys: 
                raise HTTPException(
                    status_code=412, 
                    detail=f"مرحله پیش‌نیاز '{STEP_KEY_TO_NAME_MAP.get(previous_step_key.value)}' انجام نشده است."
                )
    except ValueError: 
        raise HTTPException(status_code=400, detail=f"مرحله '{step_to_add.value}' یک مرحله دستی مجاز نیست.")
    
    new_s = Step(project_id=project_id, name_key=step_to_add)
    
    try:
        db.add(new_s)
        db.commit()
        db.refresh(new_s)
        
        await manager.broadcast({"type": "update", "project_id": project_id})
        logger.info(f"Step '{step_to_add.value}' added to project {project_id}.")
        return StepOut.model_validate(new_s)
    except Exception as e:
        db.rollback()
        logger.error(f"Error adding step to project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در افزودن مرحله.")

@app.delete("/projects/{project_id}/steps/{step_name}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_step_from_project(
    project_id: int, 
    step_name: StepNameKey, # ✅ بهبود: استفاده از Enum برای اعتبارسنجی خودکار
    db: Session = Depends(get_db)
):
    """حذف یک مرحله از پروژه."""
    project_orm = db.query(Project).options(joinedload(Project.steps)).filter(Project.id == project_id).first()
    if not project_orm: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
        
    step_to_delete = db.query(Step).filter(
        Step.project_id == project_id, 
        Step.name_key == step_name
    ).first()
    
    if not step_to_delete: 
        raise HTTPException(status_code=404, detail=f"مرحله '{STEP_KEY_TO_NAME_MAP.get(step_name.value)}' برای این پروژه یافت نشد.")
        
    completed_keys = {s.name_key for s in project_orm.steps}
    try:
        all_ordered_keys = ORDERED_MANUAL_STEP_KEYS + [StepNameKey.EXIT_PANEL]
        current_idx = all_ordered_keys.index(step_name)
        if current_idx < len(all_ordered_keys) - 1:
            next_step_key = all_ordered_keys[current_idx + 1]
            if next_step_key in completed_keys: 
                raise HTTPException(
                    status_code=409, 
                    detail=f"ابتدا باید مرحله بعدی ('{STEP_KEY_TO_NAME_MAP.get(next_step_key.value)}') را لغو کنید."
                )
    except ValueError: 
        pass 
        
    try:
        db.delete(step_to_delete)
        db.commit()
        
        await manager.broadcast({"type": "update", "project_id": project_id})
        logger.info(f"Step '{step_name.value}' deleted from project {project_id}.")
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting step from project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در حذف مرحله.")

# خروج با بارکد
@app.post("/projects/exit-by-barcode/", response_model=StepOut, status_code=status.HTTP_201_CREATED)
async def exit_project_by_barcode(payload: BarcodeExitPayload, db: Session = Depends(get_db)):
    """ثبت خروج تابلو با اسکن بارکد (شماره درخواست)."""
    barcode_data = payload.barcode_data
    proj = db.query(Project).options(joinedload(Project.steps)).filter(Project.request_id == barcode_data).first()
    if not proj: 
        raise HTTPException(status_code=404, detail=f"پروژه با شماره درخواست '{barcode_data}' یافت نشد.")
    
    completed_step_keys = {s.name_key for s in proj.steps}
    required_step_keys = set(ORDERED_MANUAL_STEP_KEYS)
    missing_prerequisites = required_step_keys - completed_step_keys
    
    if missing_prerequisites:
        missing_names = sorted(
            [STEP_KEY_TO_NAME_MAP.get(s.value) for s in missing_prerequisites], 
            key=lambda x: [STEP_KEY_TO_NAME_MAP.get(step.value) for step in ORDERED_MANUAL_STEP_KEYS].index(x)
        )
        raise HTTPException(
            status_code=400, 
            detail=f"تمام پیش‌نیازها انجام نشده است. مراحل باقی‌مانده: {', '.join(missing_names)}"
        )
    
    if StepNameKey.EXIT_PANEL in completed_step_keys: 
        raise HTTPException(status_code=409, detail="خروج قبلاً برای این پروژه ثبت شده است.")
    
    exit_s = Step(project_id=proj.id, name_key=StepNameKey.EXIT_PANEL)
    
    try:
        db.add(exit_s)
        db.commit()
        db.refresh(exit_s)
        
        await manager.broadcast({"type": "update", "project_id": proj.id})
        logger.info(f"Project {proj.id} exited by barcode: {barcode_data}")
        return StepOut.model_validate(exit_s)
    except Exception as e:
        db.rollback()
        logger.error(f"Error exiting project by barcode {barcode_data}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در ثبت خروج.")

# جزئیات مونتاژ
@app.put("/projects/{project_id}/assembly-details/", response_model=ProjectOut)
async def update_assembly_details(
    project_id: int, 
    details: AssemblyDetailsUpdate, 
    db: Session = Depends(get_db)
):
    """بروزرسانی جزئیات مونتاژ (نوع تابلو و مونتاژکاران)."""
    p_orm = db.query(Project).get(project_id)
    if not p_orm: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    if p_orm.steps: 
        raise HTTPException(status_code=403, detail="امکان تغییر اطلاعات مونتاژ پس از شروع مراحل وجود ندارد.")
    
    p_orm.panel_type_key = details.panel_type_key.value
    p_orm.panel_code = details.panel_type_key.value
    p_orm.assembler_1 = details.assembler_1.strip()
    p_orm.assembler_2 = details.assembler_2.strip() if details.assembler_2 else None
    
    try:
        db.commit()
        db.refresh(p_orm)
        
        await manager.broadcast({"type": "update", "project_id": p_orm.id})
        logger.info(f"Assembly details updated for project {project_id}.")
        return convert_project_orm_to_pydantic(p_orm)
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating assembly details for project {project_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="خطا در بروزرسانی جزئیات مونتاژ.")

# اعتبارسنجی
@app.post("/projects/{project_id}/validate-branches", status_code=status.HTTP_202_ACCEPTED)
async def validate_branches_endpoint(
    project_id: int, 
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db)
):
    """اعتبارسنجی انشعابات پروژه با سرویس خارجی (در پس‌زمینه)."""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    purchase_list = [{"name": eq.item_name, "quantity": eq.quantity} for eq in project.equipment]
    background_tasks.add_task(run_branch_validation, project.id, project.request_id, purchase_list, manager)
    logger.info(f"Branch validation requested for project {project_id}")
    return {"message": "درخواست بررسی مغایرت انشعاب در پس‌زمینه ثبت شد."}

@app.post("/projects/{project_id}/validate-purchases", status_code=status.HTTP_202_ACCEPTED)
async def validate_purchases_endpoint(
    project_id: int, 
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db)
):
    """اعتبارسنجی خریدها با سرویس خارجی (در پس‌زمینه)."""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")

    purchase_list = [{"name": eq.item_name, "quantity": eq.quantity} for eq in project.equipment]
    background_tasks.add_task(run_purchase_validation, project.id, project.request_id, purchase_list, manager)
    logger.info(f"Purchase validation requested for project {project_id}")
    return {"message": "درخواست بررسی مغایرت تجهیزات در پس‌زمینه ثبت شد."}

# محاسبه مجدد کدهای تابلو
@app.post("/projects/recalculate-all-panel-codes/", status_code=200)
async def recalculate_all_panel_codes(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """محاسبه مجدد کدهای تابلو برای تمام پروژه‌ها بر اساس تجهیزات فعلی."""
    all_projects = db.query(Project).options(joinedload(Project.equipment)).all()
    updated_count = 0
    
    for project in all_projects:
        equipment_pydantic = [
            EquipmentItemBase(item_name=eq.item_name, quantity=eq.quantity) 
            for eq in project.equipment
        ]
        panel_code, panel_type_key = _find_panel_details_from_equipment(equipment_pydantic)
        
        if project.panel_code != panel_code or project.panel_type_key != panel_type_key:
            project.panel_code = panel_code
            project.panel_type_key = panel_type_key
            updated_count += 1
    
    if updated_count > 0:
        try:
            db.commit()
            await manager.broadcast({"type": "update", "source": "recalculation"})
            logger.info(f"Panel codes recalculated by '{current_user.username}': {updated_count} projects updated.")
        except Exception as e:
            db.rollback()
            logger.error(f"Error recalculating panel codes: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="خطا در محاسبه مجدد کدهای تابلو.")
    
    return {"message": f"عملیات با موفقیت انجام شد. {updated_count} پروژه بروزرسانی شد."}

# آمار مونتاژکاران
@app.get("/reports/assembler-stats/", response_model=Dict[str, AssemblerStatsOut])
def get_assembler_stats(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """دریافت آمار عملکرد مونتاژکاران."""
    query = db.query(Project).filter(
        Project.assembler_1.isnot(None), 
        Project.panel_type_key.isnot(None)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects = query.all()
    stats = defaultdict(lambda: {"total_panels": 0, "panels_by_type": defaultdict(int)})
    
    for p in projects:
        if not p.panel_type_key or not isinstance(p.panel_type_key, str): 
            continue
        
        panel_key = re.sub(r'\s+', '', p.panel_type_key).strip().upper()
        if not panel_key: 
            continue
        
        for assembler_name in [p.assembler_1, p.assembler_2]:
            if assembler_name and assembler_name.strip():
                clean_name = assembler_name.strip()
                stats[clean_name]["panels_by_type"][panel_key] += 1
    
    final_stats = {}
    for assembler, data in stats.items():
        total = sum(data["panels_by_type"].values())
        final_stats[assembler] = {
            "total_panels": total, 
            "panels_by_type": dict(data["panels_by_type"])
        }
    
    return final_stats
    
# ==============================================================================
# بخش ۱۷.۵: اندپوینت‌های انبارداری
# ==============================================================================

# --- مدیریت انبارها ---
@app.post("/warehouses/", response_model=WarehouseOut, status_code=status.HTTP_201_CREATED)
def create_warehouse(
    warehouse: WarehouseCreate,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """ایجاد یک انبار جدید."""
    db_warehouse = Warehouse(**warehouse.model_dump())
    db.add(db_warehouse)
    db.commit()
    db.refresh(db_warehouse)
    logger.info(f"Warehouse '{warehouse.name}' created by '{current_user.username}'.")
    return db_warehouse

@app.get("/warehouses/", response_model=List[WarehouseOut])
def list_warehouses(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت لیست تمام انبارها."""
    return db.query(Warehouse).all()

# --- مدیریت کالاها در انبار ---
@app.post("/warehouse-items/", response_model=WarehouseItemOut, status_code=status.HTTP_201_CREATED)
def create_warehouse_item(
    item: WarehouseItemCreate,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """تعریف یک کالای جدید در سیستم انبار."""
    if db.query(WarehouseItem).filter(WarehouseItem.item_name == item.item_name).first():
        raise HTTPException(status_code=409, detail="این کالا قبلاً تعریف شده است.")
    db_item = WarehouseItem(**item.model_dump())
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    logger.info(f"Warehouse item '{item.item_name}' created by '{current_user.username}'.")
    return db_item

@app.get("/warehouse-items/", response_model=List[WarehouseItemOut])
def list_warehouse_items(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت لیست تمام کالاهای تعریف‌شده."""
    return db.query(WarehouseItem).order_by(WarehouseItem.item_name).all()
    
# --- ثبت تراکنش‌ها ---
@app.post("/inventory/in/", response_model=InventoryTransactionOut)
async def log_inventory_in(
    transaction_in: InventoryTransactionIn,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """ثبت ورود کالا به انبار."""
    item = db.query(WarehouseItem).filter(WarehouseItem.item_name == transaction_in.item_name).first()
    if not item:
        raise HTTPException(status_code=404, detail=f"کالایی با نام '{transaction_in.item_name}' یافت نشد. ابتدا آن را تعریف کنید.")

    transaction = InventoryTransaction(
        warehouse_id=transaction_in.warehouse_id,
        item_id=item.id,
        quantity=transaction_in.quantity,
        transaction_type=TransactionType.IN,
        user_id=current_user.id,
        notes=transaction_in.notes
    )
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    logger.info(f"{transaction_in.quantity} of '{item.item_name}' logged IN by '{current_user.username}'.")
    # ✅ بهبود: برگرداندن مدل کامل برای نمایش در UI
    return db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    ).filter(InventoryTransaction.id == transaction.id).first()


@app.post("/inventory/out/", response_model=InventoryTransactionOut)
async def log_inventory_out(
    transaction_out: InventoryTransactionOutManual,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """ثبت خروج دستی کالا از انبار."""
    item = db.query(WarehouseItem).filter(WarehouseItem.item_name == transaction_out.item_name).first()
    if not item:
        raise HTTPException(status_code=404, detail=f"کالایی با نام '{transaction_out.item_name}' یافت نشد.")

    transaction = InventoryTransaction(
        warehouse_id=transaction_out.warehouse_id,
        item_id=item.id,
        quantity=transaction_out.quantity,
        transaction_type=TransactionType.OUT,
        user_id=current_user.id,
        project_id=transaction_out.project_id,
        notes=transaction_out.notes
    )
    db.add(transaction)
    db.commit()
    await check_stock_and_alert(db, item, transaction_out.warehouse_id)
    db.refresh(transaction)
    logger.info(f"{transaction_out.quantity} of '{item.item_name}' logged OUT by '{current_user.username}'.")
    return db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    ).filter(InventoryTransaction.id == transaction.id).first()

# --- مشاهده و جستجوی موجودی ---
@app.get("/inventory/items/", response_model=List[CurrentStockItem])
def get_current_inventory(
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت موجودی فعلی تمام کالاها در تمام انبارها."""
    stock_level_expr = func.sum(
        case(
            (InventoryTransaction.transaction_type == TransactionType.IN.value, InventoryTransaction.quantity),
            else_=-InventoryTransaction.quantity
        )
    ).label("current_stock")

    # ✅ بهبود: این کوئری تمام کالاها را در تمام انبارها با موجودی‌شان برمی‌گرداند.
    results = db.query(
        WarehouseItem.id.label("item_id"),
        WarehouseItem.item_name,
        Warehouse.id.label("warehouse_id"),
        Warehouse.name.label("warehouse_name"),
        WarehouseItem.min_stock_level,
        stock_level_expr
    ).join(InventoryTransaction, InventoryTransaction.item_id == WarehouseItem.id)\
     .join(Warehouse, Warehouse.id == InventoryTransaction.warehouse_id)\
     .group_by(WarehouseItem.id, Warehouse.id)\
     .all()

    return [CurrentStockItem(**row._asdict()) for row in results]
    
@app.get("/inventory/history/", response_model=List[InventoryTransactionOut])
def get_inventory_history(
    db: Session = Depends(get_db),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    transaction_type: Optional[TransactionType] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None,
    item_id: Optional[int] = None,
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت تاریخچه تراکنش‌های انبار با قابلیت فیلتر."""
    query = db.query(InventoryTransaction).options(
        joinedload(InventoryTransaction.warehouse),
        joinedload(InventoryTransaction.item),
        joinedload(InventoryTransaction.user)
    )
    if start_date:
        query = query.filter(InventoryTransaction.timestamp >= datetime.combine(start_date, time.min))
    if end_date:
        query = query.filter(InventoryTransaction.timestamp <= datetime.combine(end_date, time.max))
    if transaction_type:
        query = query.filter(InventoryTransaction.transaction_type == transaction_type)
    if project_id:
        query = query.filter(InventoryTransaction.project_id == project_id)
    if user_id:
        query = query.filter(InventoryTransaction.user_id == user_id)
    if item_id:
        query = query.filter(InventoryTransaction.item_id == item_id)
    
    return query.order_by(InventoryTransaction.timestamp.desc()).all()


# --- مدیریت تجهیزات مورد نیاز هر تابلو ---
@app.post("/inventory/panel-items/", status_code=status.HTTP_201_CREATED)
def define_panel_code_items(
    definition: PanelCodeItemsDefinition,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_supervisor)
):
    """تعریف یا بازنویسی لیست تجهیزات (BOM) برای یک کد تابلو."""
    # حذف تعاریف قبلی برای این کد تابلو
    db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == definition.panel_code).delete()
    
    for item in definition.items:
        # اطمینان از وجود کالا در انبار
        if not db.query(WarehouseItem).filter(WarehouseItem.item_name == item.item_name).first():
            raise HTTPException(status_code=400, detail=f"کالای '{item.item_name}' در انبار تعریف نشده است.")
        
        db_item = PanelCodeItems(
            panel_code=definition.panel_code,
            item_name=item.item_name,
            quantity_required=item.quantity_required
        )
        db.add(db_item)
    
    db.commit()
    logger.info(f"BOM for panel code '{definition.panel_code}' defined/updated by '{current_user.username}'.")
    return {"message": f"تجهیزات برای کد تابلو '{definition.panel_code}' با موفقیت تعریف شد."}

@app.get("/inventory/panel-items/{panel_code}", response_model=List[PanelCodeItemOut])
def get_panel_code_items(
    panel_code: str,
    db: Session = Depends(get_db),
    current_user: Personnel = Depends(get_current_employee)
):
    """دریافت لیست تجهیزات (BOM) برای یک کد تابلو."""
    return db.query(PanelCodeItems).filter(PanelCodeItems.panel_code == panel_code).all()    

# ==============================================================================
# بخش ۱۸: گزارش‌ها
# ==============================================================================

# گزارش خلاصه خرید
@app.get("/reports/procurement-summary/excel", response_class=StreamingResponse)
def get_procurement_summary_report(
    start_date: date, 
    end_date: date, 
    db: Session = Depends(get_db)
):
    """گزارش خلاصه خرید به صورت اکسل"""
    query = db.query(Project).options(joinedload(Project.equipment)).filter(
        Project.created_at >= datetime.combine(start_date, time.min),
        Project.created_at <= datetime.combine(end_date, time.max)
    )
    projects = query.all()

    if not projects:
        raise HTTPException(status_code=404, detail="هیچ پروژه‌ای در این بازه زمانی یافت نشد.")

    summary_by_panel = defaultdict(lambda: defaultdict(int))
    overall_summary = defaultdict(int)

    for project in projects:
        panel_key = project.panel_code or project.panel_type_key or "نامشخص"
        
        for item in project.equipment:
            normalized_name = normalize_text(item.item_name)
            summary_by_panel[panel_key][normalized_name] += item.quantity
            overall_summary[normalized_name] += item.quantity

    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    summary_sheet = workbook.create_sheet(title="خلاصه کلی تجهیزات")
    summary_sheet.sheet_view.rightToLeft = True
    summary_sheet.append(["نام تجهیز", "تعداد کل مورد نیاز"])
    
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for item_name, total_quantity in sorted(overall_summary.items()):
        summary_sheet.append([item_name, total_quantity])
    
    summary_sheet.column_dimensions['A'].width = 60
    summary_sheet.column_dimensions['B'].width = 20

    for panel_key, items in sorted(summary_by_panel.items()):
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', panel_key)[:30]
        sheet = workbook.create_sheet(title=safe_sheet_name)
        sheet.sheet_view.rightToLeft = True
        
        sheet.append([f"تجهیزات مورد نیاز برای تابلو: {panel_key}"])
        sheet.merge_cells('A1:B1')
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = center_align
        
        sheet.append(["نام تجهیز", "تعداد مورد نیاز"])
        for cell in sheet[2]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        for item_name, quantity in sorted(items.items()):
            row = [item_name, quantity]
            sheet.append(row)
            for cell in sheet[sheet.max_row]:
                cell.border = thin_border
        
        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['B'].width = 20

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Procurement_Summary_{start_date}_to_{end_date}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    
    logger.info(f"Procurement summary report generated: {filename}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

# گزارش‌های خروج
@app.get("/reports/exited-panels/simple-excel", response_class=StreamingResponse)
def get_exited_panels_simple_report(report_date: date, db: Session = Depends(get_db)):
    """گزارش ساده تابلوهای خروجی در یک تاریخ مشخص"""
    start_of_day, end_of_day = datetime.combine(report_date, time.min), datetime.combine(report_date, time.max)
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    headers = ["ردیف", "نام و نام خانوادگی", "شماره تقاضا", "کد تابلو", "تاریخ و ساعت خروج"]
    sheet.append(headers)
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for idx, project in enumerate(exited_projects, 1):
        exit_step = next((step for step in project.steps if step.name_key == StepNameKey.EXIT_PANEL), None)
        exit_time_str = jdatetime.datetime.fromgregorian(datetime=exit_step.timestamp).strftime('%Y/%m/%d - %H:%M:%S') if exit_step else "نامشخص"
        row_data = [
            idx, 
            project.customer_name, 
            project.request_id, 
            project.panel_code or project.panel_type_key or "-", 
            exit_time_str
        ]
        sheet.append(row_data)
        
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 25
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Simple_Exited_Report_{report_date.strftime('%Y-%m-%d')}.xlsx"
    logger.info(f"Simple exit report generated for date: {report_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/exited-panels/simple-excel-range", response_class=StreamingResponse)
def get_exited_panels_simple_report_range(start_date: date, end_date: date, db: Session = Depends(get_db)):
    """گزارش ساده تابلوهای خروجی در بازه زمانی"""
    start_of_day = datetime.combine(start_date, time.min)
    end_of_day = datetime.combine(end_date, time.max)
    
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    headers = ["ردیف", "نام و نام خانوادگی", "شماره تقاضا", "کد تابلو", "تاریخ و ساعت خروج"]
    sheet.append(headers)
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for idx, project in enumerate(exited_projects, 1):
        exit_step = next((step for step in project.steps if step.name_key == StepNameKey.EXIT_PANEL), None)
        exit_time_str = jdatetime.datetime.fromgregorian(datetime=exit_step.timestamp).strftime('%Y/%m/%d - %H:%M:%S') if exit_step else "نامشخص"
        row_data = [
            idx, 
            project.customer_name, 
            project.request_id, 
            project.panel_code or project.panel_type_key or "-", 
            exit_time_str
        ]
        sheet.append(row_data)
        
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 25
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Exited_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    
    logger.info(f"Exit report generated for range: {start_date} to {end_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

# گزارش‌های آماده تحویل
@app.get("/reports/ready-for-delivery/detailed-excel", response_class=StreamingResponse)
def get_detailed_delivery_report_excel(direction: str, db: Session = Depends(get_db)):
    """گزارش تفصیلی تابلوهای آماده تحویل"""
    if direction not in ["west", "east"]: 
        raise HTTPException(status_code=400, detail="جهت باید 'west' یا 'east' باشد.")
    
    company_name = WEST_COMPANY if direction == "west" else EAST_COMPANY
    jalali_date_str = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    ready_projects = get_projects_by_status(
        db, 
        required_steps=set(ORDERED_MANUAL_STEP_KEYS), 
        forbidden_steps={StepNameKey.EXIT_PANEL}
    )
    filtered_projects = [p for p in ready_projects if get_direction(p.request_id) == direction]
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    header_font_large = Font(bold=True, name='B Titr', size=14)
    header_font_small = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.merge_cells('A1:U1')
    sheet['A1'].value = "شرکت فرداد سازه گلشن"
    sheet['A1'].font = header_font_large
    sheet['A1'].alignment = center_align

    sheet.merge_cells('A2:U2')
    sheet['A2'].value = f"شرکت تحویل گیرنده: {company_name}"
    sheet['A2'].font = header_font_small
    sheet['A2'].alignment = center_align

    sheet.merge_cells('S3:U3')
    sheet['S3'].value = f"تاریخ گزارش: {jalali_date_str}"
    sheet['S3'].font = Font(name='B Nazanin', size=11)
    sheet['S3'].alignment = center_align

    table_start_row = 5
    
    header_font = Font(bold=True, name='B Nazanin', size=11)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    headers_l1 = {
        'A': 'ردیف', 
        'B': 'نام و نام خانوادگی متقاضی', 
        'C': 'شماره تقاضا', 
        'D': 'نوع تابلو', 
        'E': 'تعداد کنتور', 
        'G': 'تعداد فیوز', 
        'J': 'تعداد مودم', 
        'K': 'کنتور پیش بینی', 
        'M': 'فیوز پیش بینی', 
        'O': 'سکو', 
        'P': 'تسمه استیل', 
        'Q': 'بست تسمه', 
        'R': 'پیچ', 
        'T': 'لوله نیم گرد', 
        'U': 'لوله خرطومی'
    }
    
    for col, value in headers_l1.items(): 
        sheet[f'{col}{table_start_row}'] = value
    
    sheet.merge_cells(f'E{table_start_row}:F{table_start_row}')
    sheet.merge_cells(f'G{table_start_row}:I{table_start_row}')
    sheet.merge_cells(f'K{table_start_row}:L{table_start_row}')
    sheet.merge_cells(f'M{table_start_row}:N{table_start_row}')
    sheet.merge_cells(f'R{table_start_row}:S{table_start_row}')
    
    table_header_row_2 = table_start_row + 1
    headers_l2 = {
        'E': 'تکفاز', 
        'F': 'سه فاز', 
        'G': 'تکفاز', 
        'H': 'سه فاز', 
        'I': None, 
        'K': 'تکفاز', 
        'L': 'سه فاز', 
        'M': 'تکفاز', 
        'N': 'سه فاز', 
        'R': 'یکسررزوه', 
        'S': '۱۶*۳۰۰', 
        'T': None, 
        'U': None
    }
    
    for col, value in headers_l2.items(): 
        if value:
            sheet[f'{col}{table_header_row_2}'] = value
    
    for col in ['A', 'B', 'C', 'D', 'J', 'O', 'P', 'Q', 'T', 'U']: 
        sheet.merge_cells(f'{col}{table_start_row}:{col}{table_header_row_2}')
    
    for row in sheet.iter_rows(min_row=table_start_row, max_row=table_header_row_2):
        for cell in row: 
            if cell.value:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
    
    row_num = table_header_row_2 + 1
    for idx, project in enumerate(filtered_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "مودم" in normalize_text(item.item_name)
        )
        
        total_single_phase_fuses = sum([
            summary.get('فیوز تک فاز 16', 0), 
            summary.get('فیوز تک فاز 25', 0), 
            summary.get('فیوز تک فاز 32', 0)
        ])
        total_three_phase_fuses = sum([
            summary.get('فیوز سه فاز 25', 0), 
            summary.get('فیوز سه فاز 32', 0), 
            summary.get('فیوز سه فاز 63', 0)
        ])
        
        row_data = {
            'A': idx, 
            'B': project.customer_name, 
            'C': project.request_id, 
            'D': project.panel_code or '-',
            'E': summary.get('کنتور تک فاز', 0), 
            'F': summary.get('کنتور سه فاز', 0), 
            'G': total_single_phase_fuses, 
            'H': total_three_phase_fuses, 
            'I': None, 
            'J': modem_count,
            'K': reservation.get('purchased_single_phase', 0), 
            'L': 0, 
            'M': 0, 
            'N': 0,
            'O': summary.get('سکو', 0), 
            'P': summary.get('تسمه استیل', 0), 
            'Q': summary.get('بست تسمه استیل', 0),
            'R': 0, 
            'S': summary.get('پیچ ۱۶*۳۰۰', 0), 
            'T': summary.get('لوله نیم گرد', 0), 
            'U': summary.get('لوله خرطومی', 0)
        }
        
        for col, value in row_data.items():
            cell = sheet[f'{col}{row_num}']
            cell.value = value
            cell.alignment = center_align
            cell.border = thin_border
        
        row_num += 1
    
    footer_row = sheet.max_row + 3
    sheet.cell(row=footer_row, column=3, value="تحویل دهنده:").font = header_font_small
    sheet.cell(row=footer_row, column=10, value="تحویل گیرنده:").font = header_font_small
    sheet.cell(row=footer_row, column=18, value=f"تاریخ: {jalali_date_str}").font = header_font_small

    for col, width in {'B': 35, 'C': 18, 'D': 18}.items(): 
        sheet.column_dimensions[col].width = width
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Ready_For_Delivery_{direction}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info(f"Detailed delivery report generated for direction: {direction}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

# گزارش‌های تایید ناظر
@app.get("/reports/supervisor-checklist-items", response_model=List[str])
def get_supervisor_checklist_items():
    """دریافت آیتم‌های چک‌لیست کنترل کیفیت ناظر"""
    return SUPERVISOR_CHECKLIST_ITEMS

@app.get("/reports/supervisor-approval/simple/excel", response_class=StreamingResponse)
def get_supervisor_approval_simple_report_excel(db: Session = Depends(get_db)):
    """گزارش ساده تایید ناظر به صورت اکسل"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    sheet.merge_cells('A1:F1')
    cell_a1 = sheet['A1']
    cell_a1.value = "شرکت فرداد سازه گلشن"
    cell_a1.font = Font(bold=True, size=14)
    cell_a1.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('A2:F2')
    cell_a2 = sheet['A2']
    cell_a2.value = "لیست تابلو های آماده تحویل"
    cell_a2.font = Font(bold=True, size=12)
    cell_a2.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('E3:F3')
    sheet['E3'].value = f"تاریخ: {jalali_date}"
    sheet['E3'].alignment = Alignment(horizontal='center')
    
    headers = ["ردیف", "نام و نام خانوادگی", "شماره تقاضا", "کد تابلو", "تایید", "توضیحات"]
    sheet.append(headers)
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    header_row_num = sheet.max_row
    for cell in sheet[header_row_num]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    
    for idx, p in enumerate(projects, 1):
        sheet.append([idx, p.customer_name, p.request_id, p.panel_code or '-', '√', ''])
    
    for row in sheet.iter_rows(min_row=header_row_num + 1, max_row=sheet.max_row, min_col=1, max_col=6):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            cell.alignment = right_align if col_idx == 2 else center_align
    
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    
    last_row = sheet.max_row + 3
    sheet.cell(row=last_row, column=2, value="نام و امضا مسئول کارگاه:").font = Font(bold=True)
    sheet.cell(row=last_row, column=4, value="نام و امضا مسئول دفتر نظارت:").font = Font(bold=True)
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"SupervisorApprovalReport_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info("Supervisor approval report generated")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/supervisor-approval/simple/html", response_class=HTMLResponse)
def get_supervisor_approval_simple_report_html(db: Session = Depends(get_db)):
    """گزارش ساده تایید ناظر به صورت HTML"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    template = Template(SUPERVISOR_APPROVAL_SIMPLE_TEMPLATE_STR)
    html_content = template.render(projects=projects, jalali_date=jalali_date)
    return HTMLResponse(content=html_content)

@app.get("/reports/supervisor-approval/checklist/excel", response_class=StreamingResponse)
def get_supervisor_qc_checklist_excel(db: Session = Depends(get_db)):
    """چک‌لیست کنترل کیفیت ناظر به صورت اکسل"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    sheet.merge_cells('A1:E1')
    sheet['A1'].value = "شرکت فرداد سازه گلشن"
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = center_align
    
    sheet.merge_cells('A2:E2')
    sheet['A2'].value = "چک لیست کنترل کیفیت تابلوهای آماده تحویل"
    sheet['A2'].font = Font(bold=True, size=12)
    sheet['A2'].alignment = center_align
    
    sheet.cell(row=2, column=len(projects) + 2).value = f"تاریخ: {jalali_date}"
    
    header_request_id = ["ردیف", "شرح کنترل کیفیت"] + [p.request_id for p in projects]
    header_panel_code = ["", ""] + [p.panel_code or '-' for p in projects]
    
    sheet.append(header_request_id)
    sheet.append(header_panel_code)
    
    for row_idx in range(4, 6):
        for col_idx in range(1, len(header_request_id) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
    
    for idx, item in enumerate(SUPERVISOR_CHECKLIST_ITEMS, 1):
        sheet.append([idx, item] + ([''] * len(projects)))
    
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row):
        for cell in row: 
            cell.border = thin_border
        row[1].alignment = right_align
    
    sheet.column_dimensions['B'].width = 50
    for i in range(len(projects)): 
        sheet.column_dimensions[get_column_letter(i + 3)].width = 15
    
    last_row = sheet.max_row + 3
    sheet.cell(row=last_row, column=2, value="نام و امضا مسئول کارگاه:").font = bold_font
    sheet.cell(row=last_row, column=5, value="نام و امضا مسئول دفتر نظارت:").font = bold_font
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"SupervisorChecklist_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info("Supervisor checklist report generated")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/supervisor-approval/checklist/html", response_class=HTMLResponse)
def get_supervisor_qc_checklist_html(db: Session = Depends(get_db)):
    """چک‌لیست کنترل کیفیت ناظر به صورت HTML"""
    required_steps = {
        StepNameKey.START_ASSEMBLY, 
        StepNameKey.END_ASSEMBLY, 
        StepNameKey.TEAM_LEAD_APPROVAL, 
        StepNameKey.TEST, 
        StepNameKey.QUALITY_CONTROL
    }
    projects = get_projects_by_status(
        db, 
        required_steps=required_steps, 
        forbidden_steps={StepNameKey.SUPERVISOR_APPROVAL}
    )
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    template = Template(SUPERVISOR_CHECKLIST_TEMPLATE_STR)
    html_content = template.render(
        projects=projects, 
        checklist_items=SUPERVISOR_CHECKLIST_ITEMS, 
        jalali_date=jalali_date
    )
    return HTMLResponse(content=html_content)

# گزارش KPI
@app.get("/reports/kpi-summary", response_model=KpiSummary)
def get_kpi_summary(db: Session = Depends(get_db)):
    """دریافت خلاصه شاخص‌های عملکرد (نسخه اصلاح شده)"""
    total_projects = db.query(func.count(Project.id)).scalar() or 0

    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    completed_this_month = db.query(func.count(Project.id)).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL,
        Step.timestamp >= start_of_month
    ).scalar() or 0

    all_projects_with_steps = db.query(Project).options(joinedload(Project.steps)).all()
    
    durations = defaultdict(list)
    completion_times = []

    for p in all_projects_with_steps:
        # مراحل را بر اساس زمان مرتب می‌کنیم
        sorted_steps = sorted(p.steps, key=lambda s: s.timestamp)
        if not sorted_steps:
            continue
            
        # یک تایم‌لاین از زمان ایجاد پروژه تا پایان هر مرحله می‌سازیم
        # ✅ بهبود: ساختار تایم‌لاین برای محاسبه آسان‌تر
        timeline = [(p.created_at, "PROJECT_CREATION")] + [(s.timestamp, s.name_key) for s in sorted_steps]

        # محاسبه مدت زمان بین هر دو نقطه متوالی در تایم‌لاین
        for i in range(len(timeline) - 1):
            start_time, start_key_or_event = timeline[i]
            end_time, end_key = timeline[i+1]
            
            # مدت زمان به مرحله قبلی تعلق دارد
            # مثال: زمان بین START_ASSEMBLY و END_ASSEMBLY متعلق به مرحله START_ASSEMBLY است
            # ✅ اصلاح منطق اصلی: مدت زمان به مرحله شروع تعلق می‌گیرد
            if start_key_or_event != "PROJECT_CREATION":
                duration_hours = (end_time - start_time).total_seconds() / 3600
                if duration_hours >= 0:
                     durations[start_key_or_event].append(duration_hours)

        # محاسبه زمان تکمیل کل پروژه
        exit_step = next((s for s in sorted_steps if s.name_key == StepNameKey.EXIT_PANEL), None)
        if exit_step:
            total_duration_days = (exit_step.timestamp - p.created_at).total_seconds() / (3600 * 24)
            if total_duration_days > 0:
                completion_times.append(total_duration_days)

    # محاسبه میانگین‌ها
    avg_step_durations = {}
    for key in ORDERED_MANUAL_STEP_KEYS:
        if durations.get(key):
            # استفاده از میانه (median) به جای میانگین (mean) برای کاهش تاثیر داده‌های پرت
            avg_hours = round(float(np.median(durations[key])), 2)
            avg_step_durations[key.value] = avg_hours
        else:
            avg_step_durations[key.value] = None

    # پیدا کردن گلوگاه
    bottleneck_step_key = None
    if any(v is not None for v in avg_step_durations.values()):
        bottleneck_step_key = max(
            avg_step_durations, 
            key=lambda k: avg_step_durations.get(k) or -1
        )
    bottleneck_step_name = STEP_KEY_TO_NAME_MAP.get(bottleneck_step_key, None) if bottleneck_step_key else None

    avg_completion_time = round(float(np.mean(completion_times)), 2) if completion_times else None

    logger.info("KPI summary generated successfully.")
    return KpiSummary(
        total_projects=total_projects,
        completed_this_month=completed_this_month,
        avg_completion_time_days=avg_completion_time,
        bottleneck_step=bottleneck_step_name,
        step_durations=avg_step_durations
    )    

# ==============================================================================
# بخش ۱۹: اندپوینت‌های برچسب و چاپ
# ==============================================================================

async def get_project_label_data(project_id: int, db: Session) -> Dict[str, Any]:
    """دریافت داده‌های برچسب پروژه"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    direction = get_direction(project.request_id)
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=2)
    qr.add_data(project.request_id)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    qr_img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
    
    jalali_date_str = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return {
        "project": project, 
        "qr_code_base64": qr_img_str, 
        "report_date_jalali": jalali_date_str, 
        "direction": "شرق" if direction == "east" else "غرب" if direction == "west" else None
    }

@app.get("/projects/{project_id}/qc-label", response_class=HTMLResponse, include_in_schema=False)
async def get_qc_label(request: Request, project_id: int, db: Session = Depends(get_db)):
    """برچسب کنترل کیفیت"""
    project = db.query(Project).get(project_id)
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    panel_id = (project.panel_code or "NO-ID").replace("/", "-")
    try: 
        jalali_dt = jdatetime.datetime.fromgregorian(datetime=project.created_at)
        creation_month = f"{jalali_dt.month:02d}"
    except Exception: 
        creation_month = "00"
    
    serial_number = f"09-{panel_id}-{creation_month}-{project.request_id.strip()[-10:]}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=1)
    qr.add_data(serial_number)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    qr_code_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
    
    context = {
        "request": request, 
        "serial_number": serial_number, 
        "qr_code_base64": qr_code_base64
    }
    
    return templates.TemplateResponse("qc_label_template.html", context)

@app.get("/projects/{project_id}/label", response_class=HTMLResponse, include_in_schema=False)
async def get_project_label_for_printing(project_id: int, db: Session = Depends(get_db)):
    """برچسب پروژه برای چاپ"""
    data = await get_project_label_data(project_id, db)
    data["font_path"] = "/static/Vazirmatn-regular.ttf"
    template_str_with_print = PROJECT_SLIP_TEMPLATE_STR.replace(
        "</body>", 
        "<script>window.onload = function() { window.print(); };</script></body>"
    )
    template = Template(template_str_with_print)
    return HTMLResponse(content=template.render(data))

@app.get("/projects/{project_id}/qc-checklist", response_class=HTMLResponse)
def get_individual_qc_checklist(project_id: int, db: Session = Depends(get_db)):
    """چک‌لیست کنترل کیفیت فردی"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    template = Template(INDIVIDUAL_QC_CHECKLIST_TEMPLATE_STR)
    html_content = template.render(
        project=project, 
        checklist_items=SUPERVISOR_CHECKLIST_ITEMS, 
        jalali_date=jalali_date
    )
    return HTMLResponse(content=html_content)

@app.get("/projects/{project_id}/download-pdf", response_class=StreamingResponse)
async def download_project_label_pdf(project_id: int, db: Session = Depends(get_db)):
    """دانلود برچسب پروژه به صورت PDF"""
    data = await get_project_label_data(project_id, db)
    data["font_path"] = "Vazirmatn-regular.ttf"
    template = Template(PROJECT_SLIP_TEMPLATE_STR)
    html_content = template.render(data)
    
    pdf_stream = BytesIO()
    
    def fetch_resources(uri, rel):
        static_path = resource_path(os.path.join(STATIC_DIR_NAME, uri))
        if os.path.exists(static_path): 
            return static_path
        root_path = resource_path(uri)
        if os.path.exists(root_path): 
            return root_path
        return None
    
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_stream, link_callback=fetch_resources, encoding='utf-8')
    if pisa_status.err: 
        logger.error(f"PDF generation error: {pisa_status.err}")
        raise HTTPException(status_code=500, detail=f"خطا در تولید فایل PDF: {pisa_status.err}")
    
    pdf_stream.seek(0)
    filename = f"ProjectLabel_{data['project'].request_id}.pdf"
    
    logger.info(f"PDF label generated for project {project_id}")
    return StreamingResponse(
        pdf_stream, 
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/projects/{project_id}/download-excel", response_class=StreamingResponse)
async def download_project_label_excel(project_id: int, db: Session = Depends(get_db)):
    """دانلود اطلاعات پروژه به صورت اکسل"""
    project_data = await get_project_label_data(project_id, db)
    project = project_data['project']
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    sheet.append(["نام پروژه", project.name])
    sheet.append(["شماره درخواست", project.request_id])
    sheet.append(["مشتری", project.customer_name])
    sheet.append(["تاریخ صدور", project_data['report_date_jalali']])
    
    if project.panel_code: 
        sheet.append(["کد تابلو", project.panel_code])
    if project_data['direction']: 
        sheet.append(["جهت ارسال", project_data['direction']])
    
    sheet.append([])
    sheet.append(["ردیف", "نام تجهیز", "تعداد"])
    
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for i, item in enumerate(project.equipment, start=1):
        sheet.append([i, item.item_name, item.quantity])
    
    for col in ['A', 'B', 'C']:
        sheet.column_dimensions[col].width = 20
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"ProjectEquipment_{project.request_id}.xlsx"
    logger.info(f"Excel equipment list generated for project {project_id}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/projects/{project_id}/exit-slip", response_class=HTMLResponse)
async def get_project_exit_slip(project_id: int, db: Session = Depends(get_db)):
    """برگه خروج پروژه"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    try:
        summary_data = generate_project_summary_data(project)
        data = {
            "project": project, 
            "summary_data": summary_data, 
            "report_date_jalali": jdatetime.datetime.now().strftime('%Y/%m/%d'), 
            "font_path": "/static/Vazirmatn-regular.ttf"
        }
        
        template_str_with_print = DETAILED_EXIT_SLIP_TEMPLATE_STR.replace(
            "</body>", 
            "<script>window.onload = function() { setTimeout(function(){ window.print(); }, 500); };</script></body>"
        )
        template = Template(template_str_with_print)
        return HTMLResponse(content=template.render(data))
    except Exception as e:
        logger.error(f"ERROR rendering exit slip for project {project_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"خطا در تولید برگه خروج. جزئیات خطا: {e}")

@app.get("/projects/{project_id}/exit-slip-raw", response_class=HTMLResponse)
async def get_project_exit_slip_raw_html(project_id: int, request: Request, db: Session = Depends(get_db)):
    """برگه خروج پروژه (قالب خام)"""
    project = db.query(Project).options(joinedload(Project.equipment)).filter(Project.id == project_id).first()
    if not project: 
        raise HTTPException(status_code=404, detail="پروژه یافت نشد.")
    
    summary_data = generate_project_summary_data(project)
    context = {
        "request": request, 
        "project": project, 
        "summary_data": summary_data, 
        "report_date_jalali": jdatetime.datetime.now().strftime('%Y/%m/%d')
    }
    
    return templates.TemplateResponse("slip_instance_template.html", context)

@app.get("/reports/exited-projects/excel", response_class=StreamingResponse)
def get_exited_projects_report(report_date: date, db: Session = Depends(get_db)):
    """گزارش تابلوهای خروجی"""
    start_of_day, end_of_day = datetime.combine(report_date, time.min), datetime.combine(report_date, time.max)
    exited_projects = db.query(Project).join(Project.steps).filter(
        Step.name_key == StepNameKey.EXIT_PANEL, 
        Step.timestamp >= start_of_day, 
        Step.timestamp <= end_of_day
    ).options(joinedload(Project.equipment), joinedload(Project.steps)).order_by(Step.timestamp).all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True
    
    header_font = Font(bold=True, name='B Nazanin', size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    sheet.merge_cells('A1:V1')
    cell_a1 = sheet['A1']
    cell_a1.value = "صورتجلسه تحویل تابلو"
    cell_a1.font = Font(bold=True, name='B Nazanin', size=16)
    cell_a1.alignment = center_align
    
    sheet['B2'] = "شرکت :"
    sheet['B2'].font = header_font
    sheet['B2'].alignment = right_align
    
    sheet['O2'] = f"تاریخ تحویل : {jdatetime.date.fromgregorian(date=report_date).strftime('%Y/%m/%d')}"
    sheet['O2'].font = header_font
    sheet['O2'].alignment = right_align
    
    headers_l1 = [
        "ردیف", "نام و نام خانوادگی متقاضی", "شماره تقاضا", "نوع تابلو", "تعداد کنتور", None, 
        "تعداد فیوز", None, "تعداد مودم", "کنتور پیش بینی", None, "فیوز پیش بینی", None, 
        "سکو", "تسمه استیل", "بست تسمه", "پیچ یکسررزوه", "لوله نیم گرد", "لوله خرطومی", "پیچ ۱۶*۳۰۰"
    ]
    headers_l2 = [
        None, None, None, None, "تکفاز", "سه فاز", "تکفاز", "سه فاز", None, 
        "تکفاز", "سه فاز", "تکفاز", "سه فاز", None, None, None, None, None, None, None
    ]
    
    sheet.append(headers_l1)
    sheet.append(headers_l2)
    
    sheet.merge_cells('A3:A4')
    sheet.merge_cells('B3:B4')
    sheet.merge_cells('C3:C4')
    sheet.merge_cells('D3:D4')
    sheet.merge_cells('E3:F3')
    sheet.merge_cells('G3:H3')
    sheet.merge_cells('I3:I4')
    sheet.merge_cells('J3:K3')
    sheet.merge_cells('L3:M3')
    sheet.merge_cells('N3:N4')
    sheet.merge_cells('O3:O4')
    sheet.merge_cells('P3:P4')
    sheet.merge_cells('Q3:Q4')
    sheet.merge_cells('R3:R4')
    sheet.merge_cells('S3:S4')
    sheet.merge_cells('T3:T4')
    
    for row in sheet.iter_rows(min_row=3, max_row=4):
        for cell in row: 
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
    
    start_row = 5
    for idx, project in enumerate(exited_projects, 1):
        summary = generate_project_summary_data(project)
        reservation = calculate_reservation_details(project)
        modem_count = sum(
            item.quantity for item in project.equipment 
            if "مودم" in normalize_text(item.item_name)
        )
        
        total_single_phase_fuses = sum([
            summary.get('فیوز تک فاز 16', 0), 
            summary.get('فیوز تک فاز 25', 0), 
            summary.get('فیوز تک فاز 32', 0)
        ])
        total_three_phase_fuses = sum([
            summary.get('فیوز سه فاز 25', 0), 
            summary.get('فیوز سه فاز 32', 0), 
            summary.get('فیوز سه فاز 63', 0)
        ])
        
        row_data = [
            idx, project.customer_name, project.request_id, project.panel_code or '-',
            summary.get('کنتور تک فاز', 0), summary.get('کنتور سه فاز', 0),
            total_single_phase_fuses, total_three_phase_fuses, modem_count,
            reservation.get('purchased_single_phase', 0), 0, 0, 0,
            summary.get('سکو', 0), summary.get('تسمه استیل', 0), summary.get('بست تسمه استیل', 0),
            0, summary.get('لوله نیم گرد', 0), summary.get('لوله خرطومی', 0), summary.get('پیچ ۱۶*۳۰۰', 0)
        ]
        sheet.append(row_data)
    
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
        for cell in row: 
            cell.alignment = center_align
            cell.border = thin_border
    
    footer_row = sheet.max_row + 2
    sheet.cell(row=footer_row, column=2, value="تحویل دهنده :").font = header_font
    sheet.cell(row=footer_row, column=14, value="تحویل گیرنده :").font = header_font
    
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 15
    
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    
    filename = f"Exited_Projects_Summary_{report_date.strftime('%Y-%m-%d')}.xlsx"
    logger.info(f"Exited projects report generated for date: {report_date}")
    return StreamingResponse(
        excel_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/reports/ready-for-delivery/detailed-html", response_class=HTMLResponse)
def get_detailed_delivery_report_html(direction: str, db: Session = Depends(get_db)):
    """گزارش تفصیلی آماده تحویل به صورت HTML"""
    if direction not in ["west", "east"]: 
        raise HTTPException(status_code=400, detail="جهت باید 'west' یا 'east' باشد.")
    
    company_name = WEST_COMPANY if direction == "west" else EAST_COMPANY
    jalali_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    ready_projects = get_projects_by_status(
        db, 
        required_steps=set(ORDERED_MANUAL_STEP_KEYS), 
        forbidden_steps={StepNameKey.EXIT_PANEL}
    )
    report_data = generate_detailed_report_data(ready_projects, direction)
    
    try:
        html_template = jinja_env.get_template("detailed_report_template.html")
        return HTMLResponse(content=html_template.render(
            company_name=company_name, 
            jalali_date=jalali_date, 
            report_data=report_data
        ))
    except Exception as e:
        logger.error(f"Error loading detailed report template: {e}")
        return HTMLResponse(
            content=f"<h1>خطا در بارگذاری قالب گزارش</h1><p>فایل 'detailed_report_template.html' یافت نشد یا خطایی در آن وجود دارد.</p><p>جزئیات خطا: {e}</p>", 
            status_code=500
        )

# دیباگ
@app.get("/debug-assembler-query", response_model=List[Dict[str, Any]])
def debug_assembler_query(
    db: Session = Depends(get_db), 
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None
):
    """اندپوینت دیباگ برای کوئری مونتاژکاران"""
    query = db.query(Project).filter(
        Project.assembler_1.isnot(None), 
        Project.panel_type_key.isnot(None)
    )
    
    if start_date: 
        query = query.filter(Project.created_at >= datetime.combine(start_date, time.min))
    if end_date: 
        query = query.filter(Project.created_at <= datetime.combine(end_date, time.max))
    
    projects = query.order_by(Project.created_at.desc()).all()
    result = []
    
    for p in projects:
        result.append({
            "project_id": p.id, 
            "request_id": p.request_id, 
            "panel_type_key": p.panel_type_key,
            "assembler_1": p.assembler_1, 
            "assembler_2": p.assembler_2, 
            "created_at": p.created_at.isoformat()
        })
    
    return result



# ==============================================================================
# بخش ۲۰: WebSocket Endpoint
# ==============================================================================

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, token: Optional[str] = None):
    """اتصال WebSocket برای ارتباط بلادرنگ با سرور."""
    user_id = 0
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            username: str = payload.get("sub")
            if username:
                with get_db_session() as db:
                    user = await get_user_by_username(db, username)
                    if user and user.is_active:
                        user_id = user.id
        except JWTError as e:
            logger.warning(f"WebSocket connection attempted with invalid token: {e}")
    
    await manager.connect(websocket, user_id)
    
    try:
        while True: 
            # نگه داشتن ارتباط باز. می‌توان منطق دریافت پیام را اینجا اضافه کرد.
            await websocket.receive_text()
    except WebSocketDisconnect: 
        manager.disconnect(websocket)
        logger.info(f"WebSocket with user_id {user_id} disconnected.")
    except Exception as e: 
        logger.error(f"WebSocket error for user_id {user_id}: {e}", exc_info=True)
        manager.disconnect(websocket)

# ==============================================================================
# بخش ۲۱: Static Files و Routeهای اصلی
# ==============================================================================

# ✅ بهبود: افزودن هدرهای امنیتی به فایل‌های استاتیک
class SecureStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        # کش کردن برای یک هفته جهت افزایش سرعت بارگذاری برای کاربران تکراری
        response.headers["Cache-Control"] = "public, max-age=604800, immutable" 
        return response

app.mount("/static", SecureStaticFiles(directory=STATIC_DIR, html=True), name="static")

# اضافه کردن headers امنیتی برای فایل‌های استاتیک
class SecureStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        # اضافه کردن headers امنیتی
        response.headers["Cache-Control"] = "public, max-age=3600"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        return response

app.mount("/static", SecureStaticFiles(directory=STATIC_DIR, html=True), name="static")

# --- صفحه لاگین (دسترسی عمومی) ---
@app.get("/login", response_class=FileResponse, include_in_schema=False)
async def read_login():
    """صفحه ورود به سیستم"""
    return FileResponse(os.path.join(STATIC_DIR, "login.html"))

# --- صفحات داخلی برنامه ---
@app.get("/dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_dashboard_page(): 
    """داشبورد اصلی"""
    return FileResponse(os.path.join(STATIC_DIR, "dashboard.html"))

@app.get("/manager", response_class=HTMLResponse, include_in_schema=False)
async def read_manager_ui_page(): 
    """صفحه مدیریت"""
    return FileResponse(os.path.join(STATIC_DIR, "manager.html"))

@app.get("/workshop", response_class=HTMLResponse, include_in_schema=False)
async def read_workshop_ui_page(): 
    """صفحه کارگاه"""
    return FileResponse(os.path.join(STATIC_DIR, "workshop.html"))

@app.get("/personnel-management", response_class=HTMLResponse, include_in_schema=False)
async def read_personnel_management_page():
    """صفحه مدیریت پرسنل"""
    return FileResponse(os.path.join(STATIC_DIR, "personnel_management.html"))

@app.get("/supervisor-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_supervisor_dashboard_page():
    """داشبورد ناظر"""
    return FileResponse(os.path.join(STATIC_DIR, "supervisor_dashboard.html"))

# ==========================================================
# ✅✅✅ اندپوینت جدید برای داشبورد انبار ✅✅✅
# ==========================================================
@app.get("/warehouse-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_warehouse_dashboard_page():
    """صفحه داشبورد انبار"""
    return FileResponse(os.path.join(STATIC_DIR, "warehouse_dashboard.html"))
# ==========================================================

@app.get("/reports", response_class=HTMLResponse, include_in_schema=False)
async def read_reports_page_html():
    """صفحه گزارش‌ها"""
    fp = os.path.join(STATIC_DIR, "reports.html")
    if os.path.exists(fp):
        return FileResponse(fp)
    raise HTTPException(status_code=404, detail="reports.html not found")

@app.get("/assembler-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_assembler_dashboard_page():
    """داشبورد مونتاژکار"""
    fp = os.path.join(STATIC_DIR, "assembler_dashboard.html")
    if os.path.exists(fp): 
        return FileResponse(fp)
    raise HTTPException(status_code=404, detail="assembler_dashboard.html not found")

@app.get("/daily-work-report", response_class=HTMLResponse, include_in_schema=False)
async def read_daily_work_report_page():
    """صفحه گزارش کار روزانه"""
    return FileResponse(os.path.join(STATIC_DIR, "daily_work_report.html"))

@app.get("/employee-dashboard", response_class=HTMLResponse, include_in_schema=False)
async def read_employee_dashboard_page():
    """داشبورد کارمند"""
    return FileResponse(os.path.join(STATIC_DIR, "employee_dashboard.html"))

# --- روت اصلی و catch-all برای اپلیکیشن‌های تک‌صفحه‌ای (SPA) ---
@app.get("/", response_class=FileResponse, include_in_schema=False)
async def read_index():
    """صفحه اصلی"""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="index.html not found.")
    return FileResponse(index_path)

@app.get("/{full_path:path}", response_class=FileResponse, include_in_schema=False)
async def serve_frontend_catch_all(request: Request, full_path: str):
    """مسیرهای catch-all برای SPA"""
    # این روت فایل index.html را برای مسیرهای ناشناخته برمی‌گرداند
    # تا منطق روتینگ در فرانت‌اند انجام شود.
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="index.html not found.")
    return FileResponse(index_path)

# ==============================================================================
# بخش ۲۲: راه‌اندازی سرور و هندلرهای خطا
# ==============================================================================

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """مدیریت یکپارچه خطاهای HTTP با لاگ‌گیری دقیق."""
    logger.warning(f"HTTP Exception: {exc.status_code} {exc.detail} for URL: {request.url}")
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    """مدیریت خطاهای پیش‌بینی نشده سرور (500)."""
    logger.error(f"Unhandled Exception for URL: {request.url}", exc_info=True)
    return JSONResponse(status_code=500, content={"detail": "An internal server error occurred."})

@app.on_event("startup")
async def startup_event():
    """رویدادهای زمان راه‌اندازی برنامه."""
    logger.info("Application starting up...")
    try:
        with SessionLocal() as db:
            # ✅ اصلاح: استفاده از text() برای اجرای کوئری خام جهت امنیت و استاندارد بودن.
            db.execute(text("SELECT 1"))
        logger.info("Database connection verified successfully.")
    except Exception as e:
        logger.critical(f"FATAL: Database connection failed on startup: {e}")
        # در محیط واقعی، ممکن است بخواهید برنامه در این حالت خارج شود.
        # raise SystemExit("Could not connect to the database.")

@app.on_event("shutdown")
async def shutdown_event():
    """رویدادهای زمان خاموش شدن برنامه."""
    logger.info("Application shutting down...")

if __name__ == "__main__":
    import uvicorn
    import os

    is_production = os.getenv("ENVIRONMENT") == "production"

    uvicorn.run(
        "project_monitoring_system:app",
        host="0.0.0.0",
        port=8000,
        reload=not is_production,
        log_level="info",
        workers=os.cpu_count() if is_production else 1
    )
